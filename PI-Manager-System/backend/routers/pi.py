from fastapi import APIRouter, Depends, HTTPException
from typing import List
from sqlalchemy.orm import Session
from app.database import get_db
from crud.pi import (
    create_pi_invoice, get_pi_invoice, get_pi_invoices, update_pi_status, get_price_history,
    update_pi_invoice, get_pi_invoice_detail, get_pi_invoices_with_customer
)
from schemas.pi import PIInvoiceCreate, PIInvoiceResponse, PIInvoiceUpdate

router = APIRouter(prefix="/api/pi", tags=["PI管理"])

@router.post("/")
def create_pi_api(pi: PIInvoiceCreate, db: Session = Depends(get_db)):
    try:
        result = create_pi_invoice(db, pi)
        return {
            "id": result.id,
            "dept_id": result.dept_id,
            "pi_no": result.pi_no,
            "customer_id": result.customer_id,
            "total_amount": float(result.total_amount) if result.total_amount else 0,
            "currency": result.currency or "USD",
            "status": result.status or 1,
            "created_at": result.created_at.isoformat() if result.created_at else None
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/", response_model=list[PIInvoiceResponse])
def read_pi_list(skip: int = 0, limit: int = 100, status: int = None, db: Session = Depends(get_db)):
    results = get_pi_invoices_with_customer(db, skip=skip, limit=limit, status=status)
    print(f"DEBUG - read_pi_list returned {len(results)} items")
    return results

@router.get("/detail/{pi_id}")
def read_pi_detail(pi_id: int, db: Session = Depends(get_db)):
    detail = get_pi_invoice_detail(db, pi_id)
    if detail is None:
        raise HTTPException(status_code=404, detail="PI单不存在")
    return detail

@router.get("/{pi_id}")
def read_pi(pi_id: int, db: Session = Depends(get_db)):
    db_pi = get_pi_invoice(db, pi_id)
    if not db_pi:
        raise HTTPException(status_code=404, detail="PI单不存在")
    return {
        "id": db_pi.id,
        "dept_id": db_pi.dept_id,
        "pi_no": db_pi.pi_no,
        "customer_id": db_pi.customer_id,
        "total_amount": float(db_pi.total_amount) if db_pi.total_amount else 0,
        "currency": db_pi.currency or "USD",
        "status": db_pi.status or 1,
        "created_at": db_pi.created_at.isoformat() if db_pi.created_at else None
    }

@router.delete("/{pi_id}")
def delete_pi_api(pi_id: int, db: Session = Depends(get_db)):
    try:
        delete_pi_invoice(db, pi_id)
        return {"success": True, "message": "PI已删除"}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.post("/batch-delete")
def batch_delete_pi_api(pi_ids: List[int], db: Session = Depends(get_db)):
    """批量删除PI订单"""
    deleted = 0
    errors = []
    for pi_id in pi_ids:
        try:
            delete_pi_invoice(db, pi_id)
            deleted += 1
        except ValueError as e:
            errors.append(f"ID {pi_id}: {str(e)}")
    return {"deleted": deleted, "total": len(pi_ids), "errors": errors}

@router.get("/export/{pi_id}")
def export_pi_excel(pi_id: int, db: Session = Depends(get_db)):
    from crud.pi import get_pi_invoice_detail
    import io
    from fastapi.responses import StreamingResponse
    
    detail = get_pi_invoice_detail(db, pi_id)
    if detail is None:
        raise HTTPException(status_code=404, detail="PI单不存在")
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"PI-{detail['pi_no']}"
        
        # 样式定义
        title_font = Font(name='Arial', size=16, bold=True)
        header_font = Font(name='Arial', size=11, bold=True)
        normal_font = Font(name='Arial', size=10)
        header_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
        header_font_white = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # 标题
        ws.merge_cells('A1:H1')
        ws['A1'] = f"PROFORMA INVOICE - {detail['pi_no']}"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # 基本信息
        status_map = {1: '草稿', 2: '已确认', 3: '已发货', 4: '已完成'}
        info_data = [
            ('客户编号', detail.get('customer_code', '')),
            ('客户名称', detail.get('customer_name', '')),
            ('币种', detail.get('currency', 'USD')),
            ('总金额', detail.get('total_amount', 0)),
            ('状态', status_map.get(detail.get('status', 1), '草稿')),
            ('创建日期', str(detail.get('created_at', ''))[:10] if detail.get('created_at') else ''),
        ]
        for i, (label, value) in enumerate(info_data, start=3):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = header_font
            ws[f'B{i}'] = value
            ws[f'B{i}'].font = normal_font
        
        # 明细表头
        row = 10
        headers = ['序号', '产品编号', 'OE号', '客户编码', '描述', '数量', '单价', '总价', '备注']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # 明细数据
        for idx, item in enumerate(detail.get('items', []), 1):
            r = row + idx
            values = [
                idx,
                item.get('product_code', ''),
                item.get('oe_number', ''),
                item.get('customer_code', ''),
                item.get('detail_desc', ''),
                item.get('quantity', 0),
                item.get('unit_price', 0),
                item.get('total_price', 0),
                item.get('remark', '')
            ]
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=r, column=col, value=v)
                cell.font = normal_font
                cell.border = thin_border
        
        # 合计行
        total_row = row + len(detail.get('items', [])) + 1
        ws.cell(row=total_row, column=6, value='合计').font = header_font
        ws.cell(row=total_row, column=8, value=detail.get('total_amount', 0)).font = header_font
        
        # 付款阶段
        if detail.get('payment_stages'):
            pay_row = total_row + 2
            ws.cell(row=pay_row, column=1, value='付款阶段').font = header_font
            pay_row += 1
            pay_headers = ['阶段类型', '序号', '金额', '应付日期', '状态']
            for col, h in enumerate(pay_headers, 1):
                cell = ws.cell(row=pay_row, column=col, value=h)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.border = thin_border
            
            for stage in detail['payment_stages']:
                pay_row += 1
                stage_type_map = {'deposit': '定金', 'balance': '尾款'}
                stage_status_map = {1: '待付', 2: '已付'}
                values = [
                    stage_type_map.get(stage['stage_type'], stage['stage_type']),
                    stage.get('stage_no', ''),
                    stage.get('amount', 0),
                    str(stage['due_date'])[:10] if stage.get('due_date') else '',
                    stage_status_map.get(stage['status'], '待付')
                ]
                for col, v in enumerate(values, 1):
                    cell = ws.cell(row=pay_row, column=col, value=v)
                    cell.font = normal_font
                    cell.border = thin_border
        
        # 调整列宽
        col_widths = [8, 15, 15, 15, 30, 10, 12, 12, 20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"PI_{detail['pi_no']}.xlsx"
        headers = {
            'Content-Disposition': f'attachment; filename="{filename}"',
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
        return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except ImportError:
        raise HTTPException(status_code=500, detail="服务器缺少openpyxl库，无法导出Excel")

@router.get("/{pi_id}", response_model=PIInvoiceResponse)
def read_pi(pi_id: int, db: Session = Depends(get_db)):
    db_pi = get_pi_invoice(db, pi_id)
    if db_pi is None:
        raise HTTPException(status_code=404, detail="PI单不存在")
    return db_pi

@router.put("/{pi_id}")
def update_pi_api(pi_id: int, pi_update: PIInvoiceUpdate, db: Session = Depends(get_db)):
    try:
        db_pi = update_pi_invoice(db, pi_id, pi_update)
        if db_pi is None:
            raise HTTPException(status_code=404, detail="PI单不存在")
        return {
            "id": db_pi.id,
            "dept_id": db_pi.dept_id,
            "pi_no": db_pi.pi_no,
            "customer_id": db_pi.customer_id,
            "total_amount": float(db_pi.total_amount) if db_pi.total_amount else 0,
            "currency": db_pi.currency or "USD",
            "status": db_pi.status or 1,
            "created_at": db_pi.created_at.isoformat() if db_pi.created_at else None
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.put("/{pi_id}/status")
def update_pi_status_api(pi_id: int, status_data: dict, db: Session = Depends(get_db)):
    """更新PI单状态（草稿1→已确认2→已发货3→已完成4）"""
    try:
        status = status_data.get('status')
        if status not in [1, 2, 3, 4]:
            raise HTTPException(status_code=400, detail="无效的状态值")
        db_pi = update_pi_status(db, pi_id, status)
        return {
            "id": db_pi.id,
            "pi_no": db_pi.pi_no,
            "status": db_pi.status,
            "success": True
        }
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

@router.get("/price-history/{customer_id}/{product_id}")
def read_price_history(customer_id: int, product_id: int, db: Session = Depends(get_db)):
    history = get_price_history(db, customer_id, product_id)
    if history is None:
        return {"message": "暂无历史价格记录"}
    return {
        "unit_price": history.unit_price,
        "remark": history.remark,
        "created_at": history.created_at
    }
