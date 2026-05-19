# -*- coding: utf-8 -*-
"""
PI Manager - 客户端主窗口
"""
import sys
import os
import threading
import time
import traceback
import concurrent.futures
import urllib.request
import ctypes
from datetime import datetime, timedelta
from functools import lru_cache

# 数据处理
import pandas as pd

# PySide6 导入
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
    QLabel, QStackedWidget, QMessageBox, QTableWidget, QTableWidgetItem, QDialog,
    QFormLayout, QLineEdit, QTextEdit, QComboBox, QHeaderView, QAbstractItemView,
    QGridLayout, QCheckBox, QGroupBox, QFileDialog, QProgressDialog, QTabWidget,
    QScrollArea, QDateEdit, QMenu, QFrame, QStatusBar, QSpinBox, QDoubleSpinBox,
    QSizePolicy, QProgressBar
)
from PySide6.QtCore import Qt, QTimer, QDate, QEvent, QThread, Signal, QObject
from PySide6.QtGui import (
    QIcon, QPalette, QColor, QFont, QFontDatabase, QBrush, QPixmap, QImage,
    QAction, QPainter
)

# 本地模块（添加到path）
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from cache_manager import cache_manager, CACHE_KEYS, set_user, invalidate_cache
from api.client import ApiClient
from api.cached_client import CachedApiClient
from config import Config
from product_categories import get_category_options, get_category_code, get_category_name
from widgets.action_bar import ActionBarFactory
from widgets.order_summary_edit_dialog import OrderSummaryEditDialog
from widgets.order_summary_dialogs import (
    CustomerRequirementDialog, 
    CustomerModelDialog, 
    CustomerReplyDialog
)

# 测试模块（可选）
try:
    from test_customer_reply import CustomerReplyTester
    HAS_CUSTOMER_REPLY_TEST = True
except ImportError:
    HAS_CUSTOMER_REPLY_TEST = False
    CustomerReplyTester = None

# ============================================
# 全局常量和配置
# ============================================

# 全局线程池（复用，避免重复创建）
_global_thread_pool = concurrent.futures.ThreadPoolExecutor(max_workers=8, thread_name_prefix="pi_manager")

# 图片内存缓存（LRU缓存，最多100张）
_image_cache = {}
_image_cache_lock = threading.Lock()
_MAX_IMAGE_CACHE_SIZE = 100

# 省份编码映射（静态数据）
PROVINCE_CODE_MAP = {
    "北京": "11", "天津": "12", "河北": "13", "山西": "14", "内蒙古": "15",
    "辽宁": "21", "吉林": "22", "黑龙江": "23",
    "上海": "31", "江苏": "32", "浙江": "33", "安徽": "34", "福建": "35", "江西": "36", "山东": "37",
    "河南": "41", "湖北": "42", "湖南": "43", "广东": "44", "广西": "45", "海南": "46",
    "重庆": "50", "四川": "51", "贵州": "52", "云南": "53", "西藏": "54",
    "陕西": "61", "甘肃": "62", "青海": "63", "宁夏": "64", "新疆": "65",
    "台湾": "71", "香港": "81", "澳门": "82"
}

# 城市编码映射（静态数据，模块级别只创建一次）
CITY_CODE_MAP = {
    "11": {"北京": "1"}, "12": {"天津": "1"}, "31": {"上海": "1"}, "50": {"重庆": "1"},
    "13": {"石家庄": "1", "唐山": "2", "秦皇岛": "3", "邯郸": "4", "邢台": "5", "保定": "6", "张家口": "7", "承德": "8", "沧州": "9", "廊坊": "A", "衡水": "B"},
    "14": {"太原": "1", "大同": "2", "阳泉": "3", "长治": "4", "晋城": "5", "朔州": "6", "晋中": "7", "运城": "8", "忻州": "9", "临汾": "A", "吕梁": "B"},
    "15": {"呼和浩特": "1", "包头": "2", "乌海": "3", "赤峰": "4", "通辽": "5", "鄂尔多斯": "6", "呼伦贝尔": "7", "巴彦淖尔": "8", "乌兰察布": "9", "兴安": "A", "锡林郭勒": "B", "阿拉善": "C"},
    "21": {"沈阳": "1", "大连": "2", "鞍山": "3", "抚顺": "4", "本溪": "5", "丹东": "6", "锦州": "7", "营口": "8", "阜新": "9", "辽阳": "A", "盘锦": "B", "铁岭": "C", "朝阳": "D", "葫芦岛": "E"},
    "22": {"长春": "1", "吉林": "2", "四平": "3", "辽源": "4", "通化": "5", "白山": "6", "松原": "7", "白城": "8", "延边": "9"},
    "23": {"哈尔滨": "1", "齐齐哈尔": "2", "鸡西": "3", "鹤岗": "4", "双鸭山": "5", "大庆": "6", "伊春": "7", "佳木斯": "8", "七台河": "9", "牡丹江": "A", "黑河": "B", "绥化": "C", "大兴安岭": "D"},
    "32": {"南京": "1", "无锡": "2", "徐州": "3", "常州": "4", "苏州": "5", "南通": "6", "连云港": "7", "淮安": "8", "盐城": "9", "扬州": "A", "镇江": "B", "泰州": "C", "宿迁": "D"},
    "33": {"杭州": "1", "宁波": "2", "温州": "3", "嘉兴": "4", "湖州": "5", "绍兴": "6", "金华": "7", "衢州": "8", "舟山": "9", "台州": "A", "丽水": "B"},
    "34": {"合肥": "1", "芜湖": "2", "蚌埠": "3", "淮南": "4", "马鞍山": "5", "淮北": "6", "铜陵": "7", "安庆": "8", "黄山": "9", "阜阳": "A", "宿州": "B", "滁州": "C", "六安": "D", "宣城": "E", "池州": "F", "亳州": "G"},
    "35": {"福州": "1", "厦门": "2", "莆田": "3", "三明": "4", "泉州": "5", "漳州": "6", "南平": "7", "龙岩": "8", "宁德": "9"},
    "36": {"南昌": "1", "景德镇": "2", "萍乡": "3", "九江": "4", "新余": "5", "鹰潭": "6", "赣州": "7", "吉安": "8", "宜春": "9", "抚州": "A", "上饶": "B"},
    "37": {"济南": "1", "青岛": "2", "淄博": "3", "枣庄": "4", "东营": "5", "烟台": "6", "潍坊": "7", "济宁": "8", "泰安": "9", "威海": "A", "日照": "B", "临沂": "C", "德州": "D", "滨州": "E", "菏泽": "F"},
    "41": {"郑州": "1", "开封": "2", "洛阳": "3", "平顶山": "4", "安阳": "5", "鹤壁": "6", "新乡": "7", "焦作": "8", "濮阳": "9", "许昌": "A", "漯河": "B", "三门峡": "C", "南阳": "D", "商丘": "E", "信阳": "F", "周口": "G", "驻马店": "H", "济源": "I"},
    "42": {"武汉": "1", "黄石": "2", "十堰": "3", "宜昌": "4", "襄阳": "5", "鄂州": "6", "荆门": "7", "孝感": "8", "荆州": "9", "黄冈": "A", "咸宁": "B", "随州": "C", "恩施": "D", "仙桃": "E", "潜江": "F", "天门": "G"},
    "43": {"长沙": "1", "株洲": "2", "湘潭": "3", "衡阳": "4", "邵阳": "5", "岳阳": "6", "常德": "7", "张家界": "8", "益阳": "9", "郴州": "A", "永州": "B", "怀化": "C", "娄底": "D", "湘西": "E"},
    "44": {"广州": "1", "韶关": "2", "深圳": "3", "珠海": "4", "汕头": "5", "佛山": "6", "江门": "7", "湛江": "8", "茂名": "9", "肇庆": "A", "惠州": "B", "梅州": "C", "汕尾": "D", "河源": "E", "阳江": "F", "清远": "G", "东莞": "H", "中山": "I", "潮州": "J", "揭阳": "K", "云浮": "L"},
    "45": {"南宁": "1", "柳州": "2", "桂林": "3", "梧州": "4", "北海": "5", "防城港": "6", "钦州": "7", "贵港": "8", "玉林": "9", "百色": "A", "贺州": "B", "河池": "C", "来宾": "D", "崇左": "E"},
    "46": {"海口": "1", "三亚": "2", "三沙": "3", "儋州": "4"},
    "51": {"成都": "1", "自贡": "2", "攀枝花": "3", "泸州": "4", "德阳": "5", "绵阳": "6", "广元": "7", "遂宁": "8", "内江": "9", "乐山": "A", "南充": "B", "眉山": "C", "宜宾": "D", "广安": "E", "达州": "F", "雅安": "G", "巴中": "H", "资阳": "I", "阿坝": "J", "甘孜": "K", "凉山": "L"},
    "52": {"贵阳": "1", "六盘水": "2", "遵义": "3", "安顺": "4", "毕节": "5", "铜仁": "6", "黔西南": "7", "黔东南": "8", "黔南": "9"},
    "53": {"昆明": "1", "曲靖": "2", "玉溪": "3", "保山": "4", "昭通": "5", "丽江": "6", "普洱": "7", "临沧": "8", "楚雄": "9", "红河": "A", "文山": "B", "西双版纳": "C", "大理": "D", "怒江": "E", "迪庆": "F"},
    "54": {"拉萨": "1", "日喀则": "2", "昌都": "3", "林芝": "4", "山南": "5", "那曲": "6", "阿里": "7"},
    "61": {"西安": "1", "铜川": "2", "宝鸡": "3", "咸阳": "4", "渭南": "5", "延安": "6", "汉中": "7", "榆林": "8", "安康": "9", "商洛": "A"},
    "62": {"兰州": "1", "嘉峪关": "2", "金昌": "3", "白银": "4", "天水": "5", "武威": "6", "张掖": "7", "平凉": "8", "酒泉": "9", "庆阳": "A", "定西": "B", "陇南": "C", "临夏": "D", "甘南": "E"},
    "63": {"西宁": "1", "海东": "2", "海北": "3", "黄南": "4", "海南": "5", "果洛": "6", "玉树": "7", "海西": "8"},
    "64": {"银川": "1", "石嘴山": "2", "吴忠": "3", "固原": "4", "中卫": "5"},
    "65": {"乌鲁木齐": "1", "克拉玛依": "2", "吐鲁番": "3", "哈密": "4", "昌吉": "5", "博尔塔拉": "6", "巴音郭楞": "7", "阿克苏": "8", "克孜勒苏": "9", "喀什": "A", "和田": "B", "伊犁": "C", "塔城": "D", "阿勒泰": "E", "石河子": "F", "阿拉尔": "G", "图木舒克": "H", "五家渠": "I", "北屯": "J", "铁门关": "K", "双河": "L", "可克达拉": "M", "昆玉": "N", "胡杨河": "O", "新星": "P"},
    "71": {"台北": "1", "高雄": "2", "台中": "3", "台南": "4", "新北": "5", "桃园": "6"},
    "81": {"香港": "1"}, "82": {"澳门": "1"}
}

def get_font(size=10, weight=QFont.Weight.Normal):
    font = QFont()
    font.setPointSize(size)
    font.setWeight(weight)
    # 尝试多种中文字体
    available = QFontDatabase.families()
    for family in ["Microsoft YaHei", "SimHei", "SimSun", "KaiTi", "WenQuanYi Micro Hei", "Heiti SC"]:
        if family in available:
            font.setFamily(family)
            return font
    # 如果没有中文字体，使用默认字体
    font.setFamily(QFont().defaultFamily())
    return font

def set_global_font(app):
    font = get_font(10)
    app.setFont(font)

# 部门配置映射（从配置文件获取）
DEPARTMENT_CONFIG = Config.DEPARTMENT_DB_CONFIG

class LoginWindow(QDialog):
    def __init__(self, api_client: ApiClient):
        super().__init__()
        self.api_client = api_client
        self.selected_dept = None
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("PI订单管理系统 - 登录")
        self.setFixedSize(700, 650)
        
        # 使用绝对定位
        # 标题
        title = QLabel("PI订单管理系统", self)
        title.setGeometry(50, 80, 600, 60)
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 36px; font-weight: bold;")
        
        # 版本
        version = QLabel("客户端 v1.0", self)
        version.setGeometry(50, 150, 600, 30)
        version.setAlignment(Qt.AlignCenter)
        version.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 16px; color: #666666;")
        
        # 登录模式选择
        mode_label = QLabel("登录模式：", self)
        mode_label.setGeometry(50, 220, 600, 30)
        mode_label.setAlignment(Qt.AlignCenter)
        mode_label.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 18px;")
        
        self.mode_combo = QComboBox(self)
        self.mode_combo.setGeometry(160, 260, 380, 50)
        self.mode_combo.addItem("普通用户模式", False)
        self.mode_combo.addItem("管理员模式", True)
        self.mode_combo.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 16px;")
        
        # 部门标签
        dept_label = QLabel("选择部门：", self)
        dept_label.setGeometry(50, 330, 600, 30)
        dept_label.setAlignment(Qt.AlignCenter)
        dept_label.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 18px;")
        
        # 部门下拉框
        self.dept_combo = QComboBox(self)
        self.dept_combo.setGeometry(160, 370, 380, 50)
        self.dept_combo.addItems([v["name"] for v in DEPARTMENT_CONFIG.values()])
        self.dept_combo.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 16px;")
        
        # API标签
        api_label = QLabel("API服务器地址：", self)
        api_label.setGeometry(50, 440, 600, 30)
        api_label.setAlignment(Qt.AlignCenter)
        api_label.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 18px;")
        
        # API输入框
        self.api_url_input = QLineEdit(self)
        self.api_url_input.setGeometry(150, 480, 400, 50)
        self.api_url_input.setText(Config.API_BASE_URL)
        self.api_url_input.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 16px; padding-left: 10px;")
        
        # 登录按钮
        self.login_btn = QPushButton("登录", self)
        self.login_btn.setGeometry(160, 550, 380, 55)
        self.login_btn.setStyleSheet("""
            QPushButton {
                font-family: 'Microsoft YaHei';
                font-size: 18px;
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 8px;
            }
            QPushButton:hover {
                background-color: #1d4ed8;
            }
        """)
        self.login_btn.clicked.connect(self.connect_to_server)
        
        # 状态标签
        self.status_label = QLabel("", self)
        self.status_label.setGeometry(50, 620, 600, 30)
        self.status_label.setAlignment(Qt.AlignCenter)
        self.status_label.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 16px; color: red;")

    def connect_to_server(self):
        is_admin = self.mode_combo.currentData()
        dept_name = self.dept_combo.currentText()
        api_url = self.api_url_input.text().strip()

        if not api_url:
            self.status_label.setText("请输入API服务器地址")
            return

        self.status_label.setText("正在连接...")
        self.status_label.setStyleSheet("color: #2563eb;")

        try:
            # 获取部门ID和对应的数据库配置（自动获取，不对外显示）
            dept_id = next((k for k, v in DEPARTMENT_CONFIG.items() if v["name"] == dept_name), "S")
            self.selected_dept = dept_id
            
            # 从部门配置中自动获取数据库连接参数
            dept_db_config = DEPARTMENT_CONFIG[dept_id]

            # 设置API客户端
            self.api_client.base_url = api_url
            
            # 设置当前用户信息（模拟登录）
            if hasattr(self.api_client, 'current_user'):
                self.api_client.current_user = {
                    "id": 1,
                    "username": "admin" if is_admin else "user",
                    "real_name": "管理员" if is_admin else "普通用户",
                    "is_admin": is_admin,
                    "dept_id": dept_id
                }
                # 设置缓存用户ID
                set_user(str(self.api_client.current_user["id"]))
            
            # 传递数据库配置到API（从配置文件获取，不对外显示）
            db_config = {
                "db_host": dept_db_config["db_host"],
                "db_port": dept_db_config["db_port"],
                "db_user": dept_db_config["db_user"],
                "db_password": dept_db_config["db_password"],
                "db_name": dept_db_config["db_name"]
            }
            
            # 测试连接
            products = self.api_client.get_products(db_config=db_config)
            
            self.status_label.setText("连接成功！")
            self.status_label.setStyleSheet("color: #16a34a;")
            QTimer.singleShot(500, self.accept)
        except Exception as e:
            self.status_label.setText(f"连接失败: {str(e)}")
            self.status_label.setStyleSheet("color: #dc2626;")

    def get_selected_department(self):
        return self.selected_dept


class FieldEditDialog(QDialog):
    """字段编辑对话框"""
    def __init__(self, field_name, current_value, parent=None):
        super().__init__(parent)
        self.field_name = field_name
        self.new_value = current_value
        self.setWindowTitle(f"编辑: {field_name}")
        self.setMinimumWidth(400)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # 字段名标签
        label = QLabel(f"字段: {self.field_name}")
        label.setFont(QFont("", 10, QFont.Weight.Bold))
        layout.addWidget(label)
        
        # 编辑框
        self.editor = QTextEdit()
        self.editor.setPlainText(str(self.new_value))
        self.editor.setMinimumHeight(100)
        layout.addWidget(self.editor)
        
        # 按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        cancel_btn = QPushButton("取消")
        cancel_btn.setFixedWidth(80)
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)
        
        save_btn = QPushButton("保存")
        save_btn.setFixedWidth(80)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        save_btn.clicked.connect(self._on_save)
        btn_layout.addWidget(save_btn)
        
        layout.addLayout(btn_layout)
    
    def _on_save(self):
        self.new_value = self.editor.toPlainText()
        self.accept()
    
    def get_value(self):
        return self.new_value


class OrderEditDialog(QDialog):
    """订单编辑对话框"""
    def __init__(self, order, parent=None):
        super().__init__(parent)
        self.order = order
        self.updated_order = order.copy()
        self.setWindowTitle(f"编辑订单: {order.get('order_no', '新建订单')}")
        self.setMinimumWidth(600)
        self.setMinimumHeight(500)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # 可编辑字段列表
        editable_fields = [
            ("ORDER NO.", "order_no", "line"),
            ("客户", "customer_name", "line"),
            ("订单日期", "order_date", "line"),
            ("客户产品编号", "customer_product_code", "line"),
            ("OE号", "oe_number", "line"),
            ("产品名称", "product_name", "line"),
            ("客户型号", "customer_model", "line"),
            ("数量", "quantity", "number"),
            ("单价", "unit_price", "number"),
            ("总金额", "total_amount", "number"),
            ("客户预付款", "customer_prepayment", "number"),
            ("待收尾款", "remaining_payment", "number"),
            ("采购价格", "purchase_price", "number"),
            ("运费", "shipping_fee", "number"),
            ("杂费", "misc_fee", "number"),
            ("工厂简称", "supplier_name", "line"),
            ("店铺链接", "supplier_link", "line"),
            ("交货日期", "delivery_date", "line"),
            ("客户需求备注", "customer_requirement", "text"),
            ("客户最新回复", "customer_reply", "text"),
            ("品牌", "brand", "line"),
            ("开票情况", "invoice_status", "line"),
        ]
        
        # 创建滚动区域
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_widget = QWidget()
        scroll_layout = QFormLayout(scroll_widget)
        scroll_layout.setLabelAlignment(Qt.AlignRight)
        scroll_layout.setHorizontalSpacing(20)
        scroll_layout.setVerticalSpacing(10)
        
        self.editors = {}
        for label_text, field_key, field_type in editable_fields:
            value = str(self.order.get(field_key, ''))
            
            if field_type == "line":
                editor = QLineEdit(value)
                editor.setFixedHeight(30)
            elif field_type == "number":
                editor = QDoubleSpinBox()
                editor.setRange(0, 99999999)
                editor.setDecimals(2)
                try:
                    editor.setValue(float(value) if value else 0)
                except:
                    editor.setValue(0)
            elif field_type == "text":
                editor = QTextEdit(value)
                editor.setMinimumHeight(60)
                editor.setMaximumHeight(100)
            
            self.editors[field_key] = editor
            scroll_layout.addRow(f"{label_text}:", editor)
        
        scroll.setWidget(scroll_widget)
        layout.addWidget(scroll)
        
        # 按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        cancel_btn = QPushButton("取消")
        cancel_btn.setFixedWidth(80)
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)
        
        save_btn = QPushButton("保存")
        save_btn.setFixedWidth(80)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #10b981;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
            }
            QPushButton:hover { background-color: #059669; }
        """)
        save_btn.clicked.connect(self._on_save)
        btn_layout.addWidget(save_btn)
        
        layout.addLayout(btn_layout)
    
    def _on_save(self):
        # 更新订单数据
        for field_key, editor in self.editors.items():
            if isinstance(editor, QLineEdit):
                self.updated_order[field_key] = editor.text()
            elif isinstance(editor, QDoubleSpinBox):
                self.updated_order[field_key] = editor.value()
            elif isinstance(editor, QTextEdit):
                self.updated_order[field_key] = editor.toPlainText()
        self.accept()
    
    def get_updated_order(self):
        return self.updated_order


class SettingsDialog(QDialog):
    """系统设置对话框"""
    def __init__(self, api_client, parent=None):
        super().__init__(parent)
        self.api_client = api_client
        self.setWindowTitle("系统设置")
        self.setMinimumWidth(450)
        self.init_ui()
        self.load_settings()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        # 毛利率设置
        profit_group = QGroupBox("毛利率设置")
        profit_layout = QVBoxLayout()
        
        # 说明
        profit_info = QLabel("毛利率用于自动计算产品报价基准价。\n公式: 预估美金报价 = 工厂人民币价格 × (1 + 毛利率) / 汇率")
        profit_info.setStyleSheet("color: #64748b; font-size: 12px;")
        profit_layout.addWidget(profit_info)
        
        # 输入框
        profit_input_layout = QHBoxLayout()
        profit_input_layout.addWidget(QLabel("基础毛利率:"))
        
        self.profit_margin_spin = QDoubleSpinBox()
        self.profit_margin_spin.setRange(0, 100)
        self.profit_margin_spin.setDecimals(2)
        self.profit_margin_spin.setSuffix(" %")
        self.profit_margin_spin.setFixedWidth(120)
        profit_input_layout.addWidget(self.profit_margin_spin)
        
        profit_input_layout.addStretch()
        profit_layout.addLayout(profit_input_layout)
        
        profit_group.setLayout(profit_layout)
        layout.addWidget(profit_group)
        
        # 汇率设置
        rate_group = QGroupBox("汇率设置")
        rate_layout = QVBoxLayout()
        
        # 说明
        rate_info = QLabel("人民币兑美元汇率，用于计算预估美金报价。")
        rate_info.setStyleSheet("color: #64748b; font-size: 12px;")
        rate_layout.addWidget(rate_info)
        
        # 输入框
        rate_input_layout = QHBoxLayout()
        rate_input_layout.addWidget(QLabel("USD/RMB 汇率:"))
        
        self.exchange_rate_spin = QDoubleSpinBox()
        self.exchange_rate_spin.setRange(0.01, 100)
        self.exchange_rate_spin.setDecimals(4)
        self.exchange_rate_spin.setFixedWidth(120)
        rate_input_layout.addWidget(self.exchange_rate_spin)
        
        rate_input_layout.addStretch()
        rate_layout.addLayout(rate_input_layout)
        
        rate_group.setLayout(rate_layout)
        layout.addWidget(rate_group)
        
        # 按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        cancel_btn = QPushButton("取消")
        cancel_btn.setFixedWidth(80)
        cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(cancel_btn)
        
        save_btn = QPushButton("保存")
        save_btn.setFixedWidth(80)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                padding: 8px 16px;
                border-radius: 4px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        save_btn.clicked.connect(self.save_settings)
        btn_layout.addWidget(save_btn)
        
        layout.addLayout(btn_layout)
    
    def load_settings(self):
        """加载设置（使用本地配置，无网络延迟）"""
        try:
            from config.local_settings_manager import load_local_settings
            settings = load_local_settings()
            self.profit_margin_spin.setValue(settings.get('default_profit_margin', 25.0))
            self.exchange_rate_spin.setValue(settings.get('exchange_rate', 7.24))
        except Exception as e:
            print(f"加载设置失败: {e}")
            self.profit_margin_spin.setValue(25.0)
            self.exchange_rate_spin.setValue(7.24)
    
    def save_settings(self):
        """保存设置到本地配置文件"""
        try:
            margin = self.profit_margin_spin.value()
            rate = self.exchange_rate_spin.value()
            
            # 保存到本地配置（不使用数据库）
            from config.local_settings_manager import save_local_settings
            settings = {
                'default_profit_margin': margin,
                'exchange_rate': rate
            }
            save_local_settings(settings)
            
            QMessageBox.information(self, "成功", f"设置已保存\n毛利率: {margin}%\n汇率: {rate}")
            self.accept()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"保存失败: {e}")


class ProductDialog(QDialog):
    def __init__(self, api_client: ApiClient, dept_id: str, product=None):
        super().__init__()
        self.api_client = api_client
        self.dept_id = dept_id
        self.product = product
        self.is_edit = product is not None
        self.suppliers = []
        self.customers = []
        self.supplier_schemes = []
        self.init_ui()
        QTimer.singleShot(0, self.load_data)

    def load_data(self):
        """加载供应商和客户列表"""
        try:
            self.suppliers = self.api_client.get_suppliers()
        except Exception as e:
            print(f"加载供应商失败: {e}")

        try:
            self.customers = self.api_client.get_customers()
        except Exception as e:
            print(f"加载客户失败: {e}")

        # 加载已有方案
        if self.is_edit and self.product:
            self.load_supplier_schemes()
            self.load_product_oes()
            self.load_product_customers()

    def init_ui(self):
        self.setWindowTitle("编辑产品" if self.is_edit else "新增产品")
        self.setMinimumSize(1000, 800)
        self.resize(1000, 800)

        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)
        scroll_layout.setSpacing(15)

        # ===== 顶部区域：图片 + 基本信息 =====
        top_widget = QWidget()
        top_layout = QHBoxLayout(top_widget)
        top_layout.setSpacing(20)

        # 左侧：图片区域（主图 + 副图）
        left_layout = QVBoxLayout()

        # 主图
        main_image_group = QGroupBox("主图")
        main_image_layout = QVBoxLayout()
        self.image_preview = QLabel()
        self.image_preview.setFixedSize(200, 200)
        self.image_preview.setStyleSheet("border: 2px dashed #d1d5db;")
        self.image_preview.setAlignment(Qt.AlignCenter)
        self.image_preview.setText("点击选择图片")
        self.image_preview.setCursor(Qt.PointingHandCursor)
        self.image_preview.installEventFilter(self)
        self.selected_image_path = None
        if self.product and self.product.get('default_image_url'):
            try:
                image_data = urllib.request.urlopen(self.product.get('default_image_url')).read()
                image = QImage.fromData(image_data)
                pixmap = QPixmap.fromImage(image).scaled(196, 196, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                self.image_preview.setPixmap(pixmap)
                self.image_preview.setText("")
            except Exception:
                pass
        main_image_layout.addWidget(self.image_preview)

        main_image_btns = QHBoxLayout()
        self.upload_image_btn = QPushButton("上传主图")
        self.upload_image_btn.setFixedWidth(90)
        self.upload_image_btn.clicked.connect(self.select_image)
        main_image_btns.addWidget(self.upload_image_btn)
        self.clear_image_btn = QPushButton("清除")
        self.clear_image_btn.setFixedWidth(60)
        self.clear_image_btn.clicked.connect(self.clear_image)
        main_image_btns.addWidget(self.clear_image_btn)
        main_image_layout.addLayout(main_image_btns)
        main_image_group.setLayout(main_image_layout)
        left_layout.addWidget(main_image_group)

        # 副图区域
        sub_image_group = QGroupBox("副图")
        sub_image_layout = QVBoxLayout()
        self.sub_images_container = QWidget()
        self.sub_images_layout = QHBoxLayout(self.sub_images_container)
        self.sub_images_layout.setSpacing(5)
        self.sub_image_paths = []
        if self.product:
            self.load_sub_images()
        self.sub_images_layout.addStretch()
        sub_image_layout.addWidget(self.sub_images_container)

        add_sub_img_btn = QPushButton("+ 添加副图")
        add_sub_img_btn.setFixedWidth(90)
        add_sub_img_btn.clicked.connect(self.add_sub_image)
        sub_image_layout.addWidget(add_sub_img_btn)
        sub_image_group.setLayout(sub_image_layout)
        left_layout.addWidget(sub_image_group)
        left_layout.addStretch()

        top_layout.addLayout(left_layout, 1)

        # 右侧：基本信息表单
        right_form = QFormLayout()
        right_form.setSpacing(12)

        if self.is_edit:
            self.code_label = QLabel(self.product.get('product_code', ''))
            self.code_label.setStyleSheet("font-weight: bold; color: #2563eb;")
            right_form.addRow("产品编号:", self.code_label)

        self.oe_input = QLineEdit()
        self.oe_input.setPlaceholderText("请输入OE号")
        self.oe_input.setMinimumHeight(32)
        if self.product:
            self.oe_input.setText(self.product.get('oe_number', ''))
        right_form.addRow("OE号:", self.oe_input)

        self.detail_input = QTextEdit()
        self.detail_input.setPlaceholderText("请输入产品细节描述")
        if self.product:
            self.detail_input.setText(self.product.get('detail_desc', ''))
        self.detail_input.setMaximumHeight(80)
        right_form.addRow("细节描述:", self.detail_input)

        self.category_combo = QComboBox()
        self.category_combo.setFixedHeight(35)
        for code, name in get_category_options():
            self.category_combo.addItem(name, code)
        if self.product:
            category_code = str(self.product.get('category_id', '01'))
            for i in range(self.category_combo.count()):
                if self.category_combo.itemData(i) == category_code:
                    self.category_combo.setCurrentIndex(i)
                    break
        right_form.addRow("产品类别:", self.category_combo)

        self.channel_input = QLineEdit()
        self.channel_input.setPlaceholderText("采购渠道（如1688、对私付款等）")
        self.channel_input.setMinimumHeight(32)
        if self.product:
            self.channel_input.setText(self.product.get('purchase_channel', ''))
        right_form.addRow("采购渠道:", self.channel_input)

        top_layout.addLayout(right_form, 2)
        scroll_layout.addWidget(top_widget)

        # ===== 供应商方案设计（弹窗模式） =====
        supplier_scheme_group = QGroupBox("供应商方案设计（价格、包装、重量随方案配置）")
        supplier_scheme_layout = QVBoxLayout()

        scheme_toolbar = QHBoxLayout()
        add_scheme_btn = QPushButton("+ 添加供应商方案")
        add_scheme_btn.clicked.connect(self.add_supplier_scheme)
        add_scheme_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 16px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        scheme_toolbar.addWidget(add_scheme_btn)
        scheme_toolbar.addStretch()
        supplier_scheme_layout.addLayout(scheme_toolbar)

        # 摘要表格（只读）
        self.supplier_scheme_table = QTableWidget()
        self.supplier_scheme_table.setColumnCount(6)
        self.supplier_scheme_table.setHorizontalHeaderLabels([
            "供应商", "客户", "客户产品编号", "主要价格", "备注", "操作"
        ])
        self.supplier_scheme_table.setColumnWidth(0, 180)
        self.supplier_scheme_table.setColumnWidth(1, 180)
        self.supplier_scheme_table.setColumnWidth(2, 120)
        self.supplier_scheme_table.setColumnWidth(3, 100)
        self.supplier_scheme_table.setColumnWidth(5, 120)
        self.supplier_scheme_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)
        self.supplier_scheme_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.supplier_scheme_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.supplier_scheme_table.setFixedHeight(180)
        supplier_scheme_layout.addWidget(self.supplier_scheme_table)

        supplier_scheme_group.setLayout(supplier_scheme_layout)
        scroll_layout.addWidget(supplier_scheme_group)
        
        # ===== OE号管理区域 =====
        oe_group = QGroupBox("OE号管理")
        oe_layout = QVBoxLayout()
        
        # OE列表工具栏
        oe_toolbar = QHBoxLayout()
        self.oe_table = QTableWidget()
        self.oe_table.setColumnCount(4)
        self.oe_table.setHorizontalHeaderLabels(["OE号", "是否主OE", "创建时间", "操作"])
        self.oe_table.setColumnWidth(0, 200)
        self.oe_table.setColumnWidth(1, 80)
        self.oe_table.setColumnWidth(2, 150)
        self.oe_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.oe_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.oe_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.oe_table.setFixedHeight(150)
        oe_layout.addWidget(self.oe_table)
        
        # 添加OE按钮
        add_oe_btn = QPushButton("+ 添加OE号")
        add_oe_btn.setStyleSheet("""
            QPushButton {
                background-color: #3b82f6;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 12px;
            }
            QPushButton:hover { background-color: #2563eb; }
        """)
        add_oe_btn.clicked.connect(self.add_product_oe)
        oe_toolbar.addWidget(add_oe_btn)
        oe_toolbar.addStretch()
        oe_layout.addLayout(oe_toolbar)
        
        oe_group.setLayout(oe_layout)
        scroll_layout.addWidget(oe_group)
        
        # ===== 客户产品编号管理区域 =====
        pc_group = QGroupBox("客户产品编号管理")
        pc_layout = QVBoxLayout()
        
        # 客户产品列表
        self.pc_table = QTableWidget()
        self.pc_table.setColumnCount(6)
        self.pc_table.setHorizontalHeaderLabels(["客户", "客户产品编号", "客户OE号", "USD价格", "RMB价格", "操作"])
        self.pc_table.setColumnWidth(0, 150)
        self.pc_table.setColumnWidth(1, 150)
        self.pc_table.setColumnWidth(2, 150)
        self.pc_table.setColumnWidth(3, 100)
        self.pc_table.setColumnWidth(4, 100)
        self.pc_table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.Stretch)
        self.pc_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.pc_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.pc_table.setFixedHeight(180)
        pc_layout.addWidget(self.pc_table)
        
        # 添加客户产品按钮
        add_pc_btn = QPushButton("+ 添加客户产品")
        add_pc_btn.setStyleSheet("""
            QPushButton {
                background-color: #10b981;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 12px;
            }
            QPushButton:hover { background-color: #059669; }
        """)
        add_pc_btn.clicked.connect(self.add_product_customer)
        pc_toolbar = QHBoxLayout()
        pc_toolbar.addWidget(add_pc_btn)
        pc_toolbar.addStretch()
        pc_layout.addLayout(pc_toolbar)
        
        pc_group.setLayout(pc_layout)
        scroll_layout.addWidget(pc_group)
        
        scroll_area.setWidget(scroll_widget)
        layout.addWidget(scroll_area)

        # 底部按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        save_btn = QPushButton("保存")
        save_btn.setFixedWidth(100)
        save_btn.clicked.connect(self.save_product)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        btn_layout.addWidget(save_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.setFixedWidth(100)
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        btn_layout.addWidget(cancel_btn)

        layout.addLayout(btn_layout)
        self.setLayout(layout)

    def eventFilter(self, obj, event):
        from PySide6.QtCore import QEvent
        if obj == self.image_preview and event.type() == QEvent.MouseButtonPress:
            self.select_image(event)
            return True
        return super().eventFilter(obj, event)
    
    def select_image(self, event=None):
        print("DEBUG - select_image called")
        # 如果已经有图片显示，显示选项菜单
        if self.image_preview.pixmap() is not None:
            print("DEBUG - 已有图片，显示菜单")
            from PySide6.QtWidgets import QMenu
            from PySide6.QtGui import QAction
            menu = QMenu(self)
            view_action = QAction("查看大图", self)
            change_action = QAction("更换图片", self)
            menu.addAction(view_action)
            menu.addAction(change_action)
            
            def view_image():
                print("DEBUG - 查看大图")
                from PySide6.QtWidgets import QDialog, QVBoxLayout, QLabel
                dialog = QDialog(self)
                dialog.setWindowTitle("查看图片")
                dialog.setMinimumSize(600, 600)
                layout = QVBoxLayout()
                label = QLabel()
                label.setAlignment(Qt.AlignCenter)
                if self.selected_image_path:
                    pixmap = QPixmap(self.selected_image_path).scaled(580, 580, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                else:
                    pixmap = self.image_preview.pixmap().scaled(580, 580, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                label.setPixmap(pixmap)
                layout.addWidget(label)
                dialog.setLayout(layout)
                dialog.exec()
            
            def change_image():
                print("DEBUG - 更换图片")
                file_path, _ = QFileDialog.getOpenFileName(self, "选择图片", "", "图片文件 (*.png *.jpg *.jpeg *.gif)")
                print(f"DEBUG - 选择的文件路径: {file_path}")
                if file_path:
                    self.selected_image_path = file_path
                    print(f"DEBUG - selected_image_path 设置为: {self.selected_image_path}")
                    pixmap = QPixmap(file_path).scaled(196, 196, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                    self.image_preview.setPixmap(pixmap)
            
            view_action.triggered.connect(view_image)
            change_action.triggered.connect(change_image)
            
            menu.exec(self.image_preview.mapToGlobal(event.pos()) if event else self.image_preview.mapToGlobal(self.image_preview.rect().center()))
            return
        
        # 没有图片时，直接选择图片
        print("DEBUG - 没有图片，直接选择")
        file_path, _ = QFileDialog.getOpenFileName(self, "选择图片", "", "图片文件 (*.png *.jpg *.jpeg *.gif)")
        print(f"DEBUG - 选择的文件路径: {file_path}")
        if file_path:
            self.selected_image_path = file_path
            print(f"DEBUG - selected_image_path 设置为: {self.selected_image_path}")
            pixmap = QPixmap(file_path).scaled(196, 196, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.image_preview.setPixmap(pixmap)
            self.image_preview.setText("")

    def clear_image(self):
        self.selected_image_path = None
        self.image_preview.clear()
        self.image_preview.setText("点击选择图片")

    def load_sub_images(self):
        """加载产品副图"""
        try:
            images = self.api_client.get_product_images(self.product.get('id'))
            for img in images:
                if not img.get('is_default'):
                    self.add_sub_image_thumbnail(img.get('image_url'))
        except Exception as e:
            print(f"加载副图失败: {e}")

    def add_sub_image_thumbnail(self, image_url=None, file_path=None):
        """添加副图缩略图到界面"""
        thumb = QLabel()
        thumb.setFixedSize(60, 60)
        thumb.setStyleSheet("border: 1px solid #e5e7eb;")
        thumb.setAlignment(Qt.AlignCenter)

        if image_url:
            try:
                image_data = urllib.request.urlopen(image_url).read()
                image = QImage.fromData(image_data)
                pixmap = QPixmap.fromImage(image).scaled(56, 56, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                thumb.setPixmap(pixmap)
            except Exception:
                thumb.setText("图")
        elif file_path:
            pixmap = QPixmap(file_path).scaled(56, 56, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            thumb.setPixmap(pixmap)
        else:
            thumb.setText("+")

        thumb.setCursor(Qt.PointingHandCursor)
        thumb.setProperty("image_url", image_url)
        thumb.setProperty("file_path", file_path)
        thumb.mousePressEvent = lambda event, t=thumb: self.remove_sub_image(t)
        self.sub_images_layout.insertWidget(self.sub_images_layout.count() - 1, thumb)

    def add_sub_image(self):
        """添加副图"""
        file_path, _ = QFileDialog.getOpenFileName(self, "选择副图", "", "Images (*.png *.jpg *.jpeg)")
        if file_path:
            self.sub_image_paths.append(file_path)
            self.add_sub_image_thumbnail(file_path=file_path)

    def remove_sub_image(self, thumb_label):
        """移除副图"""
        file_path = thumb_label.property("file_path")
        if file_path and file_path in self.sub_image_paths:
            self.sub_image_paths.remove(file_path)
        thumb_label.deleteLater()

    def refresh_scheme_table(self):
        """刷新供应商方案摘要表格"""
        self.supplier_scheme_table.setRowCount(len(self.supplier_schemes))
        for row, s in enumerate(self.supplier_schemes):
            self.supplier_scheme_table.setItem(row, 0, QTableWidgetItem(s.get('supplier_name', '')))
            # 客户名称：优先customer_name，否则显示"通用"
            customer_name = s.get('customer_name', '') or '通用'
            self.supplier_scheme_table.setItem(row, 1, QTableWidgetItem(customer_name))
            # 工厂编号：兼容两种字段名
            factory_code = s.get('factory_code') or s.get('customer_product_code', '')
            self.supplier_scheme_table.setItem(row, 2, QTableWidgetItem(str(factory_code)))
            # 主要价格：兼容两种字段名
            main_price = s.get('purchase_price') or s.get('exw_price_incl') or ''
            self.supplier_scheme_table.setItem(row, 3, QTableWidgetItem(str(main_price)))
            self.supplier_scheme_table.setItem(row, 4, QTableWidgetItem(s.get('remark', '')))

            # 操作按钮
            action_widget = QWidget()
            action_layout = QHBoxLayout()
            action_layout.setContentsMargins(0, 0, 0, 0)

            edit_btn = QPushButton("编辑")
            edit_btn.setFixedWidth(50)
            edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px;")
            edit_btn.clicked.connect(lambda _, r=row: self.edit_supplier_scheme(r))
            action_layout.addWidget(edit_btn)

            del_btn = QPushButton("删除")
            del_btn.setFixedWidth(50)
            del_btn.setStyleSheet("background-color: #dc2626; color: white; border: none; border-radius: 4px;")
            del_btn.clicked.connect(lambda _, r=row: self.remove_supplier_scheme(r))
            action_layout.addWidget(del_btn)

            action_widget.setLayout(action_layout)
            self.supplier_scheme_table.setCellWidget(row, 5, action_widget)

    def load_supplier_schemes(self):
        """加载供应商方案"""
        try:
            if self.product and self.product.get('id'):
                # 从API加载供应商方案
                schemes = self.api_client.get_product_suppliers(self.product.get('id'))
                print(f"DEBUG - 加载供应商方案: {schemes}")
                self.supplier_schemes = schemes if schemes else []
            else:
                self.supplier_schemes = []
            self.refresh_scheme_table()
        except Exception as e:
            print(f"加载供应商方案失败: {e}")
            self.supplier_schemes = []
            self.refresh_scheme_table()

    def add_supplier_scheme(self):
        """弹窗添加供应商方案"""
        dialog = SupplierSchemeDialog(self.api_client, self.suppliers, self.customers, parent=self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            scheme_data = dialog.get_scheme_data()
            if scheme_data:
                self.supplier_schemes.append(scheme_data)
                self.refresh_scheme_table()

    def edit_supplier_scheme(self, row):
        """弹窗编辑供应商方案"""
        if row < 0 or row >= len(self.supplier_schemes):
            return
        dialog = SupplierSchemeDialog(self.api_client, self.suppliers, self.customers, 
                                      self.supplier_schemes[row], parent=self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            scheme_data = dialog.get_scheme_data()
            if scheme_data:
                self.supplier_schemes[row] = scheme_data
                self.refresh_scheme_table()

    def remove_supplier_scheme(self, row):
        """删除供应商方案"""
        if 0 <= row < len(self.supplier_schemes):
            self.supplier_schemes.pop(row)
            self.refresh_scheme_table()
    
    # ===== OE号管理 =====
    def load_product_oes(self):
        """加载产品的OE列表"""
        try:
            if self.product and self.product.get('id'):
                self.product_oes = self.api_client.get_product_oes(self.product.get('id')) or []
            else:
                self.product_oes = []
            self.refresh_oe_table()
        except Exception as e:
            print(f"加载OE列表失败: {e}")
            self.product_oes = []
            self.refresh_oe_table()
    
    def refresh_oe_table(self):
        """刷新OE列表"""
        self.oe_table.setRowCount(len(self.product_oes))
        for row, oe in enumerate(self.product_oes):
            self.oe_table.setItem(row, 0, QTableWidgetItem(oe.get('oe_number', '')))
            is_primary = "✓ 主OE" if oe.get('is_primary') else "—"
            primary_item = QTableWidgetItem(is_primary)
            primary_item.setTextAlignment(Qt.AlignCenter)
            self.oe_table.setItem(row, 1, primary_item)
            self.oe_table.setItem(row, 2, QTableWidgetItem(str(oe.get('created_at', ''))[:19]))
            
            # 操作按钮
            action_widget = QWidget()
            action_layout = QHBoxLayout()
            action_layout.setContentsMargins(0, 0, 0, 0)
            
            if not oe.get('is_primary'):
                set_primary_btn = QPushButton("设主")
                set_primary_btn.setFixedWidth(40)
                set_primary_btn.setStyleSheet("""
                    QPushButton {
                        background-color: #3b82f6;
                        color: white;
                        border: none;
                        border-radius: 4px;
                        padding: 2px 6px;
                        font-size: 10px;
                    }
                    QPushButton:hover { background-color: #2563eb; }
                """)
                set_primary_btn.clicked.connect(lambda _, oid=oe.get('id'): self.set_product_oe_primary(oid))
                action_layout.addWidget(set_primary_btn)
            
            del_btn = QPushButton("删除")
            del_btn.setFixedWidth(40)
            del_btn.setStyleSheet("""
                QPushButton {
                    background-color: #dc2626;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 2px 6px;
                    font-size: 10px;
                }
                QPushButton:hover { background-color: #b91c1c; }
            """)
            del_btn.clicked.connect(lambda _, oid=oe.get('id'): self.delete_product_oe(oid))
            action_layout.addWidget(del_btn)
            
            action_widget.setLayout(action_layout)
            self.oe_table.setCellWidget(row, 3, action_widget)
    
    def add_product_oe(self):
        """添加OE号"""
        from widgets.product_oe_dialog import AddOEDialog
        
        if not self.product:
            QMessageBox.warning(self, "提示", "请先保存产品基本信息后再添加OE号")
            return
        
        dialog = AddOEDialog(self.product.get('id'), self.api_client, self)
        if dialog.exec() == QDialog.Accepted:
            self.load_product_oes()
    
    def set_product_oe_primary(self, oe_id):
        """设置为主OE"""
        try:
            self.api_client.set_primary_oe(self.product.get('id'), oe_id)
            self.load_product_oes()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"设置失败: {e}")
    
    def delete_product_oe(self, oe_id):
        """删除OE"""
        reply = QMessageBox.question(self, "确认", "确定删除此OE号？", 
                                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.api_client.delete_product_oe(oe_id)
                self.load_product_oes()
            except Exception as e:
                QMessageBox.warning(self, "错误", f"删除失败: {e}")
    
    # ===== 客户产品编号管理 =====
    def load_product_customers(self):
        """加载产品的客户关联"""
        try:
            if self.product and self.product.get('id'):
                self.product_customers = self.api_client.get_product_customers(self.product.get('id')) or []
            else:
                self.product_customers = []
            self.refresh_pc_table()
        except Exception as e:
            print(f"加载客户产品失败: {e}")
            self.product_customers = []
            self.refresh_pc_table()
    
    def refresh_pc_table(self):
        """刷新客户产品列表"""
        self.pc_table.setRowCount(len(self.product_customers))
        for row, pc in enumerate(self.product_customers):
            # 客户名（直接从API返回的customer_name获取，不再本地查找）
            customer_name = pc.get('customer_name', pc.get('customer_code', ''))
            self.pc_table.setItem(row, 0, QTableWidgetItem(customer_name))
            self.pc_table.setItem(row, 1, QTableWidgetItem(pc.get('customer_product_code', '')))
            self.pc_table.setItem(row, 2, QTableWidgetItem(pc.get('customer_oe_number', '')))
            # 价格显示
            price_usd = pc.get('price_usd')
            price_rmb = pc.get('price_rmb')
            self.pc_table.setItem(row, 3, QTableWidgetItem(str(price_usd) if price_usd else '-'))
            self.pc_table.setItem(row, 4, QTableWidgetItem(str(price_rmb) if price_rmb else '-'))
            
            # 操作按钮
            action_widget = QWidget()
            action_layout = QHBoxLayout()
            action_layout.setContentsMargins(0, 0, 0, 0)
            
            del_btn = QPushButton("删除")
            del_btn.setFixedWidth(40)
            del_btn.setStyleSheet("""
                QPushButton {
                    background-color: #dc2626;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 2px 6px;
                    font-size: 10px;
                }
                QPushButton:hover { background-color: #b91c1c; }
            """)
            del_btn.clicked.connect(lambda _, pcid=pc.get('id'): self.delete_product_customer(pcid))
            action_layout.addWidget(del_btn)
            
            action_widget.setLayout(action_layout)
            self.pc_table.setCellWidget(row, 5, action_widget)
    
    def add_product_customer(self):
        """添加客户产品"""
        if not self.product:
            QMessageBox.warning(self, "提示", "请先保存产品基本信息后再添加客户产品")
            return
        
        # 弹窗选择客户并输入信息
        from PySide6.QtWidgets import QDialog, QFormLayout, QLineEdit, QDoubleSpinBox, QComboBox
        
        dialog = QDialog(self)
        dialog.setWindowTitle("添加客户产品")
        dialog.setMinimumWidth(400)
        layout = QVBoxLayout(dialog)
        
        form = QFormLayout()
        
        customer_combo = QComboBox()
        customer_combo.addItem("-- 选择客户 --", None)
        for c in (self.customers or []):
            customer_combo.addItem(c.get('customer_name', ''), c.get('id'))
        form.addRow("客户:", customer_combo)
        
        code_edit = QLineEdit()
        code_edit.setPlaceholderText("客户产品编号")
        form.addRow("客户产品编号:", code_edit)
        
        oe_edit = QLineEdit()
        oe_edit.setPlaceholderText("客户OE号")
        form.addRow("客户OE号:", oe_edit)
        
        price_usd = QDoubleSpinBox()
        price_usd.setRange(0, 9999999)
        price_usd.setDecimals(2)
        price_usd.setPrefix("$ ")
        form.addRow("USD价格:", price_usd)
        
        price_rmb = QDoubleSpinBox()
        price_rmb.setRange(0, 9999999)
        price_rmb.setDecimals(2)
        price_rmb.setPrefix("¥ ")
        form.addRow("RMB价格:", price_rmb)
        
        layout.addLayout(form)
        
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(dialog.reject)
        btn_layout.addWidget(cancel_btn)
        
        save_btn = QPushButton("保存")
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #10b981;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #059669; }
        """)
        save_btn.clicked.connect(lambda: self._save_product_customer(
            customer_combo.currentData(), code_edit.text(), oe_edit.text(),
            price_usd.value(), price_rmb.value(), dialog
        ))
        btn_layout.addWidget(save_btn)
        
        layout.addLayout(btn_layout)
        dialog.exec()
    
    def _save_product_customer(self, customer_id, code, customer_oe, price_usd, price_rmb, dialog):
        """保存客户产品关联"""
        if not customer_id:
            QMessageBox.warning(dialog, "提示", "请选择客户")
            return
        
        try:
            data = {
                "product_id": self.product.get('id'),
                "customer_id": customer_id,
                "customer_product_code": code,
                "customer_oe_number": customer_oe,
                "price_usd": price_usd if price_usd > 0 else None,
                "price_rmb": price_rmb if price_rmb > 0 else None
            }
            self.api_client.create_product_customer(data)
            dialog.accept()
            self.load_product_customers()
        except Exception as e:
            QMessageBox.warning(dialog, "错误", f"保存失败: {e}")
    
    def delete_product_customer(self, pc_id):
        """删除客户产品关联"""
        reply = QMessageBox.question(self, "确认", "确定删除此客户产品关联？", 
                                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.api_client.delete_product_customer(pc_id)
                self.load_product_customers()
            except Exception as e:
                QMessageBox.warning(self, "错误", f"删除失败: {e}")
    
    def save_product(self):
        print("DEBUG - save_product called")
        
        oe_number = self.oe_input.text().strip()
        detail_desc = self.detail_input.toPlainText().strip()
        
        if not oe_number:
            QMessageBox.warning(self, "提示", "请输入OE号")
            return
        
        if not detail_desc:
            QMessageBox.warning(self, "提示", "请输入细节描述")
            return

        data = {
            "dept_id": self.dept_id,
            "oe_number": oe_number,
            "detail_desc": detail_desc,
            "purchase_channel": self.channel_input.text().strip(),
            "category_id": int(self.category_combo.currentData()),
        }

        # 在后台线程中执行保存，避免UI阻塞
        self.save_product_async(data)
    
    def save_product_async(self, data):
        """在后台线程中保存产品，避免UI阻塞"""
        import threading
        
        # 显示加载提示
        self.setEnabled(False)
        from PySide6.QtWidgets import QLabel
        self.loading_label = QLabel("正在保存...", self)
        self.loading_label.setStyleSheet("background-color: #2563eb; color: white; padding: 8px 16px; border-radius: 4px;")
        self.loading_label.move(self.width()//2 - 50, self.height()//2)
        self.loading_label.show()
        
        def do_save():
            try:
                product_id = None
                
                # 如果有选择图片，先上传图片
                if self.selected_image_path:
                    try:
                        image_url = self.api_client.upload_image(self.selected_image_path)
                        if image_url:
                            data["default_image_url"] = image_url
                    except Exception as e:
                        print(f"图片上传异常: {str(e)}")
                
                # 上传副图
                sub_image_urls = []
                for path in self.sub_image_paths:
                    try:
                        url = self.api_client.upload_image(path)
                        if url:
                            sub_image_urls.append(url)
                    except Exception as e:
                        print(f"副图上传失败: {e}")
                data["sub_images"] = sub_image_urls
                
                # 保存产品基本信息
                if self.is_edit:
                    product_id = self.product.get('id')
                    self.api_client.update_product(product_id, data)
                    if self.selected_image_path and 'default_image_url' in data:
                        self.api_client.set_product_default_image(product_id, data['default_image_url'])
                else:
                    result = self.api_client.create_product(data)
                    if result and 'id' in result:
                        product_id = result['id']
                        if self.selected_image_path:
                            self.api_client.set_product_default_image(product_id, data.get('default_image_url'))
                
                # 保存供应商方案（后台线程）
                if product_id and self.supplier_schemes:
                    self.save_supplier_schemes_async(product_id)
                
                # 在主线程中更新UI
                from PySide6.QtCore import QMetaObject, Qt, Q_ARG
                QMetaObject.invokeMethod(self, "on_save_success", Qt.QueuedConnection)
                
            except Exception as e:
                print(f"保存产品失败: {str(e)}")
                from PySide6.QtCore import QMetaObject, Qt, Q_ARG
                QMetaObject.invokeMethod(self, "on_save_error", Qt.QueuedConnection, 
                                        Q_ARG(str, str(e)))
        
        thread = threading.Thread(target=do_save, daemon=True)
        thread.start()
    
    def on_save_success(self):
        """保存成功回调（主线程）"""
        self.loading_label.hide()
        self.setEnabled(True)
        # 清除产品缓存，下次加载时重新获取
        cache_manager.delete(CACHE_KEYS['PRODUCTS'])
        cache_manager.delete(CACHE_KEYS['INVENTORY_SUMMARY'])
        print("DEBUG - 已清除产品缓存")
        QMessageBox.information(self, "成功", "产品保存成功")
        self.accept()
    
    def on_save_error(self, error_msg):
        """保存失败回调（主线程）"""
        self.loading_label.hide()
        self.setEnabled(True)
        QMessageBox.warning(self, "错误", f"保存产品失败: {error_msg}")
     
    def save_supplier_schemes_async(self, product_id):
        """异步保存供应商方案，不阻塞UI"""
        import threading
        
        def save_schemes():
            try:
                print(f"DEBUG - 开始保存供应商方案，产品ID: {product_id}")
                for scheme in self.supplier_schemes:
                    # 映射前端字段名到后端字段名
                    scheme_data = {
                        'product_id': product_id,
                        'supplier_id': scheme.get('supplier_id'),
                        'factory_code': scheme.get('factory_code') or scheme.get('customer_product_code') or '',
                        'customer_id': scheme.get('customer_id'),
                        'is_default': scheme.get('is_default', False),
                        'purchase_price': scheme.get('purchase_price') or scheme.get('exw_price_incl'),
                        'currency': scheme.get('currency', 'CNY'),
                        'units_per_carton': scheme.get('units_per_carton'),
                        'carton_length_cm': scheme.get('carton_length_cm'),
                        'carton_width_cm': scheme.get('carton_width_cm'),
                        'carton_height_cm': scheme.get('carton_height_cm'),
                        'gross_weight_kg': scheme.get('gross_weight_kg'),
                        'moq': scheme.get('moq'),
                        'lead_time_days': scheme.get('lead_time_days'),
                        'purchase_channel': scheme.get('purchase_channel'),
                        'remark': scheme.get('remark'),
                        'special_requirements': scheme.get('special_requirements')
                    }
                    
                    print(f"DEBUG - 保存方案数据: supplier_id={scheme_data['supplier_id']}, factory_code={scheme_data['factory_code']}, customer_id={scheme_data['customer_id']}")
                    
                    if scheme.get('id'):
                        # 更新现有方案
                        self.api_client.update_product_supplier(scheme['id'], scheme_data)
                        print(f"DEBUG - 更新方案 ID={scheme['id']} 成功")
                    else:
                        # 创建新方案
                        result = self.api_client.add_product_supplier_full(scheme_data)
                        print(f"DEBUG - 创建方案成功: {result}")
                print(f"DEBUG - 供应商方案保存完成")
            except Exception as e:
                import traceback
                print(f"保存供应商方案失败: {str(e)}")
                traceback.print_exc()
        
        # 在后台线程中保存
        thread = threading.Thread(target=save_schemes, daemon=True)
        thread.start()


class SupplierSchemeDialog(QDialog):
    """供应商方案编辑弹窗"""
    def __init__(self, api_client, suppliers, customers, scheme=None, parent=None):
        super().__init__(parent)
        self.api_client = api_client
        self.suppliers = suppliers
        self.customers = customers
        self.scheme = scheme or {}
        self.is_edit = bool(scheme)
        self.setWindowTitle("编辑供应商方案" if self.is_edit else "添加供应商方案")
        self.setFixedSize(650, 700)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        form_layout = QFormLayout()
        form_layout.setSpacing(12)

        # 供应商
        self.supplier_combo = QComboBox()
        self.supplier_combo.setFixedHeight(35)
        self.supplier_combo.addItem("请选择供应商", None)
        for s in self.suppliers:
            self.supplier_combo.addItem(f"{s.get('supplier_code')} - {s.get('supplier_name')}", s.get('id'))
        if self.scheme.get('supplier_id'):
            for i in range(self.supplier_combo.count()):
                if self.supplier_combo.itemData(i) == self.scheme.get('supplier_id'):
                    self.supplier_combo.setCurrentIndex(i)
                    break
        form_layout.addRow("供应商 *:", self.supplier_combo)

        # 方案类型选择
        self.scheme_type_combo = QComboBox()
        self.scheme_type_combo.setFixedHeight(35)
        self.scheme_type_combo.addItem("🏷️ 默认方案（不指定客户）", {'type': 'default', 'customer_id': None})
        self.scheme_type_combo.addItem("👤 指定客户专属方案", {'type': 'customer', 'customer_id': None})
        self.scheme_type_combo.currentIndexChanged.connect(self.on_scheme_type_changed)
        form_layout.addRow("方案类型 *:", self.scheme_type_combo)

        # 客户选择（仅在指定客户时显示）
        self.customer_combo = QComboBox()
        self.customer_combo.setFixedHeight(35)
        self.customer_combo.addItem("请选择客户", None)
        for c in self.customers:
            self.customer_combo.addItem(f"{c.get('customer_code')} - {c.get('customer_name')}", c.get('id'))
        self.customer_combo.setVisible(False)
        
        # 根据编辑数据设置初始值
        if self.is_edit:
            if self.scheme.get('customer_id'):
                self.scheme_type_combo.setCurrentIndex(1)
                self.customer_combo.setVisible(True)
                for i in range(self.customer_combo.count()):
                    if self.customer_combo.itemData(i) == self.scheme.get('customer_id'):
                        self.customer_combo.setCurrentIndex(i)
                        break
            else:
                self.scheme_type_combo.setCurrentIndex(0)
        
        form_layout.addRow("选择客户:", self.customer_combo)

        # 客户产品编号（新建时默认使用父窗口的OE号）
        self.customer_code_input = QLineEdit()
        self.customer_code_input.setPlaceholderText("客户在对方系统中的产品编号")
        # 兼容两种字段名：factory_code（API返回）和 customer_product_code（前端使用）
        default_code = self.scheme.get('factory_code') or self.scheme.get('customer_product_code', '')
        if not default_code and not self.is_edit and self.parent() and hasattr(self.parent(), 'oe_input'):
            default_code = self.parent().oe_input.text().strip()
        self.customer_code_input.setText(default_code)
        form_layout.addRow("客户产品编号:", self.customer_code_input)

        layout.addLayout(form_layout)

        # 价格信息
        price_group = QGroupBox("价格信息")
        price_layout = QGridLayout()
        price_layout.setSpacing(10)

        self.exw_incl_input = QLineEdit()
        self.exw_incl_input.setPlaceholderText("EXW含税价")
        # 兼容两种字段名
        exw_val = self.scheme.get('exw_price_incl') or self.scheme.get('purchase_price', '')
        self.exw_incl_input.setText(str(exw_val or ''))
        price_layout.addWidget(QLabel("EXW含税价:"), 0, 0)
        price_layout.addWidget(self.exw_incl_input, 0, 1)

        self.exw_excl_input = QLineEdit()
        self.exw_excl_input.setPlaceholderText("EXW不含税价")
        self.exw_excl_input.setText(str(self.scheme.get('exw_price_excl', '') or ''))
        price_layout.addWidget(QLabel("EXW不含税价:"), 0, 2)
        price_layout.addWidget(self.exw_excl_input, 0, 3)

        self.fob_incl_input = QLineEdit()
        self.fob_incl_input.setPlaceholderText("FOB含税价")
        self.fob_incl_input.setText(str(self.scheme.get('fob_price_incl', '') or ''))
        price_layout.addWidget(QLabel("FOB含税价:"), 1, 0)
        price_layout.addWidget(self.fob_incl_input, 1, 1)

        self.fob_excl_input = QLineEdit()
        self.fob_excl_input.setPlaceholderText("FOB不含税价")
        self.fob_excl_input.setText(str(self.scheme.get('fob_price_excl', '') or ''))
        price_layout.addWidget(QLabel("FOB不含税价:"), 1, 2)
        price_layout.addWidget(self.fob_excl_input, 1, 3)

        self.freight_input = QLineEdit()
        self.freight_input.setPlaceholderText("运费")
        self.freight_input.setText(str(self.scheme.get('freight', '') or ''))
        price_layout.addWidget(QLabel("运费:"), 2, 0)
        price_layout.addWidget(self.freight_input, 2, 1)

        self.packing_fee_input = QLineEdit()
        self.packing_fee_input.setPlaceholderText("包装费")
        self.packing_fee_input.setText(str(self.scheme.get('packing_fee', '') or ''))
        price_layout.addWidget(QLabel("包装费:"), 2, 2)
        price_layout.addWidget(self.packing_fee_input, 2, 3)

        price_group.setLayout(price_layout)
        layout.addWidget(price_group)

        # 包装尺寸
        size_group = QGroupBox("包装尺寸")
        size_layout = QGridLayout()
        self.carton_length_input = QLineEdit()
        self.carton_length_input.setPlaceholderText("长(cm)")
        self.carton_length_input.setText(str(self.scheme.get('carton_length_cm', '') or ''))
        size_layout.addWidget(QLabel("纸箱长(cm):"), 0, 0)
        size_layout.addWidget(self.carton_length_input, 0, 1)

        self.carton_width_input = QLineEdit()
        self.carton_width_input.setPlaceholderText("宽(cm)")
        self.carton_width_input.setText(str(self.scheme.get('carton_width_cm', '') or ''))
        size_layout.addWidget(QLabel("纸箱宽(cm):"), 0, 2)
        size_layout.addWidget(self.carton_width_input, 0, 3)

        self.carton_height_input = QLineEdit()
        self.carton_height_input.setPlaceholderText("高(cm)")
        self.carton_height_input.setText(str(self.scheme.get('carton_height_cm', '') or ''))
        size_layout.addWidget(QLabel("纸箱高(cm):"), 1, 0)
        size_layout.addWidget(self.carton_height_input, 1, 1)

        self.units_input = QLineEdit()
        self.units_input.setPlaceholderText("每箱数量")
        self.units_input.setText(str(self.scheme.get('units_per_carton', '') or ''))
        size_layout.addWidget(QLabel("每箱数量:"), 1, 2)
        size_layout.addWidget(self.units_input, 1, 3)
        size_group.setLayout(size_layout)
        layout.addWidget(size_group)

        # 重量信息
        weight_group = QGroupBox("重量信息")
        weight_layout = QHBoxLayout()
        self.gross_weight_input = QLineEdit()
        self.gross_weight_input.setPlaceholderText("毛重(kg)")
        self.gross_weight_input.setText(str(self.scheme.get('gross_weight_kg', '') or ''))
        weight_layout.addWidget(QLabel("毛重(kg):"))
        weight_layout.addWidget(self.gross_weight_input)

        self.weight_input = QLineEdit()
        self.weight_input.setPlaceholderText("净重(kg)")
        self.weight_input.setText(str(self.scheme.get('weight_kg', '') or ''))
        weight_layout.addWidget(QLabel("净重(kg):"))
        weight_layout.addWidget(self.weight_input)
        weight_group.setLayout(weight_layout)
        layout.addWidget(weight_group)

        # 备注
        self.remark_input = QTextEdit()
        self.remark_input.setPlaceholderText("备注信息")
        self.remark_input.setText(self.scheme.get('remark', ''))
        self.remark_input.setMaximumHeight(60)
        layout.addWidget(QLabel("备注:"))
        layout.addWidget(self.remark_input)
        
        # 设为默认方案
        self.is_default_checkbox = QCheckBox("设为默认供应商方案（优先使用）")
        self.is_default_checkbox.setStyleSheet("color: #2563eb; font-weight: 500;")
        if self.scheme.get('is_default'):
            self.is_default_checkbox.setChecked(True)
        layout.addWidget(self.is_default_checkbox)

        # 按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        save_btn = QPushButton("保存")
        save_btn.setFixedWidth(100)
        save_btn.clicked.connect(self.save_scheme)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        btn_layout.addWidget(save_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.setFixedWidth(100)
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        btn_layout.addWidget(cancel_btn)

        layout.addLayout(btn_layout)
        self.setLayout(layout)

    def on_scheme_type_changed(self):
        """方案类型变化时显示/隐藏客户选择"""
        scheme_data = self.scheme_type_combo.currentData()
        if scheme_data and scheme_data.get('type') == 'customer':
            self.customer_combo.setVisible(True)
        else:
            self.customer_combo.setVisible(False)

    def save_scheme(self):
        supplier_id = self.supplier_combo.currentData()
        
        if not supplier_id:
            QMessageBox.warning(self, "警告", "请选择供应商")
            return
        
        # 根据方案类型获取客户ID
        scheme_data = self.scheme_type_combo.currentData()
        customer_id = None
        if scheme_data and scheme_data.get('type') == 'customer':
            customer_id = self.customer_combo.currentData()
            if not customer_id:
                QMessageBox.warning(self, "警告", "请选择客户")
                return

        def try_float(value):
            try:
                return float(value) if value.strip() else None
            except ValueError:
                return None

        supplier_name = self.supplier_combo.currentText()
        customer_name = self.customer_combo.currentText()

        self.scheme_data = {
            "id": self.scheme.get('id') if self.is_edit else None,
            "supplier_id": supplier_id,
            "supplier_name": supplier_name,
            "customer_id": customer_id,
            "customer_name": customer_name,
            "customer_product_code": self.customer_code_input.text().strip(),
            "is_default": self.is_default_checkbox.isChecked(),
            "exw_price_incl": try_float(self.exw_incl_input.text()),
            "exw_price_excl": try_float(self.exw_excl_input.text()),
            "fob_price_incl": try_float(self.fob_incl_input.text()),
            "fob_price_excl": try_float(self.fob_excl_input.text()),
            "freight": try_float(self.freight_input.text()),
            "packing_fee": try_float(self.packing_fee_input.text()),
            "carton_length_cm": try_float(self.carton_length_input.text()),
            "carton_width_cm": try_float(self.carton_width_input.text()),
            "carton_height_cm": try_float(self.carton_height_input.text()),
            "units_per_carton": int(self.units_input.text()) if self.units_input.text().strip() else None,
            "gross_weight_kg": try_float(self.gross_weight_input.text()),
            "weight_kg": try_float(self.weight_input.text()),
            "remark": self.remark_input.toPlainText()
        }
        print(f"DEBUG - save_scheme: scheme_data = {self.scheme_data}")
        self.accept()

    def get_scheme_data(self):
        return getattr(self, 'scheme_data', None)


class CustomerDialog(QDialog):
    def __init__(self, api_client: ApiClient, customer=None):
        super().__init__()
        self.api_client = api_client
        self.customer = customer
        self.is_edit = customer is not None
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("编辑客户" if self.is_edit else "新增客户")
        self.setMinimumSize(750, 600)

        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        form_layout = QFormLayout()
        form_layout.setSpacing(12)

        self.dept_combo = QComboBox()
        self.dept_combo.addItems([
            "S - 索英普",
            "W - 维那",
            "M - 马迪那",
            "D - 银达"
        ])
        form_layout.addRow("部门:", self.dept_combo)

        if self.is_edit:
            self.code_label = QLabel(self.customer.get('customer_code', ''))
            form_layout.addRow("客户编号:", self.code_label)
        
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("请输入客户名称")
        if self.customer:
            self.name_input.setText(self.customer.get('customer_name', ''))
        form_layout.addRow("客户名称 *:", self.name_input)

        self.country_input = QLineEdit()
        self.country_input.setPlaceholderText("请输入所在国家")
        if self.customer:
            self.country_input.setText(self.customer.get('country', ''))
        form_layout.addRow("所在国家 *:", self.country_input)

        self.basic_require_input = QTextEdit()
        self.basic_require_input.setPlaceholderText("请输入通用交易条款")
        self.basic_require_input.setMaximumHeight(60)
        if self.customer:
            self.basic_require_input.setText(self.customer.get('basic_require', ''))
        form_layout.addRow("基本要求:", self.basic_require_input)

        self.special_input = QTextEdit()
        self.special_input.setPlaceholderText("请输入特殊要求，如特定包装、标签等")
        self.special_input.setMaximumHeight(60)
        if self.customer:
            self.special_input.setText(self.customer.get('special_require', ''))
        form_layout.addRow("特殊要求:", self.special_input)

        self.payment_input = QLineEdit()
        self.payment_input.setPlaceholderText("如 T/T 30天")
        if self.customer:
            self.payment_input.setText(self.customer.get('payment_terms', ''))
        form_layout.addRow("付款条款:", self.payment_input)

        contact_group = QGroupBox("联系人信息")
        contact_layout = QFormLayout()
        contact_layout.setSpacing(10)

        self.contact_name_input = QLineEdit()
        self.contact_name_input.setPlaceholderText("联系人姓名")
        if self.customer:
            self.contact_name_input.setText(self.customer.get('contact_name', ''))
        contact_layout.addRow("姓名:", self.contact_name_input)

        self.contact_phone_input = QLineEdit()
        self.contact_phone_input.setPlaceholderText("电话")
        if self.customer:
            self.contact_phone_input.setText(self.customer.get('contact_phone', ''))
        contact_layout.addRow("电话:", self.contact_phone_input)

        self.contact_email_input = QLineEdit()
        self.contact_email_input.setPlaceholderText("邮箱")
        if self.customer:
            self.contact_email_input.setText(self.customer.get('contact_email', ''))
        contact_layout.addRow("邮箱:", self.contact_email_input)

        contact_group.setLayout(contact_layout)
        layout.addLayout(form_layout)
        layout.addWidget(contact_group)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        save_btn = QPushButton("保存")
        save_btn.setFixedWidth(100)
        save_btn.clicked.connect(self.save_customer)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        btn_layout.addWidget(save_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.setFixedWidth(100)
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        btn_layout.addWidget(cancel_btn)

        layout.addLayout(btn_layout)
        self.setLayout(layout)

    def save_customer(self):
        print("save_customer called")  # 调试信息
        print(f"is_edit: {self.is_edit}")  # 调试信息
        print(f"customer: {self.customer}")  # 调试信息
        
        if not self.name_input.text().strip():
            QMessageBox.warning(self, "提示", "请输入客户名称")
            return
        if not self.country_input.text().strip():
            QMessageBox.warning(self, "提示", "请输入所在国家")
            return

        dept_map = {"S - 索英普": "S", "W - 维那": "W", "M - 马迪那": "M", "D - 银达": "D"}
        dept_id = dept_map.get(self.dept_combo.currentText(), "S")

        data = {
            "dept_id": dept_id,
            "customer_name": self.name_input.text().strip(),
            "country": self.country_input.text().strip(),
            "basic_require": self.basic_require_input.toPlainText().strip(),
            "special_require": self.special_input.toPlainText().strip(),
            "payment_terms": self.payment_input.text().strip()
        }

        print(f"准备保存数据: {data}")  # 调试信息

        contact_name = self.contact_name_input.text().strip()
        contact_phone = self.contact_phone_input.text().strip()
        contact_email = self.contact_email_input.text().strip()
        if contact_name or contact_phone or contact_email:
            data["contacts"] = [{
                "name": contact_name,
                "phone": contact_phone,
                "email": contact_email,
                "is_primary": 1
            }]

        try:
            print(f"调用API: {'update' if self.is_edit else 'create'}")  # 调试信息
            if self.is_edit:
                customer_id = self.customer.get('id', 'NOT_FOUND')
                print(f"更新客户ID: {customer_id}")  # 调试信息
                result = self.api_client.update_customer(customer_id, data)
                print(f"API返回结果: {result}")  # 调试信息
            else:
                result = self.api_client.create_customer(data)
                print(f"API返回结果: {result}")  # 调试信息
            print("API调用成功")  # 调试信息
            self.accept()
        except Exception as e:
            print(f"API调用失败: {str(e)}")  # 调试信息
            QMessageBox.warning(self, "错误", f"保存失败: {str(e)}")
            traceback.print_exc()


class CustomerDetailDialog(QDialog):
    def __init__(self, api_client: ApiClient, customer):
        super().__init__()
        self.api_client = api_client
        self.customer = customer
        self.addresses = []
        self.contacts = []
        self.pi_orders = []
        self.init_ui()
        self.load_data()

    def init_ui(self):
        self.setWindowTitle(f"客户详情 - {self.customer.get('customer_name', '')}")
        self.setMinimumSize(800, 600)

        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)

        # 标签页
        self.tab_widget = QTabWidget()
        
        # 基本信息页
        self.basic_tab = QWidget()
        self.setup_basic_tab()
        
        # 收货地址页
        self.address_tab = QWidget()
        self.setup_address_tab()
        
        # 联系人页
        self.contact_tab = QWidget()
        self.setup_contact_tab()
        
        # PI订单历史页
        self.pi_tab = QWidget()
        self.setup_pi_tab()

        self.tab_widget.addTab(self.basic_tab, "基本信息")
        self.tab_widget.addTab(self.address_tab, "收货地址")
        self.tab_widget.addTab(self.contact_tab, "联系人")
        self.tab_widget.addTab(self.pi_tab, "交易历史")

        layout.addWidget(self.tab_widget)

        # 关闭按钮
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        close_btn = QPushButton("关闭")
        close_btn.setFixedWidth(100)
        close_btn.clicked.connect(self.close)
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        btn_layout.addWidget(close_btn)
        layout.addLayout(btn_layout)

        self.setLayout(layout)

    def setup_basic_tab(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(12)

        form_layout = QFormLayout()
        form_layout.setSpacing(10)

        form_layout.addRow(QLabel("<b>客户编号:</b>"), QLabel(self.customer.get('customer_code', '')))
        form_layout.addRow(QLabel("<b>客户名称:</b>"), QLabel(self.customer.get('customer_name', '')))
        form_layout.addRow(QLabel("<b>所属部门:</b>"), QLabel(self.customer.get('dept_id', '')))
        form_layout.addRow(QLabel("<b>所在国家:</b>"), QLabel(self.customer.get('country', '')))
        
        basic_require = self.customer.get('basic_require', '')
        form_layout.addRow(QLabel("<b>基本要求:</b>"), QLabel(basic_require if basic_require else "-"))
        
        form_layout.addRow(QLabel("<b>付款条款:</b>"), QLabel(self.customer.get('payment_terms', '') or "-"))
        
        status = self.customer.get('status', 1)
        status_text = "启用" if status == 1 else "禁用"
        status_color = "#10b981" if status == 1 else "#ef4444"
        status_label = QLabel(status_text)
        status_label.setStyleSheet(f"color: {status_color}; font-weight: bold;")
        form_layout.addRow(QLabel("<b>状态:</b>"), status_label)

        layout.addLayout(form_layout)
        layout.addStretch()

        special_require = self.customer.get('special_require', '')
        if special_require:
            special_group = QGroupBox("特殊要求")
            special_group.setStyleSheet("QGroupBox { font-weight: bold; border: 2px solid #dc2626; border-radius: 5px; }")
            special_layout = QVBoxLayout()
            special_label = QLabel(special_require)
            special_label.setWordWrap(True)
            special_label.setStyleSheet("color: #dc2626; padding: 5px;")
            special_layout.addWidget(special_label)
            special_group.setLayout(special_layout)
            layout.addWidget(special_group)

        self.basic_tab.setLayout(layout)

    def setup_address_tab(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(10, 10, 10, 10)

        # 工具栏
        toolbar = QHBoxLayout()
        add_btn = QPushButton("+ 添加地址")
        add_btn.clicked.connect(self.add_address)
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 6px 12px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        toolbar.addWidget(add_btn)
        toolbar.addStretch()
        layout.addLayout(toolbar)

        # 地址列表
        self.addresses_table = QTableWidget()
        self.addresses_table.setColumnCount(6)
        self.addresses_table.setHorizontalHeaderLabels(["国家", "港口", "详细地址", "默认地址", "编辑", "删除"])
        self.addresses_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self.addresses_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        layout.addWidget(self.addresses_table)

        self.address_tab.setLayout(layout)

    def setup_contact_tab(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(10, 10, 10, 10)

        # 工具栏
        toolbar = QHBoxLayout()
        toolbar.addStretch()
        
        add_btn = QPushButton("+ 新增联系人")
        add_btn.clicked.connect(self.add_contact)
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        toolbar.addWidget(add_btn)
        
        layout.addLayout(toolbar)

        self.contacts_table = QTableWidget()
        self.contacts_table.setColumnCount(7)
        self.contacts_table.setHorizontalHeaderLabels(["姓名", "职位", "电话", "邮箱", "是否主要", "编辑", "删除"])
        self.contacts_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.contacts_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        layout.addWidget(self.contacts_table)

        self.contact_tab.setLayout(layout)

    def setup_pi_tab(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(10, 10, 10, 10)

        self.pi_table = QTableWidget()
        self.pi_table.setColumnCount(10)
        self.pi_table.setHorizontalHeaderLabels(["", "ID", "PI号", "金额", "币种", "状态", "创建时间", "操作", "完成", "导出"])
        self.pi_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.pi_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.pi_table.setColumnWidth(0, 40)
        self.pi_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.pi_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self.pi_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        self.pi_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        self.pi_table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        self.pi_table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
        self.pi_table.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeMode.ResizeToContents)
        self.pi_table.horizontalHeader().setSectionResizeMode(8, QHeaderView.ResizeMode.ResizeToContents)
        self.pi_table.horizontalHeader().setSectionResizeMode(9, QHeaderView.ResizeMode.ResizeToContents)
        self.setup_table_context_menu(self.pi_table, ["", "ID", "PI号", "金额", "币种", "状态", "创建时间", "操作", "完成", "导出"])
        layout.addWidget(self.pi_table)

        self.pi_tab.setLayout(layout)

    def load_data(self):
        try:
            # 加载地址
            self.addresses = self.api_client.get_customer_addresses(self.customer['id'])
            self.load_addresses_table()

            # 加载联系人
            self.contacts = self.api_client.get_customer_contacts(self.customer['id'])
            self.load_contacts_table()

            # 加载PI订单
            self.pi_orders = self.api_client.get_customer_pi_list(self.customer['id'])
            self.load_pi_table()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载数据失败: {str(e)}")

    def load_addresses_table(self):
        self.addresses_table.setRowCount(len(self.addresses))
        for row, addr in enumerate(self.addresses):
            self.addresses_table.setItem(row, 0, QTableWidgetItem(addr.get('country', '')))
            self.addresses_table.setItem(row, 1, QTableWidgetItem(addr.get('port', '')))
            self.addresses_table.setItem(row, 2, QTableWidgetItem(addr.get('address_detail', '')))
            
            is_default = addr.get('is_default', 0)
            default_text = "是" if is_default == 1 else "否"
            self.addresses_table.setItem(row, 3, QTableWidgetItem(default_text))

            # 编辑按钮
            edit_btn = QPushButton("编辑")
            edit_btn.setFixedWidth(50)
            edit_btn.setStyleSheet("""
                QPushButton {
                    background-color: #3b82f6;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 4px 8px;
                }
                QPushButton:hover { background-color: #2563eb; }
            """)
            edit_btn.clicked.connect(lambda _, addr=addr: self.edit_address(addr))
            self.addresses_table.setCellWidget(row, 4, edit_btn)
            
            # 删除按钮
            delete_btn = QPushButton("删除")
            delete_btn.setFixedWidth(50)
            delete_btn.setStyleSheet("""
                QPushButton {
                    background-color: #ef4444;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 4px 8px;
                }
                QPushButton:hover { background-color: #dc2626; }
            """)
            delete_btn.clicked.connect(lambda _, addr=addr: self.delete_address(addr))
            self.addresses_table.setCellWidget(row, 5, delete_btn)

    def load_contacts_table(self):
        self.contacts_table.setRowCount(len(self.contacts))
        for row, contact in enumerate(self.contacts):
            self.contacts_table.setItem(row, 0, QTableWidgetItem(contact.get('name', '')))
            self.contacts_table.setItem(row, 1, QTableWidgetItem(contact.get('position', '')))
            self.contacts_table.setItem(row, 2, QTableWidgetItem(contact.get('phone', '')))
            self.contacts_table.setItem(row, 3, QTableWidgetItem(contact.get('email', '')))
            
            is_primary = contact.get('is_primary', 0)
            primary_text = "是" if is_primary == 1 else "否"
            self.contacts_table.setItem(row, 4, QTableWidgetItem(primary_text))

            # 操作按钮
            btn_layout = QHBoxLayout()
            
            # 编辑按钮
            edit_btn = QPushButton("编辑")
            edit_btn.setFixedWidth(50)
            edit_btn.setStyleSheet("""
                QPushButton {
                    background-color: #3b82f6;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 4px 8px;
                }
                QPushButton:hover { background-color: #2563eb; }
            """)
            edit_btn.clicked.connect(lambda _, c=contact: self.edit_contact(c))
            self.contacts_table.setCellWidget(row, 5, edit_btn)
            
            # 删除按钮
            delete_btn = QPushButton("删除")
            delete_btn.setFixedWidth(50)
            delete_btn.setStyleSheet("""
                QPushButton {
                    background-color: #ef4444;
                    color: white;
                    border: none;
                    border-radius: 4px;
                    padding: 4px 8px;
                }
                QPushButton:hover { background-color: #dc2626; }
            """)
            delete_btn.clicked.connect(lambda _, c=contact: self.delete_contact(c))
            self.contacts_table.setCellWidget(row, 6, delete_btn)

    def load_pi_table(self):
        self.pi_table.setRowCount(len(self.pi_orders))
        for row, pi in enumerate(self.pi_orders):
            self.pi_table.setItem(row, 0, QTableWidgetItem(pi.get('pi_number', '')))
            self.pi_table.setItem(row, 1, QTableWidgetItem(str(pi.get('total_amount', ''))))
            
            status = pi.get('status', '')
            self.pi_table.setItem(row, 2, QTableWidgetItem(status))
            
            created_at = pi.get('created_at', '')
            if created_at:
                created_at = created_at[:19] if isinstance(created_at, str) else str(created_at)
            self.pi_table.setItem(row, 3, QTableWidgetItem(created_at))

    def add_address(self):
        dialog = AddressDialog(self.api_client, customer_id=self.customer['id'])
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_data()

    def edit_address(self, address):
        dialog = AddressDialog(self.api_client, customer_id=self.customer['id'], address=address)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_data()

    def delete_address(self, address):
        reply = QMessageBox.question(self, "确认删除", "确定要删除这个地址吗？",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.api_client.delete_customer_address(self.customer['id'], address['id'])
                self.load_data()
            except Exception as e:
                QMessageBox.warning(self, "错误", f"删除失败：{str(e)}")

    def add_contact(self):
        dialog = ContactDialog(self.api_client, customer_id=self.customer['id'])
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_data()

    def edit_contact(self, contact):
        dialog = ContactDialog(self.api_client, customer_id=self.customer['id'], contact=contact)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_data()

    def delete_contact(self, contact):
        reply = QMessageBox.question(self, "确认删除", "确定要删除这个联系人吗？",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.api_client.delete_customer_contact(self.customer['id'], contact['id'])
                self.load_data()
            except Exception as e:
                QMessageBox.warning(self, "错误", f"删除失败：{str(e)}")


class AddressDialog(QDialog):
    def __init__(self, api_client: ApiClient, customer_id, address=None):
        super().__init__()
        self.api_client = api_client
        self.customer_id = customer_id
        self.address = address
        self.is_edit = address is not None
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("编辑地址" if self.is_edit else "添加地址")
        self.setFixedSize(400, 300)

        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        form_layout = QFormLayout()
        form_layout.setSpacing(12)

        self.country_input = QLineEdit()
        if self.address:
            self.country_input.setText(self.address.get('country', ''))
        form_layout.addRow("国家:", self.country_input)

        self.port_input = QLineEdit()
        if self.address:
            self.port_input.setText(self.address.get('port', ''))
        form_layout.addRow("港口:", self.port_input)

        self.detail_input = QTextEdit()
        if self.address:
            self.detail_input.setText(self.address.get('address_detail', ''))
        self.detail_input.setMaximumHeight(80)
        form_layout.addRow("详细地址:", self.detail_input)

        self.default_checkbox = QCheckBox("设为默认地址")
        if self.address and self.address.get('is_default', 0) == 1:
            self.default_checkbox.setChecked(True)
        form_layout.addRow("", self.default_checkbox)

        layout.addLayout(form_layout)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        save_btn = QPushButton("保存")
        save_btn.setFixedWidth(100)
        save_btn.clicked.connect(self.save_address)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        btn_layout.addWidget(save_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.setFixedWidth(100)
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        btn_layout.addWidget(cancel_btn)

        layout.addLayout(btn_layout)
        self.setLayout(layout)

    def save_address(self):
        data = {
            "country": self.country_input.text().strip(),
            "port": self.port_input.text().strip(),
            "address_detail": self.detail_input.toPlainText().strip(),
            "is_default": 1 if self.default_checkbox.isChecked() else 0
        }

        try:
            if self.is_edit:
                self.api_client.update_customer_address(self.customer_id, self.address['id'], data)
            else:
                self.api_client.create_customer_address(self.customer_id, data)
            self.accept()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"保存失败：{str(e)}")


class ContactDialog(QDialog):
    def __init__(self, api_client: ApiClient, customer_id, contact=None):
        super().__init__()
        self.api_client = api_client
        self.customer_id = customer_id
        self.contact = contact
        self.is_edit = contact is not None
        self.init_ui()

    def init_ui(self):
        self.setWindowTitle("编辑联系人" if self.is_edit else "添加联系人")
        self.setFixedSize(400, 350)

        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        form_layout = QFormLayout()
        form_layout.setSpacing(12)

        self.name_input = QLineEdit()
        if self.contact:
            self.name_input.setText(self.contact.get('name', ''))
        form_layout.addRow("姓名:", self.name_input)

        self.position_input = QLineEdit()
        if self.contact:
            self.position_input.setText(self.contact.get('position', ''))
        form_layout.addRow("职位:", self.position_input)

        self.phone_input = QLineEdit()
        if self.contact:
            self.phone_input.setText(self.contact.get('phone', ''))
        form_layout.addRow("电话:", self.phone_input)

        self.email_input = QLineEdit()
        if self.contact:
            self.email_input.setText(self.contact.get('email', ''))
        form_layout.addRow("邮箱:", self.email_input)

        self.primary_checkbox = QCheckBox("设为主要联系人")
        if self.contact and self.contact.get('is_primary', 0) == 1:
            self.primary_checkbox.setChecked(True)
        form_layout.addRow("", self.primary_checkbox)

        layout.addLayout(form_layout)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        save_btn = QPushButton("保存")
        save_btn.setFixedWidth(100)
        save_btn.clicked.connect(self.save_contact)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        btn_layout.addWidget(save_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.setFixedWidth(100)
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        btn_layout.addWidget(cancel_btn)

        layout.addLayout(btn_layout)
        self.setLayout(layout)

    def save_contact(self):
        if not self.name_input.text().strip():
            QMessageBox.warning(self, "提示", "请输入联系人姓名")
            return

        data = {
            "name": self.name_input.text().strip(),
            "position": self.position_input.text().strip(),
            "phone": self.phone_input.text().strip(),
            "email": self.email_input.text().strip(),
            "is_primary": 1 if self.primary_checkbox.isChecked() else 0
        }

        try:
            if self.is_edit:
                self.api_client.update_customer_contact(self.customer_id, self.contact['id'], data)
            else:
                self.api_client.create_customer_contact(self.customer_id, data)
            self.accept()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"保存失败：{str(e)}")


class SupplierDialog(QDialog):
    def __init__(self, api_client: ApiClient, supplier=None):
        super().__init__()
        self.api_client = api_client
        self.supplier = supplier
        self.is_edit = supplier is not None
        self.provinces = []
        self.cities = []
        self.selected_city_code = ""
        self.init_ui()
        QTimer.singleShot(0, self.load_provinces)

    def load_provinces(self):
        try:
            self.provinces = self.api_client.get_provinces()
            self.province_combo.clear()
            self.province_combo.addItems(self.provinces)
            if self.supplier and self.supplier.get('region'):
                region = self.supplier.get('region', '')
                for prov in self.provinces:
                    if region.startswith(prov):
                        self.province_combo.setCurrentText(prov)
                        self.load_cities(prov)
                        city_name = region[len(prov):].strip()
                        if city_name and city_name in self.cities:
                            self.city_combo.setCurrentText(city_name)
                        break
        except Exception as e:
            print(f"加载省份失败: {e}")

    def load_cities(self, province):
        try:
            self.cities = self.api_client.get_cities(province)
            self.city_combo.clear()
            self.city_combo.addItems(self.cities)
        except Exception as e:
            print(f"加载城市失败: {e}")

    def on_province_changed(self, province):
        self.load_cities(province)

    def on_city_changed(self, city):
        province = self.province_combo.currentText()
        try:
            # 使用模块级别的静态映射（只创建一次）
            p_code = PROVINCE_CODE_MAP.get(province, "")
            c_map = CITY_CODE_MAP.get(p_code, {})
            self.selected_city_code = p_code + c_map.get(city, "0")
        except Exception as e:
            print(f"获取城市编码失败: {e}")

    def init_ui(self):
        self.setWindowTitle("编辑供应商" if self.is_edit else "新增供应商")
        self.setFixedSize(500, 480)

        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        form_layout = QFormLayout()
        form_layout.setSpacing(12)

        if self.is_edit:
            self.code_label = QLabel(self.supplier.get('supplier_code', ''))
            form_layout.addRow("供应商编号:", self.code_label)

        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("请输入供应商名称")
        if self.supplier:
            self.name_input.setText(self.supplier.get('supplier_name', ''))
        form_layout.addRow("供应商名称:", self.name_input)

        province_layout = QHBoxLayout()
        self.province_combo = QComboBox()
        self.province_combo.setFixedHeight(35)
        self.province_combo.currentTextChanged.connect(self.on_province_changed)
        province_layout.addWidget(self.province_combo)

        self.city_combo = QComboBox()
        self.city_combo.setFixedHeight(35)
        self.city_combo.currentTextChanged.connect(self.on_city_changed)
        province_layout.addWidget(self.city_combo)
        form_layout.addRow("省份/城市:", province_layout)

        self.contact_input = QLineEdit()
        self.contact_input.setPlaceholderText("请输入联系人")
        if self.supplier:
            self.contact_input.setText(self.supplier.get('contact_person', ''))
        form_layout.addRow("联系人:", self.contact_input)

        self.phone_input = QLineEdit()
        self.phone_input.setPlaceholderText("请输入联系电话")
        if self.supplier:
            self.phone_input.setText(self.supplier.get('phone', ''))
        form_layout.addRow("联系电话:", self.phone_input)

        self.email_input = QLineEdit()
        self.email_input.setPlaceholderText("请输入邮箱地址")
        if self.supplier:
            self.email_input.setText(self.supplier.get('email', ''))
        form_layout.addRow("邮箱:", self.email_input)

        self.address_input = QTextEdit()
        self.address_input.setPlaceholderText("请输入详细地址")
        if self.supplier:
            self.address_input.setText(self.supplier.get('address', ''))
        self.address_input.setMaximumHeight(80)
        form_layout.addRow("详细地址:", self.address_input)

        layout.addLayout(form_layout)

        btn_layout = QHBoxLayout()
        btn_layout.addStretch()

        save_btn = QPushButton("保存")
        save_btn.setFixedWidth(100)
        save_btn.clicked.connect(self.save_supplier)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        btn_layout.addWidget(save_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.setFixedWidth(100)
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        btn_layout.addWidget(cancel_btn)

        layout.addLayout(btn_layout)
        self.setLayout(layout)

    def save_supplier(self):
        if not self.name_input.text().strip():
            QMessageBox.warning(self, "提示", "请输入供应商名称")
            return

        province = self.province_combo.currentText()
        city = self.city_combo.currentText()
        region = f"{province} {city}" if province and city else ""

        data = {
            "supplier_name": self.name_input.text().strip(),
            "province": province,
            "city": city,
            "city_code": self.selected_city_code,
            "region": region,
            "contact_person": self.contact_input.text().strip(),
            "phone": self.phone_input.text().strip(),
            "email": self.email_input.text().strip(),
            "address": self.address_input.toPlainText().strip()
        }

        try:
            if self.is_edit:
                self.api_client.update_supplier(self.supplier['id'], data)
            else:
                self.api_client.create_supplier(data)
            self.accept()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"保存失败: {str(e)}")


class MainWindow(QMainWindow):
    def __init__(self, api_client: ApiClient, dept_id: str):
        super().__init__()
        self.api_client = api_client
        self.dept_id = dept_id
        # 全局变量
        self.default_profit_margin = 25.0  # 默认毛利率
        self.exchange_rate = 7.24          # 默认汇率
        self.init_ui()
        self.load_globals()
        self.load_data()
    
    def load_globals(self):
        """加载全局变量（使用本地配置，无网络延迟）"""
        try:
            from config.local_settings_manager import load_local_settings
            settings = load_local_settings()
            self.default_profit_margin = settings.get('default_profit_margin', 25.0)
            self.exchange_rate = settings.get('exchange_rate', 7.24)
            print(f"[INFO] 全局变量加载: 毛利率={self.default_profit_margin}%, 汇率={self.exchange_rate}")
        except Exception as e:
            print(f"[WARN] 加载全局变量失败，使用默认值: {e}")
            self.default_profit_margin = 25.0
            self.exchange_rate = 7.24
    
    def calculate_estimated_usd_price(self, factory_rmb_price):
        """计算预估美金报价
        公式: 预估美金报价 = 工厂人民币价格 × (1 + 毛利率) / 汇率
        """
        if not factory_rmb_price or factory_rmb_price <= 0:
            return 0
        margin_factor = 1 + (self.default_profit_margin / 100)
        return factory_rmb_price * margin_factor / self.exchange_rate
    
    def calculate_order_profit_margin(self, customer_usd_price, total_rmb_amount, exchange_rate=None):
        """计算订单预估毛利率
        公式: 预估毛利率 = 客户美金报价 × 汇率 / 总金额
        注意: 需要将总金额转为美金计算毛利率
        """
        if not exchange_rate:
            exchange_rate = self.exchange_rate
        if not customer_usd_price or not total_rmb_amount or total_rmb_amount <= 0:
            return 0
        # 客户总美金 = 客户美金报价 × 汇率（折算成人民币）
        customer_total_rmb = customer_usd_price * exchange_rate
        # 毛利率 = (客户人民币 - 成本人民币) / 客户人民币
        if customer_total_rmb <= 0:
            return 0
        profit_margin = (customer_total_rmb - total_rmb_amount) / customer_total_rmb * 100
        return max(0, profit_margin)  # 不返回负数

    def init_ui(self):
        self.setWindowTitle(f"PI订单管理系统 - {DEPARTMENT_CONFIG[self.dept_id]['name']}")
        self.setMinimumSize(1200, 800)
        # 默认全屏显示
        self.showMaximized()

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(0, 0, 0, 0)

        header = QWidget()
        header.setFixedHeight(70)
        header.setStyleSheet("background-color: #2563eb;")
        header_layout = QHBoxLayout()
        header_layout.setContentsMargins(20, 0, 20, 0)

        title = QLabel(f"📦 PI订单管理系统 - {DEPARTMENT_CONFIG[self.dept_id]['name']}")
        title.setFont(get_font(16, QFont.Weight.Bold))
        title.setStyleSheet("color: white;")
        header_layout.addWidget(title)

        header_layout.addStretch()

        # 用户信息和管理员模式切换
        self.user_info_label = QLabel()
        self.user_info_label.setStyleSheet("color: white; font-size: 14px;")
        header_layout.addWidget(self.user_info_label)
        
        self.admin_mode_label = QLabel()
        self.admin_mode_label.setStyleSheet("color: #fbbf24; font-size: 12px; font-weight: bold;")
        header_layout.addWidget(self.admin_mode_label)
        
        # 退出登录按钮
        logout_btn = QPushButton("退出")
        logout_btn.setStyleSheet("""
            QPushButton {
                background-color: #dc2626;
                color: white;
                border: none;
                padding: 5px 15px;
                border-radius: 4px;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #b91c1c;
            }
        """)
        logout_btn.clicked.connect(self.logout)
        header_layout.addWidget(logout_btn)

        header.setLayout(header_layout)
        main_layout.addWidget(header)

        content = QWidget()
        content_layout = QHBoxLayout()
        content_layout.setContentsMargins(0, 0, 0, 0)

        sidebar = QWidget()
        sidebar.setFixedWidth(200)
        sidebar.setStyleSheet("background-color: #1e293b;")
        sidebar_layout = QVBoxLayout()
        sidebar_layout.setContentsMargins(0, 20, 0, 0)
        sidebar_layout.setSpacing(5)

        self.tab_buttons = {}
        tabs = [
            ("产品管理", "products"),
            ("客户管理", "customers"),
            ("供应商管理", "suppliers"),
            ("报价管理", "quotes"),
            ("PI管理", "pi"),
            ("采购管理", "purchase"),
            ("出货管理", "shipment"),
            ("客户付款", "customer_payment"),
            ("供应商付款", "supplier_payment"),
            ("库存管理", "inventory"),
            ("订单总表", "order_summary")
        ]

        for name, key in tabs:
            btn = QPushButton(name)
            btn.setFixedHeight(45)
            btn.setFont(get_font(10))
            btn.setStyleSheet("""
                QPushButton {
                    color: white;
                    background-color: transparent;
                    border: none;
                    text-align: left;
                    padding-left: 20px;
                }
                QPushButton:hover {
                    background-color: #334155;
                }
                QPushButton:checked {
                    background-color: #2563eb;
                }
            """)
            btn.setCheckable(True)
            btn.clicked.connect(lambda checked, k=key: self.switch_tab(k))
            sidebar_layout.addWidget(btn)
            self.tab_buttons[key] = btn

        self.tab_buttons["products"].setChecked(True)

        sidebar_layout.addStretch()
        
        # 设置按钮
        settings_btn = QPushButton("⚙ 设置")
        settings_btn.setFixedHeight(40)
        settings_btn.setFont(get_font(10))
        settings_btn.setStyleSheet("""
            QPushButton {
                color: #94a3b8;
                background-color: transparent;
                border: none;
                text-align: left;
                padding-left: 20px;
            }
            QPushButton:hover {
                background-color: #334155;
                color: white;
            }
        """)
        settings_btn.clicked.connect(self.open_settings)
        sidebar_layout.addWidget(settings_btn)

        sidebar.setLayout(sidebar_layout)
        content_layout.addWidget(sidebar)

        self.content_stack = QStackedWidget()

        self.products_widget = self.create_products_tab()
        self.customers_widget = self.create_customers_tab()
        self.suppliers_widget = self.create_suppliers_tab()
        self.quotes_widget = self.create_quotes_tab()
        self.pi_widget = self.create_pi_tab()
        self.purchase_widget = self.create_purchase_tab()
        self.shipment_widget = self.create_shipment_tab()
        self.customer_payment_widget = self.create_customer_payment_tab()
        self.supplier_payment_widget = self.create_supplier_payment_tab()
        self.inventory_widget = self.create_inventory_tab()
        self.order_summary_widget = self.create_order_summary_tab()

        self.content_stack.addWidget(self.products_widget)
        self.content_stack.addWidget(self.customers_widget)
        self.content_stack.addWidget(self.suppliers_widget)
        self.content_stack.addWidget(self.quotes_widget)
        self.content_stack.addWidget(self.pi_widget)
        self.content_stack.addWidget(self.purchase_widget)
        self.content_stack.addWidget(self.shipment_widget)
        self.content_stack.addWidget(self.customer_payment_widget)
        self.content_stack.addWidget(self.supplier_payment_widget)
        self.content_stack.addWidget(self.inventory_widget)
        self.content_stack.addWidget(self.order_summary_widget)

        content_layout.addWidget(self.content_stack)
        content.setLayout(content_layout)
        main_layout.addWidget(content)

        central_widget.setLayout(main_layout)

    def switch_tab(self, key):
        print(f"DEBUG - switch_tab called with key: {key}")
        tab_map = {"products": 0, "customers": 1, "suppliers": 2, "quotes": 3, "pi": 4, "purchase": 5,
                   "shipment": 6, "customer_payment": 7, "supplier_payment": 8, "inventory": 9, "order_summary": 10}
        self.content_stack.setCurrentIndex(tab_map.get(key, 0))
        for k, btn in self.tab_buttons.items():
            btn.setChecked(k == key)
        
        # 切换到PI标签时强制刷新（清除缓存）
        if key == "pi":
            print("DEBUG - Loading PI orders...")
            self.load_pi_orders_async(force_refresh=True)
        else:
            refresh_map = {
                "products": self.load_products_async,
                "customers": self.load_customers_async,
                "suppliers": self.load_suppliers_async,
                "quotes": self.load_quotes_async,
                "purchase": self.load_purchase_orders_async,
                "shipment": self.load_shipments_async,
                "customer_payment": self.load_customer_payments_async,
                "supplier_payment": self.load_supplier_payments_async,
                "inventory": self.load_inventories_async,
                "order_summary": self.load_order_summary,
            }
            if key in refresh_map:
                refresh_map[key]()
    
    def _load_async(self, api_method, update_method, error_msg="加载失败"):
        """通用异步加载方法，使用QThread确保UI在主线程更新"""
        from PySide6.QtCore import QThread
        
        class LoaderThread(QThread):
            def __init__(self, api_method, parent=None):
                super().__init__(parent)
                self.api_method = api_method
                self.result_data = []
                self.error_occurred = False
            
            def run(self):
                try:
                    data = self.api_method()
                    self.result_data = data if data else []
                except Exception as e:
                    print(f"{error_msg}: {e}")
                    self.error_occurred = True
                    self.result_data = []
        
        thread = LoaderThread(api_method, self)
        thread.finished.connect(lambda: update_method(thread.result_data))
        thread.start()
        return thread
    
    def load_customers_async(self):
        """异步加载客户数据"""
        self._load_async(
            self.api_client.get_customers,
            self._update_customers_table,
            "加载客户失败"
        )
    
    def _update_customers_table(self, customers):
        """在主线程更新客户表格"""
        if not customers:
            customers = []
        self.customers_table.setRowCount(len(customers))
        for row, c in enumerate(customers):
            checkbox = QCheckBox()
            checkbox.setStyleSheet("margin-left: 50%;")
            self.customers_table.setCellWidget(row, 0, checkbox)
            self.customers_table.setItem(row, 1, QTableWidgetItem(str(c.get('id', ''))))
            self.customers_table.setItem(row, 2, QTableWidgetItem(c.get('customer_code', '')))
            self.customers_table.setItem(row, 3, QTableWidgetItem(c.get('customer_name', '')))
            self.customers_table.setItem(row, 4, QTableWidgetItem(c.get('country', '')))
            self.customers_table.setItem(row, 5, QTableWidgetItem(""))
            self.customers_table.setItem(row, 6, QTableWidgetItem(""))
            self.customers_table.setItem(row, 7, QTableWidgetItem(c.get('payment_terms', '')))
            status = c.get('status', 1)
            status_text = "启用" if status == 1 else "禁用"
            status_color = "#10b981" if status == 1 else "#ef4444"
            status_item = QTableWidgetItem(status_text)
            status_item.setForeground(QBrush(QColor(status_color)))
            self.customers_table.setItem(row, 8, status_item)
            action_bar = ActionBarFactory.create_customer_action_bar(
                edit_callback=lambda _, c=c: self.edit_customer(c),
                toggle_callback=lambda _, c=c: self.toggle_customer_status(c),
                status=status
            )
            self.customers_table.setCellWidget(row, 9, action_bar)
        self.customer_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
        # 异步加载联系人和地址
        QTimer.singleShot(0, lambda: self._load_customer_extra_info(customers))
    
    def load_suppliers_async(self):
        """异步加载供应商数据"""
        self._load_async(
            self.api_client.get_suppliers,
            self._update_suppliers_table,
            "加载供应商失败"
        )
    
    def _update_suppliers_table(self, suppliers):
        """在主线程更新供应商表格"""
        if not suppliers:
            suppliers = []
        self.suppliers_table.setRowCount(len(suppliers))
        for row, s in enumerate(suppliers):
            checkbox = QCheckBox()
            checkbox.setStyleSheet("margin-left: 15px;")
            self.suppliers_table.setCellWidget(row, 0, checkbox)
            self.suppliers_table.setItem(row, 1, QTableWidgetItem(str(s.get('id', ''))))
            self.suppliers_table.setItem(row, 2, QTableWidgetItem(s.get('supplier_code', '')))
            self.suppliers_table.setItem(row, 3, QTableWidgetItem(s.get('supplier_name', '')))
            self.suppliers_table.setItem(row, 4, QTableWidgetItem(s.get('region', '')))
            self.suppliers_table.setItem(row, 5, QTableWidgetItem(s.get('contact_person', '')))
            self.suppliers_table.setItem(row, 6, QTableWidgetItem(s.get('phone', '')))
            edit_btn = QPushButton("编辑")
            edit_btn.setFixedWidth(60)
            edit_btn.clicked.connect(lambda _, s=s: self.edit_supplier(s))
            self.suppliers_table.setCellWidget(row, 7, edit_btn)

    def create_products_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)

        toolbar = QHBoxLayout()
        
        title = QLabel("产品列表")
        title.setFont(get_font(14, QFont.Weight.Bold))
        toolbar.addWidget(title)
        
        toolbar.addStretch()

        # 搜索框
        search_layout = QHBoxLayout()
        search_layout.setSpacing(10)

        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("搜索（OE号/产品编号/工厂编号/品牌/描述）")
        self.search_input.setFixedWidth(250)
        search_layout.addWidget(self.search_input)

        self.category_filter = QComboBox()
        self.category_filter.addItem("全部分类", 0)
        search_layout.addWidget(self.category_filter)
        
        # 加载产品类别
        self.load_product_categories()

        search_btn = QPushButton("搜索")
        search_btn.clicked.connect(self.search_products)
        search_btn.setStyleSheet("""
            QPushButton {
                background-color: #3b82f6;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 6px 12px;
            }
            QPushButton:hover { background-color: #2563eb; }
        """)
        search_layout.addWidget(search_btn)

        reset_btn = QPushButton("重置")
        reset_btn.clicked.connect(self.reset_search)
        reset_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 6px 12px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        search_layout.addWidget(reset_btn)

        toolbar.addLayout(search_layout)

        add_btn = QPushButton("+ 新增产品")
        add_btn.clicked.connect(self.add_product)
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        toolbar.addWidget(add_btn)

        refresh_btn = QPushButton("刷新")
        refresh_btn.clicked.connect(self.load_products)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        toolbar.addWidget(refresh_btn)

        import_btn = QPushButton("批量导入")
        import_btn.clicked.connect(self.import_products)
        import_btn.setStyleSheet("""
            QPushButton {
                background-color: #10b981;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #059669; }
        """)
        toolbar.addWidget(import_btn)

        # 批量操作按钮
        batch_layout = QHBoxLayout()
        batch_layout.setSpacing(10)
        
        self.select_all_checkbox = QCheckBox("全选")
        self.select_all_checkbox.stateChanged.connect(self.toggle_select_all_products)
        batch_layout.addWidget(self.select_all_checkbox)
        
        batch_disable_btn = QPushButton("批量禁用")
        batch_disable_btn.clicked.connect(self.batch_toggle_product_status)
        batch_disable_btn.setStyleSheet("""
            QPushButton {
                background-color: #f59e0b;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 6px 12px;
            }
            QPushButton:hover { background-color: #d97706; }
        """)
        batch_layout.addWidget(batch_disable_btn)
        
        batch_delete_btn = QPushButton("批量删除")
        batch_delete_btn.clicked.connect(self.batch_delete_products)
        batch_delete_btn.setStyleSheet("""
            QPushButton {
                background-color: #dc2626;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 6px 12px;
            }
            QPushButton:hover { background-color: #b91c1c; }
        """)
        batch_layout.addWidget(batch_delete_btn)
        
        toolbar.addLayout(batch_layout)

        layout.addLayout(toolbar)

        self.products_table = QTableWidget()
        self.products_table.setColumnCount(11)
        self.products_table.setHorizontalHeaderLabels([
            "", "客户产品编号", "OE号", "图片", "产品名称", 
            "客户型号", "客户号", "产品特性", "数量", "报价", "编辑"
        ])
        self.products_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.products_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.products_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        # 设置图片列宽度和行高
        self.products_table.setColumnWidth(3, 70)
        self.products_table.setColumnWidth(0, 30)  # 复选框列（无表头，仅显示checkbox）
        self.products_table.setColumnWidth(10, 50)  # 编辑列
        self.products_table.verticalHeader().setDefaultSectionSize(70)
        self.products_table.doubleClicked.connect(self.on_product_double_click)
        self.setup_table_context_menu(self.products_table, ["", "客户产品编号", "OE号", "图片", "产品名称", "客户型号", "客户号", "产品特性", "数量", "报价", "编辑"])
        layout.addWidget(self.products_table)

        widget.setLayout(layout)
        return widget

    def create_customers_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)

        toolbar = QHBoxLayout()
        title = QLabel("客户列表")
        title.setFont(get_font(14, QFont.Weight.Bold))
        toolbar.addWidget(title)
        toolbar.addStretch()

        self.customer_select_all_checkbox = QCheckBox("全选")
        self.customer_select_all_checkbox.clicked.connect(self.toggle_select_all_customers)
        toolbar.addWidget(self.customer_select_all_checkbox)

        self.customer_search_input = QLineEdit()
        self.customer_search_input.setPlaceholderText("搜索客户名称/编号...")
        self.customer_search_input.setFixedWidth(200)
        self.customer_search_input.returnPressed.connect(self.search_customers)
        toolbar.addWidget(self.customer_search_input)

        self.customer_country_filter = QComboBox()
        self.customer_country_filter.addItem("全部国家", 0)
        self.customer_country_filter.setFixedWidth(150)
        toolbar.addWidget(self.customer_country_filter)

        search_btn = QPushButton("搜索")
        search_btn.clicked.connect(self.search_customers)
        search_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        toolbar.addWidget(search_btn)

        add_btn = QPushButton("+ 新增客户")
        add_btn.clicked.connect(self.add_customer)
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        toolbar.addWidget(add_btn)

        refresh_btn = QPushButton("刷新")
        refresh_btn.clicked.connect(self.load_customers)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        toolbar.addWidget(refresh_btn)

        layout.addLayout(toolbar)

        self.customers_table = QTableWidget()
        self.customers_table.setColumnCount(10)
        self.customers_table.setHorizontalHeaderLabels(["选择", "ID", "客户编号", "客户名称", "国家", "默认联系人", "默认地址", "付款条款", "状态", "操作"])
        self.customers_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.customers_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.customers_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.customers_table.doubleClicked.connect(self.on_customer_double_click)
        self.setup_table_context_menu(self.customers_table, ["选择", "ID", "客户编号", "客户名称", "国家", "默认联系人", "默认地址", "付款条款", "状态", "操作"])
        layout.addWidget(self.customers_table)

        widget.setLayout(layout)
        return widget

    def create_suppliers_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)

        toolbar = QHBoxLayout()
        title = QLabel("供应商列表")
        title.setFont(get_font(14, QFont.Weight.Bold))
        toolbar.addWidget(title)
        toolbar.addStretch()

        self.select_all_checkbox = QCheckBox("全选")
        self.select_all_checkbox.clicked.connect(self.toggle_select_all_suppliers)
        toolbar.addWidget(self.select_all_checkbox)

        delete_btn = QPushButton("删除选中")
        delete_btn.clicked.connect(self.delete_selected_suppliers)
        delete_btn.setStyleSheet("""
            QPushButton {
                background-color: #dc2626;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #b91c1c; }
        """)
        toolbar.addWidget(delete_btn)

        import_btn = QPushButton("批量导入")
        import_btn.clicked.connect(self.import_suppliers)
        import_btn.setStyleSheet("""
            QPushButton {
                background-color: #16a34a;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #15803d; }
        """)
        toolbar.addWidget(import_btn)

        add_btn = QPushButton("+ 新增供应商")
        add_btn.clicked.connect(self.add_supplier)
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        toolbar.addWidget(add_btn)

        refresh_btn = QPushButton("刷新")
        refresh_btn.clicked.connect(self.load_suppliers)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        toolbar.addWidget(refresh_btn)

        layout.addLayout(toolbar)

        self.suppliers_table = QTableWidget()
        self.suppliers_table.setColumnCount(8)
        self.suppliers_table.setHorizontalHeaderLabels(["", "ID", "供应商编号", "供应商名称", "地区", "联系人", "电话", "操作"])
        self.suppliers_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.suppliers_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.suppliers_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.suppliers_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)
        self.suppliers_table.setColumnWidth(0, 40)
        self.suppliers_table.doubleClicked.connect(self.on_supplier_double_click)
        self.setup_table_context_menu(self.suppliers_table, ["", "ID", "供应商编号", "供应商名称", "地区", "联系人", "电话", "操作"])
        layout.addWidget(self.suppliers_table)

        widget.setLayout(layout)
        return widget

    def toggle_select_all_suppliers(self, checked):
        for row in range(self.suppliers_table.rowCount()):
            checkbox = self.suppliers_table.cellWidget(row, 0)
            if checkbox:
                checkbox.setChecked(checked)

    def delete_selected_suppliers(self):
        selected_ids = []
        for row in range(self.suppliers_table.rowCount()):
            checkbox = self.suppliers_table.cellWidget(row, 0)
            if checkbox and checkbox.isChecked():
                supplier_id = int(self.suppliers_table.item(row, 1).text())
                selected_ids.append(supplier_id)

        if not selected_ids:
            QMessageBox.warning(self, "提示", "请先选择要删除的供应商")
            return

        reply = QMessageBox.question(self, "确认删除", f"确定要删除选中的 {len(selected_ids)} 个供应商吗？",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.No:
            return

        for supplier_id in selected_ids:
            try:
                self.api_client.delete_supplier(supplier_id)
            except Exception as e:
                QMessageBox.warning(self, "错误", f"删除供应商失败: {str(e)}")

        self.load_suppliers()
        self.select_all_checkbox.setChecked(False)

    def create_pi_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)

        toolbar = QHBoxLayout()
        title = QLabel("PI订单列表")
        title.setFont(get_font(14, QFont.Weight.Bold))
        toolbar.addWidget(title)
        toolbar.addStretch()

        add_btn = QPushButton("+ 新建PI")
        add_btn.clicked.connect(self.add_pi)
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        toolbar.addWidget(add_btn)
        
        self.pi_batch_delete_btn = QPushButton("批量删除")
        self.pi_batch_delete_btn.clicked.connect(self.batch_delete_pi)
        self.pi_batch_delete_btn.setStyleSheet("""
            QPushButton {
                background-color: #dc2626;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #b91c1c; }
        """)
        toolbar.addWidget(self.pi_batch_delete_btn)
        
        self.pi_batch_export_btn = QPushButton("批量导出")
        self.pi_batch_export_btn.clicked.connect(self.batch_export_pi)
        self.pi_batch_export_btn.setStyleSheet("""
            QPushButton {
                background-color: #059669;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #047857; }
        """)
        toolbar.addWidget(self.pi_batch_export_btn)
        
        refresh_btn = QPushButton("刷新")
        refresh_btn.clicked.connect(self.load_pi_orders_async)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        toolbar.addWidget(refresh_btn)

        layout.addLayout(toolbar)

        self.pi_table = QTableWidget()
        self.pi_table.setColumnCount(10)
        self.pi_table.setHorizontalHeaderLabels(["", "ID", "PI号", "金额", "币种", "状态", "创建时间", "操作", "完成", "导出"])
        self.pi_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.pi_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.pi_table.setColumnWidth(0, 40)
        self.pi_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.pi_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self.pi_table.doubleClicked.connect(self.on_pi_double_click)
        self.setup_table_context_menu(self.pi_table, ["", "ID", "PI号", "金额", "币种", "状态", "创建时间", "操作", "完成", "导出"])
        layout.addWidget(self.pi_table)

        widget.setLayout(layout)
        return widget

    def create_purchase_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)

        toolbar = QHBoxLayout()
        title = QLabel("采购订单列表")
        title.setFont(get_font(14, QFont.Weight.Bold))
        toolbar.addWidget(title)
        toolbar.addStretch()

        add_btn = QPushButton("+ 新建采购单")
        add_btn.clicked.connect(self.add_purchase)
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        toolbar.addWidget(add_btn)

        refresh_btn = QPushButton("刷新")
        refresh_btn.clicked.connect(self.load_purchase_orders)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        toolbar.addWidget(refresh_btn)

        layout.addLayout(toolbar)

        self.purchase_table = QTableWidget()
        self.purchase_table.setColumnCount(9)
        self.purchase_table.setHorizontalHeaderLabels(["ID", "采购单号", "PI号", "供应商", "金额", "状态", "操作", "确认", "入库"])
        self.purchase_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.purchase_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.purchase_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.setup_table_context_menu(self.purchase_table, ["ID", "采购单号", "PI号", "供应商", "金额", "状态", "操作"])
        layout.addWidget(self.purchase_table)

        widget.setLayout(layout)
        return widget

    def create_quotes_tab(self):
        """报价管理标签页"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)

        toolbar = QHBoxLayout()
        title = QLabel("报价管理")
        title.setFont(get_font(14, QFont.Weight.Bold))
        toolbar.addWidget(title)
        toolbar.addStretch()

        add_btn = QPushButton("+ 新建报价单")
        add_btn.clicked.connect(self.add_quote)
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        toolbar.addWidget(add_btn)
        
        batch_delete_btn = QPushButton("批量删除")
        batch_delete_btn.clicked.connect(self.batch_delete_quotes)
        batch_delete_btn.setStyleSheet("""
            QPushButton {
                background-color: #dc2626;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #b91c1c; }
        """)
        toolbar.addWidget(batch_delete_btn)
        
        batch_export_btn = QPushButton("批量导出")
        batch_export_btn.clicked.connect(self.batch_export_quotes)
        batch_export_btn.setStyleSheet("""
            QPushButton {
                background-color: #059669;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #047857; }
        """)
        toolbar.addWidget(batch_export_btn)
        
        refresh_btn = QPushButton("刷新")
        refresh_btn.clicked.connect(self.load_quotes_async)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #6b7280;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #4b5563; }
        """)
        toolbar.addWidget(refresh_btn)

        layout.addLayout(toolbar)

        # 报价单列表
        self.quote_table = QTableWidget()
        self.quote_table.setColumnCount(10)
        self.quote_table.setHorizontalHeaderLabels(["", "ID", "报价单号", "客户", "金额", "币种", "状态", "有效期", "备注", "操作"])
        self.quote_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.quote_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.quote_table.setColumnHidden(1, True)  # 隐藏ID列
        self.quote_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.quote_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.quote_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self.quote_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.quote_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        self.quote_table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        self.quote_table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
        self.quote_table.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeMode.ResizeToContents)
        self.quote_table.horizontalHeader().setSectionResizeMode(8, QHeaderView.ResizeMode.Stretch)
        self.quote_table.horizontalHeader().setSectionResizeMode(9, QHeaderView.ResizeMode.ResizeToContents)
        layout.addWidget(self.quote_table)

        widget.setLayout(layout)
        return widget

    def load_quotes_async(self):
        """异步加载报价单"""
        self._load_async(
            self.api_client.get_quotes,
            self._update_quote_table,
            "加载报价单失败"
        )

    def _update_quote_table(self, quotes):
        """更新报价单表格"""
        try:
            self.quote_table.setRowCount(len(quotes))
            status_map = {1: "草稿", 2: "已发送", 3: "已接受", 4: "已拒绝"}
            
            for row, q in enumerate(quotes):
                # 复选框
                checkbox = QTableWidgetItem()
                checkbox.setFlags(checkbox.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                checkbox.setCheckState(Qt.CheckState.Unchecked)
                checkbox.setData(Qt.ItemDataRole.UserRole, q.get('id'))
                self.quote_table.setItem(row, 0, checkbox)
                
                self.quote_table.setItem(row, 1, QTableWidgetItem(str(q.get('id', ''))))
                self.quote_table.setItem(row, 2, QTableWidgetItem(q.get('quote_no', '')))
                self.quote_table.setItem(row, 3, QTableWidgetItem(q.get('customer_name', '')))
                self.quote_table.setItem(row, 4, QTableWidgetItem(f"${q.get('total_amount', 0):,.2f}"))
                self.quote_table.setItem(row, 5, QTableWidgetItem(q.get('currency', 'USD')))
                self.quote_table.setItem(row, 6, QTableWidgetItem(status_map.get(q.get('status', 1), '草稿')))
                valid_until = q.get('valid_until')
                if valid_until:
                    valid_until = str(valid_until)[:10]
                self.quote_table.setItem(row, 7, QTableWidgetItem(valid_until or '-'))
                self.quote_table.setItem(row, 8, QTableWidgetItem(q.get('remark', '') or '-'))
                
                # 操作按钮
                btn_widget = QWidget()
                btn_layout = QHBoxLayout()
                btn_layout.setContentsMargins(0, 0, 0, 0)
                
                edit_btn = QPushButton("编辑")
                edit_btn.setFixedWidth(50)
                edit_btn.clicked.connect(lambda _, qid=q.get('id'): self.edit_quote(qid))
                btn_layout.addWidget(edit_btn)
                
                pi_btn = QPushButton("转PI")
                pi_btn.setFixedWidth(50)
                pi_btn.setStyleSheet("color: #10b981;")
                pi_btn.clicked.connect(lambda _, qid=q.get('id'): self.convert_quote_to_pi(qid))
                btn_layout.addWidget(pi_btn)
                
                btn_widget.setLayout(btn_layout)
                self.quote_table.setCellWidget(row, 9, btn_widget)
        except Exception as e:
            print(f"更新报价单表格失败: {e}")
        finally:
            self.quote_table.viewport().update()
    
    def get_selected_quote_ids(self):
        """获取选中的报价单ID列表"""
        ids = []
        for row in range(self.quote_table.rowCount()):
            item = self.quote_table.item(row, 0)
            if item and item.checkState() == Qt.CheckState.Checked:
                quote_id = item.data(Qt.ItemDataRole.UserRole)
                if quote_id:
                    ids.append(quote_id)
        return ids
    
    def toggle_all_quotes(self, state):
        """全选/取消全选"""
        check_state = Qt.CheckState.Checked if state else Qt.CheckState.Unchecked
        for row in range(self.quote_table.rowCount()):
            item = self.quote_table.item(row, 0)
            if item:
                item.setCheckState(check_state)
    
    def batch_delete_quotes(self):
        """批量删除报价单"""
        selected_ids = self.get_selected_quote_ids()
        if not selected_ids:
            QMessageBox.warning(self, "提示", "请先选择要删除的报价单")
            return
        
        reply = QMessageBox.question(
            self, "确认删除", 
            f"确定要删除选中的 {len(selected_ids)} 个报价单吗？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        try:
            result = self.api_client.batch_delete_quotes(selected_ids)
            deleted = result.get('deleted', 0)
            errors = result.get('errors', [])
            if errors:
                QMessageBox.warning(self, "部分删除失败", f"成功删除 {deleted} 个\n失败: {len(errors)} 个\n{errors}")
            else:
                QMessageBox.information(self, "成功", f"已删除 {deleted} 个报价单")
            self.load_quotes_async()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"删除失败: {str(e)}")
            self.load_quotes_async()
    
    def batch_export_quotes(self):
        """批量导出报价单"""
        selected_ids = self.get_selected_quote_ids()
        if not selected_ids:
            QMessageBox.warning(self, "提示", "请先选择要导出的报价单")
            return
        
        try:
            quotes_data = []
            for quote_id in selected_ids:
                quote = self.api_client.get_quote(quote_id)
                quotes_data.append(quote)
            
            if not quotes_data:
                QMessageBox.information(self, "提示", "没有可导出的数据")
                return
            
            # 构建导出数据
            export_rows = []
            for q in quotes_data:
                for item in q.get('items', []):
                    export_rows.append({
                        '报价单号': q.get('quote_no', ''),
                        '客户': q.get('customer_name', ''),
                        '币种': q.get('currency', 'USD'),
                        '总金额': q.get('total_amount', 0),
                        '有效期': q.get('valid_until', ''),
                        '状态': ['草稿', '已发送', '已接受', '已拒绝'][q.get('status', 1) - 1] if q.get('status', 1) <= 4 else '',
                        '备注': q.get('remark', ''),
                        '产品编号': item.get('product_id', ''),
                        'OE号': item.get('oe_number', ''),
                        '客户编号': item.get('customer_code', ''),
                        '产品描述': item.get('detail_desc', ''),
                        '数量': item.get('quantity', 0),
                        '单价': item.get('unit_price', 0),
                        '总价': item.get('total_price', 0),
                        '明细备注': item.get('remark', ''),
                    })
            
            # 创建DataFrame并导出
            df = pd.DataFrame(export_rows)
            
            # 选择保存路径
            from PySide6.QtWidgets import QFileDialog
            file_path, _ = QFileDialog.getSaveFileName(
                self, "保存报价单", 
                f"报价单导出_{len(selected_ids)}个.xlsx",
                "Excel Files (*.xlsx)"
            )
            if file_path:
                df.to_excel(file_path, index=False, engine='openpyxl')
                QMessageBox.information(self, "成功", f"已导出 {len(export_rows)} 条明细到:\n{file_path}")
        except Exception as e:
            QMessageBox.warning(self, "错误", f"导出失败: {str(e)}")

    def add_quote(self):
        """新建报价单"""
        dialog = QuoteDialog(self, self.api_client, self.dept_id)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            QMessageBox.information(self, "成功", "报价单已保存")
            self.load_quotes_async()

    def edit_quote(self, quote_id):
        """编辑报价单"""
        try:
            quote = self.api_client.get_quote(quote_id)
            dialog = QuoteDialog(self, self.api_client, self.dept_id, quote)
            if dialog.exec() == QDialog.DialogCode.Accepted:
                self.load_quotes_async()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"获取报价单失败: {str(e)}")

    def convert_quote_to_pi(self, quote_id):
        """将报价单转为PI"""
        reply = QMessageBox.question(self, "确认", "确定要将此报价单转为PI吗？")
        if reply == QMessageBox.StandardButton.Yes:
            try:
                result = self.api_client.convert_quote_to_pi(quote_id)
                QMessageBox.information(self, "成功", f"报价单已转为PI\nPI单号: {result.get('pi_no', '')}")
                self.load_quotes_async()
            except Exception as e:
                QMessageBox.warning(self, "错误", f"转换失败: {str(e)}")

    def create_shipment_tab(self):
        """出货管理标签页 - 主从表设计（参考库存管理）"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)

        toolbar = QHBoxLayout()
        title = QLabel("出货管理")
        title.setFont(get_font(14, QFont.Weight.Bold))
        toolbar.addWidget(title)
        toolbar.addStretch()

        add_btn = QPushButton("+ 新建出货")
        add_btn.clicked.connect(self.add_shipment)
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        toolbar.addWidget(add_btn)

        refresh_btn = QPushButton("刷新")
        refresh_btn.clicked.connect(self.load_shipments_async)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        toolbar.addWidget(refresh_btn)

        layout.addLayout(toolbar)

        # 主表：出货汇总
        self.shipment_table = QTableWidget()
        self.shipment_table.setColumnCount(9)
        self.shipment_table.setHorizontalHeaderLabels(["ID", "PI号", "总金额", "总箱数", "付款状态", "出货状态", "阶段数", "操作", "确认"])
        self.shipment_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.shipment_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.shipment_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.setup_table_context_menu(self.shipment_table, ["ID", "PI号", "总金额", "总箱数", "付款状态", "出货状态", "阶段数", "操作"])
        layout.addWidget(self.shipment_table)

        # 详情标签
        detail_label = QLabel("📋 出货阶段详情（请点击上方记录查看）")
        detail_label.setFont(get_font(12, QFont.Weight.Bold))
        layout.addWidget(detail_label)

        # 从表：出货阶段明细
        self.shipment_stage_table = QTableWidget()
        self.shipment_stage_table.setColumnCount(9)
        self.shipment_stage_table.setHorizontalHeaderLabels(["阶段", "出货日期", "柜号", "提单号", "数量", "库存", "存放位置", "付款状态", "操作"])
        self.shipment_stage_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.shipment_stage_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.shipment_stage_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.shipment_stage_table.setMaximumHeight(250)
        layout.addWidget(self.shipment_stage_table)

        # 绑定点击事件
        try:
            self.shipment_table.cellClicked.disconnect()
        except RuntimeError:
            pass
        self.shipment_table.cellClicked.connect(self.show_shipment_stages)

        widget.setLayout(layout)
        return widget

    def create_customer_payment_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)

        toolbar = QHBoxLayout()
        title = QLabel("客户付款列表")
        title.setFont(get_font(14, QFont.Weight.Bold))
        toolbar.addWidget(title)
        toolbar.addStretch()

        add_btn = QPushButton("+ 新建付款记录")
        add_btn.clicked.connect(self.add_customer_payment)
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        toolbar.addWidget(add_btn)

        refresh_btn = QPushButton("刷新")
        refresh_btn.clicked.connect(self.load_customer_payments)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        toolbar.addWidget(refresh_btn)

        layout.addLayout(toolbar)

        self.customer_payment_table = QTableWidget()
        self.customer_payment_table.setColumnCount(6)
        self.customer_payment_table.setHorizontalHeaderLabels(["ID", "PI号", "付款日期", "金额", "付款方式", "操作"])
        self.customer_payment_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.customer_payment_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.customer_payment_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.setup_table_context_menu(self.customer_payment_table, ["ID", "PI号", "付款日期", "金额", "付款方式", "操作"])
        layout.addWidget(self.customer_payment_table)

        widget.setLayout(layout)
        return widget

    def create_supplier_payment_tab(self):
        """供应商付款标签页 - 主从表设计（参考库存管理）"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)

        toolbar = QHBoxLayout()
        title = QLabel("供应商付款管理")
        title.setFont(get_font(14, QFont.Weight.Bold))
        toolbar.addWidget(title)
        toolbar.addStretch()

        add_btn = QPushButton("+ 新建付款记录")
        add_btn.clicked.connect(self.add_supplier_payment)
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        toolbar.addWidget(add_btn)

        refresh_btn = QPushButton("刷新")
        refresh_btn.clicked.connect(self.load_supplier_payments_async)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        toolbar.addWidget(refresh_btn)

        layout.addLayout(toolbar)

        # 主表：付款汇总
        self.supplier_payment_table = QTableWidget()
        self.supplier_payment_table.setColumnCount(8)
        self.supplier_payment_table.setHorizontalHeaderLabels(["ID", "供应商", "采购单", "总金额", "已付金额", "未付金额", "状态", "操作"])
        self.supplier_payment_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.supplier_payment_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.supplier_payment_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.setup_table_context_menu(self.supplier_payment_table, ["ID", "供应商", "采购单", "总金额", "已付金额", "未付金额", "状态", "操作"])
        layout.addWidget(self.supplier_payment_table)

        # 详情标签
        detail_label = QLabel("📋 付款阶段详情（请点击上方记录查看）")
        detail_label.setFont(get_font(12, QFont.Weight.Bold))
        layout.addWidget(detail_label)

        # 从表：付款阶段明细
        self.supplier_payment_stage_table = QTableWidget()
        self.supplier_payment_stage_table.setColumnCount(7)
        self.supplier_payment_stage_table.setHorizontalHeaderLabels(["阶段", "应付金额", "已付金额", "状态", "付款日期", "凭证", "操作"])
        self.supplier_payment_stage_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.supplier_payment_stage_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.supplier_payment_stage_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.supplier_payment_stage_table.setMaximumHeight(200)
        layout.addWidget(self.supplier_payment_stage_table)

        # 绑定点击事件
        try:
            self.supplier_payment_table.cellClicked.disconnect()
        except RuntimeError:
            pass
        self.supplier_payment_table.cellClicked.connect(self.show_supplier_payment_stages)

        widget.setLayout(layout)
        return widget

    def create_inventory_tab(self):
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)

        toolbar = QHBoxLayout()
        title = QLabel("库存管理（按OE号分组）")
        title.setFont(get_font(14, QFont.Weight.Bold))
        toolbar.addWidget(title)
        toolbar.addStretch()

        add_btn = QPushButton("+ 新建库存")
        add_btn.clicked.connect(self.add_inventory)
        add_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        toolbar.addWidget(add_btn)

        refresh_btn = QPushButton("刷新")
        refresh_btn.clicked.connect(self.load_inventories)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        toolbar.addWidget(refresh_btn)

        layout.addLayout(toolbar)

        # 搜索框
        search_layout = QHBoxLayout()
        self.inventory_search_input = QLineEdit()
        self.inventory_search_input.setPlaceholderText("搜索OE号...")
        self.inventory_search_input.setFixedHeight(35)
        self.inventory_search_input.returnPressed.connect(self.search_inventory)
        search_layout.addWidget(self.inventory_search_input)
        
        search_btn = QPushButton("搜索")
        search_btn.clicked.connect(self.search_inventory)
        search_btn.setStyleSheet("""
            QPushButton {
                background-color: #3b82f6;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
        """)
        search_layout.addWidget(search_btn)
        layout.addLayout(search_layout)

        # 主表格：显示OE号（产品维度）
        self.inventory_product_table = QTableWidget()
        self.inventory_product_table.setColumnCount(10)
        self.inventory_product_table.setHorizontalHeaderLabels(["OE号", "产品编号", "总库存", "供应商", "客户", "状态分布", "最近入库供应商", "最近出库客户", "最近变更", "操作"])
        self.inventory_product_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)
        self.inventory_product_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)
        self.inventory_product_table.setColumnWidth(0, 150)
        self.inventory_product_table.setColumnWidth(1, 110)
        self.inventory_product_table.setColumnWidth(2, 60)
        self.inventory_product_table.setColumnWidth(3, 60)
        self.inventory_product_table.setColumnWidth(4, 60)
        self.inventory_product_table.setColumnWidth(5, 100)
        self.inventory_product_table.setColumnWidth(6, 100)
        self.inventory_product_table.setColumnWidth(7, 100)
        self.inventory_product_table.setColumnWidth(8, 120)
        self.inventory_product_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.inventory_product_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        layout.addWidget(QLabel("📦 产品库存汇总（双击展开查看详情）"))
        layout.addWidget(self.inventory_product_table)
        
        # 分隔线
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setStyleSheet("background-color: #e5e7eb;")
        line.setFixedHeight(2)
        layout.addWidget(line)
        
        # 子表格：显示选中OE号的详细库存记录
        self.inventory_detail_label = QLabel("📋 库存详情（请先选择上方产品）")
        layout.addWidget(self.inventory_detail_label)
        
        self.inventory_detail_table = QTableWidget()
        self.inventory_detail_table.setColumnCount(9)
        self.inventory_detail_table.setHorizontalHeaderLabels(["ID", "供应商", "客户", "数量", "库位", "状态", "备注", "创建时间", "操作"])
        self.inventory_detail_table.setColumnWidth(0, 40)
        self.inventory_detail_table.setColumnWidth(1, 100)
        self.inventory_detail_table.setColumnWidth(2, 100)
        self.inventory_detail_table.setColumnWidth(3, 60)
        self.inventory_detail_table.setColumnWidth(4, 80)
        self.inventory_detail_table.setColumnWidth(5, 50)
        self.inventory_detail_table.setColumnWidth(6, 120)
        self.inventory_detail_table.setColumnWidth(7, 130)
        self.inventory_detail_table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeMode.Stretch)
        self.inventory_detail_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.inventory_detail_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        layout.addWidget(self.inventory_detail_table)

        status_label = QLabel("● 黄色: 采购在途 | ● 蓝色: 待入库 | ● 绿色: 已入库 | ● 黑色: 历史库存")
        status_label.setStyleSheet("color: #6b7280; padding: 5px;")
        layout.addWidget(status_label)

        widget.setLayout(layout)
        return widget

    def load_data(self):
        """异步加载所有数据，避免阻塞UI"""
        self.update_user_info()
        
        # 显示预加载状态提示
        self._show_loading_tip("正在同步服务器数据...")
        
        # 所有模块都使用异步加载，同时触发
        QTimer.singleShot(0, self.load_products_async)
        QTimer.singleShot(0, self.load_customers_async)
        QTimer.singleShot(0, self.load_suppliers_async)
        QTimer.singleShot(0, self.load_pi_orders_async)
        QTimer.singleShot(0, self.load_purchase_orders_async)
        QTimer.singleShot(0, self.load_shipments_async)
        QTimer.singleShot(0, self.load_customer_payments_async)
        QTimer.singleShot(0, self.load_supplier_payments_async)
        QTimer.singleShot(0, self.load_inventories_async)
    
    def _show_loading_tip(self, message: str):
        """显示加载提示（状态栏或临时覆盖层）"""
        # 在主窗口底部状态栏显示提示
        if not hasattr(self, '_status_label'):
            self._status_label = QLabel()
            self._status_label.setStyleSheet("""
                QLabel {
                    background-color: #2563eb;
                    color: white;
                    padding: 8px 20px;
                    font-size: 12px;
                }
            """)
            self.statusBar().addPermanentWidget(self._status_label)
        self._status_label.setText(f"  ⏳ {message} ")
        # 3秒后自动隐藏
        QTimer.singleShot(3000, self._hide_loading_tip)
    
    def _hide_loading_tip(self):
        """隐藏加载提示"""
        if hasattr(self, '_status_label'):
            self._status_label.setText("")
    
    def _test_customer_reply(self):
        """测试客户回复API"""
        from PySide6.QtWidgets import QMessageBox
        
        if not HAS_CUSTOMER_REPLY_TEST or not CustomerReplyTester:
            QMessageBox.information(self, "提示", "测试模块未导入成功")
            return
        
        print("\n" + "="*60)
        print("开始测试客户回复API...")
        print("="*60)
        
        try:
            tester = CustomerReplyTester(self.api_client)
            results = tester.run_all_tests()
            tester.print_summary()
            
            # 显示结果
            msg = f"测试完成!\n\n通过: {results['passed']}/{results['total']}"
            if results['failed'] > 0:
                msg += f"\n失败: {results['failed']}"
            QMessageBox.information(self, "测试结果", msg)
        except Exception as e:
            QMessageBox.warning(self, "测试失败", f"测试过程中出错:\n{str(e)}")
            import traceback
            traceback.print_exc()
    
    def _update_work_status(self, task_id: str, text: str, progress: int = -1):
        """更新右下角工作状态"""
        if not hasattr(self, '_work_status_manager'):
            self._work_status_manager = {}
        
        self._work_status_manager[task_id] = {
            "text": text,
            "progress": progress,
            "timestamp": time.time()
        }
        
        self._refresh_work_status_display()
    
    def _remove_work_status(self, task_id: str):
        """移除工作状态"""
        if hasattr(self, '_work_status_manager') and task_id in self._work_status_manager:
            del self._work_status_manager[task_id]
            self._refresh_work_status_display()
    
    def _refresh_work_status_display(self):
        """刷新工作状态显示"""
        if not hasattr(self, '_work_status_label'):
            self._work_status_label = QLabel()
            self._work_status_label.setStyleSheet("""
                QLabel {
                    background-color: #059669;
                    color: white;
                    padding: 5px 15px;
                    font-size: 11px;
                }
            """)
            self.statusBar().insertPermanentWidget(1, self._work_status_label)
        
        if not hasattr(self, '_work_status_manager') or not self._work_status_manager:
            self._work_status_label.setText("  ✓ 就绪 ")
            return
        
        parts = []
        for item in self._work_status_manager.values():
            text = item["text"]
            progress = item["progress"]
            if progress >= 0:
                parts.append(f"{text} ({progress}%)")
            else:
                parts.append(text)
        
        # 限制显示长度
        display_text = " | ".join(parts[:3])
        if len(parts) > 3:
            display_text += f" (+{len(parts)-3} 更多)"
        
        self._work_status_label.setText(f"  🔄 {display_text} ")
        
        # 清理过期状态（超过5分钟）
        now = time.time()
        expired = [k for k, v in self._work_status_manager.items() 
                   if now - v["timestamp"] > 300]
        for k in expired:
            del self._work_status_manager[k]
        if expired:
            self._refresh_work_status_display()

    def update_user_info(self):
        """更新用户信息显示"""
        if hasattr(self.api_client, 'current_user') and self.api_client.current_user:
            user = self.api_client.current_user
            self.user_info_label.setText(f"👤 {user.get('real_name', '用户')}")
            
            if user.get('is_admin'):
                self.admin_mode_label.setText("🔑 管理员模式")
            else:
                self.admin_mode_label.setText("👤 普通用户")
        else:
            self.user_info_label.setText("👤 未登录")
            self.admin_mode_label.setText("")
    
    def logout(self):
        """退出登录"""
        reply = QMessageBox.question(
            self, 
            "确认退出", 
            "确定要退出登录吗？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            if hasattr(self.api_client, 'logout'):
                self.api_client.logout()
            self.close()
    
    def open_settings(self):
        """打开设置对话框"""
        dialog = SettingsDialog(self.api_client, self)
        if dialog.exec():
            # 刷新全局变量
            self.load_globals()
    
    def create_order_summary_tab(self):
        """创建订单总表Tab - 汇总所有模块数据（上下两部分布局）"""
        widget = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(15, 15, 15, 15)
        
        # ==================== 工具栏 ====================
        toolbar = QHBoxLayout()
        title = QLabel("📊 订单管理总表")
        title.setFont(get_font(14, QFont.Weight.Bold))
        toolbar.addWidget(title)
        toolbar.addStretch()
        
        # 导入按钮
        import_btn = QPushButton("📥 导入")
        import_btn.clicked.connect(self._import_order_summary)
        import_btn.setStyleSheet("""
            QPushButton {
                background-color: #6366f1;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #4f46e5; }
        """)
        toolbar.addWidget(import_btn)
        
        # 刷新按钮
        refresh_btn = QPushButton("🔄 刷新")
        refresh_btn.clicked.connect(self.load_order_summary)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        toolbar.addWidget(refresh_btn)
        
        # 导出按钮
        export_btn = QPushButton("📤 导出")
        export_btn.clicked.connect(self.export_order_summary)
        export_btn.setStyleSheet("""
            QPushButton {
                background-color: #10b981;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #059669; }
        """)
        toolbar.addWidget(export_btn)
        
        layout.addLayout(toolbar)
        
        # ==================== 上半部分：订单详情 ====================
        self.order_detail_group = QGroupBox("📋 订单详情（点击下方列表查看）")
        detail_layout = QVBoxLayout()
        
        # 详情表格 - 保持原有的41列完整格式
        self.order_detail_table = QTableWidget()
        self.order_detail_table.setColumnCount(41)
        self.order_detail_table.setHorizontalHeaderLabels([
            "订单日期", "ORDER NO.", "客户产品编号", "OE号", "客户需求备注",
            "产品名称", "图片", "客户型号", "OE号.1", "数量",
            "报价(USD/RMB)", "合计金额", "客户最新回复", "客户预付款", "待收尾款",
            "预估美金报价", "预估毛利率", "采购价格", "运费", "杂费",
            "总金额", "工厂简称", "店铺链接", "交货日期", "是否已收货",
            "工厂订金", "工厂尾款", "入库操作", "入库数量", "包装方式",
            "采购选项/名称", "产品细节", "工厂编号", "纸箱尺寸", "打包规格",
            "箱数", "预估体积", "整箱毛重", "总重量", "品牌", "开票情况"
        ])
        self.order_detail_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        # 设置关键列宽度（大幅加宽）
        self.order_detail_table.setColumnWidth(0, 80)    # 序号/标记列
        self.order_detail_table.setColumnWidth(1, 200)   # ORDER NO.
        self.order_detail_table.setColumnWidth(2, 200)   # 客户产品编号/客户
        self.order_detail_table.setColumnWidth(3, 180)   # OE号/总金额
        self.order_detail_table.setColumnWidth(4, 250)   # 备注/状态
        self.order_detail_table.setColumnWidth(5, 250)   # 产品名称/预付款
        self.order_detail_table.setColumnWidth(6, 100)  # 图片列
        self.order_detail_table.setColumnWidth(7, 180)   # 客户型号
        self.order_detail_table.setColumnWidth(8, 180)   # OE号.1
        self.order_detail_table.setColumnWidth(9, 100)  # 数量列
        self.order_detail_table.setColumnWidth(10, 120) # 报价列
        self.order_detail_table.setColumnWidth(11, 150) # 合计金额
        self.order_detail_table.setColumnWidth(12, 300) # 客户最新回复
        self.order_detail_table.setColumnWidth(13, 150) # 客户预付款
        self.order_detail_table.setColumnWidth(14, 150) # 待收尾款
        self.order_detail_table.setColumnWidth(15, 150) # 预估美金报价
        self.order_detail_table.setColumnWidth(16, 120) # 预估毛利率
        self.order_detail_table.setColumnWidth(17, 150) # 采购价格
        self.order_detail_table.setColumnWidth(18, 120) # 运费
        self.order_detail_table.setColumnWidth(19, 100) # 杂费
        self.order_detail_table.setColumnWidth(20, 150) # 总金额
        self.order_detail_table.setColumnWidth(21, 180) # 工厂简称
        self.order_detail_table.setColumnWidth(22, 300) # 店铺链接
        self.order_detail_table.setColumnWidth(23, 150) # 交货日期
        self.order_detail_table.setColumnWidth(24, 100) # 是否已收货
        self.order_detail_table.setColumnWidth(25, 150) # 工厂订金
        self.order_detail_table.setColumnWidth(26, 150) # 工厂尾款
        self.order_detail_table.setColumnWidth(27, 120) # 入库操作
        self.order_detail_table.setColumnWidth(28, 100) # 入库数量
        self.order_detail_table.setColumnWidth(29, 120) # 包装方式
        self.order_detail_table.setColumnWidth(30, 180) # 采购选项/名称
        self.order_detail_table.setColumnWidth(31, 200) # 产品细节
        self.order_detail_table.setColumnWidth(32, 150) # 工厂编号
        self.order_detail_table.setColumnWidth(33, 150) # 纸箱尺寸
        self.order_detail_table.setColumnWidth(34, 150) # 打包规格
        self.order_detail_table.setColumnWidth(35, 80)  # 箱数
        self.order_detail_table.setColumnWidth(36, 120) # 预估体积
        self.order_detail_table.setColumnWidth(37, 120) # 整箱毛重
        self.order_detail_table.setColumnWidth(38, 120) # 总重量
        self.order_detail_table.setColumnWidth(39, 120) # 品牌
        self.order_detail_table.setColumnWidth(40, 150) # 开票情况
        self.order_detail_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.order_detail_table.setAlternatingRowColors(True)
        self.order_detail_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        # 双击编辑
        self.order_detail_table.cellDoubleClicked.connect(self._on_order_detail_double_click)
        self.order_detail_table.setStyleSheet("""
            QTableWidget {
                background-color: #f9fafb;
                border: 1px solid #e5e7eb;
                border-radius: 4px;
            }
        """)
        detail_layout.addWidget(self.order_detail_table)
        
        # 提示标签
        self.order_detail_hint = QLabel("⬆ 点击下方订单列表中的订单查看详情")
        self.order_detail_hint.setStyleSheet("color: #9ca3af; font-size: 12px; padding: 5px;")
        self.order_detail_hint.setAlignment(Qt.AlignCenter)
        detail_layout.addWidget(self.order_detail_hint)
        
        self.order_detail_group.setLayout(detail_layout)
        layout.addWidget(self.order_detail_group)
        
        # ==================== 分割线 ====================
        splitter = QFrame()
        splitter.setFrameShape(QFrame.Shape.HLine)
        splitter.setStyleSheet("background-color: #e5e7eb; max-height: 1px;")
        layout.addWidget(splitter)
        
        # ==================== 下半部分：订单列表 ====================
        list_group = QGroupBox("📑 订单列表")
        list_layout = QVBoxLayout()
        
        # 搜索栏
        search_layout = QHBoxLayout()
        self.order_summary_search = QLineEdit()
        self.order_summary_search.setPlaceholderText("🔍 搜索订单号/OE号/客户/产品...")
        self.order_summary_search.setFixedHeight(35)
        self.order_summary_search.textChanged.connect(self.filter_order_summary)
        search_layout.addWidget(self.order_summary_search)
        
        # 状态筛选
        self.order_status_filter = QComboBox()
        self.order_status_filter.addItems(["全部状态", "进行中", "已完成", "已取消"])
        self.order_status_filter.currentTextChanged.connect(self.filter_order_summary)
        search_layout.addWidget(self.order_status_filter)
        
        # 清除筛选
        clear_btn = QPushButton("🗑 清除")
        clear_btn.setFixedWidth(70)
        clear_btn.clicked.connect(self._clear_order_summary_filter)
        search_layout.addWidget(clear_btn)
        
        list_layout.addLayout(search_layout)
        
        # 订单列表表格（主索引 - 按订单号分组）
        self.order_list_table = QTableWidget()
        self.order_list_table.setColumnCount(7)
        self.order_list_table.setHorizontalHeaderLabels([
            "选择", "ORDER NO.", "客户", "订单日期", "产品数", "总金额", "状态"
        ])
        self.order_list_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.order_list_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.order_list_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.order_list_table.setAlternatingRowColors(True)
        self.order_list_table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #e5e7eb;
                border-radius: 4px;
            }
        """)
        self.order_list_table.setColumnWidth(0, 50)  # 选择列
        self.order_list_table.setColumnWidth(3, 100)  # 日期列
        self.order_list_table.setColumnWidth(4, 60)   # 产品数列
        self.order_list_table.setColumnWidth(6, 80)   # 状态列
        # 绑定点击事件
        self.order_list_table.cellClicked.connect(self._on_order_list_click)
        # 双击编辑
        self.order_list_table.cellDoubleClicked.connect(self._on_order_list_double_click)
        list_layout.addWidget(self.order_list_table)
        
        # 状态栏
        self._order_summary_status = QLabel("准备就绪")
        self._order_summary_status.setStyleSheet("color: #6b7280; font-size: 12px;")
        list_layout.addWidget(self._order_summary_status)
        
        list_group.setLayout(list_layout)
        layout.addWidget(list_group)
        
        # 存储所有订单的完整数据
        self._order_summary_orders = []
        self._order_summary_filtered = []
        self._selected_order_index = None
        
        widget.setLayout(layout)
        
        # 延迟加载数据
        QTimer.singleShot(100, self.load_order_summary)
        
        return widget
    
    def _on_order_list_click(self, row, column):
        """点击订单列表行，显示详情"""
        if row < 0 or row >= len(self._order_summary_filtered):
            return
        
        # 如果点击的是选择列，复选框逻辑
        if column == 0:
            return  # 让复选框自行处理
        
        # 获取订单数据
        order = self._order_summary_filtered[row]
        self._selected_order_index = row
        self._show_order_detail(order)
    
    def _on_order_list_double_click(self, row, column):
        """双击订单列表行，打开编辑对话框"""
        if row < 0 or row >= len(self._order_summary_filtered):
            return
        
        # 跳过选择列
        if column == 0:
            return
        
        order = self._order_summary_filtered[row]
        self._selected_order_index = row
        self._open_order_edit_dialog(order)
    
    def _on_order_detail_double_click(self, row, column):
        """双击订单详情单元格，打开编辑对话框"""
        if self._selected_order_index is None:
            QMessageBox.information(self, "提示", "请先选择一个订单")
            return
        
        order = self._order_summary_filtered[self._selected_order_index]
        
        # 开票情况列（40）特殊处理 - 点击切换状态
        if column == 40:
            current_status = order.get('invoice_status', '未上传')
            if '已上传' in current_status or '已开' in current_status:
                new_status = '未上传'
            else:
                new_status = '已上传'
            
            # 更新数据
            order['invoice_status'] = new_status
            
            # 更新显示
            invoice_item = QTableWidgetItem(new_status)
            if '已上传' in new_status or '已开' in new_status:
                invoice_item.setForeground(QBrush(QColor("#10b981")))
            else:
                invoice_item.setForeground(QBrush(QColor("#ef4444")))
            self.order_detail_table.setItem(row, column, invoice_item)
            
            # 同步更新列表数据
            self._order_summary_filtered[self._selected_order_index] = order
            return
        
        # 其他列正常打开编辑对话框
        headers = [self.order_detail_table.horizontalHeaderItem(i).text() 
                   for i in range(self.order_detail_table.columnCount())]
        if column < len(headers):
            field_name = headers[column]
            current_value = self.order_detail_table.item(row, column).text() if self.order_detail_table.item(row, column) else ""
            self._open_field_edit_dialog(order, field_name, current_value, row, column)
    
    def _open_order_edit_dialog(self, order):
        """打开订单编辑对话框"""
        dialog = OrderEditDialog(order, self)
        if dialog.exec():
            # 保存编辑后的数据
            updated_order = dialog.get_updated_order()
            # 更新列表中的数据
            idx = self._selected_order_index
            if idx is not None and idx < len(self._order_summary_filtered):
                self._order_summary_filtered[idx] = updated_order
                # 如果在原始列表中也存在
                pi_id = updated_order.get('id')
                for i, o in enumerate(self._order_summary_orders):
                    if o.get('id') == pi_id:
                        self._order_summary_orders[i] = updated_order
                        break
                # 刷新显示
                self._populate_order_list_table(self._order_summary_filtered)
                self._show_order_detail(updated_order)
                self._order_summary_status.setText(f"已更新订单: {updated_order.get('order_no', '')}")
    
    def _open_field_edit_dialog(self, order, field_name, current_value, row, column):
        """打开字段编辑对话框"""
        dialog = FieldEditDialog(field_name, current_value, self)
        if dialog.exec():
            new_value = dialog.get_value()
            # 更新订单数据
            order[field_name] = new_value
            # 刷新详情显示
            self.order_detail_table.setItem(row, column, QTableWidgetItem(str(new_value)))
            # 更新列表
            self._order_summary_filtered[self._selected_order_index] = order
    
    def _show_order_detail(self, order):
        """显示订单详情 - 支持多产品显示"""
        self.order_detail_hint.hide()
        
        # 更新标题显示订单信息
        items = order.get('items', [])
        currency = order.get('currency', 'USD')
        item_count = len(items) if items else 1
        title = f"📋 订单: {order.get('order_no', '')} | 客户: {order.get('customer_name', '')} | 共 {item_count} 个产品 | 总金额: {order.get('total_amount', 0)} {currency}"
        self.order_detail_group.setTitle(title)
        
        # 清空详情表格
        self.order_detail_table.setRowCount(0)
        self.order_detail_table.setSortingEnabled(False)
        
        # 如果没有items，显示订单基本信息作为一行
        if not items:
            row = self.order_detail_table.rowCount()
            self.order_detail_table.insertRow(row)
            
            # 显示订单基本信息
            self.order_detail_table.setItem(row, 0, QTableWidgetItem("📋"))
            self.order_detail_table.setItem(row, 1, QTableWidgetItem(order.get('order_no', '')))
            self.order_detail_table.setItem(row, 2, QTableWidgetItem(order.get('customer_name', '')))
            self.order_detail_table.setItem(row, 3, QTableWidgetItem(str(order.get('total_amount', 0))))
            self.order_detail_table.setItem(row, 4, QTableWidgetItem(order.get('status', '进行中')))
            self.order_detail_table.setItem(row, 5, QTableWidgetItem(f"{order.get('customer_prepayment', 0)}"))
            self.order_detail_table.setItem(row, 6, QTableWidgetItem(f"{order.get('remaining_payment', 0)}"))
            
            # 预估计算
            factory_rmb = order.get('purchase_price', 0) or 0
            if factory_rmb > 0:
                estimated_usd = self.calculate_estimated_usd_price(factory_rmb)
                self.order_detail_table.setItem(row, 15, QTableWidgetItem(f"{estimated_usd:.2f} USD"))
                total_amount = order.get('total_amount', 0) or 0
                if total_amount > 0:
                    profit_margin = self.calculate_order_profit_margin(total_amount, factory_rmb)
                    self.order_detail_table.setItem(row, 16, QTableWidgetItem(f"{profit_margin:.1f}%"))
            self.order_detail_table.setSortingEnabled(True)
            return
        
        # 为每个产品显示一行
        for idx, item in enumerate(items):
            row = self.order_detail_table.rowCount()
            self.order_detail_table.insertRow(row)
            
            # 产品序号
            self.order_detail_table.setItem(row, 0, QTableWidgetItem(f"#{idx + 1}"))
            
            # 产品基本信息
            self.order_detail_table.setItem(row, 1, QTableWidgetItem(order.get('order_no', '')))                    # ORDER NO.
            self.order_detail_table.setItem(row, 2, QTableWidgetItem(item.get('customer_code', '')))               # 客户产品编号
            self.order_detail_table.setItem(row, 3, QTableWidgetItem(item.get('oe_number', '')))                  # OE号
            self.order_detail_table.setItem(row, 4, QTableWidgetItem(item.get('remark', '')))                     # 客户需求备注
            
            # 产品名称
            self.order_detail_table.setItem(row, 5, QTableWidgetItem(item.get('product_name', '')))
            
            # 图片列
            from PySide6.QtWidgets import QLabel
            image_label = QLabel()
            image_label.setFixedSize(74, 74)
            image_label.setAlignment(Qt.AlignCenter)
            image_label.setStyleSheet("border: 1px solid #e5e7eb;")
            image_url = item.get('image_url', '') or item.get('image', '')
            if image_url:
                try:
                    image_data = urllib.request.urlopen(image_url).read()
                    image = QImage.fromData(image_data)
                    pixmap = QPixmap.fromImage(image).scaled(68, 68, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                    image_label.setPixmap(pixmap)
                except:
                    image_label.setText("暂无")
            else:
                image_label.setText("暂无")
            self.order_detail_table.setCellWidget(row, 6, image_label)
            
            self.order_detail_table.setItem(row, 7, QTableWidgetItem(item.get('customer_model', '')))              # 客户型号
            self.order_detail_table.setItem(row, 8, QTableWidgetItem(item.get('oe_number', '')))                  # OE号.1
            
            # 数量和价格
            quantity = item.get('quantity', 0) or 0
            unit_price = item.get('unit_price', 0) or 0
            self.order_detail_table.setItem(row, 9, QTableWidgetItem(str(quantity)))                               # 数量
            self.order_detail_table.setItem(row, 10, QTableWidgetItem(f"{unit_price} {currency}"))                # 报价
            self.order_detail_table.setItem(row, 11, QTableWidgetItem(str(quantity * unit_price)))                # 合计金额
            
            # 客户回复
            self.order_detail_table.setItem(row, 12, QTableWidgetItem(item.get('customer_reply', '')))
            
            # 采购信息
            self.order_detail_table.setItem(row, 17, QTableWidgetItem(str(item.get('purchase_price', 0))))         # 采购价格
            self.order_detail_table.setItem(row, 18, QTableWidgetItem(str(item.get('shipping_fee', 0))))          # 运费
            
            # 品牌
            self.order_detail_table.setItem(row, 39, QTableWidgetItem(item.get('brand', '')))
            
            # 开票情况 - 显示状态并支持点击切换
            invoice_status = order.get('invoice_status', '未上传')
            if not invoice_status:
                invoice_status = '未上传'
            invoice_item = QTableWidgetItem(invoice_status)
            if '已上传' in invoice_status or '已开' in invoice_status:
                invoice_item.setForeground(QBrush(QColor("#10b981")))  # 绿色
            else:
                invoice_item.setForeground(QBrush(QColor("#ef4444")))  # 红色
            self.order_detail_table.setItem(row, 40, invoice_item)
            
            # 预估计算（基于采购价格）
            factory_rmb = item.get('purchase_price', 0) or 0
            if factory_rmb > 0:
                # 预估美金报价
                estimated_usd = self.calculate_estimated_usd_price(factory_rmb)
                self.order_detail_table.setItem(row, 15, QTableWidgetItem(f"{estimated_usd:.2f} USD"))
                
                # 预估毛利率
                item_total = quantity * unit_price
                if item_total > 0:
                    profit_margin = self.calculate_order_profit_margin(item_total, factory_rmb)
                    profit_item = QTableWidgetItem(f"{profit_margin:.1f}%")
                    if profit_margin > 0:
                        profit_item.setForeground(QBrush(QColor("#10b981")))
                    self.order_detail_table.setItem(row, 16, profit_item)
        
        self.order_detail_table.setSortingEnabled(True)
    
    def _import_order_summary(self):
        """导入订单总表（占位功能）"""
        print("[INFO] 导入订单总表功能")
        
        # ============================================================
        # TODO: 导入功能实现
        # ============================================================
        # 示例代码：
        #
        # file_path, _ = QFileDialog.getOpenFileName(
        #     self,
        #     "选择导入文件",
        #     "",
        #     "Excel文件 (*.xlsx *.xls);;CSV文件 (*.csv)"
        # )
        # if not file_path:
        #     return
        #
        # try:
        #     # 读取Excel/CSV文件
        #     df = pd.read_excel(file_path) if file_path.endswith('.xlsx') else pd.read_csv(file_path)
        #
        #     # 字段识别示例（需要根据实际表格结构调整）
        #     field_mapping = {
        #         '订单号': 'order_no',
        #         '客户': 'customer_name',
        #         '产品编号': 'customer_product_code',
        #         'OE号': 'oe_number',
        #         '数量': 'quantity',
        #         '单价': 'unit_price',
        #         '总价': 'total_amount',
        #         '日期': 'order_date',
        #         '备注': 'customer_requirement',
        #     }
        #
        #     # 处理每行数据
        #     for _, row in df.iterrows():
        #         order_data = {}
        #         for col_name, field_name in field_mapping.items():
        #             if col_name in df.columns:
        #                 order_data[field_name] = row[col_name]
        #
        #         # 调用API保存
        #         self.api_client.create_order_summary_item(order_data)
        #
        #     QMessageBox.information(self, "成功", f"已导入 {len(df)} 条记录")
        #     self.load_order_summary()
        # except Exception as e:
        #     QMessageBox.warning(self, "错误", f"导入失败: {e}")
        # ============================================================
        
        QMessageBox.information(
            self, 
            "导入功能", 
            "导入功能开发中...\n\n字段识别规则待配置。\n请查看代码中的注释示例。"
        )
    
    def load_order_summary(self):
        """加载订单总表数据"""
        print("[INFO] 订单总表: 开始加载...")
        self._show_loading_tip("正在加载订单总表...")
        
        def fetch():
            try:
                # 获取PI订单
                pi_list = self.api_client.get_pi_orders() or []
                print(f"[INFO] 订单总表: 获取到 {len(pi_list)} 条PI订单")
                
                # 获取采购订单
                purchase_list = self.api_client.get_purchase_orders() or []
                print(f"[INFO] 订单总表: 获取到 {len(purchase_list)} 条采购订单")
                
                # 获取出货记录
                shipment_list = self.api_client.get_shipments() or []
                print(f"[INFO] 订单总表: 获取到 {len(shipment_list)} 条出货记录")
                
                # 获取客户付款
                customer_payment_list = self.api_client.get_customer_payments() or []
                print(f"[INFO] 订单总表: 获取到 {len(customer_payment_list)} 条客户付款")
                
                # 获取供应商付款
                supplier_payment_list = self.api_client.get_supplier_payments() or []
                print(f"[INFO] 订单总表: 获取到 {len(supplier_payment_list)} 条供应商付款")
                
                return {
                    'pi_list': pi_list,
                    'purchase_list': purchase_list,
                    'shipment_list': shipment_list,
                    'customer_payment_list': customer_payment_list,
                    'supplier_payment_list': supplier_payment_list
                }
            except Exception as e:
                print(f"[ERROR] 订单总表: 获取数据失败: {e}")
                import traceback
                traceback.print_exc()
                return None
        
        class LoaderThread(QThread):
            data_ready = Signal(object)
            
            def run(self):
                data = fetch()
                self.data_ready.emit(data)
        
        self._loader_thread = LoaderThread(self)
        self._loader_thread.data_ready.connect(self._on_order_summary_data_ready)
        self._loader_thread.start()
    
    def _on_order_summary_data_ready(self, data):
        """数据加载完成回调"""
        print("[INFO] 订单总表: 数据加载完成，开始处理...")
        
        if data is None:
            print("[ERROR] 订单总表: 数据为空")
            self._hide_loading_tip()
            return
        
        try:
            # 保存缓存
            self._order_summary_cache = data
            pi_list = data.get('pi_list', [])
            purchase_list = data.get('purchase_list', [])
            shipment_list = data.get('shipment_list', [])
            customer_payment_list = data.get('customer_payment_list', [])
            supplier_payment_list = data.get('supplier_payment_list', [])
            
            print(f"[INFO] 订单总表: 处理 {len(pi_list)} 条PI订单")
            
            # 获取客户、产品、供应商映射
            customers_raw = self.api_client.get_customers() or []
            products_raw = self.api_client.get_products() or []
            suppliers_raw = self.api_client.get_suppliers() or []
            
            customers = {c.get('id'): c for c in customers_raw if c.get('id')}
            products = {p.get('id'): p for p in products_raw if p.get('id')}
            suppliers = {s.get('id'): s for s in suppliers_raw if s.get('id')}
            
            print(f"[INFO] 订单总表: 映射 {len(customers)} 客户, {len(products)} 产品, {len(suppliers)} 供应商")
            
            # 构建订单列表
            orders = []
            for pi in pi_list:
                try:
                    # 获取PI详情
                    pi_id = pi.get('id')
                    pi_detail = None
                    if pi_id:
                        try:
                            pi_detail = self.api_client.get_pi_detail(pi_id)
                        except Exception as e:
                            print(f"[WARN] 获取PI详情失败: {e}")
                    
                    order = self._build_order_summary_row(
                        pi_detail or pi, 
                        purchase_list, 
                        shipment_list, 
                        customer_payment_list, 
                        supplier_payment_list, 
                        customers, 
                        products,
                        suppliers
                    )
                    self._calculate_order_estimates(order)
                    orders.append(order)
                except Exception as e:
                    print(f"[WARN] 构建订单行失败: {e}")
                    continue
            
            print(f"[INFO] 订单总表: 构建完成, 共 {len(orders)} 条")
            
            # 更新UI（在主线程）
            self._order_summary_orders = orders
            self._order_summary_filtered = orders
            self._populate_order_list_table(orders)
            
            self._order_summary_status.setText(f"共 {len(orders)} 条订单")
            self._hide_loading_tip()
            print("[INFO] 订单总表: 加载完成")
            
        except Exception as e:
            print(f"[ERROR] 订单总表: 处理失败: {e}")
            import traceback
            traceback.print_exc()
            self._hide_loading_tip()
    
    def _calculate_order_estimates(self, order):
        """前端计算订单的预估美金报价和毛利率"""
        try:
            factory_rmb = order.get('purchase_price', 0) or 0
            total_amount = order.get('total_amount', 0) or 0
            
            # 预估美金报价 = 工厂人民币价格 × (1 + 毛利率) / 汇率
            if factory_rmb > 0:
                margin_factor = 1 + (self.default_profit_margin / 100)
                estimated_usd = factory_rmb * margin_factor / self.exchange_rate
                order['estimated_usd_price'] = round(estimated_usd, 2)
            else:
                order['estimated_usd_price'] = 0
            
            # 预估毛利率 = (客户总价 - 成本总价) / 客户总价
            if total_amount > 0 and factory_rmb > 0:
                customer_total_rmb = total_amount
                if customer_total_rmb > 0:
                    profit_margin = (customer_total_rmb - factory_rmb) / customer_total_rmb * 100
                    order['profit_margin'] = round(max(0, profit_margin), 1)
                else:
                    order['profit_margin'] = 0
            else:
                order['profit_margin'] = 0
        except Exception as e:
            print(f"[WARN] 计算预估失败: {e}")
    
    def _populate_order_list_table(self, orders):
        """填充订单列表表格（下半部分）"""
        self.order_list_table.setRowCount(0)
        self.order_list_table.setSortingEnabled(False)
        
        for order in orders:
            row = self.order_list_table.rowCount()
            self.order_list_table.insertRow(row)
            
            # 0: 选择复选框
            checkbox = QCheckBox()
            checkbox.setStyleSheet("margin-left: 50%;")
            self.order_list_table.setCellWidget(row, 0, checkbox)
            
            # 1: ORDER NO.
            self.order_list_table.setItem(row, 1, QTableWidgetItem(order.get('order_no', '')))
            
            # 2: 客户
            self.order_list_table.setItem(row, 2, QTableWidgetItem(order.get('customer_name', '')))
            
            # 3: 订单日期
            self.order_list_table.setItem(row, 3, QTableWidgetItem(order.get('order_date', '')))
            
            # 4: 产品数量
            items = order.get('items', [])
            product_count = len(items) if items else (1 if order.get('product_name') else 0)
            count_item = QTableWidgetItem(str(product_count))
            count_item.setTextAlignment(Qt.AlignCenter)
            self.order_list_table.setItem(row, 4, count_item)
            
            # 5: 总金额
            currency = order.get('currency', 'USD')
            total = order.get('total_amount', 0)
            amount_item = QTableWidgetItem(f"{total} {currency}")
            amount_item.setTextAlignment(Qt.AlignRight)
            self.order_list_table.setItem(row, 5, amount_item)
            
            # 6: 状态
            status = order.get('status', '进行中')
            status_item = QTableWidgetItem(status)
            if '完成' in status:
                status_item.setForeground(QBrush(QColor("#10b981")))  # 绿色
            elif '取消' in status:
                status_item.setForeground(QBrush(QColor("#ef4444")))  # 红色
            else:
                status_item.setForeground(QBrush(QColor("#3b82f6")))  # 蓝色
            self.order_list_table.setItem(row, 6, status_item)
        
        self.order_list_table.setSortingEnabled(True)
    
    def _build_order_summary_row(self, pi, purchase_list, shipment_list, customer_payment_list, supplier_payment_list, customers=None, products=None, suppliers=None):
        """构建订单总表单行数据"""
        if customers is None:
            customers = {}
        if products is None:
            products = {}
        if suppliers is None:
            suppliers = {}
        
        pi_id = pi.get('id', 'N/A')
        pi_no = pi.get('pi_no', 'N/A')
        print(f"[DEBUG] 订单总表: 构建PI_ID={pi_id}, PI_NO={pi_no}")
        
        # 基础信息
        order_date = pi.get('created_at', '')[:10] if pi.get('created_at') else ''
        order_no = pi.get('pi_no', '')
        customer_code = pi.get('customer_code', '')
        
        # 从 items 中获取产品信息（PI detail 包含 items）
        items = pi.get('items', [])
        first_item = items[0] if items else {}
        product_id = first_item.get('product_id')
        quantity = first_item.get('quantity', 0)
        unit_price = first_item.get('unit_price', 0)
        oe_number = first_item.get('oe_number', '')
        remark = first_item.get('remark', '')
        
        # 使用传入的字典快速查找产品
        product = products.get(product_id) if product_id else None
        if product:
            print(f"[DEBUG] 订单总表: 找到产品ID={product_id}, 产品名={product.get('name', 'N/A')}")
        else:
            print(f"[DEBUG] 订单总表: 未找到产品ID={product_id}, products字典keys={list(products.keys())}")
        
        # 获取客户信息
        customer_id = pi.get('customer_id')
        print(f"[DEBUG] 订单总表: PI客户ID={customer_id}, 客户字典keys={list(customers.keys())}")
        customer = customers.get(customer_id) if customer_id else None
        if customer:
            print(f"[DEBUG] 订单总表: 找到客户ID={customer_id}, 客户={customer.get('name', 'N/A')}")
        
        # 查找采购信息
        purchase = next((p for p in purchase_list if p.get('pi_id') == pi_id), None)
        if purchase:
            print(f"[DEBUG] 订单总表: 找到采购订单ID={purchase.get('id')}")
        else:
            print(f"[DEBUG] 订单总表: 未找到采购订单")
        
        # 查找供应商
        supplier_id = purchase.get('supplier_id') if purchase else None
        supplier = None
        if supplier_id:
            suppliers = self.api_client.get_suppliers() or []
            supplier_dict = {s['id']: s for s in suppliers}
            supplier = supplier_dict.get(supplier_id)
        
        # 查找出货信息
        shipment = next((s for s in shipment_list if s.get('pi_id') == pi_id), None)
        if shipment:
            print(f"[DEBUG] 订单总表: 找到出货记录ID={shipment.get('id')}")
        
        # 查找客户付款
        customer_payment = next((cp for cp in customer_payment_list if cp.get('pi_id') == pi_id), None)
        
        # 查找供应商付款
        purchase_id_for_supplier = purchase.get('id') if purchase else None
        supplier_payment = next((sp for sp in supplier_payment_list if sp.get('purchase_id') == purchase_id_for_supplier), None)
        
        # 获取产品OE列表
        oe_list = []
        try:
            if product_id:
                oe_list = self.api_client.get_product_oes(product_id) or []
                print(f"[DEBUG] 订单总表: 获取产品OE列表，共{len(oe_list)}个")
        except Exception as e:
            print(f"[DEBUG] 订单总表: 获取OE列表失败: {e}")
        
        # 获取产品-客户关联信息
        customer_product_info = None
        try:
            if product_id and customer_id:
                customer_product_info = self.api_client.get_product_customer(product_id, customer_id)
                print(f"[DEBUG] 订单总表: 获取客户产品关联: {customer_product_info}")
        except Exception as e:
            print(f"[DEBUG] 订单总表: 获取客户产品关联失败: {e}")
        
        # 状态判断
        is_completed = pi.get('status') == 4
        print(f"[DEBUG] 订单总表: PI状态={pi.get('status')}, 是否完成={is_completed}")
        
        # 构建OE显示文本
        primary_oe = next((oe for oe in oe_list if oe.get('is_primary')), None)
        if primary_oe:
            display_oe = primary_oe.get('oe_number', '')
        elif oe_list:
            display_oe = "多OE号"  # 有多个OE但没有主OE
        else:
            display_oe = oe_number or ''
        
        # 构建客户产品编号显示（取客户号后面的字符）
        customer_code_display = customer_code
        if customer_product_info:
            full_code = customer_product_info.get('customer_product_code', '')
            if full_code:
                # 去掉客户号前缀
                customer_code_display = full_code
        else:
            customer_code_display = customer_code
        
        # 构建订单数据 - 包含编辑所需的所有字段
        order_data = {
            'order_date': order_date,
            'order_no': order_no,
            'customer_product_code': customer_code_display,  # 显示用
            'customer_product_code_full': customer_product_info.get('customer_product_code', '') if customer_product_info else '',
            'oe_number': display_oe,  # 显示用
            'oe_number_list': oe_list,  # 完整OE列表，用于弹窗显示
            'oe_count': len(oe_list),  # OE数量
            'customer_requirement': remark,
            # 产品名使用 product_code 或 detail_desc
            'product_name': product.get('product_code', '') if product else (product.get('detail_desc', '') if product else ''),
            'product_code': product.get('product_code', '') if product else '',
            'product_detail_desc': product.get('detail_desc', '') if product else '',
            'image': product.get('default_image_url', '') if product else '',
            'customer_model': customer_code,
            # 客户名使用 customer_name
            'customer_name': customer.get('customer_name', '') if customer else '',
            'customer_contact': customer.get('contact', '') if customer else '',
            'customer_phone': customer.get('phone', '') if customer else '',
            'customer_address': customer.get('address', '') if customer else '',
            'quantity': quantity,
            'unit_price': unit_price,
            'total_amount': pi.get('total_amount', 0),
            'currency': pi.get('currency', 'USD'),
            'customer_prepayment': customer_payment.get('amount', 0) if customer_payment else 0,
            'remaining_payment': (pi.get('total_amount', 0) or 0) - (customer_payment.get('amount', 0) if customer_payment else 0),
            'purchase_price': purchase.get('total_amount', 0) if purchase else 0,
            'shipping_fee': purchase.get('shipping_fee', 0) if purchase else 0,
            'misc_fee': purchase.get('misc_fee', 0) if purchase else 0,
            # 供应商名使用 supplier_name
            'supplier_name': supplier.get('supplier_name', '') if supplier else '',
            'supplier_link': supplier.get('shop_link', '') if supplier else '',
            'supplier_code': supplier.get('code', '') if supplier else '',
            'delivery_date': purchase.get('expected_date', '')[:10] if purchase and purchase.get('expected_date') else '',
            'is_received': '是' if shipment else '否',
            'supplier_deposit': supplier_payment.get('deposit_amount', 0) if supplier_payment else 0,
            'supplier_balance': supplier_payment.get('balance_amount', 0) if supplier_payment else 0,
            'storage_status': '已入库' if shipment else '未入库',
            'storage_quantity': shipment.get('quantity', 0) if shipment else 0,
            'packaging': '',
            'factory_code': supplier.get('code', '') if supplier else '',
            'carton_size': '',
            'packing_spec': '',
            'carton_count': 0,
            'estimated_volume': '',
            'carton_gross_weight': '',
            'total_weight': '',
            'brand': product.get('brand', '') if product else '',
            'product_spec': product.get('specifications', '') if product else '',
            'invoice_status': '',
            'status': '已完成' if is_completed else '进行中',
            # 客户最新回复（从API获取）
            'customer_reply': '',
            # 额外信息用于编辑
            'pi_id': pi_id,
            'product_id': product_id,
            'customer_id': customer_id,
            'purchase_id': purchase.get('id') if purchase else None,
        }
        
        # 获取客户最新回复
        try:
            latest_reply = self.api_client.get_latest_customer_reply(pi_id)
            if latest_reply:
                order_data['customer_reply'] = latest_reply.get('reply_content', '')
        except Exception as e:
            print(f"[DEBUG] 订单总表: 获取客户回复失败: {e}")
        
        return order_data
    
    def _populate_order_summary_table(self, orders):
        """填充订单总表数据"""
        self.order_summary_table.setRowCount(0)
        self.order_summary_table.setSortingEnabled(False)
        
        for order in orders:
            row = self.order_summary_table.rowCount()
            self.order_summary_table.insertRow(row)
            
            # 0: 订单日期
            self.order_summary_table.setItem(row, 0, QTableWidgetItem(order['order_date']))
            
            # 1: ORDER NO.
            self.order_summary_table.setItem(row, 1, QTableWidgetItem(order['order_no']))
            
            # 2: 客户产品编号
            self.order_summary_table.setItem(row, 2, QTableWidgetItem(order['customer_product_code']))
            
            # 3: OE号 - 根据OE数量决定显示方式
            oe_count = order.get('oe_count', 0)
            if oe_count > 1:
                # 多个OE号，显示可点击的按钮样式
                from PySide6.QtWidgets import QPushButton, QWidget, QHBoxLayout
                btn = QPushButton("多OE号")
                btn.setStyleSheet("""
                    QPushButton {
                        background-color: #3b82f6;
                        color: white;
                        border: none;
                        border-radius: 4px;
                        padding: 4px 8px;
                        font-size: 11px;
                    }
                    QPushButton:hover { background-color: #2563eb; }
                """)
                btn.clicked.connect(lambda checked, r=row: self._show_oe_list_dialog(r))
                self.order_summary_table.setCellWidget(row, 3, btn)
            else:
                # 单个OE号，直接显示
                self.order_summary_table.setItem(row, 3, QTableWidgetItem(order['oe_number']))
            
            # 4: 客户需求备注
            self.order_summary_table.setItem(row, 4, QTableWidgetItem(order['customer_requirement']))
            
            # 5: 产品名称
            self.order_summary_table.setItem(row, 5, QTableWidgetItem(order['product_name']))
            
            # 6: 图片 - 异步加载显示
            img_label = QLabel("📷")
            img_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            img_label.setFixedSize(56, 56)
            self.order_summary_table.setCellWidget(row, 6, img_label)
            if order.get('image'):
                # 异步加载图片
                self.load_image_async(img_label, order['image'])
            
            # 7: 客户型号
            self.order_summary_table.setItem(row, 7, QTableWidgetItem(order['customer_model']))
            
            # 8: OE号.1 (同OE号)
            self.order_summary_table.setItem(row, 8, QTableWidgetItem(order['oe_number']))
            
            # 9: 数量
            qty_item = QTableWidgetItem(str(order['quantity']))
            qty_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.order_summary_table.setItem(row, 9, qty_item)
            
            # 10: 报价(USD/RMB)
            price_text = f"{order['unit_price']} {order['currency']}"
            price_item = QTableWidgetItem(price_text)
            price_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)
            self.order_summary_table.setItem(row, 10, price_item)
            
            # 11: 合计金额
            total_item = QTableWidgetItem(f"{order['total_amount']:.2f}")
            total_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)
            self.order_summary_table.setItem(row, 11, total_item)
            
            # 填充完所有列后设置行高（图片需要更大的行高）
            self.order_summary_table.setRowHeight(row, 60)
            
            # 12: 客户最新回复（从API获取）
            customer_reply = order.get('customer_reply', '')
            reply_item = QTableWidgetItem(customer_reply[:50] + "..." if len(customer_reply) > 50 else customer_reply)
            reply_item.setForeground(QBrush(QColor("#059669")))  # 绿色显示
            self.order_summary_table.setItem(row, 12, reply_item)
            
            # 13: 客户预付款
            prepay_item = QTableWidgetItem(f"{order['customer_prepayment']:.2f}")
            prepay_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)
            self.order_summary_table.setItem(row, 13, prepay_item)
            
            # 14: 待收尾款
            remaining_item = QTableWidgetItem(f"{order['remaining_payment']:.2f}")
            remaining_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)
            self.order_summary_table.setItem(row, 14, remaining_item)
            
            # 15-19: 采购相关
            self.order_summary_table.setItem(row, 15, QTableWidgetItem(""))  # 预估美金报价
            self.order_summary_table.setItem(row, 16, QTableWidgetItem(""))  # 预估毛利率
            purchase_price_item = QTableWidgetItem(f"{order['purchase_price']:.2f}")
            purchase_price_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)
            self.order_summary_table.setItem(row, 17, purchase_price_item)
            shipping_item = QTableWidgetItem(f"{order['shipping_fee']:.2f}")
            shipping_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)
            self.order_summary_table.setItem(row, 18, shipping_item)
            misc_item = QTableWidgetItem(f"{order['misc_fee']:.2f}")
            misc_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)
            self.order_summary_table.setItem(row, 19, misc_item)
            
            # 20-24: 工厂相关
            total_amount = order['purchase_price'] + order['shipping_fee'] + order['misc_fee']
            total_item = QTableWidgetItem(f"{total_amount:.2f}")
            total_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)
            self.order_summary_table.setItem(row, 20, total_item)
            self.order_summary_table.setItem(row, 21, QTableWidgetItem(order['supplier_name']))
            self.order_summary_table.setItem(row, 22, QTableWidgetItem(order['supplier_link']))
            self.order_summary_table.setItem(row, 23, QTableWidgetItem(order['delivery_date']))
            self.order_summary_table.setItem(row, 24, QTableWidgetItem(order['is_received']))
            
            # 25-29: 付款和入库
            deposit_item = QTableWidgetItem(f"{order['supplier_deposit']:.2f}")
            deposit_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)
            self.order_summary_table.setItem(row, 25, deposit_item)
            balance_item = QTableWidgetItem(f"{order['supplier_balance']:.2f}")
            balance_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)
            self.order_summary_table.setItem(row, 26, balance_item)
            self.order_summary_table.setItem(row, 27, QTableWidgetItem(order['storage_status']))
            storage_qty_item = QTableWidgetItem(str(order['storage_quantity']))
            storage_qty_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.order_summary_table.setItem(row, 28, storage_qty_item)
            self.order_summary_table.setItem(row, 29, QTableWidgetItem(order['packaging']))
            
            # 30-40: 其他信息
            self.order_summary_table.setItem(row, 30, QTableWidgetItem(order['factory_code']))
            self.order_summary_table.setItem(row, 31, QTableWidgetItem(""))
            self.order_summary_table.setItem(row, 32, QTableWidgetItem(order['factory_code']))
            self.order_summary_table.setItem(row, 33, QTableWidgetItem(order['carton_size']))
            self.order_summary_table.setItem(row, 34, QTableWidgetItem(order['packing_spec']))
            carton_item = QTableWidgetItem(str(order['carton_count']))
            carton_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.order_summary_table.setItem(row, 35, carton_item)
            self.order_summary_table.setItem(row, 36, QTableWidgetItem(order['estimated_volume']))
            self.order_summary_table.setItem(row, 37, QTableWidgetItem(order['carton_gross_weight']))
            self.order_summary_table.setItem(row, 38, QTableWidgetItem(order['total_weight']))
            self.order_summary_table.setItem(row, 39, QTableWidgetItem(order['brand']))
            self.order_summary_table.setItem(row, 40, QTableWidgetItem(order['invoice_status']))
        
        self.order_summary_table.setSortingEnabled(True)
    
    def filter_order_summary(self):
        """筛选订单列表"""
        search_text = self.order_summary_search.text().lower()
        status_filter = self.order_status_filter.currentText()
        
        # 从原始数据筛选
        filtered = []
        for order in self._order_summary_orders:
            match = True
            
            # 搜索文本匹配
            if search_text:
                match = False
                search_fields = [
                    order.get('order_no', ''),
                    order.get('customer_name', ''),
                    order.get('oe_number', ''),
                    order.get('product_name', ''),
                    order.get('customer_product_code', ''),
                ]
                for field in search_fields:
                    if search_text in str(field).lower():
                        match = True
                        break
            
            # 状态筛选
            if match and status_filter != "全部状态":
                status = order.get('status', '进行中')
                if status_filter == "进行中" and "进行中" not in status:
                    match = False
                elif status_filter == "已完成" and "完成" not in status:
                    match = False
                elif status_filter == "已取消" and "取消" not in status:
                    match = False
            
            if match:
                filtered.append(order)
        
        # 更新筛选后的列表
        self._order_summary_filtered = filtered
        self._populate_order_list_table(filtered)
    
    def _clear_order_summary_filter(self):
        """清除筛选条件"""
        self.order_summary_search.clear()
        self.order_status_filter.setCurrentText("全部状态")
        self.filter_order_summary()
    
    def _on_order_summary_double_click(self, row, column):
        """订单总表双击编辑"""
        from widgets.order_summary_edit_dialog import OrderSummaryEditDialog
        from PySide6.QtWidgets import QMessageBox
        
        print(f"[DEBUG] 订单总表: 双击行 {row}, 列 {column}")
        
        # 检查是否被筛选隐藏
        if self.order_summary_table.isRowHidden(row):
            QMessageBox.warning(self, "提示", "该行已被筛选隐藏，请先清除筛选")
            return
        
        # 使用存储的完整数据
        if not hasattr(self, '_order_summary_orders') or row >= len(self._order_summary_orders):
            QMessageBox.warning(self, "提示", "订单数据未加载完成")
            return
        
        order_data = self._order_summary_orders[row]
        
        print(f"[DEBUG] 订单总表: 编辑订单 PI_NO={order_data.get('order_no', 'N/A')}")
        print(f"[DEBUG] 订单总表: 产品名={order_data.get('product_name', 'N/A')}")
        print(f"[DEBUG] 订单总表: 客户名={order_data.get('customer_name', 'N/A')}")
        print(f"[DEBUG] 订单总表: 供应商={order_data.get('supplier_name', 'N/A')}")
        
        # 打开编辑对话框
        dialog = OrderSummaryEditDialog(order_data, self.api_client, self)
        dialog.saved.connect(lambda data: self._on_order_summary_saved(data, row))
        dialog.exec()
    
    def _on_order_summary_saved(self, data, row):
        """订单总表保存后的处理"""
        from PySide6.QtWidgets import QMessageBox
        print(f"[DEBUG] 订单总表: 保存数据到行 {row}")
        # 根据需要更新表格
        # 这里可以实现保存到后端的逻辑
        QMessageBox.information(self, "成功", "订单已保存")
    
    def _on_order_summary_cell_click(self, row, column):
        """订单总表单元格点击事件 - 特定列打开独立窗口"""
        # 列索引: 4=客户需求备注, 7=客户型号, 12=客户最新回复
        special_columns = {4: "客户需求备注", 7: "客户型号", 12: "客户最新回复"}
        
        if column not in special_columns:
            return
        
        print(f"[DEBUG] 订单总表: 点击行 {row}, 列 {column} ({special_columns[column]})")
        
        # 检查数据是否加载
        if not hasattr(self, '_order_summary_orders') or row >= len(self._order_summary_orders):
            return
        
        order_data = self._order_summary_orders[row]
        pi_id = order_data.get('pi_id')
        pi_no = order_data.get('order_no', '')
        
        if column == 4:  # 客户需求备注
            self._open_requirement_dialog(order_data, row)
        elif column == 7:  # 客户型号
            self._open_customer_model_dialog(order_data, row)
        elif column == 12:  # 客户最新回复
            self._open_customer_reply_dialog(pi_id, pi_no, order_data, row)
    
    def _open_requirement_dialog(self, order_data, row):
        """打开客户需求备注对话框"""
        current_value = order_data.get('customer_requirement', '')
        pi_no = order_data.get('order_no', '')
        
        dialog = CustomerRequirementDialog(current_value, pi_no, self)
        dialog.saved.connect(lambda text: self._on_requirement_saved(text, row))
        dialog.exec()
    
    def _on_requirement_saved(self, text, row):
        """客户需求备注保存后更新"""
        print(f"[DEBUG] 订单总表: 客户需求备注已保存到行 {row}")
        # 更新表格显示
        self.order_summary_table.setItem(row, 4, QTableWidgetItem(text))
        # 更新存储的数据
        if hasattr(self, '_order_summary_orders') and row < len(self._order_summary_orders):
            self._order_summary_orders[row]['customer_requirement'] = text
    
    def _open_customer_model_dialog(self, order_data, row):
        """打开客户型号对话框"""
        current_value = order_data.get('customer_model', '')
        pi_no = order_data.get('order_no', '')
        
        dialog = CustomerModelDialog(current_value, pi_no, self)
        dialog.saved.connect(lambda text: self._on_customer_model_saved(text, row))
        dialog.exec()
    
    def _on_customer_model_saved(self, text, row):
        """客户型号保存后更新"""
        print(f"[DEBUG] 订单总表: 客户型号已保存到行 {row}")
        # 更新表格显示
        self.order_summary_table.setItem(row, 7, QTableWidgetItem(text))
        # 更新存储的数据
        if hasattr(self, '_order_summary_orders') and row < len(self._order_summary_orders):
            self._order_summary_orders[row]['customer_model'] = text
    
    def _open_customer_reply_dialog(self, pi_id, pi_no, order_data, row):
        """打开客户回复对话框"""
        current_reply = order_data.get('customer_reply', '')
        
        dialog = CustomerReplyDialog(pi_id, pi_no, self.api_client, current_reply, self)
        dialog.saved.connect(lambda data: self._on_customer_reply_saved(data, row))
        dialog.exec()
    
    def _on_customer_reply_saved(self, reply_data, row):
        """客户回复保存后更新"""
        print(f"[DEBUG] 订单总表: 客户回复已保存到行 {row}")
        # 更新表格显示最新回复
        content = reply_data.get('reply_content', '')[:30]
        self.order_summary_table.setItem(row, 12, QTableWidgetItem(f"{content}..."))
        # 更新存储的数据
        if hasattr(self, '_order_summary_orders') and row < len(self._order_summary_orders):
            self._order_summary_orders[row]['customer_reply'] = reply_data.get('reply_content', '')
    
    def _show_oe_list_dialog(self, row):
        """显示OE号列表弹窗"""
        if not hasattr(self, '_order_summary_orders') or row >= len(self._order_summary_orders):
            return
        
        order_data = self._order_summary_orders[row]
        oe_list = order_data.get('oe_number_list', [])
        product_id = order_data.get('product_id')
        pi_no = order_data.get('order_no', '')
        
        from widgets.product_oe_dialog import ProductOEDialog
        
        dialog = ProductOEDialog(product_id, oe_list, self.api_client, self)
        dialog.exec()
    
    def _show_product_oe_dialog(self, product_id, oe_list):
        """产品列表中的OE号弹窗"""
        from widgets.product_oe_dialog import ProductOEDialog
        
        dialog = ProductOEDialog(product_id, oe_list, self.api_client, self)
        if dialog.exec() == QDialog.Accepted:
            # 刷新产品列表
            self.load_products()
    
    def export_order_summary(self):
        """导出订单总表为Excel"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from PySide6.QtWidgets import QFileDialog
            
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "保存订单总表",
                f"订单总表_{QDate.currentDate().toString('yyyyMMdd')}.xlsx",
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "订单总表"
            
            # 写入表头
            headers = [
                "订单日期", "ORDER NO.", "客户产品编号", "OE号", "客户需求备注",
                "产品名称", "图片", "客户型号", "OE号.1", "数量",
                "报价(USD/RMB)", "合计金额", "客户最新回复", "客户预付款", "待收尾款",
                "预估美金报价", "预估毛利率", "采购价格", "运费", "杂费",
                "总金额", "工厂简称", "店铺链接", "交货日期", "是否已收货",
                "工厂订金", "工厂尾款", "入库操作", "入库数量", "包装方式",
                "采购选项/名称", "产品细节", "工厂编号", "纸箱尺寸", "打包规格",
                "箱数", "预估体积", "整箱毛重", "总重量", "品牌", "开票情况"
            ]
            
            # 表头样式
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=10)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # 写入数据
            for row in range(self.order_summary_table.rowCount()):
                if self.order_summary_table.isRowHidden(row):
                    continue
                for col in range(self.order_summary_table.columnCount()):
                    item = self.order_summary_table.item(row, col)
                    value = item.text() if item else ""
                    ws.cell(row=row + 2, column=col + 1, value=value)
            
            # 设置列宽
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 12
            
            wb.save(file_path)
            QMessageBox.information(self, "导出成功", f"订单总表已导出到:\n{file_path}")
            
        except Exception as e:
            QMessageBox.warning(self, "导出失败", f"导出Excel时出错:\n{str(e)}")

    def load_image_async(self, label, image_url):
        """异步加载图片 - 使用全局线程池和图片缓存"""
        global _image_cache, _image_cache_lock
        
        # 先检查内存缓存
        with _image_cache_lock:
            if image_url in _image_cache:
                pixmap = _image_cache[image_url]
                label.setPixmap(pixmap)
                return
        
        # 显示加载中占位符
        label.setText("...")
        
        def fetch_image():
            try:
                # 使用缓存避免重复下载
                cache_key = f"img_{hash(image_url)}"
                cached_data = cache_manager.get(cache_key, max_age=3600)  # 图片缓存1小时
                if cached_data:
                    return cached_data
                
                image_data = urllib.request.urlopen(image_url, timeout=3).read()
                cache_manager.set(cache_key, image_data)
                return image_data
            except Exception as e:
                print(f"图片加载失败: {e}")
                return None
        
        def on_done(future):
            image_data = future.result()
            if image_data:
                try:
                    image = QImage.fromData(image_data)
                    pixmap = QPixmap.fromImage(image).scaled(56, 56, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                    
                    # 存入内存缓存
                    with _image_cache_lock:
                        if len(_image_cache) >= _MAX_IMAGE_CACHE_SIZE:
                            # 简单LRU：删除最早的
                            oldest_key = next(iter(_image_cache))
                            del _image_cache[oldest_key]
                        _image_cache[image_url] = pixmap
                    
                    label.setPixmap(pixmap)
                    return
                except Exception as e:
                    print(f"图片处理失败: {e}")
            label.setText("暂无图片")
        
        # 使用全局线程池执行
        future = _global_thread_pool.submit(fetch_image)
        future.add_done_callback(on_done)

    def load_products(self, use_cache: bool = True):
        try:
            # 尝试从缓存加载产品
            products = None
            if use_cache:
                products = cache_manager.get(CACHE_KEYS['PRODUCTS'], max_age=300)
            
            if products is None:
                print("DEBUG - 从API加载产品...")
                products = self.api_client.get_products()
                # 保存到缓存
                cache_manager.set(CACHE_KEYS['PRODUCTS'], products)
            else:
                print("DEBUG - 从缓存加载产品")
            
            # 尝试从缓存加载库存汇总
            inventory_summary = cache_manager.get(CACHE_KEYS['INVENTORY_SUMMARY'], max_age=60)
            if inventory_summary is None:
                try:
                    print("DEBUG - 从API加载库存...")
                    inventory_summary = self.api_client.get_all_inventory_summary()
                    cache_manager.set(CACHE_KEYS['INVENTORY_SUMMARY'], inventory_summary)
                except Exception as e:
                    print(f"获取库存失败: {e}")
                    inventory_summary = {}
            else:
                print("DEBUG - 从缓存加载库存")
            
            if products is None:
                products = []
            print(f"[DEBUG] 产品列表: 共加载 {len(products)} 个产品")
            self.products_table.setRowCount(len(products))
            
            # 批量获取OE和客户产品数据（优化性能）
            product_ids = [p.get('id') for p in products if p.get('id')]
            all_oes = []
            all_customer_products = []
            try:
                if product_ids:
                    all_oes = self.api_client.get_product_oes_batch(product_ids) or []
                    all_customer_products = self.api_client.get_product_customers_batch(product_ids) or []
                    print(f"[DEBUG] 产品列表: 批量获取OE={len(all_oes)}条, 客户产品={len(all_customer_products)}条")
            except Exception as e:
                print(f"[DEBUG] 产品列表: 批量获取数据失败: {e}")
            
            # 按product_id分组
            oes_by_product = {}
            for oe in all_oes:
                pid = oe.get('product_id')
                if pid not in oes_by_product:
                    oes_by_product[pid] = []
                oes_by_product[pid].append(oe)
            
            pc_by_product = {}
            for pc in all_customer_products:
                pid = pc.get('product_id')
                if pid not in pc_by_product:
                    pc_by_product[pid] = []
                pc_by_product[pid].append(pc)
            
            for row, p in enumerate(products):
                product_id = p.get('id')
                print(f"[DEBUG] 产品列表: 处理第{row}行, ID={product_id}")
                
                # 0: 复选框
                checkbox = QCheckBox()
                checkbox.setCheckState(Qt.CheckState.Unchecked)
                checkbox.setStyleSheet("margin-left: 50%;")
                self.products_table.setCellWidget(row, 0, checkbox)
                
                # 从预加载的数据中获取
                oe_list = oes_by_product.get(product_id, [])
                customer_product_list = pc_by_product.get(product_id, [])
                print(f"[DEBUG] 产品列表: 行{row} - OE数量={len(oe_list)}, 客户产品数量={len(customer_product_list)}")
                
                # 1: 客户产品编号（从产品-客户关联获取，显示第一个客户的编号）
                customer_product_code = ""
                if customer_product_list:
                    # 取第一个客户的编号
                    first_pc = customer_product_list[0]
                    full_code = first_pc.get('customer_product_code', '')
                    customer_code = first_pc.get('customer_code', '')
                    print(f"[DEBUG] 产品列表: 行{row} - full_code={full_code}, customer_code={customer_code}")
                    if full_code and customer_code:
                        # 去掉客户号前缀
                        customer_product_code = full_code.replace(customer_code, "", 1).lstrip("-")
                    else:
                        customer_product_code = full_code or ""
                    print(f"[DEBUG] 产品列表: 行{row} - 显示的客户产品编号={customer_product_code}")
                self.products_table.setItem(row, 1, QTableWidgetItem(customer_product_code))
                
                # 2: OE号（显示主OE或"多OE号"按钮）
                primary_oe = next((oe for oe in oe_list if oe.get('is_primary')), None)
                if len(oe_list) > 1:
                    # 多个OE，显示按钮
                    btn = QPushButton("多OE号")
                    btn.setStyleSheet("""
                        QPushButton {
                            background-color: #3b82f6;
                            color: white;
                            border: none;
                            border-radius: 4px;
                            padding: 4px 8px;
                            font-size: 11px;
                        }
                        QPushButton:hover { background-color: #2563eb; }
                    """)
                    btn.clicked.connect(lambda checked, pid=product_id, oes=oe_list: self._show_product_oe_dialog(pid, oes))
                    self.products_table.setCellWidget(row, 2, btn)
                elif primary_oe:
                    self.products_table.setItem(row, 2, QTableWidgetItem(primary_oe.get('oe_number', '')))
                else:
                    self.products_table.setItem(row, 2, QTableWidgetItem(p.get('oe_number', '') or '-'))
                
                # 3: 图片
                image_label = QLabel()
                image_label.setFixedSize(60, 60)
                image_label.setStyleSheet("border: 1px solid #e5e7eb;")
                image_label.setAlignment(Qt.AlignCenter)
                
                image_url = p.get('default_image_url')
                if image_url:
                    self.load_image_async(image_label, image_url)
                else:
                    image_label.setText("暂无图片")
                
                image_label.mousePressEvent = lambda event, prod=p: self.view_product_images(prod.get('id'))
                image_label.setCursor(Qt.PointingHandCursor)
                self.products_table.setCellWidget(row, 3, image_label)
                
                # 4: 产品名称（detail_desc）
                self.products_table.setItem(row, 4, QTableWidgetItem(p.get('detail_desc', '')))
                
                # 5: 客户型号（从产品-客户关联获取，显示客户的）
                customer_model = ""
                if customer_product_list:
                    customer_model = customer_product_list[0].get('customer_model', '') or ""
                self.products_table.setItem(row, 5, QTableWidgetItem(customer_model))
                
                # 6: 客户号（显示客户名）
                customer_name = ""
                if customer_product_list:
                    # 需要通过customer_id获取客户名
                    pass  # 暂不显示，等待确认
                self.products_table.setItem(row, 6, QTableWidgetItem(""))  # 客户号列
                
                # 7: 产品特性（品牌）
                self.products_table.setItem(row, 7, QTableWidgetItem(p.get('brand', '') or '-'))
                
                # 8: 数量（库存）
                stock_qty = inventory_summary.get(product_id, 0)
                stock_item = QTableWidgetItem(str(int(stock_qty)) if stock_qty else '0')
                if stock_qty > 0:
                    stock_item.setForeground(QBrush(QColor("#10b981")))
                else:
                    stock_item.setForeground(QBrush(QColor("#6b7280")))
                stock_item.setTextAlignment(Qt.AlignCenter)
                self.products_table.setItem(row, 8, stock_item)
                
                # 9: 报价（EXW含税价）
                price = p.get('exw_price_incl', 0)
                price_text = f"{price} USD" if price else "-"
                price_item = QTableWidgetItem(price_text)
                price_item.setTextAlignment(Qt.AlignRight)
                self.products_table.setItem(row, 9, price_item)
                
                # 10: 编辑按钮
                def create_edit_callback(product):
                    return lambda: self.edit_product(product)
                
                action_widget = QWidget()
                action_layout = QHBoxLayout()
                action_layout.setContentsMargins(0, 0, 0, 0)
                
                edit_btn = QPushButton("编辑")
                edit_btn.setFixedWidth(50)
                edit_btn.clicked.connect(create_edit_callback(p))
                action_layout.addWidget(edit_btn)
                
                action_widget.setLayout(action_layout)
                self.products_table.setCellWidget(row, 10, action_widget)
                
            self.select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载产品数据失败: {str(e)}")

    def view_product_images(self, product):
        """查看产品图片"""
        product_id = product.get('id')
        product_code = product.get('product_code', '')
        image_url = product.get('default_image_url')
        
        if not image_url:
            QMessageBox.information(self, "图片查看", f"产品 {product_code} 暂无图片")
            return
        
        try:
            # 创建图片查看对话框
            dialog = QDialog(self)
            dialog.setWindowTitle(f"产品图片 - {product_code}")
            dialog.setMinimumSize(600, 600)
            
            layout = QVBoxLayout()
            
            # 图片标签
            image_label = QLabel()
            image_label.setAlignment(Qt.AlignCenter)
            
            # 加载图片
            image_data = urllib.request.urlopen(image_url).read()
            image = QImage.fromData(image_data)
            pixmap = QPixmap.fromImage(image).scaled(580, 580, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            image_label.setPixmap(pixmap)
            
            layout.addWidget(image_label)
            dialog.setLayout(layout)
            dialog.exec()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载图片失败: {str(e)}")
    
    def confirm_product_import(self, product):
        """确认产品导入"""
        print(f"DEBUG - confirm_product_import called for product: {product.get('product_code')}")

        product_id = product.get('id')
        product_code = product.get('product_code', '')

        # 检查产品是否已经导入
        if product.get('is_imported') == 1:
            QMessageBox.warning(self, "提示", f"产品 {product_code} 已经确认导入，无需重复操作")
            return

        reply = QMessageBox.question(
            self, "确认导入",
            f"确定要确认产品 {product_code} 已导入吗？确认后将无法修改（普通用户）",
            QMessageBox.Ok | QMessageBox.Cancel
        )

        if reply == QMessageBox.Ok:
            try:
                print(f"DEBUG - Calling API to confirm import for product_id: {product_id}")
                result = self.api_client.confirm_product_import(product_id)
                print(f"DEBUG - API response: {result}")
                QMessageBox.information(self, "成功", "产品导入已确认")
                self.load_products()
            except Exception as e:
                print(f"DEBUG - Confirm import failed: {str(e)}")
                QMessageBox.warning(self, "错误", f"确认导入失败: {str(e)}")
    
    def cancel_product_import(self, product):
        """取消产品导入确认（仅管理员）"""
        print(f"DEBUG - cancel_product_import called for product: {product.get('product_code')}")
        
        product_id = product.get('id')
        product_code = product.get('product_code', '')
        
        reply = QMessageBox.question(
            self, "取消导入", 
            f"确定要取消产品 {product_code} 的导入确认吗？",
            QMessageBox.Ok | QMessageBox.Cancel
        )
        
        if reply == QMessageBox.Ok:
            try:
                print(f"DEBUG - Calling API to cancel import for product_id: {product_id}")
                result = self.api_client.cancel_product_import(product_id)
                print(f"DEBUG - API response: {result}")
                QMessageBox.information(self, "成功", "产品导入确认已取消")
                self.load_products()
            except Exception as e:
                print(f"DEBUG - Cancel import failed: {str(e)}")
                QMessageBox.warning(self, "错误", f"取消导入失败: {str(e)}")

    def batch_confirm_import_products(self):
        """批量确认产品导入"""
        print("DEBUG - batch_confirm_import_products called")
        
        selected_products = []
        already_imported = []
        
        # 获取选中的产品
        for row in range(self.products_table.rowCount()):
            checkbox = self.products_table.cellWidget(row, 0)
            if checkbox and checkbox.isChecked():
                # 获取产品数据
                product_id = int(self.products_table.item(row, 2).text())
                product_code = self.products_table.item(row, 9)  # 产品编号在第10列（索引9）
                if not product_code:
                    product_code = self.products_table.item(row, 10).text()  # 或者从隐藏列获取
                else:
                    product_code = product_code.text()
                
                # 检查是否已导入（从数据源获取）
                # 这里需要重新获取产品数据来检查is_imported状态
                try:
                    product_detail = self.api_client.get_product_detail(product_id)
                    if product_detail.get('is_imported') == 1:
                        already_imported.append(product_code)
                    else:
                        selected_products.append({'id': product_id, 'code': product_code})
                except Exception as e:
                    print(f"DEBUG - Failed to get product detail for {product_id}: {str(e)}")
                    selected_products.append({'id': product_id, 'code': product_code})
        
        if not selected_products and not already_imported:
            QMessageBox.warning(self, "提示", "请先选择要确认导入的产品")
            return
        
        # 如果有已导入的产品，显示提示
        message = ""
        if already_imported:
            message = f"以下产品已确认导入，将跳过：\n{', '.join(already_imported)}\n\n"
        
        if not selected_products:
            QMessageBox.information(self, "提示", message + "没有需要确认导入的产品")
            return
        
        reply = QMessageBox.question(
            self, "批量确认导入",
            f"{message}确定要确认选中的 {len(selected_products)} 个产品已导入吗？\n确认后将无法修改（普通用户）",
            QMessageBox.Ok | QMessageBox.Cancel
        )
        
        if reply == QMessageBox.Ok:
            success_count = 0
            failed_count = 0
            failed_list = []
            
            for product in selected_products:
                try:
                    print(f"DEBUG - Calling API to confirm import for product_id: {product['id']}")
                    result = self.api_client.confirm_product_import(product['id'])
                    print(f"DEBUG - API response: {result}")
                    success_count += 1
                except Exception as e:
                    print(f"DEBUG - Confirm import failed for {product['code']}: {str(e)}")
                    failed_count += 1
                    failed_list.append(product['code'])
            
            # 显示结果
            result_msg = f"批量确认导入完成\n成功：{success_count} 个\n失败：{failed_count} 个"
            if failed_list:
                result_msg += f"\n失败产品：{', '.join(failed_list)}"
            
            QMessageBox.information(self, "批量确认导入结果", result_msg)
            self.load_products()

    def search_products(self):
        keyword = self.search_input.text().strip()
        category_id = self.category_filter.currentData()
        category_id = category_id if category_id != 0 else None
        
        try:
            # 搜索时显示产品编号列
            self.products_table.setColumnHidden(10, False)
            print(f"搜索参数 - 关键词: '{keyword}', 类别ID: {category_id}")
            products = self.api_client.search_products(keyword, category_id)
            print(f"搜索结果数量: {len(products)}")
            self.products_table.setRowCount(len(products))
            for row, p in enumerate(products):
                product_id = p.get('id')
                
                # 0: 复选框
                checkbox = QCheckBox()
                checkbox.setCheckState(Qt.CheckState.Unchecked)
                checkbox.setStyleSheet("margin-left: 50%;")
                self.products_table.setCellWidget(row, 0, checkbox)
                
                # 获取OE和客户关联
                oe_list = []
                customer_product_list = []
                try:
                    oe_list = self.api_client.get_product_oes(product_id) or []
                    customer_product_list = self.api_client.get_product_customers(product_id) or []
                except:
                    pass
                
                # 1: 客户产品编号
                customer_product_code = ""
                if customer_product_list:
                    first_pc = customer_product_list[0]
                    full_code = first_pc.get('customer_product_code', '')
                    customer_code = first_pc.get('customer_code', '')
                    if full_code and customer_code:
                        customer_product_code = full_code.replace(customer_code, "", 1).lstrip("-")
                    else:
                        customer_product_code = full_code or ""
                self.products_table.setItem(row, 1, QTableWidgetItem(customer_product_code))
                
                # 2: OE号
                primary_oe = next((oe for oe in oe_list if oe.get('is_primary')), None)
                if len(oe_list) > 1:
                    btn = QPushButton("多OE号")
                    btn.setStyleSheet("""
                        QPushButton {
                            background-color: #3b82f6;
                            color: white;
                            border: none;
                            border-radius: 4px;
                            padding: 4px 8px;
                            font-size: 11px;
                        }
                        QPushButton:hover { background-color: #2563eb; }
                    """)
                    btn.clicked.connect(lambda checked, pid=product_id, oes=oe_list: self._show_product_oe_dialog(pid, oes))
                    self.products_table.setCellWidget(row, 2, btn)
                elif primary_oe:
                    self.products_table.setItem(row, 2, QTableWidgetItem(primary_oe.get('oe_number', '')))
                else:
                    self.products_table.setItem(row, 2, QTableWidgetItem(p.get('oe_number', '') or '-'))
                
                # 3: 图片
                image_label = QLabel()
                image_label.setFixedSize(60, 60)
                image_label.setStyleSheet("border: 1px solid #e5e7eb;")
                image_label.setAlignment(Qt.AlignCenter)
                
                image_url = p.get('default_image_url')
                if image_url:
                    try:
                        image_data = urllib.request.urlopen(image_url).read()
                        image = QImage.fromData(image_data)
                        pixmap = QPixmap.fromImage(image).scaled(56, 56, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                        image_label.setPixmap(pixmap)
                    except Exception as e:
                        image_label.setText("暂无图片")
                else:
                    image_label.setText("暂无图片")
                
                image_label.setCursor(Qt.PointingHandCursor)
                self.products_table.setCellWidget(row, 3, image_label)
                
                # 4: 产品名称
                self.products_table.setItem(row, 4, QTableWidgetItem(p.get('detail_desc', '')))
                
                # 5: 客户型号
                customer_model = ""
                if customer_product_list:
                    customer_model = customer_product_list[0].get('customer_model', '') or ""
                self.products_table.setItem(row, 5, QTableWidgetItem(customer_model))
                
                # 6: 客户号（留空）
                self.products_table.setItem(row, 6, QTableWidgetItem(""))
                
                # 7: 产品特性（品牌）
                self.products_table.setItem(row, 7, QTableWidgetItem(p.get('brand', '') or '-'))
                
                # 8: 数量（库存）- 搜索时显示units_per_carton
                units = p.get('units_per_carton', 0)
                units_item = QTableWidgetItem(str(units) if units else '0')
                units_item.setTextAlignment(Qt.AlignCenter)
                self.products_table.setItem(row, 8, units_item)
                
                # 9: 报价
                unit_price = p.get('exw_price_incl', 0) or p.get('fob_price_excl', 0) or p.get('exw_price_excl', 0) or ''
                price_text = f"{unit_price} USD" if unit_price else "-"
                price_item = QTableWidgetItem(str(price_text))
                price_item.setTextAlignment(Qt.AlignRight)
                self.products_table.setItem(row, 9, price_item)
                
                # 10: 编辑按钮
                edit_btn = QPushButton("编辑")
                edit_btn.setFixedWidth(50)
                edit_btn.clicked.connect(lambda _, prod=p: self.edit_product(prod))
                self.products_table.setCellWidget(row, 10, edit_btn)
                
            self.select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
        except Exception as e:
            print(f"搜索错误: {str(e)}")
            QMessageBox.warning(self, "错误", f"搜索产品失败: {str(e)}")

    def reset_search(self):
        self.search_input.clear()
        self.category_filter.setCurrentIndex(0)
        self.load_products()

    def load_product_categories(self):
        """从数据库加载产品类别"""
        try:
            categories = self.api_client.get("/product-categories")
            # 清空现有选项（保留"全部分类"）
            self.category_filter.clear()
            self.category_filter.addItem("全部分类", 0)
            for cat in categories:
                self.category_filter.addItem(cat.get('name', ''), cat.get('id', 0))
        except Exception as e:
            print(f"加载产品类别失败: {str(e)}")

    def toggle_select_all_products(self, state):
        for row in range(self.products_table.rowCount()):
            checkbox = self.products_table.cellWidget(row, 0)
            if checkbox:
                checkbox.setCheckState(Qt.CheckState(state))

    def get_selected_product_ids(self):
        selected_ids = []
        for row in range(self.products_table.rowCount()):
            checkbox = self.products_table.cellWidget(row, 0)
            if checkbox and checkbox.isChecked():
                id_item = self.products_table.item(row, 2)
                if id_item:
                    selected_ids.append(int(id_item.text()))
        return selected_ids

    def batch_toggle_product_status(self):
        selected_ids = self.get_selected_product_ids()
        if not selected_ids:
            QMessageBox.warning(self, "警告", "请先选择要操作的产品")
            return

        reply = QMessageBox.question(
            self, "确认操作", 
            f"确定要切换选中的 {len(selected_ids)} 个产品的状态吗？",
            QMessageBox.Ok | QMessageBox.Cancel
        )

        if reply == QMessageBox.Ok:
            try:
                for product_id in selected_ids:
                    self.api_client.update_product_status(product_id, None)
                QMessageBox.information(self, "成功", f"已成功切换 {len(selected_ids)} 个产品的状态")
                self.load_products()
            except Exception as e:
                QMessageBox.warning(self, "错误", f"批量操作失败: {str(e)}")

    def batch_delete_products(self):
        selected_ids = self.get_selected_product_ids()
        if not selected_ids:
            QMessageBox.warning(self, "警告", "请先选择要删除的产品")
            return

        reply = QMessageBox.question(
            self, "确认删除", 
            f"确定要删除选中的 {len(selected_ids)} 个产品吗？此操作不可恢复！",
            QMessageBox.Ok | QMessageBox.Cancel
        )

        if reply == QMessageBox.Ok:
            try:
                for product_id in selected_ids:
                    self.api_client.delete_product(product_id)
                QMessageBox.information(self, "成功", f"已成功删除 {len(selected_ids)} 个产品")
                self.load_products()
            except Exception as e:
                QMessageBox.warning(self, "错误", f"批量删除失败: {str(e)}")

    def view_product_images(self, product_id):
        try:
            images = self.api_client.get_product_images(product_id)
            
            if not images:
                QMessageBox.information(self, "提示", "该产品暂无图片")
                return

            dialog = QDialog(self)
            dialog.setWindowTitle("产品图片")
            dialog.setMinimumSize(800, 600)
            
            layout = QVBoxLayout()
            
            scroll_area = QScrollArea()
            scroll_area.setWidgetResizable(True)
            scroll_content = QWidget()
            scroll_layout = QVBoxLayout(scroll_content)
            
            for i, image_data in enumerate(images, 1):
                image_url = image_data.get('image_url', '')
                if image_url:
                    try:
                        response = requests.get(image_url)
                        response.raise_for_status()
                        
                        pixmap = QPixmap()
                        pixmap.loadFromData(response.content)
                        
                        if not pixmap.isNull():
                            label = QLabel(f"图片 {i}")
                            label.setFont(get_font(12, QFont.Weight.Bold))
                            scroll_layout.addWidget(label)
                            
                            image_label = QLabel()
                            image_label.setPixmap(pixmap.scaled(700, 500, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))
                            image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
                            scroll_layout.addWidget(image_label)
                        else:
                            scroll_layout.addWidget(QLabel(f"图片 {i} 加载失败"))
                    except Exception as e:
                        scroll_layout.addWidget(QLabel(f"图片 {i} 加载失败: {str(e)}"))
            
            scroll_area.setWidget(scroll_content)
            layout.addWidget(scroll_area)
            
            close_btn = QPushButton("关闭")
            close_btn.clicked.connect(dialog.close)
            layout.addWidget(close_btn)
            
            dialog.setLayout(layout)
            dialog.exec()
            
        except Exception as e:
            QMessageBox.warning(self, "错误", f"获取产品图片失败: {str(e)}")

    def load_customers(self):
        try:
            customers = self.api_client.get_customers()
            if customers is None:
                customers = []
            self.customers_table.setRowCount(len(customers))
            
            countries = sorted(set([c.get('country', '') for c in customers if c.get('country')]))
            current_country = self.customer_country_filter.currentText()
            self.customer_country_filter.clear()
            self.customer_country_filter.addItem("全部国家", 0)
            for country in countries:
                self.customer_country_filter.addItem(country, country)
            index = self.customer_country_filter.findText(current_country)
            if index >= 0:
                self.customer_country_filter.setCurrentIndex(index)
            
            # 先显示基本信息（不等待联系人和地址）
            for row, c in enumerate(customers):
                checkbox = QCheckBox()
                checkbox.setStyleSheet("margin-left: 50%;")
                self.customers_table.setCellWidget(row, 0, checkbox)
                
                self.customers_table.setItem(row, 1, QTableWidgetItem(str(c.get('id', ''))))
                self.customers_table.setItem(row, 2, QTableWidgetItem(c.get('customer_code', '')))
                self.customers_table.setItem(row, 3, QTableWidgetItem(c.get('customer_name', '')))
                self.customers_table.setItem(row, 4, QTableWidgetItem(c.get('country', '')))
                self.customers_table.setItem(row, 5, QTableWidgetItem(""))  # 默认联系人（稍后填充）
                self.customers_table.setItem(row, 6, QTableWidgetItem(""))  # 默认地址（稍后填充）
                self.customers_table.setItem(row, 7, QTableWidgetItem(c.get('payment_terms', '')))
                
                status = c.get('status', 1)
                status_text = "启用" if status == 1 else "禁用"
                status_color = "#10b981" if status == 1 else "#ef4444"
                status_item = QTableWidgetItem(status_text)
                status_item.setForeground(QBrush(QColor(status_color)))
                self.customers_table.setItem(row, 8, status_item)

                action_bar = ActionBarFactory.create_customer_action_bar(
                    edit_callback=lambda _, c=c: self.edit_customer(c),
                    toggle_callback=lambda _, c=c: self.toggle_customer_status(c),
                    status=status
                )
                self.customers_table.setCellWidget(row, 9, action_bar)
            
            self.customer_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
            
            # 异步加载联系人和地址（不阻塞UI）
            QTimer.singleShot(0, lambda: self._load_customer_extra_info(customers))
            
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载客户数据失败: {str(e)}")
    
    def _load_customer_extra_info(self, customers=None):
        """异步加载客户的联系人和地址信息"""
        try:
            # 如果没有传入客户列表，则从API获取
            if customers is None:
                customers = self.api_client.get_customers()
            if not customers:
                return
            
            customer_contacts = {}
            customer_addresses = {}
            
            for c in customers:
                try:
                    contacts = self.api_client.get_customer_contacts(c['id'])
                    if contacts:
                        primary_contact = None
                        for contact in contacts:
                            if contact.get('is_primary', 0) == 1:
                                primary_contact = contact
                                break
                        if not primary_contact and contacts:
                            primary_contact = contacts[0]
                        if primary_contact:
                            customer_contacts[c['id']] = primary_contact.get('name', '')
                except Exception:
                    pass
                
                try:
                    addresses = self.api_client.get_customer_addresses(c['id'])
                    if addresses:
                        default_addr = None
                        for addr in addresses:
                            if addr.get('is_default', 0) == 1:
                                default_addr = addr
                                break
                        if not default_addr and addresses:
                            default_addr = addresses[0]
                        if default_addr:
                            addr_text = f"{default_addr.get('country', '')}-{default_addr.get('port', '')}"
                            customer_addresses[c['id']] = addr_text
                except Exception:
                    pass
            
            # 更新表格中的联系人和地址列
            for row in range(self.customers_table.rowCount()):
                id_item = self.customers_table.item(row, 1)
                if id_item:
                    cid = int(id_item.text())
                    contact = customer_contacts.get(cid, '')
                    address = customer_addresses.get(cid, '')
                    if contact:
                        self.customers_table.setItem(row, 5, QTableWidgetItem(contact))
                    if address:
                        self.customers_table.setItem(row, 6, QTableWidgetItem(address))
        except Exception as e:
            print(f"加载客户额外信息失败: {e}")

    def search_customers(self):
        keyword = self.customer_search_input.text().strip()
        country = self.customer_country_filter.currentData()
        country = country if country != 0 else None
        
        try:
            customers = self.api_client.search_customers(keyword, country)
            self.customers_table.setRowCount(len(customers))
            
            # 先显示基本信息
            for row, c in enumerate(customers):
                checkbox = QCheckBox()
                checkbox.setStyleSheet("margin-left: 50%;")
                self.customers_table.setCellWidget(row, 0, checkbox)
                
                self.customers_table.setItem(row, 1, QTableWidgetItem(str(c.get('id', ''))))
                self.customers_table.setItem(row, 2, QTableWidgetItem(c.get('customer_code', '')))
                self.customers_table.setItem(row, 3, QTableWidgetItem(c.get('customer_name', '')))
                self.customers_table.setItem(row, 4, QTableWidgetItem(c.get('country', '')))
                self.customers_table.setItem(row, 5, QTableWidgetItem(""))  # 默认联系人（稍后填充）
                self.customers_table.setItem(row, 6, QTableWidgetItem(""))  # 默认地址（稍后填充）
                self.customers_table.setItem(row, 7, QTableWidgetItem(c.get('payment_terms', '')))
                
                status = c.get('status', 1)
                status_text = "启用" if status == 1 else "禁用"
                status_color = "#10b981" if status == 1 else "#ef4444"
                status_item = QTableWidgetItem(status_text)
                status_item.setForeground(QBrush(QColor(status_color)))
                self.customers_table.setItem(row, 8, status_item)

                action_bar = ActionBarFactory.create_customer_action_bar(
                    edit_callback=lambda _, c=c: self.edit_customer(c),
                    toggle_callback=lambda _, c=c: self.toggle_customer_status(c),
                    status=status
                )
                self.customers_table.setCellWidget(row, 9, action_bar)
            
            self.customer_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)
            
            # 异步加载联系人和地址
            QTimer.singleShot(0, lambda: self._load_customer_extra_info(customers))
        except Exception as e:
            QMessageBox.warning(self, "错误", f"搜索客户失败: {str(e)}")

    def toggle_customer_status(self, customer):
        try:
            result = self.api_client.toggle_customer_status(customer['id'])
            if result:
                self.load_customers()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"操作失败: {str(e)}")

    def on_customer_double_click(self, index):
        row = index.row()
        customer_id = self.customers_table.item(row, 1).text()
        try:
            customer = self.api_client.get_customer_detail(int(customer_id))
            self.view_customer_detail(customer)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载客户详情失败: {str(e)}")

    def on_pi_double_click(self, index):
        """双击PI行查看详情"""
        row = index.row()
        pi_id = self.pi_table.item(row, 1).text()
        try:
            pi_detail = self.api_client.get_pi_detail(int(pi_id))
            self.view_pi_detail(pi_detail)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载PI详情失败: {str(e)}")

    def view_pi_detail(self, pi):
        """查看PI详情（只读）"""
        try:
            pi_id = pi.get('id')
            pi_detail = self.api_client.get_pi_detail(pi_id)
            dialog = PIDialog(self.api_client, self.dept_id, pi_detail, readonly=True)
            dialog.exec()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"查看PI详情失败: {str(e)}")

    def view_customer_detail(self, customer):
        dialog = CustomerDetailDialog(self.api_client, customer)
        dialog.exec()

    def setup_table_context_menu(self, table, headers):
        def show_context_menu(point):
            menu = QMenu()
            row = table.rowAt(point.y())
            if row < 0:
                return
            copy_cn = menu.addAction("复制中文信息")
            copy_en = menu.addAction("Copy English Info")
            menu.addSeparator()
            copy_row = menu.addAction("复制整行数据")
            action = menu.exec(table.mapToGlobal(point))
            if action == copy_cn or action == copy_en:
                texts = []
                for col in range(table.columnCount()):
                    item = table.item(row, col)
                    if item:
                        texts.append(str(item.text()))
                    else:
                        widget = table.cellWidget(row, col)
                        if widget:
                            if isinstance(widget, QPushButton):
                                texts.append(widget.text())
                            else:
                                texts.append("")
                        else:
                            texts.append("")
                text = " | ".join(texts) if action == copy_cn else " | ".join(texts)
                from PySide6.QtGui import QGuiApplication
                clipboard = QGuiApplication.clipboard()
                clipboard.setText(text)
            elif action == copy_row:
                row_data = []
                for col in range(table.columnCount()):
                    item = table.item(row, col)
                    if item:
                        row_data.append(str(item.text()))
                    else:
                        widget = table.cellWidget(row, col)
                        if widget and isinstance(widget, QPushButton):
                            row_data.append(widget.text())
                        else:
                            row_data.append("")
                from PySide6.QtGui import QGuiApplication
                clipboard = QGuiApplication.clipboard()
                clipboard.setText("\t".join(row_data))
        table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        table.customContextMenuRequested.connect(show_context_menu)

    def toggle_select_all_customers(self, state):
        for row in range(self.customers_table.rowCount()):
            checkbox = self.customers_table.cellWidget(row, 0)
            if checkbox:
                checkbox.setCheckState(Qt.CheckState(state))

    def load_suppliers(self):
        try:
            suppliers = self.api_client.get_suppliers()
            if suppliers is None:
                suppliers = []
            self.suppliers_table.setRowCount(len(suppliers))
            for row, s in enumerate(suppliers):
                checkbox = QCheckBox()
                checkbox.setStyleSheet("margin-left: 15px;")
                self.suppliers_table.setCellWidget(row, 0, checkbox)

                self.suppliers_table.setItem(row, 1, QTableWidgetItem(str(s.get('id', ''))))
                self.suppliers_table.setItem(row, 2, QTableWidgetItem(s.get('supplier_code', '')))
                self.suppliers_table.setItem(row, 3, QTableWidgetItem(s.get('supplier_name', '')))
                self.suppliers_table.setItem(row, 4, QTableWidgetItem(s.get('region', '')))
                self.suppliers_table.setItem(row, 5, QTableWidgetItem(s.get('contact_person', '')))
                self.suppliers_table.setItem(row, 6, QTableWidgetItem(s.get('phone', '')))

                edit_btn = QPushButton("编辑")
                edit_btn.setFixedWidth(60)
                edit_btn.clicked.connect(lambda _, s=s: self.edit_supplier(s))
                self.suppliers_table.setCellWidget(row, 7, edit_btn)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载供应商数据失败: {str(e)}")

    def load_pi_orders(self):
        try:
            pi_orders = self.api_client.get_pi_orders()
            if pi_orders is None:
                pi_orders = []
            self.pi_table.setRowCount(len(pi_orders))
            status_map = {1: "草稿", 2: "已确认", 3: "已发货", 4: "已完成"}
            for row, pi in enumerate(pi_orders):
                status = pi.get('status', 1)
                is_completed = status == 4  # 已完成状态不可操作
                
                # 选择框（已完成PI不可选）
                if is_completed:
                    checkbox = QTableWidgetItem("✓")
                    checkbox.setFlags(checkbox.flags() & ~Qt.ItemFlag.ItemIsSelectable)
                    checkbox.setForeground(QColor("#9ca3af"))  # 灰色
                    self.pi_table.setItem(row, 0, checkbox)
                else:
                    checkbox = QTableWidgetItem()
                    checkbox.setFlags(checkbox.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                    checkbox.setCheckState(Qt.CheckState.Unchecked)
                    checkbox.setData(Qt.ItemDataRole.UserRole, pi.get('id'))
                    self.pi_table.setItem(row, 0, checkbox)
                
                self.pi_table.setItem(row, 1, QTableWidgetItem(str(pi.get('id', ''))))
                self.pi_table.setItem(row, 2, QTableWidgetItem(pi.get('pi_no', '')))
                self.pi_table.setItem(row, 3, QTableWidgetItem(f"{pi.get('total_amount', 0):,.2f}"))
                self.pi_table.setItem(row, 4, QTableWidgetItem(pi.get('currency', 'USD')))
                status_text = status_map.get(status, "未知")
                status_item = QTableWidgetItem(status_text)
                if is_completed:
                    status_item.setForeground(QColor("#6b7280"))  # 灰色
                self.pi_table.setItem(row, 5, status_item)
                self.pi_table.setItem(row, 6, QTableWidgetItem(str(pi.get('created_at', ''))[:19] if pi.get('created_at') else ''))

                # 操作列(7)、完成列(8)、导出列(9)
                if is_completed:
                    # 已完成PI：操作列显示"-"，完成列显示"✓"，导出列显示导出按钮
                    self.pi_table.setItem(row, 7, QTableWidgetItem("-"))
                    self.pi_table.item(row, 7).setForeground(QColor("#9ca3af"))
                    self.pi_table.setItem(row, 8, QTableWidgetItem("✓"))
                    self.pi_table.item(row, 8).setForeground(QColor("#10b981"))
                    export_btn = QPushButton("导出")
                    export_btn.setFixedWidth(50)
                    export_btn.setStyleSheet("background-color: #10b981; color: white; border: none; border-radius: 4px; padding: 2px 4px;")
                    export_btn.clicked.connect(lambda _, p=pi: self.export_pi(p))
                    self.pi_table.setCellWidget(row, 9, export_btn)
                else:
                    # 未完成PI：操作列显示编辑按钮，完成列显示完成按钮，导出列显示导出按钮
                    edit_btn = QPushButton("编辑")
                    edit_btn.setFixedWidth(50)
                    edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 4px;")
                    edit_btn.clicked.connect(lambda _, p=pi: self.edit_pi(p))
                    self.pi_table.setCellWidget(row, 7, edit_btn)
                    
                    complete_btn = QPushButton("完成")
                    complete_btn.setFixedWidth(40)
                    complete_btn.setStyleSheet("background-color: #8b5cf6; color: white; border: none; border-radius: 4px; padding: 2px 4px;")
                    complete_btn.clicked.connect(lambda _, p=pi: self.complete_pi(p))
                    self.pi_table.setCellWidget(row, 8, complete_btn)
                    
                    export_btn = QPushButton("导出")
                    export_btn.setFixedWidth(40)
                    export_btn.setStyleSheet("background-color: #10b981; color: white; border: none; border-radius: 4px; padding: 2px 4px;")
                    export_btn.clicked.connect(lambda _, p=pi: self.export_pi(p))
                    self.pi_table.setCellWidget(row, 9, export_btn)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载PI订单数据失败: {str(e)}")

    def load_purchase_orders(self):
        try:
            purchase_orders = self.api_client.get_purchase_orders()
            if purchase_orders is None:
                purchase_orders = []
            self.purchase_table.setRowCount(len(purchase_orders))
            status_map = {1: "草稿", 2: "已确认", 3: "已入库"}
            for row, po in enumerate(purchase_orders):
                self.purchase_table.setItem(row, 0, QTableWidgetItem(str(po.get('id', ''))))
                self.purchase_table.setItem(row, 1, QTableWidgetItem(po.get('po_no', '')))
                self.purchase_table.setItem(row, 2, QTableWidgetItem(po.get('pi_no', '')))
                self.purchase_table.setItem(row, 3, QTableWidgetItem(po.get('supplier_name', '')))
                self.purchase_table.setItem(row, 4, QTableWidgetItem(str(po.get('total_amount', ''))))
                self.purchase_table.setItem(row, 5, QTableWidgetItem(status_map.get(po.get('status', 1), "未知")))

                edit_btn = QPushButton("编辑")
                edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 8px;")
                edit_btn.clicked.connect(lambda _, p=po: self.edit_purchase(p))
                self.purchase_table.setCellWidget(row, 6, edit_btn)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载采购订单数据失败: {str(e)}")

    def load_shipments(self):
        try:
            shipments = self.api_client.get_shipments()
            if shipments is None:
                shipments = []
            self.shipment_table.setRowCount(len(shipments))
            status_map = {1: "待出货", 2: "出货中", 3: "已出货", 4: "已到达"}
            for row, s in enumerate(shipments):
                self.shipment_table.setItem(row, 0, QTableWidgetItem(str(s.get('id', ''))))
                self.shipment_table.setItem(row, 1, QTableWidgetItem(s.get('pi_no', '')))
                self.shipment_table.setItem(row, 2, QTableWidgetItem(str(s.get('shipment_date', ''))[:10] if s.get('shipment_date') else ''))
                self.shipment_table.setItem(row, 3, QTableWidgetItem(s.get('container_no', '')))
                self.shipment_table.setItem(row, 4, QTableWidgetItem(s.get('bl_no', '')))
                self.shipment_table.setItem(row, 5, QTableWidgetItem(status_map.get(s.get('status', 1), "未知")))

                edit_btn = QPushButton("编辑")
                edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 8px;")
                edit_btn.clicked.connect(lambda _, s=s: self.edit_shipment(s))
                self.shipment_table.setCellWidget(row, 6, edit_btn)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载出货数据失败: {str(e)}")

    def load_customer_payments(self):
        try:
            payments = self.api_client.get_customer_payments()
            if payments is None:
                payments = []
            self.customer_payment_table.setRowCount(len(payments))
            for row, p in enumerate(payments):
                self.customer_payment_table.setItem(row, 0, QTableWidgetItem(str(p.get('id', ''))))
                self.customer_payment_table.setItem(row, 1, QTableWidgetItem(p.get('pi_no', '')))
                self.customer_payment_table.setItem(row, 2, QTableWidgetItem(str(p.get('payment_date', ''))[:10] if p.get('payment_date') else ''))
                self.customer_payment_table.setItem(row, 3, QTableWidgetItem(str(p.get('actual_amount', ''))))
                self.customer_payment_table.setItem(row, 4, QTableWidgetItem(p.get('payment_method', '')))

                edit_btn = QPushButton("编辑")
                edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 8px;")
                edit_btn.clicked.connect(lambda _, p=p: self.edit_customer_payment(p))
                self.customer_payment_table.setCellWidget(row, 5, edit_btn)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载客户付款数据失败: {str(e)}")

    def load_supplier_payments(self):
        try:
            payments = self.api_client.get_supplier_payments()
            if payments is None:
                payments = []
            self.supplier_payment_table.setRowCount(len(payments))
            for row, p in enumerate(payments):
                self.supplier_payment_table.setItem(row, 0, QTableWidgetItem(str(p.get('id', ''))))
                self.supplier_payment_table.setItem(row, 1, QTableWidgetItem(p.get('supplier_name', '')))
                self.supplier_payment_table.setItem(row, 2, QTableWidgetItem(f"PO-{p.get('po_id', '')}"))
                self.supplier_payment_table.setItem(row, 3, QTableWidgetItem(p.get('payment_stage', '')))
                self.supplier_payment_table.setItem(row, 4, QTableWidgetItem(str(p.get('payment_date', ''))[:10] if p.get('payment_date') else ''))
                self.supplier_payment_table.setItem(row, 5, QTableWidgetItem(str(p.get('actual_amount', ''))))

                edit_btn = QPushButton("编辑")
                edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 8px;")
                edit_btn.clicked.connect(lambda _, p=p: self.edit_supplier_payment(p))
                self.supplier_payment_table.setCellWidget(row, 6, edit_btn)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载供应商付款数据失败: {str(e)}")

    # ========== 异步加载方法 ==========

    def _show_loading_indicator(self, table, message="加载中..."):
        """显示加载状态指示器"""
        table.setRowCount(1)
        table.setItem(0, 0, QTableWidgetItem(""))
        table.setItem(0, 1, QTableWidgetItem(message))
        for col in range(2, table.columnCount()):
            table.setItem(0, col, QTableWidgetItem(""))
        table.setEnabled(False)

    def _hide_loading_indicator(self, table):
        """隐藏加载状态指示器"""
        table.setEnabled(True)

    def load_pi_orders_async(self, force_refresh=False):
        """异步加载PI订单（带缓存和加载指示器）"""
        print("DEBUG - load_pi_orders_async started")
        # 显示加载状态（在主线程）
        self._show_loading_indicator(self.pi_table, "正在加载PI订单...")
        
        # 使用QThread来确保信号在主线程处理
        from PySide6.QtCore import QThread
        
        class PiLoaderThread(QThread):
            def __init__(self, api_client, force_refresh, parent=None):
                super().__init__(parent)
                self.api_client = api_client
                self.force_refresh = force_refresh
                self.result_data = []
                self.error_msg = None
            
            def run(self):
                try:
                    if self.force_refresh:
                        cache_manager.delete(CACHE_KEYS['PI_LIST'])
                    data = self.api_client.get_pi_orders()
                    if data:
                        cache_manager.set(CACHE_KEYS['PI_LIST'], data)
                    self.result_data = data if data else []
                except Exception as e:
                    print(f"加载PI订单失败: {e}")
                    self.error_msg = str(e)
                    self.result_data = []
        
        self._pi_loader_thread = PiLoaderThread(self.api_client, force_refresh, self)
        self._pi_loader_thread.finished.connect(
            lambda: self._on_pi_load_finished(self._pi_loader_thread.result_data)
        )
        self._pi_loader_thread.start()
        print("DEBUG - thread started")
    
    def _on_pi_load_finished(self, data):
        """PI加载完成回调"""
        print(f"DEBUG - _on_pi_load_finished with {len(data)} items")
        self._update_pi_table(data)

    def load_pi_orders(self):
        """同步加载PI订单（首次进入）"""
        try:
            # 先尝试从缓存加载
            cached = cache_manager.get(CACHE_KEYS['PI_LIST'], max_age=120)
            if cached is not None:
                self._update_pi_table(cached)
                return
            # 缓存不存在，从API加载
            pi_orders = self.api_client.get_pi_orders()
            if pi_orders:
                cache_manager.set(CACHE_KEYS['PI_LIST'], pi_orders)
            self._update_pi_table(pi_orders if pi_orders else [])
        except Exception as e:
            print(f"加载PI订单失败: {e}")
            self._update_pi_table([])

    def _update_pi_table(self, pi_orders):
        try:
            print(f"DEBUG - _update_pi_table called with {len(pi_orders) if pi_orders else 0} orders")
            print(f"DEBUG - pi_orders type: {type(pi_orders)}")
            if pi_orders and len(pi_orders) > 0:
                print(f"DEBUG - first order: {pi_orders[0]}")
            
            # 隐藏加载指示器
            self._hide_loading_indicator(self.pi_table)
            
            # 确保数据是列表
            if not pi_orders:
                pi_orders = []
            
            self.pi_table.setRowCount(len(pi_orders))
            status_map = {1: "草稿", 2: "已确认", 3: "已发货", 4: "已完成"}
            for row, pi in enumerate(pi_orders):
                status = pi.get('status', 1)
                is_completed = status == 4
                
                # 选择框（已完成PI不可选）
                if is_completed:
                    checkbox = QTableWidgetItem("✓")
                    checkbox.setFlags(checkbox.flags() & ~Qt.ItemFlag.ItemIsSelectable)
                    checkbox.setForeground(QColor("#9ca3af"))
                    self.pi_table.setItem(row, 0, checkbox)
                else:
                    checkbox = QTableWidgetItem()
                    checkbox.setFlags(checkbox.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                    checkbox.setCheckState(Qt.CheckState.Unchecked)
                    checkbox.setData(Qt.ItemDataRole.UserRole, pi.get('id'))
                    self.pi_table.setItem(row, 0, checkbox)
                
                self.pi_table.setItem(row, 1, QTableWidgetItem(str(pi.get('id', ''))))
                self.pi_table.setItem(row, 2, QTableWidgetItem(pi.get('pi_no', '')))
                self.pi_table.setItem(row, 3, QTableWidgetItem(f"{pi.get('total_amount', 0):,.2f}"))
                self.pi_table.setItem(row, 4, QTableWidgetItem(pi.get('currency', 'USD')))
                status_text = status_map.get(status, "未知")
                status_item = QTableWidgetItem(status_text)
                if is_completed:
                    status_item.setForeground(QColor("#6b7280"))
                self.pi_table.setItem(row, 5, status_item)
                self.pi_table.setItem(row, 6, QTableWidgetItem(str(pi.get('created_at', ''))[:19] if pi.get('created_at') else ''))
                
                # 操作列(7)、完成列(8)、导出列(9)
                if is_completed:
                    # 已完成PI：操作列显示"-"，完成列显示"✓"，导出列显示导出按钮
                    self.pi_table.setItem(row, 7, QTableWidgetItem("-"))
                    self.pi_table.item(row, 7).setForeground(QColor("#9ca3af"))
                    self.pi_table.setItem(row, 8, QTableWidgetItem("✓"))
                    self.pi_table.item(row, 8).setForeground(QColor("#10b981"))
                    export_btn = QPushButton("导出")
                    export_btn.setFixedWidth(50)
                    export_btn.setStyleSheet("background-color: #10b981; color: white; border: none; border-radius: 4px; padding: 2px 4px;")
                    export_btn.clicked.connect(lambda _, p=pi: self.export_pi(p))
                    self.pi_table.setCellWidget(row, 9, export_btn)
                else:
                    # 未完成PI：操作列显示编辑按钮，完成列显示完成按钮，导出列显示导出按钮
                    edit_btn = QPushButton("编辑")
                    edit_btn.setFixedWidth(50)
                    edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 4px;")
                    edit_btn.clicked.connect(lambda _, p=pi: self.edit_pi(p))
                    self.pi_table.setCellWidget(row, 7, edit_btn)
                    
                    complete_btn = QPushButton("完成")
                    complete_btn.setFixedWidth(40)
                    complete_btn.setStyleSheet("background-color: #8b5cf6; color: white; border: none; border-radius: 4px; padding: 2px 4px;")
                    complete_btn.clicked.connect(lambda _, p=pi: self.complete_pi(p))
                    self.pi_table.setCellWidget(row, 8, complete_btn)
                    
                    export_btn = QPushButton("导出")
                    export_btn.setFixedWidth(40)
                    export_btn.setStyleSheet("background-color: #10b981; color: white; border: none; border-radius: 4px; padding: 2px 4px;")
                    export_btn.clicked.connect(lambda _, p=pi: self.export_pi(p))
                    self.pi_table.setCellWidget(row, 9, export_btn)
        except Exception as e:
            print(f"更新PI订单表格失败: {e}")
            import traceback
            traceback.print_exc()

    def load_purchase_orders_async(self):
        """异步加载采购订单"""
        self._load_async(
            self.api_client.get_purchase_orders,
            self._update_purchase_table,
            "加载采购订单失败"
        )

    def _update_purchase_table(self, purchase_orders):
        try:
            self.purchase_table.setRowCount(len(purchase_orders))
            status_map = {1: "草稿", 2: "已确认", 3: "已入库"}
            for row, po in enumerate(purchase_orders):
                self.purchase_table.setItem(row, 0, QTableWidgetItem(str(po.get('id', ''))))
                self.purchase_table.setItem(row, 1, QTableWidgetItem(po.get('po_no', '')))
                self.purchase_table.setItem(row, 2, QTableWidgetItem(po.get('pi_no', '')))
                self.purchase_table.setItem(row, 3, QTableWidgetItem(po.get('supplier_name', '')))
                self.purchase_table.setItem(row, 4, QTableWidgetItem(str(po.get('total_amount', ''))))
                status = po.get('status', 1)
                status_item = QTableWidgetItem(status_map.get(status, "未知"))
                if status == 3:
                    status_item.setForeground(QBrush(QColor("#10b981")))
                elif status == 2:
                    status_item.setForeground(QBrush(QColor("#f59e0b")))
                self.purchase_table.setItem(row, 5, status_item)
                edit_btn = QPushButton("编辑")
                edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 8px;")
                edit_btn.clicked.connect(lambda _, p=po: self.edit_purchase(p))
                self.purchase_table.setCellWidget(row, 6, edit_btn)
                # 确认按钮
                confirm_btn = QPushButton("确认")
                confirm_btn.setStyleSheet("background-color: #f59e0b; color: white; border: none; border-radius: 4px; padding: 2px 8px;")
                confirm_btn.clicked.connect(lambda _, p=po: self.confirm_purchase_order(p))
                self.purchase_table.setCellWidget(row, 7, confirm_btn)
                # 入库按钮
                inbound_btn = QPushButton("入库")
                inbound_btn.setStyleSheet("background-color: #10b981; color: white; border: none; border-radius: 4px; padding: 2px 8px;")
                inbound_btn.clicked.connect(lambda _, p=po: self.inbound_purchase_order(p))
                self.purchase_table.setCellWidget(row, 8, inbound_btn)
        except Exception as e:
            print(f"更新采购订单表格失败: {e}")

    def load_shipments_async(self):
        """异步加载出货数据"""
        self._load_async(
            self.api_client.get_shipments,
            self._update_shipment_table,
            "加载出货数据失败"
        )

    def _update_shipment_table(self, shipments):
        """更新出货主表（汇总信息）"""
        try:
            self.shipment_table.setRowCount(len(shipments))
            status_map = {1: "待出货", 2: "出货中", 3: "已出货", 4: "已到达"}
            payment_status_map = {1: "未收款", 2: "部分收款", 3: "已收齐"}
            
            for row, s in enumerate(shipments):
                # ID
                self.shipment_table.setItem(row, 0, QTableWidgetItem(str(s.get('id', ''))))
                # PI号
                self.shipment_table.setItem(row, 1, QTableWidgetItem(s.get('pi_no', '')))
                # 总金额
                total_amount = s.get('total_amount', 0) or 0
                self.shipment_table.setItem(row, 2, QTableWidgetItem(f"{float(total_amount):,.2f}"))
                # 总箱数
                total_cartons = s.get('total_cartons', 0) or 0
                self.shipment_table.setItem(row, 3, QTableWidgetItem(str(total_cartons)))
                # 付款状态
                payment_status = s.get('payment_status', 1)
                payment_item = QTableWidgetItem(payment_status_map.get(payment_status, "未知"))
                if payment_status == 3:
                    payment_item.setForeground(QBrush(QColor("#10b981")))
                elif payment_status == 2:
                    payment_item.setForeground(QBrush(QColor("#f59e0b")))
                self.shipment_table.setItem(row, 4, payment_item)
                # 出货状态
                status = s.get('status', 1)
                status_item = QTableWidgetItem(status_map.get(status, "未知"))
                if status == 3:
                    status_item.setForeground(QBrush(QColor("#10b981")))
                elif status == 2:
                    status_item.setForeground(QBrush(QColor("#3b82f6")))
                self.shipment_table.setItem(row, 5, status_item)
                # 阶段数
                stages_count = s.get('stages_count', 0) or 0
                self.shipment_table.setItem(row, 6, QTableWidgetItem(str(stages_count)))
                # 操作按钮
                edit_btn = QPushButton("编辑")
                edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 8px;")
                edit_btn.clicked.connect(lambda _, s=s: self.edit_shipment(s))
                self.shipment_table.setCellWidget(row, 7, edit_btn)
                # 确认出货按钮
                confirm_btn = QPushButton("确认出货")
                confirm_btn.setStyleSheet("background-color: #10b981; color: white; border: none; border-radius: 4px; padding: 2px 8px;")
                confirm_btn.clicked.connect(lambda _, s=s: self.confirm_shipment_order(s))
                self.shipment_table.setCellWidget(row, 8, confirm_btn)
            
            # 清空阶段表
            self.shipment_stage_table.setRowCount(0)
        except Exception as e:
            print(f"更新出货表格失败: {e}")

    def show_shipment_stages(self, row, column):
        """显示选中出货记录的阶段明细"""
        try:
            shipment_id = int(self.shipment_table.item(row, 0).text())
            # 获取出货详情（包含stages）
            shipment = self.api_client.get_shipment(shipment_id)
            if not shipment:
                return
            stages = shipment.get('stages', [])
            self._update_shipment_stage_table(stages)
        except Exception as e:
            print(f"显示出货阶段失败: {e}")

    def _update_shipment_stage_table(self, stages):
        """更新出货阶段从表"""
        try:
            self.shipment_stage_table.setRowCount(len(stages))
            payment_status_map = {1: "未收款", 2: "部分收款", 3: "已收齐"}
            
            for row, s in enumerate(stages):
                # 阶段名称
                self.shipment_stage_table.setItem(row, 0, QTableWidgetItem(s.get('stage_name', '')))
                # 出货日期
                shipment_date = s.get('shipment_date', '')
                if shipment_date:
                    shipment_date = str(shipment_date)[:10]
                self.shipment_stage_table.setItem(row, 1, QTableWidgetItem(shipment_date))
                # 柜号
                self.shipment_stage_table.setItem(row, 2, QTableWidgetItem(s.get('container_no', '')))
                # 提单号
                self.shipment_stage_table.setItem(row, 3, QTableWidgetItem(s.get('bl_no', '')))
                # 数量
                quantity = s.get('quantity', 0) or 0
                self.shipment_stage_table.setItem(row, 4, QTableWidgetItem(f"{float(quantity):,.0f}"))
                # 库存
                inv_qty = s.get('inventory_quantity', 0) or 0
                inv_item = QTableWidgetItem(f"{float(inv_qty):,.0f}")
                if inv_qty > 0:
                    inv_item.setForeground(QBrush(QColor("#10b981")))
                self.shipment_stage_table.setItem(row, 5, inv_item)
                # 存放位置
                self.shipment_stage_table.setItem(row, 6, QTableWidgetItem(s.get('storage_location', '-')))
                # 付款状态
                payment_status = s.get('payment_status', 1)
                pay_item = QTableWidgetItem(payment_status_map.get(payment_status, "未知"))
                if payment_status == 3:
                    pay_item.setForeground(QBrush(QColor("#10b981")))
                elif payment_status == 2:
                    pay_item.setForeground(QBrush(QColor("#f59e0b")))
                self.shipment_stage_table.setItem(row, 7, pay_item)
                # 操作按钮
                btn_widget = QWidget()
                btn_layout = QHBoxLayout()
                btn_layout.setContentsMargins(0, 0, 0, 0)
                
                ci_btn = QPushButton("CI")
                ci_btn.setFixedWidth(40)
                ci_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 3px;")
                ci_btn.clicked.connect(lambda _, s=s: self.view_ci_document(s))
                btn_layout.addWidget(ci_btn)
                
                pl_btn = QPushButton("PL")
                pl_btn.setFixedWidth(40)
                pl_btn.setStyleSheet("background-color: #10b981; color: white; border: none; border-radius: 3px;")
                pl_btn.clicked.connect(lambda _, s=s: self.view_pl_document(s))
                btn_layout.addWidget(pl_btn)
                
                btn_widget.setLayout(btn_layout)
                self.shipment_stage_table.setCellWidget(row, 8, btn_widget)
        except Exception as e:
            print(f"更新出货阶段表格失败: {e}")

    def view_ci_document(self, stage):
        """查看CI文档"""
        ci_doc = stage.get('ci_document')
        if ci_doc:
            # TODO: 打开CI文档查看器
            QMessageBox.information(self, "CI文档", f"CI文档路径: {ci_doc}")
        else:
            QMessageBox.information(self, "CI文档", "该阶段暂无CI文档")

    def view_pl_document(self, stage):
        """查看PL文档"""
        pl_doc = stage.get('pl_document')
        if pl_doc:
            # TODO: 打开PL文档查看器
            QMessageBox.information(self, "PL文档", f"PL文档路径: {pl_doc}")
        else:
            QMessageBox.information(self, "PL文档", "该阶段暂无PL文档")

    def load_customer_payments_async(self):
        """异步加载客户付款"""
        self._load_async(
            self.api_client.get_customer_payments,
            self._update_customer_payment_table,
            "加载客户付款失败"
        )

    def _update_customer_payment_table(self, payments):
        try:
            self.customer_payment_table.setRowCount(len(payments))
            for row, p in enumerate(payments):
                self.customer_payment_table.setItem(row, 0, QTableWidgetItem(str(p.get('id', ''))))
                self.customer_payment_table.setItem(row, 1, QTableWidgetItem(p.get('pi_no', '')))
                self.customer_payment_table.setItem(row, 2, QTableWidgetItem(str(p.get('payment_date', ''))[:10] if p.get('payment_date') else ''))
                self.customer_payment_table.setItem(row, 3, QTableWidgetItem(str(p.get('actual_amount', ''))))
                self.customer_payment_table.setItem(row, 4, QTableWidgetItem(p.get('payment_method', '')))
                edit_btn = QPushButton("编辑")
                edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 8px;")
                edit_btn.clicked.connect(lambda _, p=p: self.edit_customer_payment(p))
                self.customer_payment_table.setCellWidget(row, 5, edit_btn)
        except Exception as e:
            print(f"更新客户付款表格失败: {e}")

    def load_supplier_payments_async(self):
        """异步加载供应商付款"""
        self._load_async(
            self.api_client.get_supplier_payments,
            self._update_supplier_payment_table,
            "加载供应商付款失败"
        )

    def _update_supplier_payment_table(self, payments):
        """更新供应商付款主表（汇总信息）"""
        try:
            self.supplier_payment_table.setRowCount(len(payments))
            for row, p in enumerate(payments):
                # ID
                self.supplier_payment_table.setItem(row, 0, QTableWidgetItem(str(p.get('id', ''))))
                # 供应商
                self.supplier_payment_table.setItem(row, 1, QTableWidgetItem(p.get('supplier_name', '')))
                # 采购单
                self.supplier_payment_table.setItem(row, 2, QTableWidgetItem(f"PO-{p.get('po_id', '')}"))
                # 总金额
                total = p.get('total_amount', 0) or 0
                self.supplier_payment_table.setItem(row, 3, QTableWidgetItem(f"{float(total):,.2f}"))
                # 已付金额
                paid = p.get('paid_amount', 0) or 0
                paid_item = QTableWidgetItem(f"{float(paid):,.2f}")
                if paid > 0:
                    paid_item.setForeground(QBrush(QColor("#10b981")))
                self.supplier_payment_table.setItem(row, 4, paid_item)
                # 未付金额
                unpaid = p.get('unpaid_amount', 0) or 0
                unpaid_item = QTableWidgetItem(f"{float(unpaid):,.2f}")
                if unpaid > 0:
                    unpaid_item.setForeground(QBrush(QColor("#ef4444")))
                self.supplier_payment_table.setItem(row, 5, unpaid_item)
                # 状态
                status_map = {1: "待付款", 2: "部分付款", 3: "已付清"}
                status = p.get('status', 1)
                status_text = status_map.get(status, "未知")
                status_item = QTableWidgetItem(status_text)
                if status == 3:
                    status_item.setForeground(QBrush(QColor("#10b981")))
                elif status == 2:
                    status_item.setForeground(QBrush(QColor("#f59e0b")))
                self.supplier_payment_table.setItem(row, 6, status_item)
                # 操作按钮
                edit_btn = QPushButton("编辑")
                edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 8px;")
                edit_btn.clicked.connect(lambda _, p=p: self.edit_supplier_payment(p))
                self.supplier_payment_table.setCellWidget(row, 7, edit_btn)
            # 清空阶段表
            self.supplier_payment_stage_table.setRowCount(0)
        except Exception as e:
            print(f"更新供应商付款表格失败: {e}")

    def show_supplier_payment_stages(self, row, column):
        """显示选中付款记录的阶段明细"""
        try:
            payment_id = int(self.supplier_payment_table.item(row, 0).text())
            # 获取付款详情（包含stages）
            payment = self.api_client.get_supplier_payment(payment_id)
            if not payment:
                return
            stages = payment.get('stages', [])
            self._update_supplier_payment_stage_table(stages)
        except Exception as e:
            print(f"显示付款阶段失败: {e}")

    def _update_supplier_payment_stage_table(self, stages):
        """更新付款阶段从表"""
        try:
            self.supplier_payment_stage_table.setRowCount(len(stages))
            status_map = {1: "待付", 2: "部分付", 3: "已付清"}
            for row, s in enumerate(stages):
                # 阶段名称
                self.supplier_payment_stage_table.setItem(row, 0, QTableWidgetItem(s.get('stage_name', '')))
                # 应付金额
                amount = s.get('amount', 0) or 0
                self.supplier_payment_stage_table.setItem(row, 1, QTableWidgetItem(f"{float(amount):,.2f}"))
                # 已付金额
                paid = s.get('paid_amount', 0) or 0
                paid_item = QTableWidgetItem(f"{float(paid):,.2f}")
                if paid > 0:
                    paid_item.setForeground(QBrush(QColor("#10b981")))
                self.supplier_payment_stage_table.setItem(row, 2, paid_item)
                # 状态
                status = s.get('status', 1)
                status_item = QTableWidgetItem(status_map.get(status, "未知"))
                if status == 3:
                    status_item.setForeground(QBrush(QColor("#10b981")))
                elif status == 2:
                    status_item.setForeground(QBrush(QColor("#f59e0b")))
                self.supplier_payment_stage_table.setItem(row, 3, status_item)
                # 付款日期
                payment_date = s.get('payment_date', '')
                if payment_date:
                    payment_date = str(payment_date)[:10]
                self.supplier_payment_stage_table.setItem(row, 4, QTableWidgetItem(payment_date))
                # 凭证
                has_proof = "有" if s.get('payment_proof') else "无"
                self.supplier_payment_stage_table.setItem(row, 5, QTableWidgetItem(has_proof))
                # 操作按钮
                pay_btn = QPushButton("付款")
                pay_btn.setStyleSheet("background-color: #10b981; color: white; border: none; border-radius: 4px; padding: 2px 8px;")
                pay_btn.clicked.connect(lambda _, s=s: self.pay_supplier_stage(s))
                self.supplier_payment_stage_table.setCellWidget(row, 6, pay_btn)
        except Exception as e:
            print(f"更新付款阶段表格失败: {e}")

    def load_inventories_async(self):
        """异步加载库存数据"""
        self._load_async(
            self.api_client.get_inventories,
            self._populate_inventories,
            "加载库存失败"
        )

    def load_inventories(self):
        """加载库存数据 - 按OE号分组显示（参考供应商方案模式）"""
        try:
            inventories = self.api_client.get_inventories()
            if inventories is None:
                inventories = []
            self._populate_inventories(inventories)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载库存数据失败: {str(e)}")

    def _populate_inventories(self, inventories):
        """填充库存数据到表格"""
        try:
            # 按OE号分组统计
            self.inventory_data = {}
            for inv in inventories:
                oe = inv.get('oe_number') or inv.get('product_code', '未知')
                if oe not in self.inventory_data:
                    self.inventory_data[oe] = []
                self.inventory_data[oe].append(inv)
            
            self._load_inventory_product_table()
            self.inventory_detail_table.setRowCount(0)
            self.inventory_detail_label.setText("📋 库存详情（请点击上方产品查看）")
        except Exception as e:
            print(f"填充库存数据失败: {e}")

    def load_products_async(self):
        """异步加载产品数据"""
        self._load_async(
            self.api_client.get_products,
            self.load_products_callback,
            "加载产品失败"
        )

    def load_products_callback(self, products):
        """产品数据加载完成后更新UI"""
        try:
            self.load_products_with_data(products)
        except Exception as e:
            print(f"更新产品表格失败: {e}")

    def load_products_with_data(self, products):
        """用已有数据加载产品列表（异步回调使用）"""
        try:
            inventory_summary = {}
            try:
                inventory_summary = self.api_client.get_all_inventory_summary()
            except Exception:
                pass

            if products is None:
                products = []
            self.products_table.setRowCount(len(products))
            for row, p in enumerate(products):
                product_id = p.get('id')
                
                # 0: 复选框
                checkbox = QCheckBox()
                checkbox.setStyleSheet("margin-left: 50%;")
                self.products_table.setCellWidget(row, 0, checkbox)
                
                # 获取OE和客户关联
                oe_list = []
                customer_product_list = []
                try:
                    oe_list = self.api_client.get_product_oes(product_id) or []
                    customer_product_list = self.api_client.get_product_customers(product_id) or []
                except:
                    pass
                
                # 1: 客户产品编号
                customer_product_code = ""
                if customer_product_list:
                    first_pc = customer_product_list[0]
                    full_code = first_pc.get('customer_product_code', '')
                    customer_code = first_pc.get('customer_code', '')
                    if full_code and customer_code:
                        customer_product_code = full_code.replace(customer_code, "", 1).lstrip("-")
                    else:
                        customer_product_code = full_code or ""
                self.products_table.setItem(row, 1, QTableWidgetItem(customer_product_code))
                
                # 2: OE号
                primary_oe = next((oe for oe in oe_list if oe.get('is_primary')), None)
                if len(oe_list) > 1:
                    btn = QPushButton("多OE号")
                    btn.setStyleSheet("""
                        QPushButton {
                            background-color: #3b82f6;
                            color: white;
                            border: none;
                            border-radius: 4px;
                            padding: 4px 8px;
                            font-size: 11px;
                        }
                        QPushButton:hover { background-color: #2563eb; }
                    """)
                    btn.clicked.connect(lambda checked, pid=product_id, oes=oe_list: self._show_product_oe_dialog(pid, oes))
                    self.products_table.setCellWidget(row, 2, btn)
                elif primary_oe:
                    self.products_table.setItem(row, 2, QTableWidgetItem(primary_oe.get('oe_number', '')))
                else:
                    self.products_table.setItem(row, 2, QTableWidgetItem(p.get('oe_number', '') or '-'))
                
                # 3: 图片
                image_label = QLabel()
                image_label.setFixedSize(60, 60)
                image_label.setStyleSheet("border: 1px solid #e5e7eb;")
                image_label.setAlignment(Qt.AlignCenter)
                image_url = p.get('default_image_url')
                if image_url:
                    self.load_image_async(image_label, image_url)
                else:
                    image_label.setText("暂无图片")
                image_label.setCursor(Qt.CursorShape.PointingHandCursor)
                self.products_table.setCellWidget(row, 3, image_label)
                
                # 4: 产品名称
                self.products_table.setItem(row, 4, QTableWidgetItem(p.get('detail_desc', '')))
                
                # 5: 客户型号
                customer_model = ""
                if customer_product_list:
                    customer_model = customer_product_list[0].get('customer_model', '') or ""
                self.products_table.setItem(row, 5, QTableWidgetItem(customer_model))
                
                # 6: 客户号（留空）
                self.products_table.setItem(row, 6, QTableWidgetItem(""))
                
                # 7: 产品特性（品牌）
                self.products_table.setItem(row, 7, QTableWidgetItem(p.get('brand', '') or '-'))
                
                # 8: 数量（库存）
                qty = inventory_summary.get(product_id, 0)
                qty_item = QTableWidgetItem(str(int(qty)) if qty else '0')
                if qty > 0:
                    qty_item.setForeground(QBrush(QColor("#10b981")))
                else:
                    qty_item.setForeground(QBrush(QColor("#6b7280")))
                qty_item.setTextAlignment(Qt.AlignCenter)
                self.products_table.setItem(row, 8, qty_item)
                # 9: 报价
                price = p.get('exw_price_incl', 0)
                price_text = f"{price} USD" if price else "-"
                price_item = QTableWidgetItem(price_text)
                price_item.setTextAlignment(Qt.AlignRight)
                self.products_table.setItem(row, 9, price_item)
                # 10: 操作
                action_widget = QWidget()
                action_layout = QHBoxLayout()
                action_layout.setContentsMargins(0, 0, 0, 0)
                
                edit_btn = QPushButton("编辑")
                edit_btn.setFixedWidth(50)
                edit_btn.clicked.connect(lambda _, prod=p: self.edit_product(prod))
                action_layout.addWidget(edit_btn)
                
                action_widget.setLayout(action_layout)
                self.products_table.setCellWidget(row, 10, action_widget)  # 修复：编辑按钮应该在列10
        except Exception as e:
            print(f"更新产品表格失败: {e}")

    def _load_inventory_product_table(self, data=None):
        """加载产品库存汇总表"""
        if data is None:
            data = self.inventory_data
        
        # 获取最近变更日志
            product_logs = {}
            try:
                product_logs = self.api_client.get_product_logs() or {}
                print(f"DEBUG - 获取到 {len(product_logs)} 条产品日志")
            except Exception as e:
                print(f"DEBUG - 获取产品日志失败: {e}")
            
            self.inventory_product_table.setRowCount(len(data))
        
        for row, (oe_number, records) in enumerate(data.items()):
            # 获取产品ID
            product_id = records[0].get('product_id')
            print(f"DEBUG - 产品ID={product_id}, OE={oe_number}")
            
            # OE号（缩短显示）
            oe_display = str(oe_number)
            if len(oe_display) > 18:
                oe_display = oe_display[:15] + '...'
            oe_item = QTableWidgetItem(oe_display)
            oe_item.setToolTip(str(oe_number))  # 悬停显示完整OE号
            self.inventory_product_table.setItem(row, 0, oe_item)
            
            # 产品编号（取第一条记录的product_code）
            product_code = records[0].get('product_code', '') or ''
            if len(str(product_code)) > 14:
                product_code = str(product_code)[:11] + '...'
            self.inventory_product_table.setItem(row, 1, QTableWidgetItem(str(product_code)))
            
            # 总库存量
            total_qty = sum(float(r.get('total_quantity', 0) or 0) for r in records)
            qty_item = QTableWidgetItem(str(int(total_qty)))
            # 库存为0时标红
            if total_qty == 0:
                qty_item.setForeground(QBrush(QColor("#ef4444")))
            self.inventory_product_table.setItem(row, 2, qty_item)
            
            # 供应商数
            suppliers = set(r.get('supplier_name') for r in records if r.get('supplier_name'))
            self.inventory_product_table.setItem(row, 3, QTableWidgetItem(str(len(suppliers))))
            
            # 客户数
            customers = set(r.get('customer_name') for r in records if r.get('customer_name'))
            self.inventory_product_table.setItem(row, 4, QTableWidgetItem(str(len(customers))))
            
            # 状态分布（用颜色点表示）
            status_counts = {}
            for r in records:
                st = r.get('stock_type', 1)
                if isinstance(st, str):
                    try: st = int(st)
                    except ValueError: st = 1
                status_counts[st] = status_counts.get(st, 0) + 1
            
            status_parts = []
            if 1 in status_counts:
                status_parts.append(f"🟡{status_counts[1]}")
            if 2 in status_counts:
                status_parts.append(f"🔵{status_counts[2]}")
            if 3 in status_counts:
                status_parts.append(f"🟢{status_counts[3]}")
            if 4 in status_counts:
                status_parts.append(f"⚫{status_counts[4]}")
            self.inventory_product_table.setItem(row, 5, QTableWidgetItem(" ".join(status_parts) if status_parts else "-"))
            
            # 最近入库供应商
            # product_id 是整数，但 product_logs 的键是字符串
            log_info = product_logs.get(str(product_id), {}) or product_logs.get(product_id, {})
            print(f"DEBUG - 产品{product_id}的日志: {log_info}")
            supplier_name = log_info.get('supplier_name', '') or ''
            if len(supplier_name) > 10:
                supplier_name = supplier_name[:8] + '..'
            self.inventory_product_table.setItem(row, 6, QTableWidgetItem(supplier_name or '-'))
            
            # 最近出库客户
            customer_name = log_info.get('customer_name', '') or ''
            if len(customer_name) > 10:
                customer_name = customer_name[:8] + '..'
            self.inventory_product_table.setItem(row, 7, QTableWidgetItem(customer_name or '-'))
            
            # 最近变更时间
            last_change = log_info.get('last_change_time', '') or ''
            if last_change and len(str(last_change)) > 16:
                last_change = str(last_change)[:16].replace('T', ' ')
            self.inventory_product_table.setItem(row, 8, QTableWidgetItem(last_change))
            
            # 操作按钮（展开 + 添加）
            op_widget = QWidget()
            op_layout = QHBoxLayout(op_widget)
            op_layout.setContentsMargins(2, 2, 2, 2)
            op_layout.setSpacing(4)
            
            expand_btn = QPushButton("展开")
            expand_btn.setFixedWidth(50)
            expand_btn.setStyleSheet("""
                QPushButton {
                    background-color: #10b981; color: white; border: none;
                    border-radius: 4px; padding: 4px 8px; font-size: 12px;
                }
                QPushButton:hover { background-color: #059669; }
            """)
            expand_btn.clicked.connect(lambda _, oe=oe_number: self.show_inventory_detail(oe))
            op_layout.addWidget(expand_btn)
            
            add_btn = QPushButton("+添加")
            add_btn.setFixedWidth(50)
            add_btn.setStyleSheet("""
                QPushButton {
                    background-color: #3b82f6; color: white; border: none;
                    border-radius: 4px; padding: 4px 8px; font-size: 12px;
                }
                QPushButton:hover { background-color: #2563eb; }
            """)
            add_btn.clicked.connect(lambda _, oe=oe_number, recs=records: self.add_inventory_for_oe(oe, recs))
            op_layout.addWidget(add_btn)
            
            self.inventory_product_table.setCellWidget(row, 9, op_widget)
        
        # 双击展开详情
        try:
            self.inventory_product_table.cellDoubleClicked.disconnect()
        except TypeError:
            pass
        self.inventory_product_table.cellDoubleClicked.connect(
            lambda r, c: self.show_inventory_detail(list(data.keys())[r])
        )
    
    def show_inventory_detail(self, oe_number):
        """显示指定OE号的库存详情"""
        records = self.inventory_data.get(oe_number, [])
        self.inventory_detail_label.setText(f"📋 '{oe_number}' 的库存详情（共 {len(records)} 条记录）")
        
        self.inventory_detail_table.setRowCount(len(records))
        
        for row, inv in enumerate(records):
            # ID
            self.inventory_detail_table.setItem(row, 0, QTableWidgetItem(str(inv.get('id', ''))))
            # 供应商（缩短显示）
            supplier_name = inv.get('supplier_name', '-') or '-'
            if len(str(supplier_name)) > 10:
                supplier_name = str(supplier_name)[:8] + '..'
            supplier_item = QTableWidgetItem(supplier_name)
            supplier_item.setToolTip(inv.get('supplier_name', '-') or '-')
            self.inventory_detail_table.setItem(row, 1, supplier_item)
            # 客户（缩短显示）
            customer_name = inv.get('customer_name', '-') or '-'
            if len(str(customer_name)) > 10:
                customer_name = str(customer_name)[:8] + '..'
            customer_item = QTableWidgetItem(customer_name)
            customer_item.setToolTip(inv.get('customer_name', '-') or '-')
            self.inventory_detail_table.setItem(row, 2, customer_item)
            # 数量
            self.inventory_detail_table.setItem(row, 3, QTableWidgetItem(str(int(inv.get('total_quantity', 0) or 0))))
            # 库位
            self.inventory_detail_table.setItem(row, 4, QTableWidgetItem(inv.get('current_location', '') or ''))
            # 状态颜色
            stock_type = inv.get('stock_type', 1)
            if isinstance(stock_type, str):
                try:
                    stock_type = int(stock_type)
                except (ValueError, TypeError):
                    stock_type = 1
            color_map = {1: "#ffff00", 2: "#0000ff", 3: "#00cc00", 4: "#000000"}
            color_hex = color_map.get(stock_type, "#cccccc")
            color_label = QLabel()
            color_label.setFixedSize(20, 20)
            color_label.setStyleSheet(f"background-color: {color_hex}; border: 1px solid #999; border-radius: 3px;")
            type_text_map = {1: "采购在途", 2: "待入库", 3: "已入库", 4: "历史库存"}
            color_label.setToolTip(type_text_map.get(stock_type, str(stock_type)))
            self.inventory_detail_table.setCellWidget(row, 5, color_label)
            # 备注
            remark_item = QTableWidgetItem(inv.get('remark', '') or '')
            self.inventory_detail_table.setItem(row, 6, remark_item)
            # 创建时间
            created_at = inv.get('created_at', '') or ''
            if created_at and len(str(created_at)) > 16:
                created_at = str(created_at)[:16].replace('T', ' ')
            self.inventory_detail_table.setItem(row, 7, QTableWidgetItem(str(created_at)))
            # 编辑按钮
            edit_btn = QPushButton("编辑")
            edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 8px;")
            edit_btn.clicked.connect(lambda _, inv=inv: self.edit_inventory(inv))
            self.inventory_detail_table.setCellWidget(row, 8, edit_btn)
    
    def add_inventory_for_oe(self, oe_number, existing_records):
        """为指定OE号添加库存记录"""
        # 获取产品信息
        product_id = existing_records[0].get('product_id') if existing_records else None
        if product_id:
            self.add_inventory_with_product(product_id, oe_number)
        else:
            self.add_inventory()
    
    def search_inventory(self):
        """搜索库存"""
        keyword = self.inventory_search_input.text().strip().lower()
        if not keyword:
            self._load_inventory_product_table()
            return
        
        # 过滤数据
        filtered_data = {}
        for oe_number, records in self.inventory_data.items():
            if keyword in str(oe_number).lower():
                filtered_data[oe_number] = records
            else:
                for r in records:
                    if (keyword in str(r.get('supplier_name', '')).lower() or 
                        keyword in str(r.get('customer_name', '')).lower()):
                        filtered_data[oe_number] = records
                        break
        
        self._load_inventory_product_table(filtered_data)
    
    def add_inventory_with_product(self, product_id, oe_number):
        """为指定产品添加库存"""
        dialog = InventoryDialog(self.api_client, self.dept_id, product_id=product_id, oe_number=oe_number)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_inventories()

    def toggle_product_status(self, product):
        product_id = product.get('id')
        product_code = product.get('product_code', '')
        status = product.get('status', 1)
        status_text = "禁用" if status == 1 else "启用"
        
        reply = QMessageBox.question(
            self, "确认操作", 
            f"确定要{status_text}产品 {product_code} 吗？",
            QMessageBox.Ok | QMessageBox.Cancel
        )
        
        if reply == QMessageBox.Ok:
            try:
                self.api_client.toggle_product_status(product_id)
                QMessageBox.information(self, "成功", f"产品已{status_text}")
                self.load_products()
            except Exception as e:
                QMessageBox.warning(self, "错误", f"{status_text}产品失败: {str(e)}")

    def delete_product(self, product):
        product_id = product.get('id')
        product_code = product.get('product_code', '')
        
        reply = QMessageBox.question(
            self, "确认删除", 
            f"确定要删除产品 {product_code} 吗？此操作不可恢复！",
            QMessageBox.Ok | QMessageBox.Cancel
        )
        
        if reply == QMessageBox.Ok:
            try:
                self.api_client.delete_product(product_id)
                QMessageBox.information(self, "成功", "产品已删除")
                self.load_products()
            except Exception as e:
                QMessageBox.warning(self, "错误", f"删除产品失败: {str(e)}")

    def add_product(self):
        print("add_product called, dept_id:", self.dept_id)
        try:
            dialog = ProductDialog(self.api_client, self.dept_id)
            print("ProductDialog created successfully")
            if dialog.exec() == QDialog.DialogCode.Accepted:
                self.load_products()
        except Exception as e:
            print("Error in add_product:", str(e))
            QMessageBox.warning(self, "错误", f"打开新增产品对话框失败: {str(e)}")

    def on_product_double_click(self, index):
        row = index.row()
        product_id = self.products_table.item(row, 2).text()
        try:
            product = self.api_client.get_product_detail(int(product_id))
            self.edit_product(product)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载产品信息失败: {str(e)}")

    def edit_product(self, product):
        dialog = ProductDialog(self.api_client, self.dept_id, product)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_products()

    def import_products(self):
        print("DEBUG - import_products called")
        
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择导入文件", "", "Excel文件 (*.xlsx *.xls)"
        )
        
        print(f"DEBUG - Selected file: {file_path}")
        
        if not file_path:
            return
        
        try:
            df = pd.read_excel(file_path)
            print(f"DEBUG - Excel file loaded, rows: {len(df)}")
            
            products_data = []
            for idx, row in df.iterrows():
                product = {
                    "oe_number": str(row.get("OE号", "")),
                    "factory_code": str(row.get("工厂编号", "")),
                    "brand": str(row.get("品牌", "")),
                    "detail_desc": str(row.get("细节描述", "")),
                    "category_id": int(row.get("类别ID", 1)),
                    "supplier_id": int(row.get("供应商ID", 0)) if pd.notna(row.get("供应商ID")) else None,
                    "exw_price_incl": float(row.get("EXW含税价", 0)) if pd.notna(row.get("EXW含税价")) else None,
                    "exw_price_excl": float(row.get("EXW不含税价", 0)) if pd.notna(row.get("EXW不含税价")) else None,
                    "fob_price_incl": float(row.get("FOB含税价", 0)) if pd.notna(row.get("FOB含税价")) else None,
                    "fob_price_excl": float(row.get("FOB不含税价", 0)) if pd.notna(row.get("FOB不含税价")) else None,
                    "freight": float(row.get("运费", 0)) if pd.notna(row.get("运费")) else None,
                    "packing_fee": float(row.get("包装费", 0)) if pd.notna(row.get("包装费")) else None,
                    "purchase_channel": str(row.get("采购渠道", "")),
                    "carton_length_cm": float(row.get("纸箱长(cm)", 0)) if pd.notna(row.get("纸箱长(cm)")) else None,
                    "carton_width_cm": float(row.get("纸箱宽(cm)", 0)) if pd.notna(row.get("纸箱宽(cm)")) else None,
                    "carton_height_cm": float(row.get("纸箱高(cm)", 0)) if pd.notna(row.get("纸箱高(cm)")) else None,
                    "carton_volume_cbm": float(row.get("纸箱体积(CBM)", 0)) if pd.notna(row.get("纸箱体积(CBM)")) else None,
                    "carton_weight_kg": float(row.get("纸箱重量(KG)", 0)) if pd.notna(row.get("纸箱重量(KG)")) else None,
                    "pieces_per_carton": int(row.get("每箱数量", 0)) if pd.notna(row.get("每箱数量")) else None,
                    "unit": str(row.get("单位", "件")),
                    "moq": int(row.get("最小起订量", 0)) if pd.notna(row.get("最小起订量")) else None,
                }
                products_data.append(product)
            
            print(f"DEBUG - Prepared {len(products_data)} products for import")
            print(f"DEBUG - First product data: {products_data[0] if products_data else 'None'}")
            
            result = self.api_client.import_products(products_data)
            print(f"DEBUG - Import API result: {result}")
            
            if result.get("success"):
                QMessageBox.information(self, "成功", f"成功导入 {result.get('count', 0)} 个产品")
            else:
                QMessageBox.warning(self, "导入结果", f"导入完成，部分失败: {result.get('message', '')}")
            
            self.load_products()
            
        except Exception as e:
            print(f"DEBUG - Import failed: {str(e)}")
            QMessageBox.warning(self, "错误", f"导入失败: {str(e)}")

    def add_customer(self):
        dialog = CustomerDialog(self.api_client)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            cache_manager.delete(CACHE_KEYS['CUSTOMERS'])
            self.load_customers()

    def edit_customer(self, customer):
        dialog = CustomerDialog(self.api_client, customer)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            cache_manager.delete(CACHE_KEYS['CUSTOMERS'])
            self.load_customers()

    def import_suppliers(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择导入文件", "", 
            "Excel文件 (*.xlsx *.xls);;CSV文件 (*.csv)"
        )
        
        if not file_path:
            return
            
        try:
            if file_path.endswith('.csv'):
                df = pd.read_csv(file_path, encoding='utf-8')
            else:
                df = pd.read_excel(file_path)
            
            supplier_list = []
            province_codes = self.PROVINCE_CODES
            city_data = self.CITY_DATA
            
            for _, row in df.iterrows():
                supplier_data = {}
                
                for col in df.columns:
                    col_lower = str(col).strip().lower()
                    value = row[col]
                    
                    if pd.isna(value):
                        value = None
                    else:
                        value = str(value).strip()
                    
                    if '供应商' in col_lower or '名称' in col_lower:
                        supplier_data['supplier_name'] = value
                    elif '省份' in col_lower:
                        supplier_data['province'] = value
                    elif '城市' in col_lower or '市' in col_lower:
                        supplier_data['city'] = value
                    elif '联系人' in col_lower:
                        supplier_data['contact_person'] = value
                    elif '电话' in col_lower or '手机' in col_lower:
                        supplier_data['phone'] = value
                    elif '邮箱' in col_lower or 'email' in col_lower:
                        supplier_data['email'] = value
                    elif '地址' in col_lower:
                        supplier_data['address'] = value
                
                if 'supplier_name' in supplier_data and supplier_data['supplier_name']:
                    province = supplier_data.get('province')
                    city = supplier_data.get('city')
                    
                    if province and city:
                        province_code = province_codes.get(province)
                        if province_code:
                            cities = city_data.get(province_code, {})
                            city_code = cities.get(city)
                            if city_code:
                                supplier_data['city_code'] = province_code + city_code
                    
                    supplier_list.append(supplier_data)
            
            if not supplier_list:
                QMessageBox.warning(self, "警告", "未找到有效的供应商数据")
                return
                
            progress = QProgressDialog("正在导入供应商...", "取消", 0, len(supplier_list), self)
            progress.setWindowModality(2)
            progress.show()
            
            def import_task():
                try:
                    result = self.api_client.post("/suppliers/batch", {"suppliers": supplier_list})
                    return result
                except Exception as e:
                    return {"error": str(e)}
            
            result = import_task()
            
            progress.close()
            
            if "error" in result:
                QMessageBox.critical(self, "导入失败", f"导入过程中发生错误: {result['error']}")
            else:
                success = result.get("success", 0)
                failed = result.get("failed", 0)
                msg = f"导入完成！\n成功: {success} 条\n失败: {failed} 条"
                if failed > 0:
                    failed_items = result.get("failed_items", [])
                    for item in failed_items[:5]:
                        msg += f"\n- {item['supplier_name']}: {item['error']}"
                    if len(failed_items) > 5:
                        msg += f"\n... 还有 {len(failed_items) - 5} 条失败记录"
                QMessageBox.information(self, "导入完成", msg)
                self.load_suppliers()
                
        except Exception as e:
            QMessageBox.critical(self, "导入失败", f"读取文件时发生错误: {str(e)}")

    def add_supplier(self):
        dialog = SupplierDialog(self.api_client)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            cache_manager.delete(CACHE_KEYS['SUPPLIERS'])
            self.load_suppliers()

    def on_supplier_double_click(self, index):
        row = index.row()
        supplier_id = self.suppliers_table.item(row, 1).text()
        try:
            supplier = self.api_client.get_supplier_detail(int(supplier_id))
            self.edit_supplier(supplier)
        except Exception as e:
            QMessageBox.warning(self, "错误", f"加载供应商信息失败: {str(e)}")

    def edit_supplier(self, supplier):
        dialog = SupplierDialog(self.api_client, supplier)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            cache_manager.delete(CACHE_KEYS['SUPPLIERS'])
            self.load_suppliers()

    def add_pi(self):
        dialog = PIDialog(self.api_client, self.dept_id)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_pi_orders_async()

    def edit_pi(self, pi):
        """编辑PI订单"""
        try:
            # 获取完整PI详情
            pi_id = pi.get('id')
            if not pi_id:
                QMessageBox.warning(self, "错误", "无法获取PI ID")
                return
            pi_detail = self.api_client.get_pi_detail(pi_id)
            dialog = PIDialog(self.api_client, self.dept_id, pi_detail)
            if dialog.exec() == QDialog.DialogCode.Accepted:
                self.load_pi_orders_async()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"打开PI失败: {str(e)}")

    def get_selected_pi_ids(self):
        """获取选中的PI ID列表"""
        ids = []
        for row in range(self.pi_table.rowCount()):
            item = self.pi_table.item(row, 0)
            if item and item.checkState() == Qt.CheckState.Checked:
                pi_id = item.data(Qt.ItemDataRole.UserRole)
                if pi_id:
                    ids.append(pi_id)
        return ids

    def batch_delete_pi(self):
        """批量删除PI订单"""
        selected_ids = self.get_selected_pi_ids()
        if not selected_ids:
            QMessageBox.warning(self, "提示", "请先选择要删除的PI订单")
            return
        
        reply = QMessageBox.question(
            self, "确认删除", 
            f"确定要删除选中的 {len(selected_ids)} 个PI订单吗？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        try:
            result = self.api_client.batch_delete_pi(selected_ids)
            deleted = result.get('deleted', 0)
            errors = result.get('errors', [])
            if errors:
                QMessageBox.warning(self, "部分删除失败", f"成功删除 {deleted} 个\n失败: {len(errors)} 个\n{errors}")
            else:
                QMessageBox.information(self, "成功", f"已删除 {deleted} 个PI订单")
            self.load_pi_orders_async()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"删除失败: {str(e)}")
            self.load_pi_orders_async()

    def batch_export_pi(self):
        """批量导出PI订单"""
        selected_ids = self.get_selected_pi_ids()
        if not selected_ids:
            QMessageBox.warning(self, "提示", "请先选择要导出的PI订单")
            return
        
        try:
            pi_data_list = []
            for pi_id in selected_ids:
                pi_detail = self.api_client.get_pi_detail(pi_id)
                pi_data_list.append(pi_detail)
            
            if not pi_data_list:
                QMessageBox.information(self, "提示", "没有可导出的数据")
                return
            
            # 构建导出数据
            export_rows = []
            for pi in pi_data_list:
                for item in pi.get('items', []):
                    export_rows.append({
                        'PI号': pi.get('pi_no', ''),
                        '客户ID': pi.get('customer_id', ''),
                        '总金额': pi.get('total_amount', 0),
                        '币种': pi.get('currency', 'USD'),
                        '状态': ['草稿', '已确认', '已发货', '已完成'][pi.get('status', 1) - 1] if pi.get('status', 1) <= 4 else '',
                        '创建时间': pi.get('created_at', ''),
                        '产品编号': item.get('product_id', ''),
                        'OE号': item.get('oe_number', ''),
                        '客户编号': item.get('customer_code', ''),
                        '产品描述': item.get('detail_desc', ''),
                        '数量': item.get('quantity', 0),
                        '单价': item.get('unit_price', 0),
                        '总价': item.get('total_price', 0),
                        '备注': item.get('remark', ''),
                    })
            
            df = pd.DataFrame(export_rows)
            from PySide6.QtWidgets import QFileDialog
            file_path, _ = QFileDialog.getSaveFileName(
                self, "保存PI订单", 
                f"PI订单导出_{len(selected_ids)}个.xlsx",
                "Excel Files (*.xlsx)"
            )
            if file_path:
                df.to_excel(file_path, index=False, engine='openpyxl')
                QMessageBox.information(self, "成功", f"已导出 {len(export_rows)} 条明细到:\n{file_path}")
        except Exception as e:
            QMessageBox.warning(self, "错误", f"导出失败: {str(e)}")

    def export_pi(self, pi):
        """导出PI为Excel"""
        try:
            import tempfile
            content = self.api_client.export_pi_excel(pi.get('id'))
            # 保存到临时文件
            filename = f"PI_{pi.get('pi_no', pi.get('id'))}.xlsx"
            filepath = tempfile.gettempdir() + "\\" + filename
            with open(filepath, 'wb') as f:
                f.write(content)
            # 打开文件
            import os
            os.startfile(filepath)
            QMessageBox.information(self, "成功", f"PI已导出: {filename}")
        except Exception as e:
            QMessageBox.warning(self, "错误", f"导出失败: {str(e)}")

    def complete_pi(self, pi):
        """将PI标记为已完成"""
        reply = QMessageBox.question(
            self, "确认完成",
            f"确定要将PI单 {pi.get('pi_no', '')} 标记为已完成吗？\n完成后将不可编辑。",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        if reply == QMessageBox.StandardButton.Yes:
            try:
                self.api_client.update_pi_status(pi.get('id'), 4)
                cache_manager.delete(CACHE_KEYS['PI_LIST'])
                self.load_pi_orders_async()
                QMessageBox.information(self, "成功", f"PI单 {pi.get('pi_no', '')} 已标记为完成")
            except Exception as e:
                QMessageBox.warning(self, "错误", f"操作失败: {str(e)}")

    def add_purchase(self):
        dialog = PurchaseDialog(self.api_client, self.dept_id)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_purchase_orders()

    def edit_purchase(self, purchase):
        dialog = PurchaseDialog(self.api_client, self.dept_id, purchase)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_purchase_orders_async()

    def confirm_purchase_order(self, purchase):
        """确认采购单"""
        try:
            self.api_client.confirm_purchase(purchase.get('id'))
            QMessageBox.information(self, "成功", "采购单已确认")
            self.load_purchase_orders_async()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"确认失败: {str(e)}")

    def inbound_purchase_order(self, purchase):
        """采购单入库"""
        try:
            self.api_client.inbound_purchase_order(purchase.get('id'))
            QMessageBox.information(self, "成功", "采购单已入库")
            self.load_purchase_orders_async()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"入库失败: {str(e)}")

    def add_shipment(self):
        dialog = ShipmentDialog(self.api_client, self.dept_id)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_shipments()

    def edit_shipment(self, shipment):
        dialog = ShipmentDialog(self.api_client, self.dept_id, shipment)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_shipments_async()

    def confirm_shipment_order(self, shipment):
        """确认出货"""
        try:
            self.api_client.confirm_shipment(shipment.get('id'))
            QMessageBox.information(self, "成功", "出货已确认")
            self.load_shipments_async()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"确认失败: {str(e)}")

    def add_customer_payment(self):
        dialog = CustomerPaymentDialog(self.api_client, self.dept_id)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_customer_payments()

    def edit_customer_payment(self, payment):
        dialog = CustomerPaymentDialog(self.api_client, self.dept_id, payment)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_customer_payments()

    def add_supplier_payment(self):
        dialog = SupplierPaymentDialog(self.api_client, self.dept_id)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_supplier_payments()

    def edit_supplier_payment(self, payment):
        dialog = SupplierPaymentDialog(self.api_client, self.dept_id, payment)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_supplier_payments()

    def add_inventory(self):
        dialog = InventoryDialog(self.api_client, self.dept_id)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_inventories()

    def edit_inventory(self, inventory):
        dialog = InventoryDialog(self.api_client, self.dept_id, inventory)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.load_inventories()


class QuoteDialog(QDialog):
    """报价单对话框"""
    def __init__(self, parent, api_client, dept_id, quote=None):
        super().__init__(parent)
        self.api_client = api_client
        self.dept_id = dept_id
        self.quote = quote
        self.is_edit = quote is not None
        self.customers = []
        self.products = []
        self.items = []
        self.init_ui()
        QTimer.singleShot(0, self.load_data)

    def load_data(self):
        try:
            # 先加载客户列表
            self.customers = self.api_client.get_customers()
            self.customer_combo.clear()
            self.customer_combo.addItem("", "")
            for c in self.customers:
                self.customer_combo.addItem(f"{c.get('customer_code')} - {c.get('customer_name')}", c.get('id'))
            
            if self.quote:
                # 回填客户
                customer_id = self.quote.get('customer_id')
                idx = self.customer_combo.findData(customer_id)
                if idx >= 0:
                    self.customer_combo.setCurrentIndex(idx)
                
                # 回填币种
                currency = self.quote.get('currency', 'USD')
                idx = self.currency_combo.findText(currency)
                if idx >= 0:
                    self.currency_combo.setCurrentIndex(idx)
                
                # 回填有效期
                valid_until = self.quote.get('valid_until')
                if valid_until:
                    parts = str(valid_until)[:10].split('-')
                    if len(parts) == 3:
                        from PySide6.QtCore import QDate
                        self.valid_until_input.setDate(QDate(int(parts[0]), int(parts[1]), int(parts[2])))
                
                # 回填备注
                remark = self.quote.get('remark', '')
                if remark:
                    self.remark_input.setText(str(remark))
                
                # 回填产品明细
                if 'items' in self.quote and self.quote['items']:
                    self.items = self.quote['items']
                    self.refresh_items_table()
        except Exception as e:
            print(f"加载数据失败: {e}")

    def init_ui(self):
        self.setWindowTitle("编辑报价单" if self.is_edit else "新建报价单")
        self.setMinimumSize(900, 600)
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        basic_group = QGroupBox("基本信息")
        basic_layout = QFormLayout()
        basic_layout.setSpacing(10)

        self.customer_combo = QComboBox()
        self.customer_combo.setFixedHeight(35)
        self.customer_combo.currentIndexChanged.connect(self.on_customer_changed)
        basic_layout.addRow("客户:", self.customer_combo)

        self.currency_combo = QComboBox()
        self.currency_combo.setFixedHeight(35)
        self.currency_combo.addItems(["USD", "EUR", "CNY", "GBP"])
        basic_layout.addRow("币种:", self.currency_combo)

        self.valid_until_input = QDateEdit()
        self.valid_until_input.setCalendarPopup(True)
        self.valid_until_input.setFixedHeight(35)
        self.valid_until_input.setDate(QDate.currentDate().addDays(30))
        basic_layout.addRow("有效期至:", self.valid_until_input)

        self.remark_input = QLineEdit()
        self.remark_input.setFixedHeight(35)
        basic_layout.addRow("备注:", self.remark_input)

        basic_group.setLayout(basic_layout)
        layout.addWidget(basic_group)

        import_group = QGroupBox("产品明细")
        import_layout = QVBoxLayout()

        import_toolbar = QHBoxLayout()
        import_toolbar.addStretch()

        import_btn = QPushButton("导入历史采购")
        import_btn.clicked.connect(self.import_customer_products)
        import_btn.setStyleSheet("""
            QPushButton { background-color: #10b981; color: white; border: none; border-radius: 4px; padding: 8px 16px; }
            QPushButton:hover { background-color: #059669; }
        """)
        import_toolbar.addWidget(import_btn)

        add_product_btn = QPushButton("+ 添加产品")
        add_product_btn.clicked.connect(self.add_product)
        add_product_btn.setStyleSheet("""
            QPushButton { background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 8px 16px; }
            QPushButton:hover { background-color: #2563eb; }
        """)
        import_toolbar.addWidget(add_product_btn)
        import_layout.addLayout(import_toolbar)

        self.items_table = QTableWidget()
        self.items_table.setColumnCount(8)
        self.items_table.setHorizontalHeaderLabels(["产品编号", "OE号", "客户编号", "产品描述", "数量", "单价", "总价", "操作"])
        self.items_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.items_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.items_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self.items_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.items_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.items_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.items_table.setMaximumHeight(250)
        import_layout.addWidget(self.items_table)

        summary_layout = QHBoxLayout()
        summary_layout.addStretch()
        self.total_label = QLabel("总金额: $0.00")
        self.total_label.setStyleSheet("font-weight: bold; font-size: 14px; color: #2563eb;")
        summary_layout.addWidget(self.total_label)
        import_layout.addLayout(summary_layout)

        import_group.setLayout(import_layout)
        layout.addWidget(import_group)

        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()
        save_btn = QPushButton("保存")
        save_btn.clicked.connect(self.save_quote)
        save_btn.setStyleSheet("""
            QPushButton { background-color: #2563eb; color: white; border: none; border-radius: 6px; padding: 8px 24px; }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        buttons_layout.addWidget(save_btn)
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            QPushButton { background-color: #e5e7eb; color: #374151; border: none; border-radius: 6px; padding: 8px 24px; }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        buttons_layout.addWidget(cancel_btn)
        layout.addLayout(buttons_layout)
        self.setLayout(layout)

    def on_customer_changed(self):
        customer_id = self.customer_combo.currentData()
        if customer_id:
            for c in self.customers:
                if c.get('id') == customer_id:
                    currency = c.get('currency')
                    if currency:
                        idx = self.currency_combo.findText(currency)
                        if idx >= 0:
                            self.currency_combo.setCurrentIndex(idx)
                    break

    def import_customer_products(self):
        customer_id = self.customer_combo.currentData()
        if not customer_id:
            QMessageBox.warning(self, "警告", "请先选择客户")
            return
        try:
            products = self.api_client.get_customer_products(customer_id)
            if not products:
                QMessageBox.information(self, "提示", "该客户没有采购历史记录")
                return
            for p in products:
                item = {
                    'product_id': p.get('product_id'),
                    'product_code': p.get('product_code'),
                    'oe_number': p.get('oe_number'),
                    'customer_code': p.get('customer_code'),
                    'detail_desc': p.get('detail_desc'),
                    'quantity': p.get('last_quantity') or 1,
                    'unit_price': p.get('unit_price') or 0,
                    'total_price': (p.get('last_quantity') or 1) * (p.get('unit_price') or 0)
                }
                self.items.append(item)
            self.refresh_items_table()
            QMessageBox.information(self, "成功", f"已导入 {len(products)} 个产品")
        except Exception as e:
            QMessageBox.warning(self, "错误", f"导入失败: {str(e)}")

    def add_product(self):
        customer = self.customer_combo.currentData()
        dialog = QuoteProductDialog(self, self.api_client, customer)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            product = dialog.get_product()
            if product:
                self.items.append(product)
                self.refresh_items_table()

    def edit_item(self, index):
        if index < 0 or index >= len(self.items):
            return
        item = self.items[index]
        dialog = QuoteProductDialog(self, self.api_client, None, item)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.items[index] = dialog.get_product()
            self.refresh_items_table()

    def delete_item(self, index):
        if index < 0 or index >= len(self.items):
            return
        reply = QMessageBox.question(self, "确认", "确定要删除此产品吗？")
        if reply == QMessageBox.StandardButton.Yes:
            self.items.pop(index)
            self.refresh_items_table()

    def refresh_items_table(self):
        self.items_table.setRowCount(len(self.items))
        total = 0
        for row, item in enumerate(self.items):
            self.items_table.setItem(row, 0, QTableWidgetItem(item.get('product_code', '')))
            self.items_table.setItem(row, 1, QTableWidgetItem(item.get('oe_number', '')))
            self.items_table.setItem(row, 2, QTableWidgetItem(item.get('customer_code', '')))
            self.items_table.setItem(row, 3, QTableWidgetItem(item.get('detail_desc', '')))
            self.items_table.setItem(row, 4, QTableWidgetItem(str(item.get('quantity', 0))))
            self.items_table.setItem(row, 5, QTableWidgetItem(f"${item.get('unit_price', 0):.2f}"))
            total_price = item.get('quantity', 0) * item.get('unit_price', 0)
            total += total_price
            self.items_table.setItem(row, 6, QTableWidgetItem(f"${total_price:.2f}"))
            btn_widget = QWidget()
            btn_layout = QHBoxLayout()
            btn_layout.setContentsMargins(0, 0, 0, 0)
            edit_btn = QPushButton("编辑")
            edit_btn.setFixedWidth(40)
            edit_btn.clicked.connect(lambda _, r=row: self.edit_item(r))
            btn_layout.addWidget(edit_btn)
            del_btn = QPushButton("删除")
            del_btn.setFixedWidth(40)
            del_btn.setStyleSheet("color: #ef4444;")
            del_btn.clicked.connect(lambda _, r=row: self.delete_item(r))
            btn_layout.addWidget(del_btn)
            btn_widget.setLayout(btn_layout)
            self.items_table.setCellWidget(row, 7, btn_widget)
        self.total_label.setText(f"总金额: ${total:,.2f}")

    def save_quote(self):
        customer_id = self.customer_combo.currentData()
        if not customer_id:
            QMessageBox.warning(self, "警告", "请选择客户")
            return
        if not self.items:
            QMessageBox.warning(self, "警告", "请至少添加一个产品")
            return
        quote_data = {
            "dept_id": self.dept_id,
            "customer_id": customer_id,
            "currency": self.currency_combo.currentText(),
            "valid_until": self.valid_until_input.date().toString("yyyy-MM-dd"),
            "remark": self.remark_input.text().strip(),
            "items": [{"product_id": item.get('product_id'), "oe_number": item.get('oe_number'),
                "customer_code": item.get('customer_code'), "detail_desc": item.get('detail_desc'),
                "quantity": item.get('quantity', 0), "unit_price": item.get('unit_price', 0), "remark": ""}
                for item in self.items]
        }
        try:
            if self.is_edit:
                self.api_client.update_quote(self.quote.get('id'), quote_data)
            else:
                self.api_client.create_quote(quote_data)
            QMessageBox.information(self, "成功", "报价单已保存")
            self.accept()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"保存失败: {str(e)}")


class QuoteProductDialog(QDialog):
    """报价产品对话框"""
    def __init__(self, parent, api_client, customer, item=None):
        super().__init__(parent)
        self.api_client = api_client
        self.customer = customer
        self.item = item or {}
        self.products = []
        self.init_ui()
        self.load_products()

    def load_products(self):
        try:
            self.products = self.api_client.get_products()
            self.product_combo.clear()
            self.product_combo.addItem("", None)
            for p in self.products:
                self.product_combo.addItem(f"{p.get('product_code')} - {p.get('description', '')[:20]}", p)
            if self.item.get('product_id'):
                idx = self.product_combo.findData(self.item.get('product_id'))
                if idx >= 0:
                    self.product_combo.setCurrentIndex(idx)
        except Exception as e:
            print(f"加载产品失败: {e}")

    def init_ui(self):
        self.setWindowTitle("编辑产品" if self.item else "添加产品")
        self.setFixedSize(500, 400)
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)
        form_layout = QFormLayout()
        form_layout.setSpacing(10)
        self.product_combo = QComboBox()
        self.product_combo.setFixedHeight(35)
        self.product_combo.currentIndexChanged.connect(self.on_product_selected)
        form_layout.addRow("产品:", self.product_combo)
        self.oe_number_input = QLineEdit()
        self.oe_number_input.setFixedHeight(35)
        form_layout.addRow("OE号:", self.oe_number_input)
        self.customer_code_input = QLineEdit()
        self.customer_code_input.setFixedHeight(35)
        if self.customer:
            self.customer_code_input.setText(self.customer.get('customer_code', ''))
        form_layout.addRow("客户编号:", self.customer_code_input)
        self.detail_desc_input = QLineEdit()
        self.detail_desc_input.setFixedHeight(35)
        form_layout.addRow("产品描述:", self.detail_desc_input)
        self.quantity_input = QLineEdit()
        self.quantity_input.setFixedHeight(35)
        self.quantity_input.setText(str(self.item.get('quantity', 1)))
        form_layout.addRow("数量:", self.quantity_input)
        self.unit_price_input = QLineEdit()
        self.unit_price_input.setFixedHeight(35)
        self.unit_price_input.setText(str(self.item.get('unit_price', 0)))
        form_layout.addRow("单价:", self.unit_price_input)
        layout.addLayout(form_layout)
        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()
        ok_btn = QPushButton("确定")
        ok_btn.clicked.connect(self.validate_and_accept)
        ok_btn.setStyleSheet("""
            QPushButton { background-color: #2563eb; color: white; border: none; border-radius: 6px; padding: 8px 24px; }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        buttons_layout.addWidget(ok_btn)
        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            QPushButton { background-color: #e5e7eb; color: #374151; border: none; border-radius: 6px; padding: 8px 24px; }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        buttons_layout.addWidget(cancel_btn)
        layout.addLayout(buttons_layout)
        self.setLayout(layout)

    def on_product_selected(self):
        product = self.product_combo.currentData()
        if product:
            if not self.oe_number_input.text():
                self.oe_number_input.setText(product.get('oe_number', ''))
            if not self.detail_desc_input.text():
                self.detail_desc_input.setText(product.get('description', ''))
            if self.customer and product.get('id'):
                try:
                    price_info = self.api_client.get_latest_price(self.customer.get('id'), product.get('id'))
                    if price_info and price_info.get('unit_price'):
                        self.unit_price_input.setText(str(price_info['unit_price']))
                except Exception:
                    pass

    def validate_and_accept(self):
        try:
            qty = float(self.quantity_input.text())
            price = float(self.unit_price_input.text())
            if qty <= 0:
                QMessageBox.warning(self, "警告", "数量必须大于0")
                return
            if price < 0:
                QMessageBox.warning(self, "警告", "单价不能为负")
                return
            self.accept()
        except ValueError:
            QMessageBox.warning(self, "警告", "数量和单价必须是数字")
            return

    def get_product(self):
        product = self.product_combo.currentData()
        return {
            'product_id': product.get('id') if product else None,
            'product_code': product.get('product_code', '') if product else '',
            'oe_number': self.oe_number_input.text().strip(),
            'customer_code': self.customer_code_input.text().strip(),
            'detail_desc': self.detail_desc_input.text().strip(),
            'quantity': float(self.quantity_input.text() or 0),
            'unit_price': float(self.unit_price_input.text() or 0),
            'total_price': float(self.quantity_input.text() or 0) * float(self.unit_price_input.text() or 0)
        }


class PIDialog(QDialog):
    def __init__(self, api_client, dept_id, pi=None, readonly=False):
        super().__init__()
        self.api_client = api_client
        self.dept_id = dept_id
        self.pi = pi
        self.is_edit = pi is not None
        self.readonly = readonly
        self.customers = []
        self.products = []
        self.items = []
        self.init_ui()
        QTimer.singleShot(0, self.load_data)
    
    def load_data(self):
        """加载数据（客户同步加载保证回填，产品异步加载）"""
        # 客户列表需要同步加载，因为后面要回填选中项
        try:
            cached_customers = cache_manager.get(CACHE_KEYS['CUSTOMERS'], max_age=300)
            if cached_customers is not None:
                self.customers = cached_customers
            else:
                self.customers = self.api_client.get_customers()
                cache_manager.set(CACHE_KEYS['CUSTOMERS'], self.customers)
            
            self.customer_combo.clear()
            self.customer_combo.addItem("", "")
            for c in self.customers:
                self.customer_combo.addItem(f"{c.get('customer_code')} - {c.get('customer_name')}", c.get('id'))
            
            if self.is_edit and not self.readonly:
                self.customer_combo.setEnabled(True)
                self.currency_combo.setEnabled(True)
        except Exception as e:
            print(f"加载客户失败: {e}")
        
        # 产品列表异步加载（不需要回填）
        self.load_products()
        
        # 编辑模式：回填现有数据（在客户加载完成后执行）
        if self.is_edit and self.pi:
            self._fill_existing_data()
    
    def _fill_existing_data(self):
        """回填编辑模式下的现有PI数据"""
        # 回填客户
        customer_id = self.pi.get('customer_id')
        if customer_id:
            idx = self.customer_combo.findData(customer_id)
            if idx >= 0:
                self.customer_combo.setCurrentIndex(idx)
        
        # 回填币种
        currency = self.pi.get('currency', 'USD')
        idx = self.currency_combo.findText(currency)
        if idx >= 0:
            self.currency_combo.setCurrentIndex(idx)
        
        # 回填产品明细
        items = self.pi.get('items', [])
        if items:
            self.items = []
            for item in items:
                self.items.append({
                    "product_id": item.get('product_id'),
                    "product_code": item.get('product_code', ''),
                    "oe_number": item.get('oe_number', ''),
                    "quantity": item.get('quantity', 1),
                    "unit_price": item.get('unit_price', 0),
                    "customer_code": item.get('customer_code', ''),
                    "detail_desc": item.get('detail_desc', ''),
                    "remark": item.get('remark', '')
                })
            self.update_items_table()

    def load_customers(self):
        try:
            self.customers = self.api_client.get_customers()
            self.customer_combo.clear()
            self.customer_combo.addItem("", "")
            for c in self.customers:
                self.customer_combo.addItem(f"{c.get('customer_code')} - {c.get('customer_name')}", c.get('id'))
            
            # 编辑模式下启用客户和币种选择
            if self.is_edit and not self.readonly:
                self.customer_combo.setEnabled(True)
                self.currency_combo.setEnabled(True)
        except Exception as e:
            print(f"加载客户失败: {e}")

    def load_products(self):
        """异步加载产品列表"""
        # 先显示加载中
        self.product_combo.blockSignals(True)
        self.product_combo.clear()
        self.product_combo.addItem("加载中...", None)
        self.product_combo.blockSignals(False)
        
        def fetch():
            try:
                # 尝试从缓存加载
                products = cache_manager.get(CACHE_KEYS['PRODUCTS'], max_age=300)
                if products is None:
                    products = self.api_client.get_products()
                    cache_manager.set(CACHE_KEYS['PRODUCTS'], products)
                
                self.products = products
                self.all_products = products.copy() if products else []
                
                # 用QTimer切换到主线程更新UI
                QTimer.singleShot(0, lambda: self.update_product_combo(self.all_products))
            except Exception as e:
                print(f"加载产品失败: {e}")
                QTimer.singleShot(0, lambda: self.product_combo.setItemText(0, "加载失败"))
        
        # 使用全局线程池异步加载
        _global_thread_pool.submit(fetch)
    
    def update_product_combo(self, products):
        """更新产品下拉框（分批加载避免卡顿）"""
        self.product_combo.blockSignals(True)
        self.product_combo.clear()
        self.product_combo.addItem("", None)
        
        # 限制显示数量，避免过多产品导致卡顿
        max_display = 100
        display_products = products[:max_display] if len(products) > max_display else products
        
        for p in display_products:
            product_code = p.get('product_code', '')
            oe_number = p.get('oe_number', '')
            # 截断长文本
            oe_display = oe_number[:15] + "..." if len(oe_number) > 15 else oe_number
            self.product_combo.addItem(f"{product_code} - {oe_display}", p)
        
        if len(products) > max_display:
            self.product_combo.addItem(f"...还有 {len(products) - max_display} 个产品，请使用搜索", None)
        
        self.product_combo.blockSignals(False)
    
    def filter_products(self, text):
        """根据搜索关键词过滤产品"""
        if not text:
            self.update_product_combo(self.all_products)
            return
        text = text.lower()
        filtered = [
            p for p in self.all_products 
            if text in str(p.get('product_code', '')).lower() 
            or text in str(p.get('oe_number', '')).lower()
        ]
        self.update_product_combo(filtered)
    
    def on_product_selected(self, index):
        """产品选择变化时更新图片预览和显示供应商方案"""
        product = self.product_combo.currentData()
        if product:
            # 更新图片预览
            image_url = product.get('default_image_url') or product.get('image_url')
            if image_url:
                try:
                    import urllib.request
                    image_data = urllib.request.urlopen(image_url).read()
                    image = QImage.fromData(image_data)
                    pixmap = QPixmap.fromImage(image).scaled(46, 46, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                    self.product_image_preview.setPixmap(pixmap)
                except Exception:
                    self.product_image_preview.setText("❌")
            else:
                self.product_image_preview.setText("图")
            
            # 加载供应商方案
            self.load_product_schemes(product.get('id'))
        else:
            self.product_image_preview.setText("图")
            self.scheme_row.setEnabled(False)
    
    def load_product_schemes(self, product_id):
        """加载产品的供应商方案（从PrdProductSupplier表读取）"""
        try:
            schemes = self.api_client.get_product_schemes(product_id)
            self.scheme_combo.clear()
            self.scheme_combo.addItem("-- 选择供应商方案 --", None)
            if schemes:
                for s in schemes:
                    price = s.get('purchase_price', 0) or 0
                    customer = s.get('customer_name', '通用') or '通用'
                    label = (
                        f"【{s.get('supplier_name', '供应商')}】"
                        f" ({customer})"
                        f" 价格:{price:.2f}"
                    )
                    self.scheme_combo.addItem(label, s)
                self.scheme_row.setEnabled(True)
            else:
                self.scheme_row.setEnabled(False)
        except Exception as e:
            print(f"加载供应商方案失败: {e}")
            self.scheme_row.setEnabled(False)
    
    def on_scheme_selected(self, index):
        """供应商方案选择变化"""
        pass  # 可以在这里添加预览功能
    
    def apply_scheme(self):
        """使用选中的供应商方案填充单价"""
        scheme = self.scheme_combo.currentData()
        if not scheme:
            QMessageBox.warning(self, "提示", "请先选择一个供应商方案")
            return
        
        # 使用purchase_price作为单价
        price = scheme.get('purchase_price', 0) or 0
        if price:
            self.unit_price_input.setText(f"{price:.2f}")
            QMessageBox.information(self, "成功", f"已使用【{scheme.get('supplier_name', '供应商')}】的方案\n采购价: {price:.2f} 已填入")
        else:
            QMessageBox.warning(self, "提示", "该方案尚未设置价格，请在产品管理中设置")
    
    def create_new_scheme(self):
        """为当前产品新建供应商方案（直接弹出SupplierSchemeDialog）"""
        product = self.product_combo.currentData()
        if not product:
            QMessageBox.warning(self, "警告", "请先选择一个产品")
            return
        
        try:
            # 加载供应商和客户列表
            suppliers = self.api_client.get_suppliers()
            customers = self.api_client.get_customers()
            
            # 直接弹出供应商方案编辑弹窗
            dialog = SupplierSchemeDialog(self.api_client, suppliers, customers, parent=self)
            if dialog.exec() == QDialog.DialogCode.Accepted:
                scheme_data = dialog.get_scheme_data()
                if scheme_data:
                    # 通过API创建供应商方案
                    scheme_data['product_id'] = product.get('id')
                    self.api_client.create_product_scheme(product.get('id'), scheme_data)
                    QMessageBox.information(self, "成功", "供应商方案已创建")
                    # 刷新方案列表
                    self.load_product_schemes(product.get('id'))
        except Exception as e:
            QMessageBox.warning(self, "错误", f"创建供应商方案失败: {str(e)}")

    def init_ui(self):
        self.setWindowTitle("查看PI单" if self.readonly else ("编辑PI单" if self.is_edit else "新建PI单"))
        self.setMinimumSize(850, 750)
        self.resize(850, 750)

        layout = QVBoxLayout()
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(10)

        # ===== 上部分：基本信息 =====
        basic_group = QGroupBox("基本信息")
        basic_layout = QFormLayout()
        basic_layout.setSpacing(10)

        self.customer_combo = QComboBox()
        self.customer_combo.setFixedHeight(32)
        if self.readonly:
            self.customer_combo.setEnabled(False)
        basic_layout.addRow("客户:", self.customer_combo)

        self.currency_combo = QComboBox()
        self.currency_combo.setFixedHeight(32)
        self.currency_combo.addItems(["USD", "CNY", "EUR"])
        if self.readonly:
            self.currency_combo.setEnabled(False)
        basic_layout.addRow("货币:", self.currency_combo)
        basic_group.setLayout(basic_layout)
        layout.addWidget(basic_group)

        # ===== 中部分：添加产品区域 =====
        add_product_group = QGroupBox("添加产品")
        add_product_layout = QVBoxLayout()
        add_product_layout.setSpacing(8)

        # 产品搜索和选择行
        product_row = QHBoxLayout()
        product_row.addWidget(QLabel("搜索产品:"))
        
        self.product_search = QLineEdit()
        self.product_search.setPlaceholderText("输入产品号/OE号...")
        self.product_search.setFixedWidth(150)
        self.product_search.textChanged.connect(self.filter_products)
        product_row.addWidget(self.product_search)
        
        self.product_combo = QComboBox()
        self.product_combo.setMinimumWidth(180)
        self.product_combo.currentIndexChanged.connect(self.on_product_selected)
        product_row.addWidget(self.product_combo)
        
        self.product_image_preview = QLabel()
        self.product_image_preview.setFixedSize(50, 50)
        self.product_image_preview.setStyleSheet("border: 1px solid #d1d5db; border-radius: 4px; background-color: #f9fafb;")
        self.product_image_preview.setAlignment(Qt.AlignCenter)
        self.product_image_preview.setText("图")
        product_row.addWidget(self.product_image_preview)
        
        product_row.addWidget(QLabel("数量:"))
        self.quantity_input = QLineEdit()
        self.quantity_input.setPlaceholderText("0")
        self.quantity_input.setFixedWidth(80)
        product_row.addWidget(self.quantity_input)
        
        product_row.addWidget(QLabel("单价:"))
        self.unit_price_input = QLineEdit()
        self.unit_price_input.setPlaceholderText("0.00")
        self.unit_price_input.setFixedWidth(80)
        product_row.addWidget(self.unit_price_input)
        
        if not self.readonly:
            add_item_btn = QPushButton("+ 添加")
            add_item_btn.setStyleSheet("background-color: #2563eb; color: white; border: none; border-radius: 4px; padding: 6px 16px;")
            add_item_btn.clicked.connect(self.add_item)
            product_row.addWidget(add_item_btn)
        
        product_row.addStretch()
        add_product_layout.addLayout(product_row)

        # 供应商方案选择行（选择产品后显示）
        self.scheme_row = QHBoxLayout()
        self.scheme_row.addWidget(QLabel("供应商方案:"))
        self.scheme_combo = QComboBox()
        self.scheme_combo.setMinimumWidth(200)
        self.scheme_combo.currentIndexChanged.connect(self.on_scheme_selected)
        self.scheme_row.addWidget(self.scheme_combo)
        
        use_scheme_btn = QPushButton("使用方案")
        use_scheme_btn.setStyleSheet("background-color: #059669; color: white; border: none; border-radius: 4px; padding: 6px 12px;")
        use_scheme_btn.clicked.connect(self.apply_scheme)
        self.scheme_row.addWidget(use_scheme_btn)
        
        new_scheme_btn = QPushButton("+ 新建方案")
        new_scheme_btn.setStyleSheet("background-color: #f59e0b; color: white; border: none; border-radius: 4px; padding: 6px 12px;")
        new_scheme_btn.clicked.connect(self.create_new_scheme)
        self.scheme_row.addWidget(new_scheme_btn)
        
        self.scheme_row.addStretch()
        add_product_layout.addLayout(self.scheme_row)
        
        self.scheme_row_widget = None  # 用于控制显示
        add_product_group.setLayout(add_product_layout)
        layout.addWidget(add_product_group)

        # ===== 下部分：产品明细列表 =====
        items_group = QGroupBox("产品明细列表")
        items_layout = QVBoxLayout()

        self.items_table = QTableWidget()
        self.items_table.setColumnCount(6)
        self.items_table.setHorizontalHeaderLabels(["产品编号", "OE号", "数量", "单价", "总价", "操作"])
        self.items_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.items_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.items_table.setMinimumHeight(200)
        if self.readonly:
            self.items_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        items_layout.addWidget(self.items_table)
        items_group.setLayout(items_layout)
        layout.addWidget(items_group)

        # 底部按钮
        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        if not self.readonly:
            save_btn = QPushButton("保存")
            save_btn.setFixedWidth(100)
            save_btn.clicked.connect(self.save_pi)
            save_btn.setStyleSheet("""
                QPushButton { background-color: #2563eb; color: white; border: none; border-radius: 6px; padding: 8px 24px; }
                QPushButton:hover { background-color: #1d4ed8; }
            """)
            buttons_layout.addWidget(save_btn)

        close_btn = QPushButton("关闭" if self.readonly else "取消")
        close_btn.setFixedWidth(100)
        close_btn.clicked.connect(self.accept if self.readonly else self.reject)
        close_btn.setStyleSheet("""
            QPushButton { background-color: #e5e7eb; color: #374151; border: none; border-radius: 6px; padding: 8px 24px; }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        buttons_layout.addWidget(close_btn)

        layout.addLayout(buttons_layout)
        self.setLayout(layout)

    def add_item(self):
        product = self.product_combo.currentData()
        quantity = self.quantity_input.text().strip()
        unit_price = self.unit_price_input.text().strip()

        if not product or not quantity or not unit_price:
            QMessageBox.warning(self, "警告", "请填写完整信息")
            return

        try:
            quantity = int(quantity)
            unit_price = float(unit_price)
        except ValueError:
            QMessageBox.warning(self, "警告", "数量和单价必须是数字")
            return

        self.items.append({
            "product_id": product.get('id'),
            "product_code": product.get('product_code'),
            "oe_number": product.get('oe_number'),
            "quantity": quantity,
            "unit_price": unit_price,
            "customer_code": "",
            "detail_desc": "",
            "remark": ""
        })

        self.update_items_table()

    def update_items_table(self):
        self.items_table.setRowCount(len(self.items))
        for row, item in enumerate(self.items):
            quantity = item.get('quantity', 0)
            unit_price = item.get('unit_price', 0)
            total_price = quantity * unit_price
            
            self.items_table.setItem(row, 0, QTableWidgetItem(item.get('product_code', '')))
            self.items_table.setItem(row, 1, QTableWidgetItem(item.get('oe_number', '')))
            self.items_table.setItem(row, 2, QTableWidgetItem(str(quantity)))
            self.items_table.setItem(row, 3, QTableWidgetItem(f"{unit_price:.2f}"))
            self.items_table.setItem(row, 4, QTableWidgetItem(f"{total_price:.2f}"))

            delete_btn = QPushButton("删除")
            delete_btn.setFixedWidth(50)
            delete_btn.setStyleSheet("background-color: #dc2626; color: white; border: none; border-radius: 4px; padding: 2px;")
            delete_btn.clicked.connect(lambda _, r=row: self.remove_item(r))
            self.items_table.setCellWidget(row, 5, delete_btn)

    def remove_item(self, row):
        del self.items[row]
        self.update_items_table()

    def save_pi(self):
        """异步保存PI单"""
        customer_id = self.customer_combo.currentData()
        currency = self.currency_combo.currentText()

        if not customer_id:
            QMessageBox.warning(self, "警告", "请选择客户")
            return

        if not self.items:
            QMessageBox.warning(self, "警告", "请添加产品明细")
            return

        # 禁用保存按钮，防止重复提交
        self.setEnabled(False)
        
        pi_data = {
            "dept_id": self.dept_id,
            "customer_id": customer_id,
            "currency": currency,
            "items": self.items,
            "payment_stages": []
        }

        def do_save():
            try:
                if self.is_edit:
                    result = self.api_client.update_pi(self.pi.get('id'), pi_data)
                else:
                    result = self.api_client.create_pi(pi_data)
                
                # 清除PI列表缓存
                cache_manager.delete(CACHE_KEYS['PI_LIST'])
                
                # 在主线程更新UI
                from PySide6.QtCore import QMetaObject, Qt
                QMetaObject.invokeMethod(self, "_on_save_success",
                                        Qt.ConnectionType.QueuedConnection)
            except Exception as e:
                self._save_error_msg = str(e)
                from PySide6.QtCore import QMetaObject, Qt
                QMetaObject.invokeMethod(self, "_on_save_error",
                                        Qt.ConnectionType.QueuedConnection)
        
        # 使用全局线程池异步保存
        _global_thread_pool.submit(do_save)
    
    def _on_save_success(self):
        """保存成功回调"""
        self.setEnabled(True)
        QMessageBox.information(self, "成功", "PI单已保存")
        self.accept()
    
    def _on_save_error(self):
        """保存失败回调"""
        self.setEnabled(True)
        error_msg = getattr(self, '_save_error_msg', '未知错误')
        QMessageBox.warning(self, "错误", f"保存失败: {error_msg}")


class PurchaseDialog(QDialog):
    def __init__(self, api_client, dept_id, purchase=None):
        super().__init__()
        self.api_client = api_client
        self.dept_id = dept_id
        self.purchase = purchase
        self.is_edit = purchase is not None
        self.pi_orders = []
        self.suppliers = []
        self.products = []
        self.items = []
        self.init_ui()
        QTimer.singleShot(0, self.load_data)
    
    def load_data(self):
        self.load_pi_orders()
        self.load_suppliers()
        self.load_products()

    def load_pi_orders(self):
        try:
            self.pi_orders = self.api_client.get_pi_orders()
            self.pi_combo.clear()
            self.pi_combo.addItem("", "")
            for pi in self.pi_orders:
                self.pi_combo.addItem(f"{pi.get('pi_no')} - {pi.get('total_amount')} {pi.get('currency')}", pi)
        except Exception as e:
            print(f"加载PI订单失败: {e}")

    def load_suppliers(self):
        try:
            self.suppliers = self.api_client.get_suppliers()
            self.supplier_combo.clear()
            self.supplier_combo.addItem("", "")
            for s in self.suppliers:
                self.supplier_combo.addItem(f"{s.get('supplier_code')} - {s.get('supplier_name')}", s.get('id'))
        except Exception as e:
            print(f"加载供应商失败: {e}")

    def load_products(self):
        try:
            self.products = self.api_client.get_products()
            self.product_combo.clear()
            self.product_combo.addItem("", "")
            for p in self.products:
                self.product_combo.addItem(f"{p.get('product_code')} - {p.get('oe_number')}", p)
        except Exception as e:
            print(f"加载产品失败: {e}")

    def init_ui(self):
        self.setWindowTitle("编辑采购单" if self.is_edit else "新建采购单")
        self.setFixedSize(700, 600)

        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        form_layout = QFormLayout()
        form_layout.setSpacing(10)

        self.pi_combo = QComboBox()
        self.pi_combo.setFixedHeight(35)
        form_layout.addRow("关联PI单:", self.pi_combo)

        self.supplier_combo = QComboBox()
        self.supplier_combo.setFixedHeight(35)
        form_layout.addRow("供应商:", self.supplier_combo)

        self.currency_combo = QComboBox()
        self.currency_combo.setFixedHeight(35)
        self.currency_combo.addItems(["CNY", "USD", "EUR"])
        form_layout.addRow("货币:", self.currency_combo)

        layout.addLayout(form_layout)

        items_group = QGroupBox("采购明细")
        items_layout = QVBoxLayout()

        toolbar = QHBoxLayout()
        self.product_combo = QComboBox()
        self.product_combo.setFixedWidth(150)
        toolbar.addWidget(self.product_combo)

        self.factory_code_input = QLineEdit()
        self.factory_code_input.setPlaceholderText("工厂编号")
        self.factory_code_input.setFixedWidth(100)
        toolbar.addWidget(self.factory_code_input)

        self.color_input = QLineEdit()
        self.color_input.setPlaceholderText("颜色")
        self.color_input.setFixedWidth(80)
        toolbar.addWidget(self.color_input)

        self.quantity_input = QLineEdit()
        self.quantity_input.setPlaceholderText("数量")
        self.quantity_input.setFixedWidth(80)
        toolbar.addWidget(self.quantity_input)

        self.unit_price_input = QLineEdit()
        self.unit_price_input.setPlaceholderText("单价")
        self.unit_price_input.setFixedWidth(80)
        toolbar.addWidget(self.unit_price_input)

        add_item_btn = QPushButton("+ 添加")
        add_item_btn.clicked.connect(self.add_item)
        toolbar.addWidget(add_item_btn)

        items_layout.addLayout(toolbar)

        self.items_table = QTableWidget()
        self.items_table.setColumnCount(10)
        self.items_table.setHorizontalHeaderLabels(["产品编号", "OE号", "工厂编号", "颜色", "数量", "单价", "出厂价", "出厂含税价", "FOB价", "操作"])
        self.items_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.items_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)
        self.items_table.setColumnWidth(1, 120)
        self.items_table.setFixedHeight(200)
        items_layout.addWidget(self.items_table)

        items_group.setLayout(items_layout)
        layout.addWidget(items_group)

        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        save_btn = QPushButton("保存")
        save_btn.clicked.connect(self.save_purchase)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 24px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        buttons_layout.addWidget(save_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 24px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        buttons_layout.addWidget(cancel_btn)

        layout.addLayout(buttons_layout)

        self.setLayout(layout)

    def add_item(self):
        product = self.product_combo.currentData()
        factory_code = self.factory_code_input.text().strip()
        color = self.color_input.text().strip()
        quantity = self.quantity_input.text().strip()
        unit_price = self.unit_price_input.text().strip()

        if not product or not quantity or not unit_price:
            QMessageBox.warning(self, "警告", "请填写完整信息")
            return

        try:
            quantity = int(quantity)
            unit_price = float(unit_price)
        except ValueError:
            QMessageBox.warning(self, "警告", "数量和单价必须是数字")
            return

        self.items.append({
            "product_id": product.get('id'),
            "product_code": product.get('product_code'),
            "oe_number": product.get('oe_number'),
            "factory_code": factory_code,
            "color": color,
            "detail_requirement": '',
            "quantity": quantity,
            "unit_price": unit_price,
            "price_ex_factory": product.get('exw_price_excl') or product.get('price_ex_factory'),
            "price_ex_factory_tax": product.get('exw_price_incl') or product.get('price_ex_factory_tax'),
            "price_fob": product.get('fob_price_excl') or product.get('price_fob'),
            "price_fob_tax": product.get('fob_price_incl') or product.get('price_fob_tax'),
            "remark": ""
        })

        self.factory_code_input.clear()
        self.color_input.clear()
        self.quantity_input.clear()
        self.unit_price_input.clear()

        self.update_items_table()

    def update_items_table(self):
        self.items_table.setRowCount(len(self.items))
        for row, item in enumerate(self.items):
            self.items_table.setItem(row, 0, QTableWidgetItem(item.get('product_code', '')))
            self.items_table.setItem(row, 1, QTableWidgetItem(item.get('oe_number', '')))
            self.items_table.setItem(row, 2, QTableWidgetItem(item.get('factory_code', '')))
            self.items_table.setItem(row, 3, QTableWidgetItem(item.get('color', '')))
            self.items_table.setItem(row, 4, QTableWidgetItem(str(item.get('quantity', ''))))
            self.items_table.setItem(row, 5, QTableWidgetItem(str(item.get('unit_price', ''))))
            self.items_table.setItem(row, 6, QTableWidgetItem(str(item.get('price_ex_factory', ''))))
            self.items_table.setItem(row, 7, QTableWidgetItem(str(item.get('price_ex_factory_tax', ''))))
            self.items_table.setItem(row, 8, QTableWidgetItem(str(item.get('price_fob', ''))))

            delete_btn = QPushButton("删除")
            delete_btn.setFixedWidth(50)
            delete_btn.setStyleSheet("background-color: #dc2626; color: white; border: none; border-radius: 4px; padding: 2px;")
            delete_btn.clicked.connect(lambda _, r=row: self.remove_item(r))
            self.items_table.setCellWidget(row, 9, delete_btn)

    def remove_item(self, row):
        del self.items[row]
        self.update_items_table()

    def save_purchase(self):
        pi = self.pi_combo.currentData()
        supplier_id = self.supplier_combo.currentData()
        currency = self.currency_combo.currentText()

        if not pi:
            QMessageBox.warning(self, "警告", "请选择关联的PI单")
            return

        if not supplier_id:
            QMessageBox.warning(self, "警告", "请选择供应商")
            return

        if not self.items:
            QMessageBox.warning(self, "警告", "请添加采购明细")
            return

        purchase_data = {
            "dept_id": self.dept_id,
            "pi_id": pi.get('id'),
            "supplier_id": supplier_id,
            "currency": currency,
            "items": self.items
        }

        try:
            if self.is_edit:
                self.api_client.update_purchase(self.purchase.get('id'), purchase_data)
            else:
                self.api_client.create_purchase(purchase_data)
            QMessageBox.information(self, "成功", "采购单已保存")
            self.accept()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"保存失败: {str(e)}")


class ShipmentDialog(QDialog):
    """出货对话框 - 支持多阶段管理"""
    def __init__(self, api_client, dept_id=None, shipment=None):
        super().__init__()
        self.api_client = api_client
        self.dept_id = dept_id or "S"
        self.shipment = shipment
        self.is_edit = shipment is not None
        self.pi_orders = []
        self.stages = []  # 出货阶段列表
        self.init_ui()
        QTimer.singleShot(0, self.load_data)

    def load_data(self):
        """加载PI订单数据"""
        try:
            self.pi_orders = self.api_client.get_pi_orders()
            self.pi_combo.clear()
            self.pi_combo.addItem("", "")
            for pi in self.pi_orders:
                # 显示PI号、客户名、金额
                customer_name = pi.get('customer_name', '') or ''
                display_text = f"{pi.get('pi_no')} - {customer_name} - ${pi.get('total_amount', 0)}"
                self.pi_combo.addItem(display_text, pi)
            
            # 编辑模式：回填数据
            if self.shipment:
                idx = self.pi_combo.findData(self.shipment.get('pi_id'))
                if idx >= 0:
                    self.pi_combo.setCurrentIndex(idx)
                    self.pi_combo.setEnabled(False)  # 编辑时不能修改PI
                # 加载已有的stages（如果有）
                if 'stages' in self.shipment and self.shipment['stages']:
                    self.stages = self.shipment['stages']
                    self.refresh_stages_table()
        except Exception as e:
            print(f"加载PI订单失败: {e}")

    def init_ui(self):
        """初始化UI"""
        self.setWindowTitle("编辑出货" if self.is_edit else "新建出货")
        self.setMinimumSize(800, 600)
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        # 基本信息区域
        basic_group = QGroupBox("基本信息")
        basic_layout = QFormLayout()
        basic_layout.setSpacing(10)

        self.pi_combo = QComboBox()
        self.pi_combo.setFixedHeight(35)
        basic_layout.addRow("PI单:", self.pi_combo)

        basic_group.setLayout(basic_layout)
        layout.addWidget(basic_group)

        # 出货阶段管理区域
        stages_group = QGroupBox("出货阶段管理")
        stages_layout = QVBoxLayout()

        # 阶段列表表格
        self.stages_table = QTableWidget()
        self.stages_table.setColumnCount(6)
        self.stages_table.setHorizontalHeaderLabels(["阶段名称", "出货日期", "柜号", "提单号", "数量", "操作"])
        self.stages_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.stages_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        self.stages_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.stages_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.stages_table.setMaximumHeight(250)
        stages_layout.addWidget(self.stages_table)

        # 添加阶段按钮
        add_stage_layout = QHBoxLayout()
        add_stage_layout.addStretch()

        add_stage_btn = QPushButton("+ 添加出货阶段")
        add_stage_btn.clicked.connect(self.add_stage)
        add_stage_btn.setStyleSheet("""
            QPushButton {
                background-color: #3b82f6;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 8px 16px;
            }
            QPushButton:hover { background-color: #2563eb; }
        """)
        add_stage_layout.addWidget(add_stage_btn)
        stages_layout.addLayout(add_stage_layout)

        # 汇总信息
        summary_layout = QHBoxLayout()
        summary_layout.addStretch()
        self.total_stages_label = QLabel("阶段数: 0")
        self.total_stages_label.setStyleSheet("font-weight: bold; color: #374151;")
        summary_layout.addWidget(self.total_stages_label)
        summary_layout.addSpacing(20)
        self.total_qty_label = QLabel("总数量: 0")
        self.total_qty_label.setStyleSheet("font-weight: bold; color: #10b981;")
        summary_layout.addWidget(self.total_qty_label)
        stages_layout.addLayout(summary_layout)

        stages_group.setLayout(stages_layout)
        layout.addWidget(stages_group)

        # 按钮区域
        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        save_btn = QPushButton("保存")
        save_btn.clicked.connect(self.save_shipment)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 24px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        buttons_layout.addWidget(save_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 24px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        buttons_layout.addWidget(cancel_btn)

        layout.addLayout(buttons_layout)
        self.setLayout(layout)

    def add_stage(self):
        """添加出货阶段"""
        dialog = ShipmentStageDialog(self, len(self.stages))
        if dialog.exec() == QDialog.DialogCode.Accepted:
            stage_data = dialog.get_stage_data()
            self.stages.append(stage_data)
            self.refresh_stages_table()

    def edit_stage(self, index):
        """编辑出货阶段"""
        if index < 0 or index >= len(self.stages):
            return
        dialog = ShipmentStageDialog(self, index, self.stages[index])
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.stages[index] = dialog.get_stage_data()
            self.refresh_stages_table()

    def delete_stage(self, index):
        """删除出货阶段"""
        if index < 0 or index >= len(self.stages):
            return
        reply = QMessageBox.question(self, "确认", f"确定要删除阶段 '{self.stages[index].get('stage_name')}' 吗？")
        if reply == QMessageBox.StandardButton.Yes:
            self.stages.pop(index)
            self.refresh_stages_table()

    def refresh_stages_table(self):
        """刷新阶段表格"""
        self.stages_table.setRowCount(len(self.stages))
        
        total_qty = 0
        
        for row, stage in enumerate(self.stages):
            # 阶段名称
            self.stages_table.setItem(row, 0, QTableWidgetItem(stage.get('stage_name', '')))
            # 出货日期
            self.stages_table.setItem(row, 1, QTableWidgetItem(str(stage.get('shipment_date', ''))))
            # 柜号
            self.stages_table.setItem(row, 2, QTableWidgetItem(stage.get('container_no', '')))
            # 提单号
            self.stages_table.setItem(row, 3, QTableWidgetItem(stage.get('bl_no', '')))
            # 数量
            qty = stage.get('quantity', 0) or 0
            total_qty += float(qty)
            self.stages_table.setItem(row, 4, QTableWidgetItem(str(qty)))
            
            # 操作按钮
            btn_widget = QWidget()
            btn_layout = QHBoxLayout()
            btn_layout.setContentsMargins(0, 0, 0, 0)
            
            edit_btn = QPushButton("编辑")
            edit_btn.setFixedWidth(50)
            edit_btn.clicked.connect(lambda _, r=row: self.edit_stage(r))
            btn_layout.addWidget(edit_btn)
            
            del_btn = QPushButton("删除")
            del_btn.setFixedWidth(50)
            del_btn.setStyleSheet("color: #ef4444;")
            del_btn.clicked.connect(lambda _, r=row: self.delete_stage(r))
            btn_layout.addWidget(del_btn)
            
            btn_widget.setLayout(btn_layout)
            self.stages_table.setCellWidget(row, 5, btn_widget)
        
        # 更新汇总
        self.total_stages_label.setText(f"阶段数: {len(self.stages)}")
        self.total_qty_label.setText(f"总数量: {total_qty}")

    def save_shipment(self):
        """保存出货记录"""
        pi = self.pi_combo.currentData()
        if not pi:
            QMessageBox.warning(self, "警告", "请选择PI单")
            return

        if not self.stages:
            QMessageBox.warning(self, "警告", "请至少添加一个出货阶段")
            return

        # 构建stages数据
        stages_data = []
        for stage in self.stages:
            stages_data.append({
                'id': stage.get('id'),  # 编辑时可能有id
                'stage_name': stage.get('stage_name'),
                'shipment_date': stage.get('shipment_date'),
                'container_no': stage.get('container_no'),
                'bl_no': stage.get('bl_no'),
                'quantity': stage.get('quantity'),
                'ci_document': stage.get('ci_document'),
                'pl_document': stage.get('pl_document'),
                'storage_location': stage.get('storage_location'),
                'remark': stage.get('remark')
            })

        shipment_data = {
            "dept_id": self.dept_id,
            "pi_id": pi.get('id'),
            "stages": stages_data,
            "items": []  # 出货明细，暂时为空
        }

        try:
            if self.is_edit:
                self.api_client.update_shipment(self.shipment.get('id'), shipment_data)
            else:
                self.api_client.create_shipment(shipment_data)
            QMessageBox.information(self, "成功", "出货记录已保存")
            self.accept()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"保存失败: {str(e)}")


class ShipmentStageDialog(QDialog):
    """出货阶段对话框"""
    def __init__(self, parent, stage_no, stage_data=None):
        super().__init__(parent)
        self.stage_no = stage_no
        self.stage_data = stage_data or {}
        self.is_edit = stage_data is not None
        self.init_ui()
        
        # 编辑模式回填数据
        if self.is_edit:
            self.load_stage_data()

    def init_ui(self):
        self.setWindowTitle(f"编辑出货阶段" if self.is_edit else f"添加出货阶段 #{self.stage_no + 1}")
        self.setFixedSize(500, 450)
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        form_layout = QFormLayout()
        form_layout.setSpacing(10)

        # 阶段名称
        self.stage_name_input = QLineEdit()
        self.stage_name_input.setFixedHeight(35)
        self.stage_name_input.setText(f"出货{self.stage_no + 1}")
        form_layout.addRow("阶段名称:", self.stage_name_input)

        # 出货日期
        self.shipment_date_input = QDateEdit()
        self.shipment_date_input.setCalendarPopup(True)
        self.shipment_date_input.setFixedHeight(35)
        self.shipment_date_input.setDate(QDate.currentDate())
        form_layout.addRow("出货日期:", self.shipment_date_input)

        # 柜号
        self.container_no_input = QLineEdit()
        self.container_no_input.setFixedHeight(35)
        self.container_no_input.setPlaceholderText("如: MSKU1234567")
        form_layout.addRow("柜号:", self.container_no_input)

        # 提单号
        self.bl_no_input = QLineEdit()
        self.bl_no_input.setFixedHeight(35)
        self.bl_no_input.setPlaceholderText("如: BL123456789")
        form_layout.addRow("提单号:", self.bl_no_input)

        # 数量
        self.quantity_input = QLineEdit()
        self.quantity_input.setFixedHeight(35)
        self.quantity_input.setPlaceholderText("出货数量")
        form_layout.addRow("数量:", self.quantity_input)

        # 存放位置
        self.storage_location_input = QLineEdit()
        self.storage_location_input.setFixedHeight(35)
        self.storage_location_input.setPlaceholderText("如: 上海港")
        form_layout.addRow("存放位置:", self.storage_location_input)

        # CI文件路径
        self.ci_document_input = QLineEdit()
        self.ci_document_input.setFixedHeight(35)
        self.ci_document_input.setPlaceholderText("CI文件路径或编号")
        form_layout.addRow("CI文件:", self.ci_document_input)

        # PL文件路径
        self.pl_document_input = QLineEdit()
        self.pl_document_input.setFixedHeight(35)
        self.pl_document_input.setPlaceholderText("PL文件路径或编号")
        form_layout.addRow("PL文件:", self.pl_document_input)

        # 备注
        self.remark_input = QLineEdit()
        self.remark_input.setFixedHeight(35)
        form_layout.addRow("备注:", self.remark_input)

        layout.addLayout(form_layout)

        # 按钮
        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        save_btn = QPushButton("确定")
        save_btn.clicked.connect(self.validate_and_accept)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 24px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        buttons_layout.addWidget(save_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 24px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        buttons_layout.addWidget(cancel_btn)

        layout.addLayout(buttons_layout)
        self.setLayout(layout)

    def load_stage_data(self):
        """加载阶段数据（编辑模式）"""
        self.stage_name_input.setText(self.stage_data.get('stage_name', ''))
        if self.stage_data.get('shipment_date'):
            date = QDate.fromString(str(self.stage_data['shipment_date'])[:10], "yyyy-MM-dd")
            if date.isValid():
                self.shipment_date_input.setDate(date)
        self.container_no_input.setText(self.stage_data.get('container_no', ''))
        self.bl_no_input.setText(self.stage_data.get('bl_no', ''))
        self.quantity_input.setText(str(self.stage_data.get('quantity', '')))
        self.storage_location_input.setText(self.stage_data.get('storage_location', ''))
        self.ci_document_input.setText(self.stage_data.get('ci_document', ''))
        self.pl_document_input.setText(self.stage_data.get('pl_document', ''))
        self.remark_input.setText(self.stage_data.get('remark', ''))

    def validate_and_accept(self):
        """验证并确认"""
        if not self.stage_name_input.text().strip():
            QMessageBox.warning(self, "警告", "请输入阶段名称")
            return
        
        try:
            qty = float(self.quantity_input.text() or 0)
            if qty <= 0:
                QMessageBox.warning(self, "警告", "数量必须大于0")
                return
        except ValueError:
            QMessageBox.warning(self, "警告", "数量必须是数字")
            return

        self.accept()

    def get_stage_data(self):
        """获取阶段数据"""
        return {
            'id': self.stage_data.get('id') if self.is_edit else None,
            'stage_name': self.stage_name_input.text().strip(),
            'shipment_date': self.shipment_date_input.date().toString("yyyy-MM-dd"),
            'container_no': self.container_no_input.text().strip(),
            'bl_no': self.bl_no_input.text().strip(),
            'quantity': float(self.quantity_input.text() or 0),
            'storage_location': self.storage_location_input.text().strip(),
            'ci_document': self.ci_document_input.text().strip(),
            'pl_document': self.pl_document_input.text().strip(),
            'remark': self.remark_input.text().strip()
        }


class CustomerPaymentDialog(QDialog):
    def __init__(self, api_client, dept_id=None, payment=None):
        super().__init__()
        self.api_client = api_client
        self.dept_id = dept_id or "S"
        self.payment = payment
        self.is_edit = payment is not None
        self.pi_orders = []
        self.init_ui()
        QTimer.singleShot(0, self.load_data)

    def load_data(self):
        try:
            self.pi_orders = self.api_client.get_pi_orders()
            self.pi_combo.clear()
            self.pi_combo.addItem("", "")
            for pi in self.pi_orders:
                self.pi_combo.addItem(f"{pi.get('pi_no')} - {pi.get('total_amount')}", pi)
            if self.payment:
                idx = self.pi_combo.findData(self.payment.get('pi_id'))
                if idx >= 0:
                    self.pi_combo.setCurrentIndex(idx)
        except Exception as e:
            print(f"加载PI订单失败: {e}")

    def init_ui(self):
        self.setWindowTitle("编辑客户付款" if self.is_edit else "新建客户付款")
        self.setFixedSize(500, 350)
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        form_layout = QFormLayout()
        form_layout.setSpacing(10)

        self.pi_combo = QComboBox()
        self.pi_combo.setFixedHeight(35)
        form_layout.addRow("PI单:", self.pi_combo)

        self.payment_date_input = QDateEdit()
        self.payment_date_input.setCalendarPopup(True)
        self.payment_date_input.setFixedHeight(35)
        self.payment_date_input.setDate(QDate.currentDate())
        form_layout.addRow("付款日期:", self.payment_date_input)

        self.amount_input = QLineEdit()
        self.amount_input.setFixedHeight(35)
        form_layout.addRow("付款金额:", self.amount_input)

        self.payment_method_combo = QComboBox()
        self.payment_method_combo.setFixedHeight(35)
        self.payment_method_combo.addItems(["银行转账", "现金", "支票", "其他"])
        form_layout.addRow("付款方式:", self.payment_method_combo)

        self.remark_input = QTextEdit()
        self.remark_input.setFixedHeight(80)
        form_layout.addRow("备注:", self.remark_input)

        layout.addLayout(form_layout)

        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        save_btn = QPushButton("保存")
        save_btn.clicked.connect(self.save_payment)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 24px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        buttons_layout.addWidget(save_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 24px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        buttons_layout.addWidget(cancel_btn)

        layout.addLayout(buttons_layout)
        self.setLayout(layout)

        if self.payment:
            self.payment_date_input.setDate(QDate.fromString(self.payment.get('payment_date', ''), "yyyy-MM-dd"))
            self.amount_input.setText(str(self.payment.get('actual_amount', '')))
            method_map = {"银行转账": 0, "现金": 1, "支票": 2, "其他": 3}
            self.payment_method_combo.setCurrentIndex(method_map.get(self.payment.get('payment_method', '银行转账'), 0))
            self.remark_input.setPlainText(self.payment.get('remark', ''))

    def save_payment(self):
        pi = self.pi_combo.currentData()
        if not pi:
            QMessageBox.warning(self, "警告", "请选择PI单")
            return

        amount = self.amount_input.text().strip()
        if not amount:
            QMessageBox.warning(self, "警告", "请输入付款金额")
            return

        try:
            amount = float(amount)
        except ValueError:
            QMessageBox.warning(self, "警告", "付款金额必须是数字")
            return

        payment_data = {
            "dept_id": self.dept_id,
            "pi_id": pi.get('id'),
            "customer_id": pi.get('customer_id'),
            "amount": amount,
            "actual_amount": amount,
            "payment_date": self.payment_date_input.date().toString("yyyy-MM-dd"),
            "payment_method": self.payment_method_combo.currentText(),
            "remark": self.remark_input.toPlainText()
        }

        try:
            if self.is_edit:
                self.api_client.update_customer_payment(self.payment.get('id'), payment_data)
            else:
                self.api_client.create_customer_payment(payment_data)
            QMessageBox.information(self, "成功", "付款记录已保存")
            self.accept()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"保存失败: {str(e)}")


class SupplierPaymentDialog(QDialog):
    """供应商付款对话框 - 支持多阶段管理"""
    def __init__(self, api_client, dept_id=None, payment=None):
        super().__init__()
        self.api_client = api_client
        self.dept_id = dept_id or "S"
        self.payment = payment
        self.is_edit = payment is not None
        self.suppliers = []
        self.purchases = []
        self.stages = []  # 付款阶段列表
        self.init_ui()
        QTimer.singleShot(0, self.load_data)

    def load_data(self):
        """加载供应商和采购单数据"""
        try:
            self.suppliers = self.api_client.get_suppliers()
            self.supplier_combo.clear()
            self.supplier_combo.addItem("", "")
            for s in self.suppliers:
                self.supplier_combo.addItem(f"{s.get('supplier_code')} - {s.get('supplier_name')}", s.get('id'))
            
            # 编辑模式：回填数据
            if self.payment:
                idx = self.supplier_combo.findData(self.payment.get('supplier_id'))
                if idx >= 0:
                    self.supplier_combo.setCurrentIndex(idx)
                # 加载已有的stages
                if 'stages' in self.payment and self.payment['stages']:
                    self.stages = self.payment['stages']
                    self.refresh_stages_table()
        except Exception as e:
            print(f"加载供应商失败: {e}")

    def load_purchases(self):
        """加载供应商的采购单"""
        supplier_id = self.supplier_combo.currentData()
        if not supplier_id:
            return
        try:
            purchases = self.api_client.get_purchases_by_supplier(supplier_id)
            self.purchase_combo.clear()
            self.purchase_combo.addItem("", "")
            for p in purchases:
                self.purchase_combo.addItem(f"PO-{p.get('id')} - {p.get('total_amount', 0)}", p)
            # 编辑模式：回填采购单
            if self.payment and self.payment.get('po_id'):
                for i in range(self.purchase_combo.count()):
                    data = self.purchase_combo.itemData(i)
                    if data and data.get('id') == self.payment.get('po_id'):
                        self.purchase_combo.setCurrentIndex(i)
                        break
        except Exception as e:
            print(f"加载采购单失败: {e}")

    def init_ui(self):
        """初始化UI"""
        self.setWindowTitle("编辑供应商付款" if self.is_edit else "新建供应商付款")
        self.setMinimumSize(700, 600)
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        # 基本信息区域
        basic_group = QGroupBox("基本信息")
        basic_layout = QFormLayout()
        basic_layout.setSpacing(10)

        self.supplier_combo = QComboBox()
        self.supplier_combo.setFixedHeight(35)
        self.supplier_combo.currentIndexChanged.connect(self.load_purchases)
        basic_layout.addRow("供应商:", self.supplier_combo)

        self.purchase_combo = QComboBox()
        self.purchase_combo.setFixedHeight(35)
        basic_layout.addRow("采购单:", self.purchase_combo)

        self.payment_method_combo = QComboBox()
        self.payment_method_combo.setFixedHeight(35)
        self.payment_method_combo.addItems(["银行转账", "现金", "支票", "其他"])
        basic_layout.addRow("付款方式:", self.payment_method_combo)

        basic_group.setLayout(basic_layout)
        layout.addWidget(basic_group)

        # 付款阶段管理区域
        stages_group = QGroupBox("付款阶段管理")
        stages_layout = QVBoxLayout()

        # 阶段列表表格
        self.stages_table = QTableWidget()
        self.stages_table.setColumnCount(5)
        self.stages_table.setHorizontalHeaderLabels(["阶段名称", "应付金额", "已付金额", "状态", "操作"])
        self.stages_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.stages_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.stages_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.stages_table.setMaximumHeight(200)
        stages_layout.addWidget(self.stages_table)

        # 添加阶段按钮区域
        add_stage_layout = QHBoxLayout()
        add_stage_layout.addStretch()

        add_deposit_btn = QPushButton("+ 添加定金")
        add_deposit_btn.clicked.connect(lambda: self.add_stage('deposit'))
        add_deposit_btn.setStyleSheet("""
            QPushButton {
                background-color: #f59e0b;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 12px;
            }
            QPushButton:hover { background-color: #d97706; }
        """)
        add_stage_layout.addWidget(add_deposit_btn)

        add_balance_btn = QPushButton("+ 添加尾款")
        add_balance_btn.clicked.connect(lambda: self.add_stage('balance'))
        add_balance_btn.setStyleSheet("""
            QPushButton {
                background-color: #3b82f6;
                color: white;
                border: none;
                border-radius: 4px;
                padding: 6px 12px;
            }
            QPushButton:hover { background-color: #2563eb; }
        """)
        add_stage_layout.addWidget(add_balance_btn)

        stages_layout.addLayout(add_stage_layout)

        # 汇总信息
        summary_layout = QHBoxLayout()
        summary_layout.addStretch()
        self.total_label = QLabel("总金额: 0.00")
        self.total_label.setStyleSheet("font-weight: bold; color: #374151;")
        summary_layout.addWidget(self.total_label)
        summary_layout.addSpacing(20)
        self.paid_label = QLabel("已付: 0.00")
        self.paid_label.setStyleSheet("font-weight: bold; color: #10b981;")
        summary_layout.addWidget(self.paid_label)
        summary_layout.addSpacing(20)
        self.unpaid_label = QLabel("未付: 0.00")
        self.unpaid_label.setStyleSheet("font-weight: bold; color: #ef4444;")
        summary_layout.addWidget(self.unpaid_label)
        stages_layout.addLayout(summary_layout)

        stages_group.setLayout(stages_layout)
        layout.addWidget(stages_group)

        # 按钮区域
        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        save_btn = QPushButton("保存")
        save_btn.clicked.connect(self.save_payment)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 24px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        buttons_layout.addWidget(save_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 24px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        buttons_layout.addWidget(cancel_btn)

        layout.addLayout(buttons_layout)
        self.setLayout(layout)

        # 编辑模式：回填付款方式
        if self.payment:
            method_map = {"银行转账": 0, "现金": 1, "支票": 2, "其他": 3}
            self.payment_method_combo.setCurrentIndex(method_map.get(self.payment.get('payment_method', '银行转账'), 0))

    def add_stage(self, stage_type):
        """添加付款阶段"""
        dialog = SupplierPaymentStageDialog(self, stage_type, len(self.stages))
        if dialog.exec() == QDialog.DialogCode.Accepted:
            stage_data = dialog.get_stage_data()
            self.stages.append(stage_data)
            self.refresh_stages_table()

    def edit_stage(self, index):
        """编辑付款阶段"""
        if index < 0 or index >= len(self.stages):
            return
        dialog = SupplierPaymentStageDialog(self, self.stages[index].get('stage_type', 'balance'), 
                                           index, self.stages[index])
        if dialog.exec() == QDialog.DialogCode.Accepted:
            self.stages[index] = dialog.get_stage_data()
            self.refresh_stages_table()

    def delete_stage(self, index):
        """删除付款阶段"""
        if index < 0 or index >= len(self.stages):
            return
        reply = QMessageBox.question(self, "确认", f"确定要删除阶段 '{self.stages[index].get('stage_name')}' 吗？")
        if reply == QMessageBox.StandardButton.Yes:
            self.stages.pop(index)
            self.refresh_stages_table()

    def refresh_stages_table(self):
        """刷新阶段表格"""
        self.stages_table.setRowCount(len(self.stages))
        status_map = {1: "待付", 2: "部分付", 3: "已付清"}
        
        total = 0
        paid = 0
        
        for row, stage in enumerate(self.stages):
            # 阶段名称
            self.stages_table.setItem(row, 0, QTableWidgetItem(stage.get('stage_name', '')))
            # 应付金额
            amount = stage.get('amount', 0) or 0
            total += float(amount)
            self.stages_table.setItem(row, 1, QTableWidgetItem(f"{float(amount):,.2f}"))
            # 已付金额
            stage_paid = stage.get('paid_amount', 0) or 0
            paid += float(stage_paid)
            self.stages_table.setItem(row, 2, QTableWidgetItem(f"{float(stage_paid):,.2f}"))
            # 状态
            status = stage.get('status', 1)
            self.stages_table.setItem(row, 3, QTableWidgetItem(status_map.get(status, "未知")))
            # 操作按钮
            btn_widget = QWidget()
            btn_layout = QHBoxLayout()
            btn_layout.setContentsMargins(0, 0, 0, 0)
            
            edit_btn = QPushButton("编辑")
            edit_btn.setFixedWidth(50)
            edit_btn.clicked.connect(lambda _, r=row: self.edit_stage(r))
            btn_layout.addWidget(edit_btn)
            
            del_btn = QPushButton("删除")
            del_btn.setFixedWidth(50)
            del_btn.setStyleSheet("color: #ef4444;")
            del_btn.clicked.connect(lambda _, r=row: self.delete_stage(r))
            btn_layout.addWidget(del_btn)
            
            btn_widget.setLayout(btn_layout)
            self.stages_table.setCellWidget(row, 4, btn_widget)
        
        # 更新汇总
        unpaid = total - paid
        self.total_label.setText(f"总金额: {total:,.2f}")
        self.paid_label.setText(f"已付: {paid:,.2f}")
        self.unpaid_label.setText(f"未付: {unpaid:,.2f}")

    def save_payment(self):
        """保存付款记录"""
        supplier_id = self.supplier_combo.currentData()
        purchase = self.purchase_combo.currentData()

        if not supplier_id:
            QMessageBox.warning(self, "警告", "请选择供应商")
            return

        if not purchase:
            QMessageBox.warning(self, "警告", "请选择采购单")
            return

        if not self.stages:
            QMessageBox.warning(self, "警告", "请至少添加一个付款阶段")
            return

        # 构建stages数据
        stages_data = []
        for stage in self.stages:
            stages_data.append({
                'id': stage.get('id'),  # 编辑时可能有id
                'stage_type': stage.get('stage_type'),
                'stage_name': stage.get('stage_name'),
                'amount': stage.get('amount'),
                'paid_amount': stage.get('paid_amount', 0),
                'status': stage.get('status', 1),
                'payment_date': stage.get('payment_date'),
                'payment_proof': stage.get('payment_proof'),
                'remark': stage.get('remark')
            })

        payment_data = {
            "dept_id": self.dept_id,
            "supplier_id": supplier_id,
            "po_id": purchase.get('id'),
            "payment_method": self.payment_method_combo.currentText(),
            "stages": stages_data,
            "remark": ""
        }

        try:
            print(f"DEBUG - 保存供应商付款: {payment_data}")
            if self.is_edit:
                result = self.api_client.update_supplier_payment(self.payment.get('id'), payment_data)
            else:
                result = self.api_client.create_supplier_payment(payment_data)
            print(f"DEBUG - 保存成功: {result}")
            QMessageBox.information(self, "成功", "付款记录已保存")
            self.accept()
        except Exception as e:
            print(f"DEBUG - 保存失败: {e}")
            import traceback
            traceback.print_exc()
            QMessageBox.warning(self, "错误", f"保存失败: {str(e)}")


class SupplierPaymentStageDialog(QDialog):
    """付款阶段编辑对话框"""
    def __init__(self, parent, stage_type, index, stage_data=None):
        super().__init__(parent)
        self.stage_type = stage_type
        self.index = index
        self.stage_data = stage_data or {}
        self.is_edit = stage_data is not None
        self.init_ui()

    def init_ui(self):
        stage_type_name = "定金" if self.stage_type == 'deposit' else f"尾款{self.index}"
        self.setWindowTitle(f"编辑{stage_type_name}" if self.is_edit else f"添加{stage_type_name}")
        self.setFixedSize(400, 350)
        
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        form_layout = QFormLayout()
        form_layout.setSpacing(10)

        # 阶段名称
        self.name_input = QLineEdit()
        self.name_input.setFixedHeight(35)
        default_name = self.stage_data.get('stage_name', '')
        if not default_name:
            default_name = "定金" if self.stage_type == 'deposit' else f"尾款{self.index + 1}"
        self.name_input.setText(default_name)
        form_layout.addRow("阶段名称:", self.name_input)

        # 应付金额
        self.amount_input = QLineEdit()
        self.amount_input.setFixedHeight(35)
        self.amount_input.setText(str(self.stage_data.get('amount', '')))
        form_layout.addRow("应付金额:", self.amount_input)

        # 已付金额
        self.paid_input = QLineEdit()
        self.paid_input.setFixedHeight(35)
        self.paid_input.setText(str(self.stage_data.get('paid_amount', '0')))
        form_layout.addRow("已付金额:", self.paid_input)

        # 付款日期
        self.date_input = QDateEdit()
        self.date_input.setCalendarPopup(True)
        self.date_input.setFixedHeight(35)
        payment_date = self.stage_data.get('payment_date')
        if payment_date:
            self.date_input.setDate(QDate.fromString(str(payment_date)[:10], "yyyy-MM-dd"))
        else:
            self.date_input.setDate(QDate.currentDate())
        form_layout.addRow("付款日期:", self.date_input)

        # 备注
        self.remark_input = QTextEdit()
        self.remark_input.setFixedHeight(60)
        self.remark_input.setPlainText(self.stage_data.get('remark', ''))
        form_layout.addRow("备注:", self.remark_input)

        layout.addLayout(form_layout)

        # 按钮
        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        save_btn = QPushButton("确定")
        save_btn.clicked.connect(self.accept)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 24px;
            }
        """)
        buttons_layout.addWidget(save_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 24px;
            }
        """)
        buttons_layout.addWidget(cancel_btn)

        layout.addLayout(buttons_layout)
        self.setLayout(layout)

    def get_stage_data(self):
        """获取阶段数据"""
        amount = float(self.amount_input.text() or 0)
        paid = float(self.paid_input.text() or 0)
        
        # 自动计算状态
        status = 1  # 待付
        if paid >= amount and amount > 0:
            status = 3  # 已付清
        elif paid > 0:
            status = 2  # 部分付

        return {
            'id': self.stage_data.get('id'),  # 编辑时保留原id
            'stage_type': self.stage_type,
            'stage_name': self.name_input.text(),
            'amount': amount,
            'paid_amount': paid,
            'status': status,
            'payment_date': self.date_input.date().toString("yyyy-MM-dd"),
            'remark': self.remark_input.toPlainText()
        }


class InventoryDialog(QDialog):
    def __init__(self, api_client, dept_id=None, inventory=None, product_id=None, oe_number=None):
        super().__init__()
        self.api_client = api_client
        self.dept_id = dept_id or "S"
        self.inventory = inventory
        self.preselected_product_id = product_id  # 预选的产品ID
        self.preselected_oe_number = oe_number    # 预选的OE号
        self.is_edit = inventory is not None
        self.products = []
        self.init_ui()
        QTimer.singleShot(0, self.load_products)

    def load_products(self):
        try:
            # 尝试从缓存加载
            self.products = cache_manager.get(CACHE_KEYS['PRODUCTS'], max_age=300)
            if self.products is None:
                self.products = self.api_client.get_products()
                cache_manager.set(CACHE_KEYS['PRODUCTS'], self.products)
            
            self.product_combo.clear()
            self.product_combo.addItem("请选择产品", None)
            
            # 如果有预选的OE号，只显示匹配的产品
            if self.preselected_oe_number:
                for p in self.products:
                    if p.get('oe_number') == self.preselected_oe_number:
                        self.product_combo.addItem(f"{p.get('product_code')} - {p.get('oe_number')}", p)
                        # 自动选中
                        self.product_combo.setCurrentIndex(1)
                        break
            else:
                for p in self.products:
                    self.product_combo.addItem(f"{p.get('product_code')} - {p.get('oe_number')}", p)
            
            # 编辑模式：选中当前库存的产品
            if self.inventory:
                for i in range(self.product_combo.count()):
                    data = self.product_combo.itemData(i)
                    if data and data.get('id') == self.inventory.get('product_id'):
                        self.product_combo.setCurrentIndex(i)
                        break
            # 新建模式：有预选产品ID
            elif self.preselected_product_id:
                for i in range(self.product_combo.count()):
                    data = self.product_combo.itemData(i)
                    if data and data.get('id') == self.preselected_product_id:
                        self.product_combo.setCurrentIndex(i)
                        break
        except Exception as e:
            print(f"加载产品失败: {e}")
    
    def load_suppliers(self):
        """加载供应商列表"""
        try:
            # 尝试从缓存加载
            suppliers = cache_manager.get(CACHE_KEYS['SUPPLIERS'], max_age=300)
            if suppliers is None:
                suppliers = self.api_client.get_suppliers()
                cache_manager.set(CACHE_KEYS['SUPPLIERS'], suppliers)
            
            self.supplier_combo.clear()
            self.supplier_combo.addItem("请选择供应商", None)
            for s in suppliers:
                self.supplier_combo.addItem(f"{s.get('supplier_code')} - {s.get('supplier_name')}", s)
            
            # 如果是编辑，设置选中的供应商
            if self.inventory and self.inventory.get('supplier_id'):
                for i in range(self.supplier_combo.count()):
                    data = self.supplier_combo.itemData(i)
                    if data and data.get('id') == self.inventory.get('supplier_id'):
                        self.supplier_combo.setCurrentIndex(i)
                        break
        except Exception as e:
            print(f"加载供应商失败: {e}")
    
    def load_customers(self):
        """加载客户列表"""
        try:
            # 尝试从缓存加载
            customers = cache_manager.get(CACHE_KEYS['CUSTOMERS'], max_age=300)
            if customers is None:
                customers = self.api_client.get_customers()
                cache_manager.set(CACHE_KEYS['CUSTOMERS'], customers)
            
            self.customer_combo.clear()
            self.customer_combo.addItem("请选择客户", None)
            for c in customers:
                self.customer_combo.addItem(f"{c.get('customer_code')} - {c.get('customer_name')}", c)
            
            # 如果是编辑，设置选中的客户
            if self.inventory and self.inventory.get('customer_id'):
                for i in range(self.customer_combo.count()):
                    data = self.customer_combo.itemData(i)
                    if data and data.get('id') == self.inventory.get('customer_id'):
                        self.customer_combo.setCurrentIndex(i)
                        break
        except Exception as e:
            print(f"加载客户失败: {e}")

    def init_ui(self):
        self.setWindowTitle("编辑库存" if self.is_edit else "新建库存")
        self.setFixedSize(550, 550)
        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)

        form_layout = QFormLayout()
        form_layout.setSpacing(10)

        self.product_combo = QComboBox()
        self.product_combo.setFixedHeight(35)
        form_layout.addRow("产品 *:", self.product_combo)

        # 供应商选择
        self.supplier_combo = QComboBox()
        self.supplier_combo.setFixedHeight(35)
        form_layout.addRow("供应商 *:", self.supplier_combo)
        # 加载供应商列表
        self.load_suppliers()

        # 客户选择
        self.customer_combo = QComboBox()
        self.customer_combo.setFixedHeight(35)
        form_layout.addRow("客户 *:", self.customer_combo)
        # 加载客户列表
        self.load_customers()

        self.stock_type_combo = QComboBox()
        self.stock_type_combo.setFixedHeight(35)
        self.stock_type_combo.addItems(["采购在途", "待入库", "已入库", "历史库存"])
        form_layout.addRow("库存类型:", self.stock_type_combo)

        self.quantity_input = QLineEdit()
        self.quantity_input.setFixedHeight(35)
        form_layout.addRow("数量:", self.quantity_input)

        self.location_input = QLineEdit()
        self.location_input.setFixedHeight(35)
        form_layout.addRow("库位:", self.location_input)

        self.remark_input = QTextEdit()
        self.remark_input.setFixedHeight(80)
        form_layout.addRow("备注:", self.remark_input)

        layout.addLayout(form_layout)

        status_group = QGroupBox("库存状态颜色")
        status_layout = QHBoxLayout()
        status_layout.addWidget(QLabel("● 黄色: 采购在途"))
        status_layout.addWidget(QLabel("● 蓝色: 待入库"))
        status_layout.addWidget(QLabel("● 绿色: 已入库"))
        status_layout.addWidget(QLabel("● 黑色: 历史库存"))
        status_group.setLayout(status_layout)
        layout.addWidget(status_group)

        buttons_layout = QHBoxLayout()
        buttons_layout.addStretch()

        save_btn = QPushButton("保存")
        save_btn.clicked.connect(self.save_inventory)
        save_btn.setStyleSheet("""
            QPushButton {
                background-color: #2563eb;
                color: white;
                border: none;
                border-radius: 6px;
                padding: 8px 24px;
            }
            QPushButton:hover { background-color: #1d4ed8; }
        """)
        buttons_layout.addWidget(save_btn)

        cancel_btn = QPushButton("取消")
        cancel_btn.clicked.connect(self.reject)
        cancel_btn.setStyleSheet("""
            QPushButton {
                background-color: #e5e7eb;
                color: #374151;
                border: none;
                border-radius: 6px;
                padding: 8px 24px;
            }
            QPushButton:hover { background-color: #d1d5db; }
        """)
        buttons_layout.addWidget(cancel_btn)

        layout.addLayout(buttons_layout)
        self.setLayout(layout)

        if self.inventory:
            # 回填数量（使用 total_quantity 或 quantity）
            qty = self.inventory.get('total_quantity') or self.inventory.get('quantity', '')
            self.quantity_input.setText(str(qty))
            # 回填库位
            self.location_input.setText(self.inventory.get('current_location', '') or '')
            # 回填备注
            self.remark_input.setPlainText(self.inventory.get('remark', '') or '')
            # 回填库存类型
            stock_type = self.inventory.get('stock_type', 2)
            # stock_type 转下拉框索引
            index_map = {1: 0, 2: 1, 3: 2, 4: 3}  # 1->采购在途(0), 2->待入库(1), 3->已入库(2), 4->历史库存(3)
            if isinstance(stock_type, str):
                try:
                    stock_type = int(stock_type)
                except ValueError:
                    stock_type = 2
            combo_index = index_map.get(stock_type, 0)
            self.stock_type_combo.setCurrentIndex(combo_index)

    def save_inventory(self):
        product = self.product_combo.currentData()
        if not product:
            QMessageBox.warning(self, "警告", "请选择产品")
            return

        quantity = self.quantity_input.text().strip()
        if not quantity:
            QMessageBox.warning(self, "警告", "请输入数量")
            return

        try:
            # 先转为浮点数，再转为整数（处理后端返回的 123.0 格式）
            quantity = int(float(quantity))
        except ValueError:
            QMessageBox.warning(self, "警告", "数量必须是整数")
            return

        # 库存类型映射：下拉框文本 -> stock_type
        type_map = {"采购在途": 1, "待入库": 2, "已入库": 3, "历史库存": 4}
        stock_type_text = self.stock_type_combo.currentText()
        stock_type = type_map.get(stock_type_text, 1)

        # 获取供应商（从供应商选择下拉框）
        supplier = self.supplier_combo.currentData()
        supplier_id = supplier.get('id') if supplier else None
        
        # 获取客户（从客户选择下拉框）
        customer = self.customer_combo.currentData()
        if not customer:
            QMessageBox.warning(self, "警告", "请选择客户")
            return
        customer_id = customer.get('id')
        
        inventory_data = {
            "dept_id": self.dept_id,
            "product_id": product.get('id'),
            "supplier_id": supplier_id,
            "customer_id": customer_id,
            "quantity": quantity,
            "current_location": self.location_input.text().strip() or 'WAREHOUSE',
            "stock_type": stock_type,
            "remark": self.remark_input.toPlainText().strip()
        }

        try:
            if self.is_edit:
                self.api_client.update_inventory(self.inventory.get('id'), inventory_data)
            else:
                self.api_client.create_inventory(inventory_data)
            # 清除库存缓存
            cache_manager.delete(CACHE_KEYS['INVENTORY_SUMMARY'])
            print("DEBUG - 已清除库存缓存")
            QMessageBox.information(self, "成功", "库存记录已保存")
            self.accept()
        except Exception as e:
            QMessageBox.warning(self, "错误", f"保存失败: {str(e)}")


def main():
    ctypes.windll.kernel32.SetConsoleOutputCP(65001)
    
    # 启用高DPI缩放（使用推荐方式）
    os.environ["QT_ENABLE_HIGHDPI_SCALING"] = "1"
    os.environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "0"
    
    app = QApplication(sys.argv)
    
    # 设置全局字体
    font = QFont()
    font.setFamily("Microsoft YaHei")
    font.setPointSize(10)
    app.setFont(font)

    # 使用带缓存功能的API客户端
    api_client = CachedApiClient()

    try:
        login = LoginWindow(api_client)
        if login.exec() != QDialog.DialogCode.Accepted:
            return

        dept_id = login.get_selected_department() or "S"
        window = MainWindow(api_client, dept_id)
        window.show()

        sys.exit(app.exec())
    except Exception as e:
        print("Error during application execution:", str(e))
        traceback.print_exc()
        # 显示错误对话框
        error_dialog = QDialog()
        error_dialog.setWindowTitle("启动错误")
        error_dialog.resize(400, 200)
        layout = QVBoxLayout()
        layout.addWidget(QLabel(f"启动失败: {str(e)}"))
        layout.addWidget(QLabel("请检查日志获取详细信息"))
        error_dialog.setLayout(layout)
        error_dialog.exec()
        sys.exit(1)


if __name__ == "__main__":
    main()