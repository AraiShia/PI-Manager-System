# -*- coding: utf-8 -*-


import sys


import os


import threading


import pandas as pd


import asyncio


from cache_manager import cache_manager, CACHE_KEYS, set_user, invalidate_cache


import ctypes


import urllib.request


import traceback


import concurrent.futures


from functools import lru_cache





# 鍏ㄥ眬绾跨▼姹狅紙澶嶇敤锛岄伩鍏嶉噸澶嶅垱寤猴級


_global_thread_pool = concurrent.futures.ThreadPoolExecutor(max_workers=8, thread_name_prefix="pi_manager")





# 图片内存缓存（LRU缓存，最大200个）


_image_cache = {}


_image_cache_lock = threading.Lock()


_MAX_IMAGE_CACHE_SIZE = 100





sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))


sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'config'))





# 鐪佷唤缂栫爜鏄犲皠锛堥潤鎬佹暟鎹紝妯″潡绾у埆鍙垱寤轰竴娆★級


# 省份编码映射（全局数据，模块级别只创建一次）
# 省份编码映射（全局数据，模块级别只创建一次）
PROVINCE_CODE_MAP = {
    "北京": "11", "天津": "12", "河北": "13", "山西": "14", "内蒙古": "15",
    "辽宁": "21", "吉林": "22", "黑龙江": "23",
    "上海": "31", "江苏": "32", "浙江": "33", "安徽": "34", "福建": "35", "江西": "36", "山东": "37",
    "河南": "41", "湖北": "42", "湖南": "43", "广东": "44", "广西": "45", "海南": "46",
    "重庆": "50", "四川": "51", "贵州": "52", "云南": "53", "西藏": "54",
    "陕西": "61", "甘肃": "62", "青海": "63", "宁夏": "64", "新疆": "65",
    "台湾": "71", "香港": "81", "澳门": "82"
}


from PySide6.QtWidgets import QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QStackedWidget, QMessageBox, QTableWidget, QTableWidgetItem, QDialog, QFormLayout, QLineEdit, QTextEdit, QComboBox, QHeaderView, QAbstractItemView, QGridLayout, QCheckBox, QGroupBox, QFileDialog, QProgressDialog, QTabWidget, QScrollArea, QDateEdit, QMenu, QFrame


from PySide6.QtCore import Qt, QTimer, QDate, QModelIndex


from PySide6.QtGui import QIcon, QPalette, QColor, QFont, QFontDatabase, QBrush, QPixmap, QImage


from api.client import ApiClient


from api.cached_client import CachedApiClient


from config import Config


from product_categories import get_category_options, get_category_code, get_category_name


from widgets.action_bar import ActionBarFactory





def get_font(size=10, weight=QFont.Weight.Normal):


    font = QFont()


    font.setPointSize(size)


    font.setWeight(weight)


    # 尝试多种中文字体


    available = QFontDatabase.families()


    for family in ["Microsoft YaHei", "SimHei", "SimSun", "KaiTi", "WenQuanYi Micro Hei", "Heiti SC"]:


        if family in available:


            font.setFamily(family)


            return font


    # 如果没有中文字体，使用默认字体


    return font





def set_global_font(app):


    font = get_font(10)


    app.setFont(font)





# 閮ㄩ棬閰嶇疆鏄犲皠锛堜粠閰嶇疆鏂囦欢鑾峰彇锛?DEPARTMENT_CONFIG = Config.DEPARTMENT_DB_CONFIG





class LoginWindow(QDialog):


    def __init__(self, api_client: ApiClient):


        super().__init__()


        self.api_client = api_client


        self.selected_dept = None


        self.init_ui()





    def init_ui(self):


        self.setWindowTitle("PI璁㈠崟绠＄悊绯荤粺 - 鐧诲綍")


        self.setFixedSize(700, 650)


        


        # 浣跨敤缁濆瀹氫綅


        # 鏍囬


        title = QLabel("PI璁㈠崟绠＄悊绯荤粺", self)


        title.setGeometry(50, 80, 600, 60)


        title.setAlignment(Qt.AlignCenter)


        title.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 36px; font-weight: bold;")


        


        # 鐗堟湰


        version = QLabel("瀹㈡埛绔?v1.0", self)


        version.setGeometry(50, 150, 600, 30)


        version.setAlignment(Qt.AlignCenter)


        version.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 16px; color: #666666;")


        


        # 鐧诲綍妯″紡閫夋嫨


        mode_label = QLabel("鐧诲綍妯″紡锛?, self)


        mode_label.setGeometry(50, 220, 600, 30)


        mode_label.setAlignment(Qt.AlignCenter)


        mode_label.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 18px;")


        


        self.mode_combo = QComboBox(self)


        self.mode_combo.setGeometry(160, 260, 380, 50)


        self.mode_combo.addItem("鏅€氱敤鎴锋ā寮?, False)


        self.mode_combo.addItem("绠＄悊鍛樻ā寮?, True)


        self.mode_combo.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 16px;")


        


        # 閮ㄩ棬鏍囩


        dept_label = QLabel("閫夋嫨閮ㄩ棬锛?, self)


        dept_label.setGeometry(50, 330, 600, 30)


        dept_label.setAlignment(Qt.AlignCenter)


        dept_label.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 18px;")


        


        # 閮ㄩ棬涓嬫媺妗?        self.dept_combo = QComboBox(self)


        self.dept_combo.setGeometry(160, 370, 380, 50)


        self.dept_combo.addItems([v["name"] for v in DEPARTMENT_CONFIG.values()])


        self.dept_combo.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 16px;")


        


        # API鏍囩


        api_label = QLabel("API鏈嶅姟鍣ㄥ湴鍧€锛?, self)


        api_label.setGeometry(50, 440, 600, 30)


        api_label.setAlignment(Qt.AlignCenter)


        api_label.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 18px;")


        


        # API杈撳叆妗?        self.api_url_input = QLineEdit(self)


        self.api_url_input.setGeometry(150, 480, 400, 50)


        self.api_url_input.setText(Config.API_BASE_URL)


        self.api_url_input.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 16px; padding-left: 10px;")


        


        # 鐧诲綍鎸夐挳


        self.login_btn = QPushButton("鐧诲綍", self)


        self.login_btn.setGeometry(160, 550, 380, 55)


        self.login_btn.setStyleSheet("""


            QPushButton {


                font-family: 'Microsoft YaHei';


                font-size: 18px;


                background-color: #2563eb;


                color: white;


                border: none;


                border-radius: 8px;


            }


            QPushButton:hover {


                background-color: #1d4ed8;


            }


        """)


        self.login_btn.clicked.connect(self.connect_to_server)


        


        # 鐘舵€佹爣绛?        self.status_label = QLabel("", self)


        self.status_label.setGeometry(50, 620, 600, 30)


        self.status_label.setAlignment(Qt.AlignCenter)


        self.status_label.setStyleSheet("font-family: 'Microsoft YaHei'; font-size: 16px; color: red;")





    def connect_to_server(self):


        is_admin = self.mode_combo.currentData()


        dept_name = self.dept_combo.currentText()


        api_url = self.api_url_input.text().strip()





        if not api_url:


            self.status_label.setText("璇疯緭鍏PI鏈嶅姟鍣ㄥ湴鍧€")


            return





        self.status_label.setText("姝ｅ湪杩炴帴...")


        self.status_label.setStyleSheet("color: #2563eb;")





        try:


            # 鑾峰彇閮ㄩ棬ID鍜屽搴旂殑鏁版嵁搴撻厤缃紙鑷姩鑾峰彇锛屼笉瀵瑰鏄剧ず锛?            dept_id = next((k for k, v in DEPARTMENT_CONFIG.items() if v["name"] == dept_name), "S")


            self.selected_dept = dept_id


            


            # 浠庨儴闂ㄩ厤缃腑鑷姩鑾峰彇鏁版嵁搴撹繛鎺ュ弬鏁?            dept_db_config = DEPARTMENT_CONFIG[dept_id]





            # 璁剧疆API瀹㈡埛绔?            self.api_client.base_url = api_url


            


            # 璁剧疆褰撳墠鐢ㄦ埛淇℃伅锛堟ā鎷熺櫥褰曪級


            if hasattr(self.api_client, 'current_user'):


                self.api_client.current_user = {


                    "id": 1,


                    "username": "admin" if is_admin else "user",


                    "real_name": "绠＄悊鍛? if is_admin else "鏅€氱敤鎴?,


                    "is_admin": is_admin,


                    "dept_id": dept_id


                }


                # 璁剧疆缂撳瓨鐢ㄦ埛ID


                set_user(str(self.api_client.current_user["id"]))


            


            # 浼犻€掓暟鎹簱閰嶇疆鍒癆PI锛堜粠閰嶇疆鏂囦欢鑾峰彇锛屼笉瀵瑰鏄剧ず锛?            db_config = {


                "db_host": dept_db_config["db_host"],


                "db_port": dept_db_config["db_port"],


                "db_user": dept_db_config["db_user"],


                "db_password": dept_db_config["db_password"],


                "db_name": dept_db_config["db_name"]


            }


            


            # 娴嬭瘯杩炴帴


            products = self.api_client.get_products(db_config=db_config)


            


            self.status_label.setText("杩炴帴鎴愬姛锛?)


            self.status_label.setStyleSheet("color: #16a34a;")


            QTimer.singleShot(500, self.accept)


        except Exception as e:


            self.status_label.setText(f"杩炴帴澶辫触: {str(e)}")


            self.status_label.setStyleSheet("color: #dc2626;")





    def get_selected_department(self):


        return self.selected_dept








class ProductDialog(QDialog):


    def __init__(self, api_client: ApiClient, dept_id: str, product=None):


        super().__init__()


        self.api_client = api_client


        self.dept_id = dept_id


        self.product = product


        self.is_edit = product is not None


        self.suppliers = []


        self.customers = []


        self.supplier_schemes = []


        self.init_ui()


        QTimer.singleShot(0, self.load_data)





    def load_data(self):


        """鍔犺浇渚涘簲鍟嗗拰瀹㈡埛鍒楄〃"""


        try:


            self.suppliers = self.api_client.get_suppliers()


        except Exception as e:


            print(f"鍔犺浇渚涘簲鍟嗗け璐? {e}")





        try:


            self.customers = self.api_client.get_customers()


        except Exception as e:


            print(f"鍔犺浇瀹㈡埛澶辫触: {e}")





        # 鍔犺浇宸叉湁鏂规


        if self.is_edit and self.product:


            self.load_supplier_schemes()





    def init_ui(self):


        self.setWindowTitle("缂栬緫浜у搧" if self.is_edit else "鏂板浜у搧")


        self.setMinimumSize(1000, 800)


        self.resize(1000, 800)





        layout = QVBoxLayout()


        layout.setContentsMargins(20, 20, 20, 20)


        layout.setSpacing(12)





        scroll_area = QScrollArea()


        scroll_area.setWidgetResizable(True)


        scroll_widget = QWidget()


        scroll_layout = QVBoxLayout(scroll_widget)


        scroll_layout.setSpacing(15)





        # ===== 椤堕儴鍖哄煙锛氬浘鐗?+ 鍩烘湰淇℃伅 =====


        top_widget = QWidget()


        top_layout = QHBoxLayout(top_widget)


        top_layout.setSpacing(20)





        # 宸︿晶锛氬浘鐗囧尯鍩燂紙涓诲浘 + 鍓浘锛?        left_layout = QVBoxLayout()





        # 涓诲浘


        main_image_group = QGroupBox("涓诲浘")


        main_image_layout = QVBoxLayout()


        self.image_preview = QLabel()


        self.image_preview.setFixedSize(200, 200)


        self.image_preview.setStyleSheet("border: 2px dashed #d1d5db;")


        self.image_preview.setAlignment(Qt.AlignCenter)


        self.image_preview.setText("鐐瑰嚮閫夋嫨鍥剧墖")


        self.image_preview.setCursor(Qt.PointingHandCursor)


        self.image_preview.installEventFilter(self)


        self.selected_image_path = None


        if self.product and self.product.get('default_image_url'):


            try:


                image_data = urllib.request.urlopen(self.product.get('default_image_url')).read()


                image = QImage.fromData(image_data)


                pixmap = QPixmap.fromImage(image).scaled(196, 196, Qt.KeepAspectRatio, Qt.SmoothTransformation)


                self.image_preview.setPixmap(pixmap)


                self.image_preview.setText("")


            except Exception:


                pass


        main_image_layout.addWidget(self.image_preview)





        main_image_btns = QHBoxLayout()


        self.upload_image_btn = QPushButton("涓婁紶涓诲浘")


        self.upload_image_btn.setFixedWidth(90)


        self.upload_image_btn.clicked.connect(self.select_image)


        main_image_btns.addWidget(self.upload_image_btn)


        self.clear_image_btn = QPushButton("娓呴櫎")


        self.clear_image_btn.setFixedWidth(60)


        self.clear_image_btn.clicked.connect(self.clear_image)


        main_image_btns.addWidget(self.clear_image_btn)


        main_image_layout.addLayout(main_image_btns)


        main_image_group.setLayout(main_image_layout)


        left_layout.addWidget(main_image_group)





        # 鍓浘鍖哄煙


        sub_image_group = QGroupBox("鍓浘")


        sub_image_layout = QVBoxLayout()


        self.sub_images_container = QWidget()


        self.sub_images_layout = QHBoxLayout(self.sub_images_container)


        self.sub_images_layout.setSpacing(5)


        self.sub_image_paths = []


        if self.product:


            self.load_sub_images()


        self.sub_images_layout.addStretch()


        sub_image_layout.addWidget(self.sub_images_container)





        add_sub_img_btn = QPushButton("+ 娣诲姞鍓浘")


        add_sub_img_btn.setFixedWidth(90)


        add_sub_img_btn.clicked.connect(self.add_sub_image)


        sub_image_layout.addWidget(add_sub_img_btn)


        sub_image_group.setLayout(sub_image_layout)


        left_layout.addWidget(sub_image_group)


        left_layout.addStretch()





        top_layout.addLayout(left_layout, 1)





        # 鍙充晶锛氬熀鏈俊鎭〃鍗?        right_form = QFormLayout()


        right_form.setSpacing(12)





        if self.is_edit:


            self.code_label = QLabel(self.product.get('product_code', ''))


            self.code_label.setStyleSheet("font-weight: bold; color: #2563eb;")


            right_form.addRow("浜у搧缂栧彿:", self.code_label)





        self.oe_input = QLineEdit()


        self.oe_input.setPlaceholderText("璇疯緭鍏E鍙?)


        self.oe_input.setMinimumHeight(32)


        if self.product:


            self.oe_input.setText(self.product.get('oe_number', ''))


        right_form.addRow("OE鍙?", self.oe_input)





        self.detail_input = QTextEdit()


        self.detail_input.setPlaceholderText("璇疯緭鍏ヤ骇鍝佺粏鑺傛弿杩?)


        if self.product:


            self.detail_input.setText(self.product.get('detail_desc', ''))


        self.detail_input.setMaximumHeight(80)


        right_form.addRow("缁嗚妭鎻忚堪:", self.detail_input)





        self.category_combo = QComboBox()


        self.category_combo.setFixedHeight(35)


        for code, name in get_category_options():


            self.category_combo.addItem(name, code)


        if self.product:


            category_code = str(self.product.get('category_id', '01'))


            for i in range(self.category_combo.count()):


                if self.category_combo.itemData(i) == category_code:


                    self.category_combo.setCurrentIndex(i)


                    break


        right_form.addRow("浜у搧绫诲埆:", self.category_combo)





        self.channel_input = QLineEdit()


        self.channel_input.setPlaceholderText("閲囪喘娓犻亾锛堝1688銆佸绉佷粯娆剧瓑锛?)


        self.channel_input.setMinimumHeight(32)


        if self.product:


            self.channel_input.setText(self.product.get('purchase_channel', ''))


        right_form.addRow("閲囪喘娓犻亾:", self.channel_input)





        top_layout.addLayout(right_form, 2)


        scroll_layout.addWidget(top_widget)





        # ===== 渚涘簲鍟嗘柟妗堣璁★紙寮圭獥妯″紡锛?=====


        supplier_scheme_group = QGroupBox("渚涘簲鍟嗘柟妗堣璁★紙浠锋牸銆佸寘瑁呫€侀噸閲忛殢鏂规閰嶇疆锛?)


        supplier_scheme_layout = QVBoxLayout()





        scheme_toolbar = QHBoxLayout()


        add_scheme_btn = QPushButton("+ 娣诲姞渚涘簲鍟嗘柟妗?)


        add_scheme_btn.clicked.connect(self.add_supplier_scheme)


        add_scheme_btn.setStyleSheet("""


            QPushButton {


                background-color: #2563eb;


                color: white;


                border: none;


                border-radius: 4px;


                padding: 6px 16px;


            }


            QPushButton:hover { background-color: #1d4ed8; }


        """)


        scheme_toolbar.addWidget(add_scheme_btn)


        scheme_toolbar.addStretch()


        supplier_scheme_layout.addLayout(scheme_toolbar)





        # 鎽樿琛ㄦ牸锛堝彧璇伙級


        self.supplier_scheme_table = QTableWidget()


        self.supplier_scheme_table.setColumnCount(6)


        self.supplier_scheme_table.setHorizontalHeaderLabels([


            "渚涘簲鍟?, "瀹㈡埛", "瀹㈡埛浜у搧缂栧彿", "涓昏浠锋牸", "澶囨敞", "鎿嶄綔"


        ])


        self.supplier_scheme_table.setColumnWidth(0, 180)


        self.supplier_scheme_table.setColumnWidth(1, 180)


        self.supplier_scheme_table.setColumnWidth(2, 120)


        self.supplier_scheme_table.setColumnWidth(3, 100)


        self.supplier_scheme_table.setColumnWidth(5, 120)


        self.supplier_scheme_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.Stretch)


        self.supplier_scheme_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)


        self.supplier_scheme_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)


        self.supplier_scheme_table.setFixedHeight(180)


        supplier_scheme_layout.addWidget(self.supplier_scheme_table)





        supplier_scheme_group.setLayout(supplier_scheme_layout)


        scroll_layout.addWidget(supplier_scheme_group)





        scroll_area.setWidget(scroll_widget)


        layout.addWidget(scroll_area)





        # 搴曢儴鎸夐挳


        btn_layout = QHBoxLayout()


        btn_layout.addStretch()





        save_btn = QPushButton("淇濆瓨")


        save_btn.setFixedWidth(100)


        save_btn.clicked.connect(self.save_product)


        save_btn.setStyleSheet("""


            QPushButton {


                background-color: #2563eb;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px;


            }


            QPushButton:hover { background-color: #1d4ed8; }


        """)


        btn_layout.addWidget(save_btn)





        cancel_btn = QPushButton("鍙栨秷")


        cancel_btn.setFixedWidth(100)


        cancel_btn.clicked.connect(self.reject)


        cancel_btn.setStyleSheet("""


            QPushButton {


                background-color: #e5e7eb;


                color: #374151;


                border: none;


                border-radius: 6px;


                padding: 8px;


            }


            QPushButton:hover { background-color: #d1d5db; }


        """)


        btn_layout.addWidget(cancel_btn)





        layout.addLayout(btn_layout)


        self.setLayout(layout)





    def eventFilter(self, obj, event):


        from PySide6.QtCore import QEvent


        if obj == self.image_preview and event.type() == QEvent.MouseButtonPress:


            self.select_image(event)


            return True


        return super().eventFilter(obj, event)


    


    def select_image(self, event=None):


        print("DEBUG - select_image called")


        # 濡傛灉宸茬粡鏈夊浘鐗囨樉绀猴紝鏄剧ず閫夐」鑿滃崟


        if self.image_preview.pixmap() is not None:


            print("DEBUG - 宸叉湁鍥剧墖锛屾樉绀鸿彍鍗?)


            from PySide6.QtWidgets import QMenu


            from PySide6.QtGui import QAction


            menu = QMenu(self)


            view_action = QAction("鏌ョ湅澶у浘", self)


            change_action = QAction("鏇存崲鍥剧墖", self)


            menu.addAction(view_action)


            menu.addAction(change_action)


            


            def view_image():


                print("DEBUG - 鏌ョ湅澶у浘")


                from PySide6.QtWidgets import QDialog, QVBoxLayout, QLabel


                dialog = QDialog(self)


                dialog.setWindowTitle("鏌ョ湅鍥剧墖")


                dialog.setMinimumSize(600, 600)


                layout = QVBoxLayout()


                label = QLabel()


                label.setAlignment(Qt.AlignCenter)


                if self.selected_image_path:


                    pixmap = QPixmap(self.selected_image_path).scaled(580, 580, Qt.KeepAspectRatio, Qt.SmoothTransformation)


                else:


                    pixmap = self.image_preview.pixmap().scaled(580, 580, Qt.KeepAspectRatio, Qt.SmoothTransformation)


                label.setPixmap(pixmap)


                layout.addWidget(label)


                dialog.setLayout(layout)


                dialog.exec()


            


            def change_image():


                print("DEBUG - 鏇存崲鍥剧墖")


                file_path, _ = QFileDialog.getOpenFileName(self, "閫夋嫨鍥剧墖", "", "鍥剧墖鏂囦欢 (*.png *.jpg *.jpeg *.gif)")


                print(f"DEBUG - 閫夋嫨鐨勬枃浠惰矾寰? {file_path}")


                if file_path:


                    self.selected_image_path = file_path


                    print(f"DEBUG - selected_image_path 璁剧疆涓? {self.selected_image_path}")


                    pixmap = QPixmap(file_path).scaled(196, 196, Qt.KeepAspectRatio, Qt.SmoothTransformation)


                    self.image_preview.setPixmap(pixmap)


            


            view_action.triggered.connect(view_image)


            change_action.triggered.connect(change_image)


            


            menu.exec(self.image_preview.mapToGlobal(event.pos()) if event else self.image_preview.mapToGlobal(self.image_preview.rect().center()))


            return


        


        # 娌℃湁鍥剧墖鏃讹紝鐩存帴閫夋嫨鍥剧墖


        print("DEBUG - 娌℃湁鍥剧墖锛岀洿鎺ラ€夋嫨")


        file_path, _ = QFileDialog.getOpenFileName(self, "閫夋嫨鍥剧墖", "", "鍥剧墖鏂囦欢 (*.png *.jpg *.jpeg *.gif)")


        print(f"DEBUG - 閫夋嫨鐨勬枃浠惰矾寰? {file_path}")


        if file_path:


            self.selected_image_path = file_path


            print(f"DEBUG - selected_image_path 璁剧疆涓? {self.selected_image_path}")


            pixmap = QPixmap(file_path).scaled(196, 196, Qt.KeepAspectRatio, Qt.SmoothTransformation)


            self.image_preview.setPixmap(pixmap)


            self.image_preview.setText("")





    def clear_image(self):


        self.selected_image_path = None


        self.image_preview.clear()


        self.image_preview.setText("鐐瑰嚮閫夋嫨鍥剧墖")





    def load_sub_images(self):


        """鍔犺浇浜у搧鍓浘"""


        try:


            images = self.api_client.get_product_images(self.product.get('id'))


            for img in images:


                if not img.get('is_default'):


                    self.add_sub_image_thumbnail(img.get('image_url'))


        except Exception as e:


            print(f"鍔犺浇鍓浘澶辫触: {e}")





    def add_sub_image_thumbnail(self, image_url=None, file_path=None):


        """娣诲姞鍓浘缂╃暐鍥惧埌鐣岄潰"""


        thumb = QLabel()


        thumb.setFixedSize(60, 60)


        thumb.setStyleSheet("border: 1px solid #e5e7eb;")


        thumb.setAlignment(Qt.AlignCenter)





        if image_url:


            try:


                image_data = urllib.request.urlopen(image_url).read()


                image = QImage.fromData(image_data)


                pixmap = QPixmap.fromImage(image).scaled(56, 56, Qt.KeepAspectRatio, Qt.SmoothTransformation)


                thumb.setPixmap(pixmap)


            except Exception:


                thumb.setText("鍥?)


        elif file_path:


            pixmap = QPixmap(file_path).scaled(56, 56, Qt.KeepAspectRatio, Qt.SmoothTransformation)


            thumb.setPixmap(pixmap)


        else:


            thumb.setText("+")





        thumb.setCursor(Qt.PointingHandCursor)


        thumb.setProperty("image_url", image_url)


        thumb.setProperty("file_path", file_path)


        thumb.mousePressEvent = lambda event, t=thumb: self.remove_sub_image(t)


        self.sub_images_layout.insertWidget(self.sub_images_layout.count() - 1, thumb)





    def add_sub_image(self):


        """娣诲姞鍓浘"""


        file_path, _ = QFileDialog.getOpenFileName(self, "閫夋嫨鍓浘", "", "Images (*.png *.jpg *.jpeg)")


        if file_path:


            self.sub_image_paths.append(file_path)


            self.add_sub_image_thumbnail(file_path=file_path)





    def remove_sub_image(self, thumb_label):


        """绉婚櫎鍓浘"""


        file_path = thumb_label.property("file_path")


        if file_path and file_path in self.sub_image_paths:


            self.sub_image_paths.remove(file_path)


        thumb_label.deleteLater()





    def refresh_scheme_table(self):


        """鍒锋柊渚涘簲鍟嗘柟妗堟憳瑕佽〃鏍?""


        self.supplier_scheme_table.setRowCount(len(self.supplier_schemes))


        for row, s in enumerate(self.supplier_schemes):


            self.supplier_scheme_table.setItem(row, 0, QTableWidgetItem(s.get('supplier_name', '')))


            # 瀹㈡埛鍚嶇О锛氫紭鍏坈ustomer_name锛屽惁鍒欐樉绀?閫氱敤"


            customer_name = s.get('customer_name', '') or '閫氱敤'


            self.supplier_scheme_table.setItem(row, 1, QTableWidgetItem(customer_name))


            # 宸ュ巶缂栧彿锛氬吋瀹逛袱绉嶅瓧娈靛悕


            factory_code = s.get('factory_code') or s.get('customer_product_code', '')


            self.supplier_scheme_table.setItem(row, 2, QTableWidgetItem(str(factory_code)))


            # 涓昏浠锋牸锛氬吋瀹逛袱绉嶅瓧娈靛悕


            main_price = s.get('purchase_price') or s.get('exw_price_incl') or ''


            self.supplier_scheme_table.setItem(row, 3, QTableWidgetItem(str(main_price)))


            self.supplier_scheme_table.setItem(row, 4, QTableWidgetItem(s.get('remark', '')))





            # 鎿嶄綔鎸夐挳


            action_widget = QWidget()


            action_layout = QHBoxLayout()


            action_layout.setContentsMargins(0, 0, 0, 0)





            edit_btn = QPushButton("缂栬緫")


            edit_btn.setFixedWidth(50)


            edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px;")


            edit_btn.clicked.connect(lambda _, r=row: self.edit_supplier_scheme(r))


            action_layout.addWidget(edit_btn)





            del_btn = QPushButton("鍒犻櫎")


            del_btn.setFixedWidth(50)


            del_btn.setStyleSheet("background-color: #dc2626; color: white; border: none; border-radius: 4px;")


            del_btn.clicked.connect(lambda _, r=row: self.remove_supplier_scheme(r))


            action_layout.addWidget(del_btn)





            action_widget.setLayout(action_layout)


            self.supplier_scheme_table.setCellWidget(row, 5, action_widget)





    def load_supplier_schemes(self):


        """鍔犺浇渚涘簲鍟嗘柟妗?""


        try:


            if self.product and self.product.get('id'):


                # 浠嶢PI鍔犺浇渚涘簲鍟嗘柟妗?                schemes = self.api_client.get_product_suppliers(self.product.get('id'))


                print(f"DEBUG - 鍔犺浇渚涘簲鍟嗘柟妗? {schemes}")


                self.supplier_schemes = schemes if schemes else []


            else:


                self.supplier_schemes = []


            self.refresh_scheme_table()


        except Exception as e:


            print(f"鍔犺浇渚涘簲鍟嗘柟妗堝け璐? {e}")


            self.supplier_schemes = []


            self.refresh_scheme_table()





    def add_supplier_scheme(self):


        """寮圭獥娣诲姞渚涘簲鍟嗘柟妗?""


        dialog = SupplierSchemeDialog(self.api_client, self.suppliers, self.customers, parent=self)


        if dialog.exec() == QDialog.DialogCode.Accepted:


            scheme_data = dialog.get_scheme_data()


            if scheme_data:


                self.supplier_schemes.append(scheme_data)


                self.refresh_scheme_table()





    def edit_supplier_scheme(self, row):


        """寮圭獥缂栬緫渚涘簲鍟嗘柟妗?""


        if row < 0 or row >= len(self.supplier_schemes):


            return


        dialog = SupplierSchemeDialog(self.api_client, self.suppliers, self.customers, 


                                      self.supplier_schemes[row], parent=self)


        if dialog.exec() == QDialog.DialogCode.Accepted:


            scheme_data = dialog.get_scheme_data()


            if scheme_data:


                self.supplier_schemes[row] = scheme_data


                self.refresh_scheme_table()





    def remove_supplier_scheme(self, row):


        """鍒犻櫎渚涘簲鍟嗘柟妗?""


        if 0 <= row < len(self.supplier_schemes):


            self.supplier_schemes.pop(row)


            self.refresh_scheme_table()





    def save_product(self):


        print("DEBUG - save_product called")


        


        oe_number = self.oe_input.text().strip()


        detail_desc = self.detail_input.toPlainText().strip()


        


        if not oe_number:


            QMessageBox.warning(self, "鎻愮ず", "璇疯緭鍏E鍙?)


            return


        


        if not detail_desc:


            QMessageBox.warning(self, "鎻愮ず", "璇疯緭鍏ョ粏鑺傛弿杩?)


            return





        data = {


            "dept_id": self.dept_id,


            "oe_number": oe_number,


            "detail_desc": detail_desc,


            "purchase_channel": self.channel_input.text().strip(),


            "category_id": int(self.category_combo.currentData()),


        }





        # 鍦ㄥ悗鍙扮嚎绋嬩腑鎵ц淇濆瓨锛岄伩鍏峌I闃诲


        self.save_product_async(data)


    


    def save_product_async(self, data):


        """鍦ㄥ悗鍙扮嚎绋嬩腑淇濆瓨浜у搧锛岄伩鍏峌I闃诲"""


        import threading


        


        # 鏄剧ず鍔犺浇鎻愮ず


        self.setEnabled(False)


        from PySide6.QtWidgets import QLabel


        self.loading_label = QLabel("姝ｅ湪淇濆瓨...", self)


        self.loading_label.setStyleSheet("background-color: #2563eb; color: white; padding: 8px 16px; border-radius: 4px;")


        self.loading_label.move(self.width()//2 - 50, self.height()//2)


        self.loading_label.show()


        


        def do_save():


            try:


                product_id = None


                


                # 濡傛灉鏈夐€夋嫨鍥剧墖锛屽厛涓婁紶鍥剧墖


                if self.selected_image_path:


                    try:


                        image_url = self.api_client.upload_image(self.selected_image_path)


                        if image_url:


                            data["default_image_url"] = image_url


                    except Exception as e:


                        print(f"鍥剧墖涓婁紶寮傚父: {str(e)}")


                


                # 涓婁紶鍓浘


                sub_image_urls = []


                for path in self.sub_image_paths:


                    try:


                        url = self.api_client.upload_image(path)


                        if url:


                            sub_image_urls.append(url)


                    except Exception as e:


                        print(f"鍓浘涓婁紶澶辫触: {e}")


                data["sub_images"] = sub_image_urls


                


                # 淇濆瓨浜у搧鍩烘湰淇℃伅


                if self.is_edit:


                    product_id = self.product.get('id')


                    self.api_client.update_product(product_id, data)


                    if self.selected_image_path and 'default_image_url' in data:


                        self.api_client.set_product_default_image(product_id, data['default_image_url'])


                else:


                    result = self.api_client.create_product(data)


                    if result and 'id' in result:


                        product_id = result['id']


                        if self.selected_image_path:


                            self.api_client.set_product_default_image(product_id, data.get('default_image_url'))


                


                # 淇濆瓨渚涘簲鍟嗘柟妗堬紙鍚庡彴绾跨▼锛?                if product_id and self.supplier_schemes:


                    self.save_supplier_schemes_async(product_id)


                


                # 鍦ㄤ富绾跨▼涓洿鏂癠I


                from PySide6.QtCore import QMetaObject, Qt, Q_ARG


                QMetaObject.invokeMethod(self, "on_save_success", Qt.QueuedConnection)


                


            except Exception as e:


                print(f"淇濆瓨浜у搧澶辫触: {str(e)}")


                from PySide6.QtCore import QMetaObject, Qt, Q_ARG


                QMetaObject.invokeMethod(self, "on_save_error", Qt.QueuedConnection, 


                                        Q_ARG(str, str(e)))


        


        thread = threading.Thread(target=do_save, daemon=True)


        thread.start()


    


    def on_save_success(self):


        """淇濆瓨鎴愬姛鍥炶皟锛堜富绾跨▼锛?""


        self.loading_label.hide()


        self.setEnabled(True)


        # 娓呴櫎浜у搧缂撳瓨锛屼笅娆″姞杞芥椂閲嶆柊鑾峰彇


        cache_manager.delete(CACHE_KEYS['PRODUCTS'])


        cache_manager.delete(CACHE_KEYS['INVENTORY_SUMMARY'])


        print("DEBUG - 宸叉竻闄や骇鍝佺紦瀛?)


        QMessageBox.information(self, "鎴愬姛", "浜у搧淇濆瓨鎴愬姛")


        self.accept()


    


    def on_save_error(self, error_msg):


        """淇濆瓨澶辫触鍥炶皟锛堜富绾跨▼锛?""


        self.loading_label.hide()


        self.setEnabled(True)


        QMessageBox.warning(self, "閿欒", f"淇濆瓨浜у搧澶辫触: {error_msg}")


     


    def save_supplier_schemes_async(self, product_id):


        """寮傛淇濆瓨渚涘簲鍟嗘柟妗堬紝涓嶉樆濉濽I"""


        import threading


        


        def save_schemes():


            try:


                print(f"DEBUG - 寮€濮嬩繚瀛樹緵搴斿晢鏂规锛屼骇鍝両D: {product_id}")


                for scheme in self.supplier_schemes:


                    # 鏄犲皠鍓嶇瀛楁鍚嶅埌鍚庣瀛楁鍚?                    scheme_data = {


                        'product_id': product_id,


                        'supplier_id': scheme.get('supplier_id'),


                        'factory_code': scheme.get('factory_code') or scheme.get('customer_product_code') or '',


                        'customer_id': scheme.get('customer_id'),


                        'is_default': scheme.get('is_default', False),


                        'purchase_price': scheme.get('purchase_price') or scheme.get('exw_price_incl'),


                        'currency': scheme.get('currency', 'CNY'),


                        'units_per_carton': scheme.get('units_per_carton'),


                        'carton_length_cm': scheme.get('carton_length_cm'),


                        'carton_width_cm': scheme.get('carton_width_cm'),


                        'carton_height_cm': scheme.get('carton_height_cm'),


                        'gross_weight_kg': scheme.get('gross_weight_kg'),


                        'moq': scheme.get('moq'),


                        'lead_time_days': scheme.get('lead_time_days'),


                        'purchase_channel': scheme.get('purchase_channel'),


                        'remark': scheme.get('remark'),


                        'special_requirements': scheme.get('special_requirements')


                    }


                    


                    print(f"DEBUG - 淇濆瓨鏂规鏁版嵁: supplier_id={scheme_data['supplier_id']}, factory_code={scheme_data['factory_code']}, customer_id={scheme_data['customer_id']}")


                    


                    if scheme.get('id'):


                        # 鏇存柊鐜版湁鏂规


                        self.api_client.update_product_supplier(scheme['id'], scheme_data)


                        print(f"DEBUG - 鏇存柊鏂规 ID={scheme['id']} 鎴愬姛")


                    else:


                        # 鍒涘缓鏂版柟妗?                        result = self.api_client.add_product_supplier_full(scheme_data)


                        print(f"DEBUG - 鍒涘缓鏂规鎴愬姛: {result}")


                print(f"DEBUG - 渚涘簲鍟嗘柟妗堜繚瀛樺畬鎴?)


            except Exception as e:


                import traceback


                print(f"淇濆瓨渚涘簲鍟嗘柟妗堝け璐? {str(e)}")


                traceback.print_exc()


        


        # 鍦ㄥ悗鍙扮嚎绋嬩腑淇濆瓨


        thread = threading.Thread(target=save_schemes, daemon=True)


        thread.start()








class SupplierSchemeDialog(QDialog):


    """渚涘簲鍟嗘柟妗堢紪杈戝脊绐?""


    def __init__(self, api_client, suppliers, customers, scheme=None, parent=None):


        super().__init__(parent)


        self.api_client = api_client


        self.suppliers = suppliers


        self.customers = customers


        self.scheme = scheme or {}


        self.is_edit = bool(scheme)


        self.setWindowTitle("缂栬緫渚涘簲鍟嗘柟妗? if self.is_edit else "娣诲姞渚涘簲鍟嗘柟妗?)


        self.setFixedSize(650, 700)


        self.init_ui()





    def init_ui(self):


        layout = QVBoxLayout()


        layout.setContentsMargins(20, 20, 20, 20)


        layout.setSpacing(12)





        form_layout = QFormLayout()


        form_layout.setSpacing(12)





        # 渚涘簲鍟?        self.supplier_combo = QComboBox()


        self.supplier_combo.setFixedHeight(35)


        self.supplier_combo.addItem("璇烽€夋嫨渚涘簲鍟?, None)


        for s in self.suppliers:


            self.supplier_combo.addItem(f"{s.get('supplier_code')} - {s.get('supplier_name')}", s.get('id'))


        if self.scheme.get('supplier_id'):


            for i in range(self.supplier_combo.count()):


                if self.supplier_combo.itemData(i) == self.scheme.get('supplier_id'):


                    self.supplier_combo.setCurrentIndex(i)


                    break


        form_layout.addRow("渚涘簲鍟?*:", self.supplier_combo)





        # 鏂规绫诲瀷閫夋嫨


        self.scheme_type_combo = QComboBox()


        self.scheme_type_combo.setFixedHeight(35)


        self.scheme_type_combo.addItem("馃彿锔?榛樿鏂规锛堜笉鎸囧畾瀹㈡埛锛?, {'type': 'default', 'customer_id': None})


        self.scheme_type_combo.addItem("馃懁 鎸囧畾瀹㈡埛涓撳睘鏂规", {'type': 'customer', 'customer_id': None})


        self.scheme_type_combo.currentIndexChanged.connect(self.on_scheme_type_changed)


        form_layout.addRow("鏂规绫诲瀷 *:", self.scheme_type_combo)





        # 瀹㈡埛閫夋嫨锛堜粎鍦ㄦ寚瀹氬鎴锋椂鏄剧ず锛?        self.customer_combo = QComboBox()


        self.customer_combo.setFixedHeight(35)


        self.customer_combo.addItem("璇烽€夋嫨瀹㈡埛", None)


        for c in self.customers:


            self.customer_combo.addItem(f"{c.get('customer_code')} - {c.get('customer_name')}", c.get('id'))


        self.customer_combo.setVisible(False)


        


        # 鏍规嵁缂栬緫鏁版嵁璁剧疆鍒濆鍊?        if self.is_edit:


            if self.scheme.get('customer_id'):


                self.scheme_type_combo.setCurrentIndex(1)


                self.customer_combo.setVisible(True)


                for i in range(self.customer_combo.count()):


                    if self.customer_combo.itemData(i) == self.scheme.get('customer_id'):


                        self.customer_combo.setCurrentIndex(i)


                        break


            else:


                self.scheme_type_combo.setCurrentIndex(0)


        


        form_layout.addRow("閫夋嫨瀹㈡埛:", self.customer_combo)





        # 瀹㈡埛浜у搧缂栧彿锛堟柊寤烘椂榛樿浣跨敤鐖剁獥鍙ｇ殑OE鍙凤級


        self.customer_code_input = QLineEdit()


        self.customer_code_input.setPlaceholderText("瀹㈡埛鍦ㄥ鏂圭郴缁熶腑鐨勪骇鍝佺紪鍙?)


        # 鍏煎涓ょ瀛楁鍚嶏細factory_code锛圓PI杩斿洖锛夊拰 customer_product_code锛堝墠绔娇鐢級


        default_code = self.scheme.get('factory_code') or self.scheme.get('customer_product_code', '')


        if not default_code and not self.is_edit and self.parent() and hasattr(self.parent(), 'oe_input'):


            default_code = self.parent().oe_input.text().strip()


        self.customer_code_input.setText(default_code)


        form_layout.addRow("瀹㈡埛浜у搧缂栧彿:", self.customer_code_input)





        layout.addLayout(form_layout)





        # 浠锋牸淇℃伅


        price_group = QGroupBox("浠锋牸淇℃伅")


        price_layout = QGridLayout()


        price_layout.setSpacing(10)





        self.exw_incl_input = QLineEdit()


        self.exw_incl_input.setPlaceholderText("EXW鍚◣浠?)


        # 鍏煎涓ょ瀛楁鍚?        exw_val = self.scheme.get('exw_price_incl') or self.scheme.get('purchase_price', '')


        self.exw_incl_input.setText(str(exw_val or ''))


        price_layout.addWidget(QLabel("EXW鍚◣浠?"), 0, 0)


        price_layout.addWidget(self.exw_incl_input, 0, 1)





        self.exw_excl_input = QLineEdit()


        self.exw_excl_input.setPlaceholderText("EXW涓嶅惈绋庝环")


        self.exw_excl_input.setText(str(self.scheme.get('exw_price_excl', '') or ''))


        price_layout.addWidget(QLabel("EXW涓嶅惈绋庝环:"), 0, 2)


        price_layout.addWidget(self.exw_excl_input, 0, 3)





        self.fob_incl_input = QLineEdit()


        self.fob_incl_input.setPlaceholderText("FOB鍚◣浠?)


        self.fob_incl_input.setText(str(self.scheme.get('fob_price_incl', '') or ''))


        price_layout.addWidget(QLabel("FOB鍚◣浠?"), 1, 0)


        price_layout.addWidget(self.fob_incl_input, 1, 1)





        self.fob_excl_input = QLineEdit()


        self.fob_excl_input.setPlaceholderText("FOB涓嶅惈绋庝环")


        self.fob_excl_input.setText(str(self.scheme.get('fob_price_excl', '') or ''))


        price_layout.addWidget(QLabel("FOB涓嶅惈绋庝环:"), 1, 2)


        price_layout.addWidget(self.fob_excl_input, 1, 3)





        self.freight_input = QLineEdit()


        self.freight_input.setPlaceholderText("杩愯垂")


        self.freight_input.setText(str(self.scheme.get('freight', '') or ''))


        price_layout.addWidget(QLabel("杩愯垂:"), 2, 0)


        price_layout.addWidget(self.freight_input, 2, 1)





        self.packing_fee_input = QLineEdit()


        self.packing_fee_input.setPlaceholderText("鍖呰璐?)


        self.packing_fee_input.setText(str(self.scheme.get('packing_fee', '') or ''))


        price_layout.addWidget(QLabel("鍖呰璐?"), 2, 2)


        price_layout.addWidget(self.packing_fee_input, 2, 3)





        price_group.setLayout(price_layout)


        layout.addWidget(price_group)





        # 鍖呰灏哄


        size_group = QGroupBox("鍖呰灏哄")


        size_layout = QGridLayout()


        self.carton_length_input = QLineEdit()


        self.carton_length_input.setPlaceholderText("闀?cm)")


        self.carton_length_input.setText(str(self.scheme.get('carton_length_cm', '') or ''))


        size_layout.addWidget(QLabel("绾哥闀?cm):"), 0, 0)


        size_layout.addWidget(self.carton_length_input, 0, 1)





        self.carton_width_input = QLineEdit()


        self.carton_width_input.setPlaceholderText("瀹?cm)")


        self.carton_width_input.setText(str(self.scheme.get('carton_width_cm', '') or ''))


        size_layout.addWidget(QLabel("绾哥瀹?cm):"), 0, 2)


        size_layout.addWidget(self.carton_width_input, 0, 3)





        self.carton_height_input = QLineEdit()


        self.carton_height_input.setPlaceholderText("楂?cm)")


        self.carton_height_input.setText(str(self.scheme.get('carton_height_cm', '') or ''))


        size_layout.addWidget(QLabel("绾哥楂?cm):"), 1, 0)


        size_layout.addWidget(self.carton_height_input, 1, 1)





        self.units_input = QLineEdit()


        self.units_input.setPlaceholderText("姣忕鏁伴噺")


        self.units_input.setText(str(self.scheme.get('units_per_carton', '') or ''))


        size_layout.addWidget(QLabel("姣忕鏁伴噺:"), 1, 2)


        size_layout.addWidget(self.units_input, 1, 3)


        size_group.setLayout(size_layout)


        layout.addWidget(size_group)





        # 閲嶉噺淇℃伅


        weight_group = QGroupBox("閲嶉噺淇℃伅")


        weight_layout = QHBoxLayout()


        self.gross_weight_input = QLineEdit()


        self.gross_weight_input.setPlaceholderText("姣涢噸(kg)")


        self.gross_weight_input.setText(str(self.scheme.get('gross_weight_kg', '') or ''))


        weight_layout.addWidget(QLabel("姣涢噸(kg):"))


        weight_layout.addWidget(self.gross_weight_input)





        self.weight_input = QLineEdit()


        self.weight_input.setPlaceholderText("鍑€閲?kg)")


        self.weight_input.setText(str(self.scheme.get('weight_kg', '') or ''))


        weight_layout.addWidget(QLabel("鍑€閲?kg):"))


        weight_layout.addWidget(self.weight_input)


        weight_group.setLayout(weight_layout)


        layout.addWidget(weight_group)





        # 澶囨敞


        self.remark_input = QTextEdit()


        self.remark_input.setPlaceholderText("澶囨敞淇℃伅")


        self.remark_input.setText(self.scheme.get('remark', ''))


        self.remark_input.setMaximumHeight(60)


        layout.addWidget(QLabel("澶囨敞:"))


        layout.addWidget(self.remark_input)


        


        # 璁句负榛樿鏂规


        self.is_default_checkbox = QCheckBox("璁句负榛樿渚涘簲鍟嗘柟妗堬紙浼樺厛浣跨敤锛?)


        self.is_default_checkbox.setStyleSheet("color: #2563eb; font-weight: 500;")


        if self.scheme.get('is_default'):


            self.is_default_checkbox.setChecked(True)


        layout.addWidget(self.is_default_checkbox)





        # 鎸夐挳


        btn_layout = QHBoxLayout()


        btn_layout.addStretch()





        save_btn = QPushButton("淇濆瓨")


        save_btn.setFixedWidth(100)


        save_btn.clicked.connect(self.save_scheme)


        save_btn.setStyleSheet("""


            QPushButton {


                background-color: #2563eb;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px;


            }


            QPushButton:hover { background-color: #1d4ed8; }


        """)


        btn_layout.addWidget(save_btn)





        cancel_btn = QPushButton("鍙栨秷")


        cancel_btn.setFixedWidth(100)


        cancel_btn.clicked.connect(self.reject)


        cancel_btn.setStyleSheet("""


            QPushButton {


                background-color: #e5e7eb;


                color: #374151;


                border: none;


                border-radius: 6px;


                padding: 8px;


            }


            QPushButton:hover { background-color: #d1d5db; }


        """)


        btn_layout.addWidget(cancel_btn)





        layout.addLayout(btn_layout)


        self.setLayout(layout)





    def on_scheme_type_changed(self):


        """鏂规绫诲瀷鍙樺寲鏃舵樉绀?闅愯棌瀹㈡埛閫夋嫨"""


        scheme_data = self.scheme_type_combo.currentData()


        if scheme_data and scheme_data.get('type') == 'customer':


            self.customer_combo.setVisible(True)


        else:


            self.customer_combo.setVisible(False)





    def save_scheme(self):


        supplier_id = self.supplier_combo.currentData()


        


        if not supplier_id:


            QMessageBox.warning(self, "璀﹀憡", "璇烽€夋嫨渚涘簲鍟?)


            return


        


        # 鏍规嵁鏂规绫诲瀷鑾峰彇瀹㈡埛ID


        scheme_data = self.scheme_type_combo.currentData()


        customer_id = None


        if scheme_data and scheme_data.get('type') == 'customer':


            customer_id = self.customer_combo.currentData()


            if not customer_id:


                QMessageBox.warning(self, "璀﹀憡", "璇烽€夋嫨瀹㈡埛")


                return





        def try_float(value):


            try:


                return float(value) if value.strip() else None


            except ValueError:


                return None





        supplier_name = self.supplier_combo.currentText()


        customer_name = self.customer_combo.currentText()





        self.scheme_data = {


            "id": self.scheme.get('id') if self.is_edit else None,


            "supplier_id": supplier_id,


            "supplier_name": supplier_name,


            "customer_id": customer_id,


            "customer_name": customer_name,


            "customer_product_code": self.customer_code_input.text().strip(),


            "is_default": self.is_default_checkbox.isChecked(),


            "exw_price_incl": try_float(self.exw_incl_input.text()),


            "exw_price_excl": try_float(self.exw_excl_input.text()),


            "fob_price_incl": try_float(self.fob_incl_input.text()),


            "fob_price_excl": try_float(self.fob_excl_input.text()),


            "freight": try_float(self.freight_input.text()),


            "packing_fee": try_float(self.packing_fee_input.text()),


            "carton_length_cm": try_float(self.carton_length_input.text()),


            "carton_width_cm": try_float(self.carton_width_input.text()),


            "carton_height_cm": try_float(self.carton_height_input.text()),


            "units_per_carton": int(self.units_input.text()) if self.units_input.text().strip() else None,


            "gross_weight_kg": try_float(self.gross_weight_input.text()),


            "weight_kg": try_float(self.weight_input.text()),


            "remark": self.remark_input.toPlainText()


        }


        print(f"DEBUG - save_scheme: scheme_data = {self.scheme_data}")


        self.accept()





    def get_scheme_data(self):


        return getattr(self, 'scheme_data', None)








class CustomerDialog(QDialog):


    def __init__(self, api_client: ApiClient, customer=None):


        super().__init__()


        self.api_client = api_client


        self.customer = customer


        self.is_edit = customer is not None


        self.init_ui()





    def init_ui(self):


        self.setWindowTitle("缂栬緫瀹㈡埛" if self.is_edit else "鏂板瀹㈡埛")


        self.setMinimumSize(750, 600)





        layout = QVBoxLayout()


        layout.setContentsMargins(20, 20, 20, 20)


        layout.setSpacing(15)





        form_layout = QFormLayout()


        form_layout.setSpacing(12)





        self.dept_combo = QComboBox()


        self.dept_combo.addItems([


            "S - 绱㈣嫳鏅?,


            "W - 缁撮偅",


            "M - 椹开閭?,


            "D - 閾惰揪"


        ])


        form_layout.addRow("閮ㄩ棬:", self.dept_combo)





        if self.is_edit:


            self.code_label = QLabel(self.customer.get('customer_code', ''))


            form_layout.addRow("瀹㈡埛缂栧彿:", self.code_label)


        


        self.name_input = QLineEdit()


        self.name_input.setPlaceholderText("璇疯緭鍏ュ鎴峰悕绉?)


        if self.customer:


            self.name_input.setText(self.customer.get('customer_name', ''))


        form_layout.addRow("瀹㈡埛鍚嶇О *:", self.name_input)





        self.country_input = QLineEdit()


        self.country_input.setPlaceholderText("璇疯緭鍏ユ墍鍦ㄥ浗瀹?)


        if self.customer:


            self.country_input.setText(self.customer.get('country', ''))


        form_layout.addRow("鎵€鍦ㄥ浗瀹?*:", self.country_input)





        self.basic_require_input = QTextEdit()


        self.basic_require_input.setPlaceholderText("璇疯緭鍏ラ€氱敤浜ゆ槗鏉℃")


        self.basic_require_input.setMaximumHeight(60)


        if self.customer:


            self.basic_require_input.setText(self.customer.get('basic_require', ''))


        form_layout.addRow("鍩烘湰瑕佹眰:", self.basic_require_input)





        self.special_input = QTextEdit()


        self.special_input.setPlaceholderText("璇疯緭鍏ョ壒娈婅姹傦紝濡傜壒瀹氬寘瑁呫€佹爣绛剧瓑")


        self.special_input.setMaximumHeight(60)


        if self.customer:


            self.special_input.setText(self.customer.get('special_require', ''))


        form_layout.addRow("鐗规畩瑕佹眰:", self.special_input)





        self.payment_input = QLineEdit()


        self.payment_input.setPlaceholderText("濡?T/T 30澶?)


        if self.customer:


            self.payment_input.setText(self.customer.get('payment_terms', ''))


        form_layout.addRow("浠樻鏉℃:", self.payment_input)





        contact_group = QGroupBox("鑱旂郴浜轰俊鎭?)


        contact_layout = QFormLayout()


        contact_layout.setSpacing(10)





        self.contact_name_input = QLineEdit()


        self.contact_name_input.setPlaceholderText("鑱旂郴浜哄鍚?)


        if self.customer:


            self.contact_name_input.setText(self.customer.get('contact_name', ''))


        contact_layout.addRow("濮撳悕:", self.contact_name_input)





        self.contact_phone_input = QLineEdit()


        self.contact_phone_input.setPlaceholderText("鐢佃瘽")


        if self.customer:


            self.contact_phone_input.setText(self.customer.get('contact_phone', ''))


        contact_layout.addRow("鐢佃瘽:", self.contact_phone_input)





        self.contact_email_input = QLineEdit()


        self.contact_email_input.setPlaceholderText("閭")


        if self.customer:


            self.contact_email_input.setText(self.customer.get('contact_email', ''))


        contact_layout.addRow("閭:", self.contact_email_input)





        contact_group.setLayout(contact_layout)


        layout.addLayout(form_layout)


        layout.addWidget(contact_group)





        btn_layout = QHBoxLayout()


        btn_layout.addStretch()





        save_btn = QPushButton("淇濆瓨")


        save_btn.setFixedWidth(100)


        save_btn.clicked.connect(self.save_customer)


        save_btn.setStyleSheet("""


            QPushButton {


                background-color: #2563eb;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px;


            }


            QPushButton:hover { background-color: #1d4ed8; }


        """)


        btn_layout.addWidget(save_btn)





        cancel_btn = QPushButton("鍙栨秷")


        cancel_btn.setFixedWidth(100)


        cancel_btn.clicked.connect(self.reject)


        cancel_btn.setStyleSheet("""


            QPushButton {


                background-color: #e5e7eb;


                color: #374151;


                border: none;


                border-radius: 6px;


                padding: 8px;


            }


            QPushButton:hover { background-color: #d1d5db; }


        """)


        btn_layout.addWidget(cancel_btn)





        layout.addLayout(btn_layout)


        self.setLayout(layout)





    def save_customer(self):


        print("save_customer called")  # 璋冭瘯淇℃伅


        print(f"is_edit: {self.is_edit}")  # 璋冭瘯淇℃伅


        print(f"customer: {self.customer}")  # 璋冭瘯淇℃伅


        


        if not self.name_input.text().strip():


            QMessageBox.warning(self, "鎻愮ず", "璇疯緭鍏ュ鎴峰悕绉?)


            return


        if not self.country_input.text().strip():


            QMessageBox.warning(self, "鎻愮ず", "璇疯緭鍏ユ墍鍦ㄥ浗瀹?)


            return





        dept_map = {"S - 绱㈣嫳鏅?: "S", "W - 缁撮偅": "W", "M - 椹开閭?: "M", "D - 閾惰揪": "D"}


        dept_id = dept_map.get(self.dept_combo.currentText(), "S")





        data = {


            "dept_id": dept_id,


            "customer_name": self.name_input.text().strip(),


            "country": self.country_input.text().strip(),


            "basic_require": self.basic_require_input.toPlainText().strip(),


            "special_require": self.special_input.toPlainText().strip(),


            "payment_terms": self.payment_input.text().strip()


        }





        print(f"鍑嗗淇濆瓨鏁版嵁: {data}")  # 璋冭瘯淇℃伅





        contact_name = self.contact_name_input.text().strip()


        contact_phone = self.contact_phone_input.text().strip()


        contact_email = self.contact_email_input.text().strip()


        if contact_name or contact_phone or contact_email:


            data["contacts"] = [{


                "name": contact_name,


                "phone": contact_phone,


                "email": contact_email,


                "is_primary": 1


            }]





        try:


            print(f"璋冪敤API: {'update' if self.is_edit else 'create'}")  # 璋冭瘯淇℃伅


            if self.is_edit:


                customer_id = self.customer.get('id', 'NOT_FOUND')


                print(f"鏇存柊瀹㈡埛ID: {customer_id}")  # 璋冭瘯淇℃伅


                result = self.api_client.update_customer(customer_id, data)


                print(f"API杩斿洖缁撴灉: {result}")  # 璋冭瘯淇℃伅


            else:


                result = self.api_client.create_customer(data)


                print(f"API杩斿洖缁撴灉: {result}")  # 璋冭瘯淇℃伅


            print("API璋冪敤鎴愬姛")  # 璋冭瘯淇℃伅


            self.accept()


        except Exception as e:


            print(f"API璋冪敤澶辫触: {str(e)}")  # 璋冭瘯淇℃伅


            QMessageBox.warning(self, "閿欒", f"淇濆瓨澶辫触: {str(e)}")


            traceback.print_exc()








class CustomerDetailDialog(QDialog):


    def __init__(self, api_client: ApiClient, customer):


        super().__init__()


        self.api_client = api_client


        self.customer = customer


        self.addresses = []


        self.contacts = []


        self.pi_orders = []


        self.init_ui()


        self.load_data()





    def init_ui(self):


        self.setWindowTitle(f"瀹㈡埛璇︽儏 - {self.customer.get('customer_name', '')}")


        self.setMinimumSize(800, 600)





        layout = QVBoxLayout()


        layout.setContentsMargins(20, 20, 20, 20)





        # 鏍囩椤?        self.tab_widget = QTabWidget()


        


        # 鍩烘湰淇℃伅椤?        self.basic_tab = QWidget()


        self.setup_basic_tab()


        


        # 鏀惰揣鍦板潃椤?        self.address_tab = QWidget()


        self.setup_address_tab()


        


        # 鑱旂郴浜洪〉


        self.contact_tab = QWidget()


        self.setup_contact_tab()


        


        # PI璁㈠崟鍘嗗彶椤?        self.pi_tab = QWidget()


        self.setup_pi_tab()





        self.tab_widget.addTab(self.basic_tab, "鍩烘湰淇℃伅")


        self.tab_widget.addTab(self.address_tab, "鏀惰揣鍦板潃")


        self.tab_widget.addTab(self.contact_tab, "鑱旂郴浜?)


        self.tab_widget.addTab(self.pi_tab, "浜ゆ槗鍘嗗彶")





        layout.addWidget(self.tab_widget)





        # 鍏抽棴鎸夐挳


        btn_layout = QHBoxLayout()


        btn_layout.addStretch()


        close_btn = QPushButton("鍏抽棴")


        close_btn.setFixedWidth(100)


        close_btn.clicked.connect(self.close)


        close_btn.setStyleSheet("""


            QPushButton {


                background-color: #2563eb;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px;


            }


            QPushButton:hover { background-color: #1d4ed8; }


        """)


        btn_layout.addWidget(close_btn)


        layout.addLayout(btn_layout)





        self.setLayout(layout)





    def setup_basic_tab(self):


        layout = QVBoxLayout()


        layout.setContentsMargins(10, 10, 10, 10)


        layout.setSpacing(12)





        form_layout = QFormLayout()


        form_layout.setSpacing(10)





        form_layout.addRow(QLabel("<b>瀹㈡埛缂栧彿:</b>"), QLabel(self.customer.get('customer_code', '')))


        form_layout.addRow(QLabel("<b>瀹㈡埛鍚嶇О:</b>"), QLabel(self.customer.get('customer_name', '')))


        form_layout.addRow(QLabel("<b>鎵€灞為儴闂?</b>"), QLabel(self.customer.get('dept_id', '')))


        form_layout.addRow(QLabel("<b>鎵€鍦ㄥ浗瀹?</b>"), QLabel(self.customer.get('country', '')))


        


        basic_require = self.customer.get('basic_require', '')


        form_layout.addRow(QLabel("<b>鍩烘湰瑕佹眰:</b>"), QLabel(basic_require if basic_require else "-"))


        


        form_layout.addRow(QLabel("<b>浠樻鏉℃:</b>"), QLabel(self.customer.get('payment_terms', '') or "-"))


        


        status = self.customer.get('status', 1)


        status_text = "鍚敤" if status == 1 else "绂佺敤"


        status_color = "#10b981" if status == 1 else "#ef4444"


        status_label = QLabel(status_text)


        status_label.setStyleSheet(f"color: {status_color}; font-weight: bold;")


        form_layout.addRow(QLabel("<b>鐘舵€?</b>"), status_label)





        layout.addLayout(form_layout)


        layout.addStretch()





        special_require = self.customer.get('special_require', '')


        if special_require:


            special_group = QGroupBox("鐗规畩瑕佹眰")


            special_group.setStyleSheet("QGroupBox { font-weight: bold; border: 2px solid #dc2626; border-radius: 5px; }")


            special_layout = QVBoxLayout()


            special_label = QLabel(special_require)


            special_label.setWordWrap(True)


            special_label.setStyleSheet("color: #dc2626; padding: 5px;")


            special_layout.addWidget(special_label)


            special_group.setLayout(special_layout)


            layout.addWidget(special_group)





        self.basic_tab.setLayout(layout)





    def setup_address_tab(self):


        layout = QVBoxLayout()


        layout.setContentsMargins(10, 10, 10, 10)





        # 宸ュ叿鏍?        toolbar = QHBoxLayout()


        add_btn = QPushButton("+ 娣诲姞鍦板潃")


        add_btn.clicked.connect(self.add_address)


        add_btn.setStyleSheet("""


            QPushButton {


                background-color: #2563eb;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 6px 12px;


            }


            QPushButton:hover { background-color: #1d4ed8; }


        """)


        toolbar.addWidget(add_btn)


        toolbar.addStretch()


        layout.addLayout(toolbar)





        # 鍦板潃鍒楄〃


        self.addresses_table = QTableWidget()


        self.addresses_table.setColumnCount(6)


        self.addresses_table.setHorizontalHeaderLabels(["鍥藉", "娓彛", "璇︾粏鍦板潃", "榛樿鍦板潃", "缂栬緫", "鍒犻櫎"])


        self.addresses_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)


        self.addresses_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)


        layout.addWidget(self.addresses_table)





        self.address_tab.setLayout(layout)





    def setup_contact_tab(self):


        layout = QVBoxLayout()


        layout.setContentsMargins(10, 10, 10, 10)





        # 宸ュ叿鏍?        toolbar = QHBoxLayout()


        toolbar.addStretch()


        


        add_btn = QPushButton("+ 鏂板鑱旂郴浜?)


        add_btn.clicked.connect(self.add_contact)


        add_btn.setStyleSheet("""


            QPushButton {


                background-color: #2563eb;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #1d4ed8; }


        """)


        toolbar.addWidget(add_btn)


        


        layout.addLayout(toolbar)





        self.contacts_table = QTableWidget()


        self.contacts_table.setColumnCount(7)


        self.contacts_table.setHorizontalHeaderLabels(["濮撳悕", "鑱屼綅", "鐢佃瘽", "閭", "鏄惁涓昏", "缂栬緫", "鍒犻櫎"])


        self.contacts_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)


        self.contacts_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)


        layout.addWidget(self.contacts_table)





        self.contact_tab.setLayout(layout)





    def setup_pi_tab(self):


        layout = QVBoxLayout()


        layout.setContentsMargins(10, 10, 10, 10)





        self.pi_table = QTableWidget()


        self.pi_table.setColumnCount(10)


        self.pi_table.setHorizontalHeaderLabels(["", "ID", "PI鍙?, "閲戦", "甯佺", "鐘舵€?, "鍒涘缓鏃堕棿", "鎿嶄綔", "瀹屾垚", "瀵煎嚭"])


        self.pi_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)


        self.pi_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)


        self.pi_table.setColumnWidth(0, 40)


        self.pi_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)


        self.pi_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)


        self.pi_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)


        self.pi_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)


        self.pi_table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)


        self.pi_table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)


        self.pi_table.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeMode.ResizeToContents)


        self.pi_table.horizontalHeader().setSectionResizeMode(8, QHeaderView.ResizeMode.ResizeToContents)


        self.pi_table.horizontalHeader().setSectionResizeMode(9, QHeaderView.ResizeMode.ResizeToContents)


        self.setup_table_context_menu(self.pi_table, ["", "ID", "PI鍙?, "閲戦", "甯佺", "鐘舵€?, "鍒涘缓鏃堕棿", "鎿嶄綔", "瀹屾垚", "瀵煎嚭"])


        layout.addWidget(self.pi_table)





        self.pi_tab.setLayout(layout)





    def load_data(self):


        try:


            # 鍔犺浇鍦板潃


            self.addresses = self.api_client.get_customer_addresses(self.customer['id'])


            self.load_addresses_table()





            # 鍔犺浇鑱旂郴浜?            self.contacts = self.api_client.get_customer_contacts(self.customer['id'])


            self.load_contacts_table()





            # 鍔犺浇PI璁㈠崟


            self.pi_orders = self.api_client.get_customer_pi_list(self.customer['id'])


            self.load_pi_table()


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"鍔犺浇鏁版嵁澶辫触: {str(e)}")





    def load_addresses_table(self):


        self.addresses_table.setRowCount(len(self.addresses))


        for row, addr in enumerate(self.addresses):


            self.addresses_table.setItem(row, 0, QTableWidgetItem(addr.get('country', '')))


            self.addresses_table.setItem(row, 1, QTableWidgetItem(addr.get('port', '')))


            self.addresses_table.setItem(row, 2, QTableWidgetItem(addr.get('address_detail', '')))


            


            is_default = addr.get('is_default', 0)


            default_text = "鏄? if is_default == 1 else "鍚?


            self.addresses_table.setItem(row, 3, QTableWidgetItem(default_text))





            # 缂栬緫鎸夐挳


            edit_btn = QPushButton("缂栬緫")


            edit_btn.setFixedWidth(50)


            edit_btn.setStyleSheet("""


                QPushButton {


                    background-color: #3b82f6;


                    color: white;


                    border: none;


                    border-radius: 4px;


                    padding: 4px 8px;


                }


                QPushButton:hover { background-color: #2563eb; }


            """)


            edit_btn.clicked.connect(lambda _, addr=addr: self.edit_address(addr))


            self.addresses_table.setCellWidget(row, 4, edit_btn)


            


            # 鍒犻櫎鎸夐挳


            delete_btn = QPushButton("鍒犻櫎")


            delete_btn.setFixedWidth(50)


            delete_btn.setStyleSheet("""


                QPushButton {


                    background-color: #ef4444;


                    color: white;


                    border: none;


                    border-radius: 4px;


                    padding: 4px 8px;


                }


                QPushButton:hover { background-color: #dc2626; }


            """)


            delete_btn.clicked.connect(lambda _, addr=addr: self.delete_address(addr))


            self.addresses_table.setCellWidget(row, 5, delete_btn)





    def load_contacts_table(self):


        self.contacts_table.setRowCount(len(self.contacts))


        for row, contact in enumerate(self.contacts):


            self.contacts_table.setItem(row, 0, QTableWidgetItem(contact.get('name', '')))


            self.contacts_table.setItem(row, 1, QTableWidgetItem(contact.get('position', '')))


            self.contacts_table.setItem(row, 2, QTableWidgetItem(contact.get('phone', '')))


            self.contacts_table.setItem(row, 3, QTableWidgetItem(contact.get('email', '')))


            


            is_primary = contact.get('is_primary', 0)


            primary_text = "鏄? if is_primary == 1 else "鍚?


            self.contacts_table.setItem(row, 4, QTableWidgetItem(primary_text))





            # 鎿嶄綔鎸夐挳


            btn_layout = QHBoxLayout()


            


            # 缂栬緫鎸夐挳


            edit_btn = QPushButton("缂栬緫")


            edit_btn.setFixedWidth(50)


            edit_btn.setStyleSheet("""


                QPushButton {


                    background-color: #3b82f6;


                    color: white;


                    border: none;


                    border-radius: 4px;


                    padding: 4px 8px;


                }


                QPushButton:hover { background-color: #2563eb; }


            """)


            edit_btn.clicked.connect(lambda _, c=contact: self.edit_contact(c))


            self.contacts_table.setCellWidget(row, 5, edit_btn)


            


            # 鍒犻櫎鎸夐挳


            delete_btn = QPushButton("鍒犻櫎")


            delete_btn.setFixedWidth(50)


            delete_btn.setStyleSheet("""


                QPushButton {


                    background-color: #ef4444;


                    color: white;


                    border: none;


                    border-radius: 4px;


                    padding: 4px 8px;


                }


                QPushButton:hover { background-color: #dc2626; }


            """)


            delete_btn.clicked.connect(lambda _, c=contact: self.delete_contact(c))


            self.contacts_table.setCellWidget(row, 6, delete_btn)





    def load_pi_table(self):


        self.pi_table.setRowCount(len(self.pi_orders))


        for row, pi in enumerate(self.pi_orders):


            self.pi_table.setItem(row, 0, QTableWidgetItem(pi.get('pi_number', '')))


            self.pi_table.setItem(row, 1, QTableWidgetItem(str(pi.get('total_amount', ''))))


            


            status = pi.get('status', '')


            self.pi_table.setItem(row, 2, QTableWidgetItem(status))


            


            created_at = pi.get('created_at', '')


            if created_at:


                created_at = created_at[:19] if isinstance(created_at, str) else str(created_at)


            self.pi_table.setItem(row, 3, QTableWidgetItem(created_at))





    def add_address(self):


        dialog = AddressDialog(self.api_client, customer_id=self.customer['id'])


        if dialog.exec() == QDialog.DialogCode.Accepted:


            self.load_data()





    def edit_address(self, address):


        dialog = AddressDialog(self.api_client, customer_id=self.customer['id'], address=address)


        if dialog.exec() == QDialog.DialogCode.Accepted:


            self.load_data()





    def delete_address(self, address):


        reply = QMessageBox.question(self, "纭鍒犻櫎", "纭畾瑕佸垹闄よ繖涓湴鍧€鍚楋紵",


                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)


        if reply == QMessageBox.StandardButton.Yes:


            try:


                self.api_client.delete_customer_address(self.customer['id'], address['id'])


                self.load_data()


            except Exception as e:


                QMessageBox.warning(self, "閿欒", f"鍒犻櫎澶辫触锛歿str(e)}")





    def add_contact(self):


        dialog = ContactDialog(self.api_client, customer_id=self.customer['id'])


        if dialog.exec() == QDialog.DialogCode.Accepted:


            self.load_data()





    def edit_contact(self, contact):


        dialog = ContactDialog(self.api_client, customer_id=self.customer['id'], contact=contact)


        if dialog.exec() == QDialog.DialogCode.Accepted:


            self.load_data()





    def delete_contact(self, contact):


        reply = QMessageBox.question(self, "纭鍒犻櫎", "纭畾瑕佸垹闄よ繖涓仈绯讳汉鍚楋紵",


                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)


        if reply == QMessageBox.StandardButton.Yes:


            try:


                self.api_client.delete_customer_contact(self.customer['id'], contact['id'])


                self.load_data()


            except Exception as e:


                QMessageBox.warning(self, "閿欒", f"鍒犻櫎澶辫触锛歿str(e)}")








class AddressDialog(QDialog):


    def __init__(self, api_client: ApiClient, customer_id, address=None):


        super().__init__()


        self.api_client = api_client


        self.customer_id = customer_id


        self.address = address


        self.is_edit = address is not None


        self.init_ui()





    def init_ui(self):


        self.setWindowTitle("缂栬緫鍦板潃" if self.is_edit else "娣诲姞鍦板潃")


        self.setFixedSize(400, 300)





        layout = QVBoxLayout()


        layout.setContentsMargins(20, 20, 20, 20)


        layout.setSpacing(15)





        form_layout = QFormLayout()


        form_layout.setSpacing(12)





        self.country_input = QLineEdit()


        if self.address:


            self.country_input.setText(self.address.get('country', ''))


        form_layout.addRow("鍥藉:", self.country_input)





        self.port_input = QLineEdit()


        if self.address:


            self.port_input.setText(self.address.get('port', ''))


        form_layout.addRow("娓彛:", self.port_input)





        self.detail_input = QTextEdit()


        if self.address:


            self.detail_input.setText(self.address.get('address_detail', ''))


        self.detail_input.setMaximumHeight(80)


        form_layout.addRow("璇︾粏鍦板潃:", self.detail_input)





        self.default_checkbox = QCheckBox("璁句负榛樿鍦板潃")


        if self.address and self.address.get('is_default', 0) == 1:


            self.default_checkbox.setChecked(True)


        form_layout.addRow("", self.default_checkbox)





        layout.addLayout(form_layout)





        btn_layout = QHBoxLayout()


        btn_layout.addStretch()





        save_btn = QPushButton("淇濆瓨")


        save_btn.setFixedWidth(100)


        save_btn.clicked.connect(self.save_address)


        save_btn.setStyleSheet("""


            QPushButton {


                background-color: #2563eb;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px;


            }


            QPushButton:hover { background-color: #1d4ed8; }


        """)


        btn_layout.addWidget(save_btn)





        cancel_btn = QPushButton("鍙栨秷")


        cancel_btn.setFixedWidth(100)


        cancel_btn.clicked.connect(self.reject)


        cancel_btn.setStyleSheet("""


            QPushButton {


                background-color: #e5e7eb;


                color: #374151;


                border: none;


                border-radius: 6px;


                padding: 8px;


            }


            QPushButton:hover { background-color: #d1d5db; }


        """)


        btn_layout.addWidget(cancel_btn)





        layout.addLayout(btn_layout)


        self.setLayout(layout)





    def save_address(self):


        data = {


            "country": self.country_input.text().strip(),


            "port": self.port_input.text().strip(),


            "address_detail": self.detail_input.toPlainText().strip(),


            "is_default": 1 if self.default_checkbox.isChecked() else 0


        }





        try:


            if self.is_edit:


                self.api_client.update_customer_address(self.customer_id, self.address['id'], data)


            else:


                self.api_client.create_customer_address(self.customer_id, data)


            self.accept()


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"淇濆瓨澶辫触锛歿str(e)}")








class ContactDialog(QDialog):


    def __init__(self, api_client: ApiClient, customer_id, contact=None):


        super().__init__()


        self.api_client = api_client


        self.customer_id = customer_id


        self.contact = contact


        self.is_edit = contact is not None


        self.init_ui()





    def init_ui(self):


        self.setWindowTitle("缂栬緫鑱旂郴浜? if self.is_edit else "娣诲姞鑱旂郴浜?)


        self.setFixedSize(400, 350)





        layout = QVBoxLayout()


        layout.setContentsMargins(20, 20, 20, 20)


        layout.setSpacing(15)





        form_layout = QFormLayout()


        form_layout.setSpacing(12)





        self.name_input = QLineEdit()


        if self.contact:


            self.name_input.setText(self.contact.get('name', ''))


        form_layout.addRow("濮撳悕:", self.name_input)





        self.position_input = QLineEdit()


        if self.contact:


            self.position_input.setText(self.contact.get('position', ''))


        form_layout.addRow("鑱屼綅:", self.position_input)





        self.phone_input = QLineEdit()


        if self.contact:


            self.phone_input.setText(self.contact.get('phone', ''))


        form_layout.addRow("鐢佃瘽:", self.phone_input)





        self.email_input = QLineEdit()


        if self.contact:


            self.email_input.setText(self.contact.get('email', ''))


        form_layout.addRow("閭:", self.email_input)





        self.primary_checkbox = QCheckBox("璁句负涓昏鑱旂郴浜?)


        if self.contact and self.contact.get('is_primary', 0) == 1:


            self.primary_checkbox.setChecked(True)


        form_layout.addRow("", self.primary_checkbox)





        layout.addLayout(form_layout)





        btn_layout = QHBoxLayout()


        btn_layout.addStretch()





        save_btn = QPushButton("淇濆瓨")


        save_btn.setFixedWidth(100)


        save_btn.clicked.connect(self.save_contact)


        save_btn.setStyleSheet("""


            QPushButton {


                background-color: #2563eb;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px;


            }


            QPushButton:hover { background-color: #1d4ed8; }


        """)


        btn_layout.addWidget(save_btn)





        cancel_btn = QPushButton("鍙栨秷")


        cancel_btn.setFixedWidth(100)


        cancel_btn.clicked.connect(self.reject)


        cancel_btn.setStyleSheet("""


            QPushButton {


                background-color: #e5e7eb;


                color: #374151;


                border: none;


                border-radius: 6px;


                padding: 8px;


            }


            QPushButton:hover { background-color: #d1d5db; }


        """)


        btn_layout.addWidget(cancel_btn)





        layout.addLayout(btn_layout)


        self.setLayout(layout)





    def save_contact(self):


        if not self.name_input.text().strip():


            QMessageBox.warning(self, "鎻愮ず", "璇疯緭鍏ヨ仈绯讳汉濮撳悕")


            return





        data = {


            "name": self.name_input.text().strip(),


            "position": self.position_input.text().strip(),


            "phone": self.phone_input.text().strip(),


            "email": self.email_input.text().strip(),


            "is_primary": 1 if self.primary_checkbox.isChecked() else 0


        }





        try:


            if self.is_edit:


                self.api_client.update_customer_contact(self.customer_id, self.contact['id'], data)


            else:


                self.api_client.create_customer_contact(self.customer_id, data)


            self.accept()


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"淇濆瓨澶辫触锛歿str(e)}")








class SupplierDialog(QDialog):


    def __init__(self, api_client: ApiClient, supplier=None):


        super().__init__()


        self.api_client = api_client


        self.supplier = supplier


        self.is_edit = supplier is not None


        self.provinces = []


        self.cities = []


        self.selected_city_code = ""


        self.init_ui()


        QTimer.singleShot(0, self.load_provinces)





    def load_provinces(self):


        try:


            self.provinces = self.api_client.get_provinces()


            self.province_combo.clear()


            self.province_combo.addItems(self.provinces)


            if self.supplier and self.supplier.get('region'):


                region = self.supplier.get('region', '')


                for prov in self.provinces:


                    if region.startswith(prov):


                        self.province_combo.setCurrentText(prov)


                        self.load_cities(prov)


                        city_name = region[len(prov):].strip()


                        if city_name and city_name in self.cities:


                            self.city_combo.setCurrentText(city_name)


                        break


        except Exception as e:


            print(f"鍔犺浇鐪佷唤澶辫触: {e}")





    def load_cities(self, province):


        try:


            self.cities = self.api_client.get_cities(province)


            self.city_combo.clear()


            self.city_combo.addItems(self.cities)


        except Exception as e:


            print(f"鍔犺浇鍩庡競澶辫触: {e}")





    def on_province_changed(self, province):


        self.load_cities(province)





    def on_city_changed(self, city):


        province = self.province_combo.currentText()


        try:


            # 浣跨敤妯″潡绾у埆鐨勯潤鎬佹槧灏勶紙鍙垱寤轰竴娆★級


            p_code = PROVINCE_CODE_MAP.get(province, "")


            c_map = CITY_CODE_MAP.get(p_code, {})


            self.selected_city_code = p_code + c_map.get(city, "0")


        except Exception as e:


            print(f"鑾峰彇鍩庡競缂栫爜澶辫触: {e}")





    def init_ui(self):


        self.setWindowTitle("缂栬緫渚涘簲鍟? if self.is_edit else "鏂板渚涘簲鍟?)


        self.setFixedSize(500, 480)





        layout = QVBoxLayout()


        layout.setContentsMargins(20, 20, 20, 20)


        layout.setSpacing(15)





        form_layout = QFormLayout()


        form_layout.setSpacing(12)





        if self.is_edit:


            self.code_label = QLabel(self.supplier.get('supplier_code', ''))


            form_layout.addRow("渚涘簲鍟嗙紪鍙?", self.code_label)





        self.name_input = QLineEdit()


        self.name_input.setPlaceholderText("璇疯緭鍏ヤ緵搴斿晢鍚嶇О")


        if self.supplier:


            self.name_input.setText(self.supplier.get('supplier_name', ''))


        form_layout.addRow("渚涘簲鍟嗗悕绉?", self.name_input)





        province_layout = QHBoxLayout()


        self.province_combo = QComboBox()


        self.province_combo.setFixedHeight(35)


        self.province_combo.currentTextChanged.connect(self.on_province_changed)


        province_layout.addWidget(self.province_combo)





        self.city_combo = QComboBox()


        self.city_combo.setFixedHeight(35)


        self.city_combo.currentTextChanged.connect(self.on_city_changed)


        province_layout.addWidget(self.city_combo)


        form_layout.addRow("鐪佷唤/鍩庡競:", province_layout)





        self.contact_input = QLineEdit()


        self.contact_input.setPlaceholderText("璇疯緭鍏ヨ仈绯讳汉")


        if self.supplier:


            self.contact_input.setText(self.supplier.get('contact_person', ''))


        form_layout.addRow("鑱旂郴浜?", self.contact_input)





        self.phone_input = QLineEdit()


        self.phone_input.setPlaceholderText("璇疯緭鍏ヨ仈绯荤數璇?)


        if self.supplier:


            self.phone_input.setText(self.supplier.get('phone', ''))


        form_layout.addRow("鑱旂郴鐢佃瘽:", self.phone_input)





        self.email_input = QLineEdit()


        self.email_input.setPlaceholderText("璇疯緭鍏ラ偖绠卞湴鍧€")


        if self.supplier:


            self.email_input.setText(self.supplier.get('email', ''))


        form_layout.addRow("閭:", self.email_input)





        self.address_input = QTextEdit()


        self.address_input.setPlaceholderText("璇疯緭鍏ヨ缁嗗湴鍧€")


        if self.supplier:


            self.address_input.setText(self.supplier.get('address', ''))


        self.address_input.setMaximumHeight(80)


        form_layout.addRow("璇︾粏鍦板潃:", self.address_input)





        layout.addLayout(form_layout)





        btn_layout = QHBoxLayout()


        btn_layout.addStretch()





        save_btn = QPushButton("淇濆瓨")


        save_btn.setFixedWidth(100)


        save_btn.clicked.connect(self.save_supplier)


        save_btn.setStyleSheet("""


            QPushButton {


                background-color: #2563eb;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px;


            }


            QPushButton:hover { background-color: #1d4ed8; }


        """)


        btn_layout.addWidget(save_btn)





        cancel_btn = QPushButton("鍙栨秷")


        cancel_btn.setFixedWidth(100)


        cancel_btn.clicked.connect(self.reject)


        cancel_btn.setStyleSheet("""


            QPushButton {


                background-color: #e5e7eb;


                color: #374151;


                border: none;


                border-radius: 6px;


                padding: 8px;


            }


            QPushButton:hover { background-color: #d1d5db; }


        """)


        btn_layout.addWidget(cancel_btn)





        layout.addLayout(btn_layout)


        self.setLayout(layout)





    def save_supplier(self):


        if not self.name_input.text().strip():


            QMessageBox.warning(self, "鎻愮ず", "璇疯緭鍏ヤ緵搴斿晢鍚嶇О")


            return





        province = self.province_combo.currentText()


        city = self.city_combo.currentText()


        region = f"{province} {city}" if province and city else ""





        data = {


            "supplier_name": self.name_input.text().strip(),


            "province": province,


            "city": city,


            "city_code": self.selected_city_code,


            "region": region,


            "contact_person": self.contact_input.text().strip(),


            "phone": self.phone_input.text().strip(),


            "email": self.email_input.text().strip(),


            "address": self.address_input.toPlainText().strip()


        }





        try:


            if self.is_edit:


                self.api_client.update_supplier(self.supplier['id'], data)


            else:


                self.api_client.create_supplier(data)


            self.accept()


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"淇濆瓨澶辫触: {str(e)}")








class MainWindow(QMainWindow):


    def __init__(self, api_client: ApiClient, dept_id: str):


        super().__init__()


        self.api_client = api_client


        self.dept_id = dept_id


        self.init_ui()


        self.load_data()





    def init_ui(self):


        self.setWindowTitle(f"PI璁㈠崟绠＄悊绯荤粺 - {DEPARTMENT_CONFIG[self.dept_id]['name']}")


        self.setMinimumSize(1200, 800)


        # 榛樿鍏ㄥ睆鏄剧ず


        self.showMaximized()





        central_widget = QWidget()


        self.setCentralWidget(central_widget)


        main_layout = QVBoxLayout()


        main_layout.setContentsMargins(0, 0, 0, 0)





        header = QWidget()


        header.setFixedHeight(70)


        header.setStyleSheet("background-color: #2563eb;")


        header_layout = QHBoxLayout()


        header_layout.setContentsMargins(20, 0, 20, 0)





        title = QLabel(f"馃摝 PI璁㈠崟绠＄悊绯荤粺 - {DEPARTMENT_CONFIG[self.dept_id]['name']}")


        title.setFont(get_font(16, QFont.Weight.Bold))


        title.setStyleSheet("color: white;")


        header_layout.addWidget(title)





        header_layout.addStretch()





        # 鐢ㄦ埛淇℃伅鍜岀鐞嗗憳妯″紡鍒囨崲


        self.user_info_label = QLabel()


        self.user_info_label.setStyleSheet("color: white; font-size: 14px;")


        header_layout.addWidget(self.user_info_label)


        


        self.admin_mode_label = QLabel()


        self.admin_mode_label.setStyleSheet("color: #fbbf24; font-size: 12px; font-weight: bold;")


        header_layout.addWidget(self.admin_mode_label)


        


        # 閫€鍑虹櫥褰曟寜閽?        logout_btn = QPushButton("閫€鍑?)


        logout_btn.setStyleSheet("""


            QPushButton {


                background-color: #dc2626;


                color: white;


                border: none;


                padding: 5px 15px;


                border-radius: 4px;


                font-size: 14px;


            }


            QPushButton:hover {


                background-color: #b91c1c;


            }


        """)


        logout_btn.clicked.connect(self.logout)


        header_layout.addWidget(logout_btn)





        header.setLayout(header_layout)


        main_layout.addWidget(header)





        content = QWidget()


        content_layout = QHBoxLayout()


        content_layout.setContentsMargins(0, 0, 0, 0)





        sidebar = QWidget()


        sidebar.setFixedWidth(200)


        sidebar.setStyleSheet("background-color: #1e293b;")


        sidebar_layout = QVBoxLayout()


        sidebar_layout.setContentsMargins(0, 20, 0, 0)


        sidebar_layout.setSpacing(5)





        self.tab_buttons = {}


        tabs = [


            ("浜у搧绠＄悊", "products"),


            ("瀹㈡埛绠＄悊", "customers"),


            ("渚涘簲鍟嗙鐞?, "suppliers"),


            ("鎶ヤ环绠＄悊", "quotes"),


            ("PI绠＄悊", "pi"),


            ("閲囪喘绠＄悊", "purchase"),


            ("鍑鸿揣绠＄悊", "shipment"),


            ("瀹㈡埛浠樻", "customer_payment"),


            ("渚涘簲鍟嗕粯娆?, "supplier_payment"),


            ("搴撳瓨绠＄悊", "inventory"),


            ("璁㈠崟鎬昏〃", "order_summary")


        ]





        for name, key in tabs:


            btn = QPushButton(name)


            btn.setFixedHeight(45)


            btn.setFont(get_font(10))


            btn.setStyleSheet("""


                QPushButton {


                    color: white;


                    background-color: transparent;


                    border: none;


                    text-align: left;


                    padding-left: 20px;


                }


                QPushButton:hover {


                    background-color: #334155;


                }


                QPushButton:checked {


                    background-color: #2563eb;


                }


            """)


            btn.setCheckable(True)


            btn.clicked.connect(lambda checked, k=key: self.switch_tab(k))


            sidebar_layout.addWidget(btn)


            self.tab_buttons[key] = btn





        self.tab_buttons["products"].setChecked(True)





        sidebar_layout.addStretch()


        sidebar.setLayout(sidebar_layout)


        content_layout.addWidget(sidebar)





        self.content_stack = QStackedWidget()





        self.products_widget = self.create_products_tab()


        self.customers_widget = self.create_customers_tab()


        self.suppliers_widget = self.create_suppliers_tab()


        self.quotes_widget = self.create_quotes_tab()


        self.pi_widget = self.create_pi_tab()


        self.purchase_widget = self.create_purchase_tab()


        self.shipment_widget = self.create_shipment_tab()


        self.customer_payment_widget = self.create_customer_payment_tab()


        self.supplier_payment_widget = self.create_supplier_payment_tab()


        self.inventory_widget = self.create_inventory_tab()


        self.order_summary_widget = self.create_order_summary_tab()





        self.content_stack.addWidget(self.products_widget)


        self.content_stack.addWidget(self.customers_widget)


        self.content_stack.addWidget(self.suppliers_widget)


        self.content_stack.addWidget(self.quotes_widget)


        self.content_stack.addWidget(self.pi_widget)


        self.content_stack.addWidget(self.purchase_widget)


        self.content_stack.addWidget(self.shipment_widget)


        self.content_stack.addWidget(self.customer_payment_widget)


        self.content_stack.addWidget(self.supplier_payment_widget)


        self.content_stack.addWidget(self.inventory_widget)


        self.content_stack.addWidget(self.order_summary_widget)





        content_layout.addWidget(self.content_stack)


        content.setLayout(content_layout)


        main_layout.addWidget(content)





        central_widget.setLayout(main_layout)





    def switch_tab(self, key):


        print(f"DEBUG - switch_tab called with key: {key}")


        tab_map = {"products": 0, "customers": 1, "suppliers": 2, "quotes": 3, "pi": 4, "purchase": 5,


                   "shipment": 6, "customer_payment": 7, "supplier_payment": 8, "inventory": 9}


        self.content_stack.setCurrentIndex(tab_map.get(key, 0))


        for k, btn in self.tab_buttons.items():


            btn.setChecked(k == key)


        


        # 鍒囨崲鍒癙I鏍囩鏃跺己鍒跺埛鏂帮紙娓呴櫎缂撳瓨锛?        if key == "pi":


            print("DEBUG - Loading PI orders...")


            self.load_pi_orders_async(force_refresh=True)


        else:


            refresh_map = {


                "products": self.load_products_async,


                "customers": self.load_customers_async,


                "suppliers": self.load_suppliers_async,


                "quotes": self.load_quotes_async,


                "purchase": self.load_purchase_orders_async,


                "shipment": self.load_shipments_async,


                "customer_payment": self.load_customer_payments_async,


                "supplier_payment": self.load_supplier_payments_async,


                "inventory": self.load_inventories_async,


            }


            if key in refresh_map:


                refresh_map[key]()


    


    def _load_async(self, api_method, update_method, error_msg="鍔犺浇澶辫触"):


        """閫氱敤寮傛鍔犺浇鏂规硶锛屼娇鐢≦Thread纭繚UI鍦ㄤ富绾跨▼鏇存柊"""


        from PySide6.QtCore import QThread


        


        class LoaderThread(QThread):


            def __init__(self, api_method, parent=None):


                super().__init__(parent)


                self.api_method = api_method


                self.result_data = []


                self.error_occurred = False


            


            def run(self):


                try:


                    data = self.api_method()


                    self.result_data = data if data else []


                except Exception as e:


                    print(f"{error_msg}: {e}")


                    self.error_occurred = True


                    self.result_data = []


        


        thread = LoaderThread(api_method, self)


        thread.finished.connect(lambda: update_method(thread.result_data))


        thread.start()


        return thread


    


    def load_customers_async(self):


        """寮傛鍔犺浇瀹㈡埛鏁版嵁"""


        self._load_async(


            self.api_client.get_customers,


            self._update_customers_table,


            "鍔犺浇瀹㈡埛澶辫触"


        )


    


    def _update_customers_table(self, customers):


        """鍦ㄤ富绾跨▼鏇存柊瀹㈡埛琛ㄦ牸"""


        if not customers:


            customers = []


        self.customers_table.setRowCount(len(customers))


        for row, c in enumerate(customers):


            checkbox = QCheckBox()


            checkbox.setStyleSheet("margin-left: 50%;")


            self.customers_table.setCellWidget(row, 0, checkbox)


            self.customers_table.setItem(row, 1, QTableWidgetItem(str(c.get('id', ''))))


            self.customers_table.setItem(row, 2, QTableWidgetItem(c.get('customer_code', '')))


            self.customers_table.setItem(row, 3, QTableWidgetItem(c.get('customer_name', '')))


            self.customers_table.setItem(row, 4, QTableWidgetItem(c.get('country', '')))


            self.customers_table.setItem(row, 5, QTableWidgetItem(""))


            self.customers_table.setItem(row, 6, QTableWidgetItem(""))


            self.customers_table.setItem(row, 7, QTableWidgetItem(c.get('payment_terms', '')))


            status = c.get('status', 1)


            status_text = "鍚敤" if status == 1 else "绂佺敤"


            status_color = "#10b981" if status == 1 else "#ef4444"


            status_item = QTableWidgetItem(status_text)


            status_item.setForeground(QBrush(QColor(status_color)))


            self.customers_table.setItem(row, 8, status_item)


            action_bar = ActionBarFactory.create_customer_action_bar(


                edit_callback=lambda _, c=c: self.edit_customer(c),


                toggle_callback=lambda _, c=c: self.toggle_customer_status(c),


                status=status


            )


            self.customers_table.setCellWidget(row, 9, action_bar)


        self.customer_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)


        # 寮傛鍔犺浇鑱旂郴浜哄拰鍦板潃


        QTimer.singleShot(0, lambda: self._load_customer_extra_info(customers))


    


    def load_suppliers_async(self):


        """寮傛鍔犺浇渚涘簲鍟嗘暟鎹?""


        self._load_async(


            self.api_client.get_suppliers,


            self._update_suppliers_table,


            "鍔犺浇渚涘簲鍟嗗け璐?


        )


    


    def _update_suppliers_table(self, suppliers):


        """鍦ㄤ富绾跨▼鏇存柊渚涘簲鍟嗚〃鏍?""


        if not suppliers:


            suppliers = []


        self.suppliers_table.setRowCount(len(suppliers))


        for row, s in enumerate(suppliers):


            checkbox = QCheckBox()


            checkbox.setStyleSheet("margin-left: 15px;")


            self.suppliers_table.setCellWidget(row, 0, checkbox)


            self.suppliers_table.setItem(row, 1, QTableWidgetItem(str(s.get('id', ''))))


            self.suppliers_table.setItem(row, 2, QTableWidgetItem(s.get('supplier_code', '')))


            self.suppliers_table.setItem(row, 3, QTableWidgetItem(s.get('supplier_name', '')))


            self.suppliers_table.setItem(row, 4, QTableWidgetItem(s.get('region', '')))


            self.suppliers_table.setItem(row, 5, QTableWidgetItem(s.get('contact_person', '')))


            self.suppliers_table.setItem(row, 6, QTableWidgetItem(s.get('phone', '')))


            edit_btn = QPushButton("缂栬緫")


            edit_btn.setFixedWidth(60)


            edit_btn.clicked.connect(lambda _, s=s: self.edit_supplier(s))


            self.suppliers_table.setCellWidget(row, 7, edit_btn)





    def create_products_tab(self):


        widget = QWidget()


        layout = QVBoxLayout()


        layout.setContentsMargins(20, 20, 20, 20)





        toolbar = QHBoxLayout()


        


        title = QLabel("浜у搧鍒楄〃")


        title.setFont(get_font(14, QFont.Weight.Bold))


        toolbar.addWidget(title)


        


        toolbar.addStretch()





        # 鎼滅储妗?        search_layout = QHBoxLayout()


        search_layout.setSpacing(10)





        self.search_input = QLineEdit()


        self.search_input.setPlaceholderText("鎼滅储锛圤E鍙?浜у搧缂栧彿/宸ュ巶缂栧彿/鍝佺墝/鎻忚堪锛?)


        self.search_input.setFixedWidth(250)


        search_layout.addWidget(self.search_input)





        self.category_filter = QComboBox()


        self.category_filter.addItem("鍏ㄩ儴鍒嗙被", 0)


        search_layout.addWidget(self.category_filter)


        


        # 鍔犺浇浜у搧绫诲埆


        self.load_product_categories()





        search_btn = QPushButton("鎼滅储")


        search_btn.clicked.connect(self.search_products)


        search_btn.setStyleSheet("""


            QPushButton {


                background-color: #3b82f6;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 6px 12px;


            }


            QPushButton:hover { background-color: #2563eb; }


        """)


        search_layout.addWidget(search_btn)





        reset_btn = QPushButton("閲嶇疆")


        reset_btn.clicked.connect(self.reset_search)


        reset_btn.setStyleSheet("""


            QPushButton {


                background-color: #e5e7eb;


                color: #374151;


                border: none;


                border-radius: 6px;


                padding: 6px 12px;


            }


            QPushButton:hover { background-color: #d1d5db; }


        """)


        search_layout.addWidget(reset_btn)





        toolbar.addLayout(search_layout)





        add_btn = QPushButton("+ 鏂板浜у搧")


        add_btn.clicked.connect(self.add_product)


        add_btn.setStyleSheet("""


            QPushButton {


                background-color: #2563eb;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #1d4ed8; }


        """)


        toolbar.addWidget(add_btn)





        refresh_btn = QPushButton("鍒锋柊")


        refresh_btn.clicked.connect(self.load_products)


        refresh_btn.setStyleSheet("""


            QPushButton {


                background-color: #e5e7eb;


                color: #374151;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #d1d5db; }


        """)


        toolbar.addWidget(refresh_btn)





        import_btn = QPushButton("鎵归噺瀵煎叆")


        import_btn.clicked.connect(self.import_products)


        import_btn.setStyleSheet("""


            QPushButton {


                background-color: #10b981;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #059669; }


        """)


        toolbar.addWidget(import_btn)





        # 鎵归噺鎿嶄綔鎸夐挳


        batch_layout = QHBoxLayout()


        batch_layout.setSpacing(10)


        


        self.select_all_checkbox = QCheckBox("鍏ㄩ€?)


        self.select_all_checkbox.stateChanged.connect(self.toggle_select_all_products)


        batch_layout.addWidget(self.select_all_checkbox)


        


        batch_disable_btn = QPushButton("鎵归噺绂佺敤")


        batch_disable_btn.clicked.connect(self.batch_toggle_product_status)


        batch_disable_btn.setStyleSheet("""


            QPushButton {


                background-color: #f59e0b;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 6px 12px;


            }


            QPushButton:hover { background-color: #d97706; }


        """)


        batch_layout.addWidget(batch_disable_btn)


        


        batch_delete_btn = QPushButton("鎵归噺鍒犻櫎")


        batch_delete_btn.clicked.connect(self.batch_delete_products)


        batch_delete_btn.setStyleSheet("""


            QPushButton {


                background-color: #dc2626;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 6px 12px;


            }


            QPushButton:hover { background-color: #b91c1c; }


        """)


        batch_layout.addWidget(batch_delete_btn)


        


        toolbar.addLayout(batch_layout)





        layout.addLayout(toolbar)





        self.products_table = QTableWidget()


        self.products_table.setColumnCount(8)


        self.products_table.setHorizontalHeaderLabels(["閫夋嫨", "浜у搧缂栧彿", "鍥剧墖", "OE鍙?, "閲囪喘娓犻亾", "搴撳瓨閲?, "鐘舵€?, "鎿嶄綔"])


        self.products_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)


        self.products_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)


        self.products_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)


        # 璁剧疆鍥剧墖鍒楀搴﹀拰琛岄珮


        self.products_table.setColumnWidth(2, 70)


        self.products_table.verticalHeader().setDefaultSectionSize(70)


        self.products_table.doubleClicked.connect(self.on_product_double_click)


        self.setup_table_context_menu(self.products_table, ["閫夋嫨", "浜у搧缂栧彿", "鍥剧墖", "OE鍙?, "閲囪喘娓犻亾", "搴撳瓨閲?, "鐘舵€?, "鎿嶄綔"])


        layout.addWidget(self.products_table)





        widget.setLayout(layout)


        return widget





    def create_customers_tab(self):


        widget = QWidget()


        layout = QVBoxLayout()


        layout.setContentsMargins(20, 20, 20, 20)





        toolbar = QHBoxLayout()


        title = QLabel("瀹㈡埛鍒楄〃")


        title.setFont(get_font(14, QFont.Weight.Bold))


        toolbar.addWidget(title)


        toolbar.addStretch()





        self.customer_select_all_checkbox = QCheckBox("鍏ㄩ€?)


        self.customer_select_all_checkbox.clicked.connect(self.toggle_select_all_customers)


        toolbar.addWidget(self.customer_select_all_checkbox)





        self.customer_search_input = QLineEdit()


        self.customer_search_input.setPlaceholderText("鎼滅储瀹㈡埛鍚嶇О/缂栧彿...")


        self.customer_search_input.setFixedWidth(200)


        self.customer_search_input.returnPressed.connect(self.search_customers)


        toolbar.addWidget(self.customer_search_input)





        self.customer_country_filter = QComboBox()


        self.customer_country_filter.addItem("鍏ㄩ儴鍥藉", 0)


        self.customer_country_filter.setFixedWidth(150)


        toolbar.addWidget(self.customer_country_filter)





        search_btn = QPushButton("鎼滅储")


        search_btn.clicked.connect(self.search_customers)


        search_btn.setStyleSheet("""


            QPushButton {


                background-color: #e5e7eb;


                color: #374151;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #d1d5db; }


        """)


        toolbar.addWidget(search_btn)





        add_btn = QPushButton("+ 鏂板瀹㈡埛")


        add_btn.clicked.connect(self.add_customer)


        add_btn.setStyleSheet("""


            QPushButton {


                background-color: #2563eb;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #1d4ed8; }


        """)


        toolbar.addWidget(add_btn)





        refresh_btn = QPushButton("鍒锋柊")


        refresh_btn.clicked.connect(self.load_customers)


        refresh_btn.setStyleSheet("""


            QPushButton {


                background-color: #e5e7eb;


                color: #374151;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #d1d5db; }


        """)


        toolbar.addWidget(refresh_btn)





        layout.addLayout(toolbar)





        self.customers_table = QTableWidget()


        self.customers_table.setColumnCount(10)


        self.customers_table.setHorizontalHeaderLabels(["閫夋嫨", "ID", "瀹㈡埛缂栧彿", "瀹㈡埛鍚嶇О", "鍥藉", "榛樿鑱旂郴浜?, "榛樿鍦板潃", "浠樻鏉℃", "鐘舵€?, "鎿嶄綔"])


        self.customers_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)


        self.customers_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)


        self.customers_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)


        self.customers_table.doubleClicked.connect(self.on_customer_double_click)


        self.setup_table_context_menu(self.customers_table, ["閫夋嫨", "ID", "瀹㈡埛缂栧彿", "瀹㈡埛鍚嶇О", "鍥藉", "榛樿鑱旂郴浜?, "榛樿鍦板潃", "浠樻鏉℃", "鐘舵€?, "鎿嶄綔"])


        layout.addWidget(self.customers_table)





        widget.setLayout(layout)


        return widget





    def create_suppliers_tab(self):


        widget = QWidget()


        layout = QVBoxLayout()


        layout.setContentsMargins(20, 20, 20, 20)





        toolbar = QHBoxLayout()


        title = QLabel("渚涘簲鍟嗗垪琛?)


        title.setFont(get_font(14, QFont.Weight.Bold))


        toolbar.addWidget(title)


        toolbar.addStretch()





        self.select_all_checkbox = QCheckBox("鍏ㄩ€?)


        self.select_all_checkbox.clicked.connect(self.toggle_select_all_suppliers)


        toolbar.addWidget(self.select_all_checkbox)





        delete_btn = QPushButton("鍒犻櫎閫変腑")


        delete_btn.clicked.connect(self.delete_selected_suppliers)


        delete_btn.setStyleSheet("""


            QPushButton {


                background-color: #dc2626;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #b91c1c; }


        """)


        toolbar.addWidget(delete_btn)





        import_btn = QPushButton("鎵归噺瀵煎叆")


        import_btn.clicked.connect(self.import_suppliers)


        import_btn.setStyleSheet("""


            QPushButton {


                background-color: #16a34a;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #15803d; }


        """)


        toolbar.addWidget(import_btn)





        add_btn = QPushButton("+ 鏂板渚涘簲鍟?)


        add_btn.clicked.connect(self.add_supplier)


        add_btn.setStyleSheet("""


            QPushButton {


                background-color: #2563eb;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #1d4ed8; }


        """)


        toolbar.addWidget(add_btn)





        refresh_btn = QPushButton("鍒锋柊")


        refresh_btn.clicked.connect(self.load_suppliers)


        refresh_btn.setStyleSheet("""


            QPushButton {


                background-color: #e5e7eb;


                color: #374151;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #d1d5db; }


        """)


        toolbar.addWidget(refresh_btn)





        layout.addLayout(toolbar)





        self.suppliers_table = QTableWidget()


        self.suppliers_table.setColumnCount(8)


        self.suppliers_table.setHorizontalHeaderLabels(["", "ID", "渚涘簲鍟嗙紪鍙?, "渚涘簲鍟嗗悕绉?, "鍦板尯", "鑱旂郴浜?, "鐢佃瘽", "鎿嶄綔"])


        self.suppliers_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)


        self.suppliers_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)


        self.suppliers_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)


        self.suppliers_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Fixed)


        self.suppliers_table.setColumnWidth(0, 40)


        self.suppliers_table.doubleClicked.connect(self.on_supplier_double_click)


        self.setup_table_context_menu(self.suppliers_table, ["", "ID", "渚涘簲鍟嗙紪鍙?, "渚涘簲鍟嗗悕绉?, "鍦板尯", "鑱旂郴浜?, "鐢佃瘽", "鎿嶄綔"])


        layout.addWidget(self.suppliers_table)





        widget.setLayout(layout)


        return widget





    def toggle_select_all_suppliers(self, checked):


        for row in range(self.suppliers_table.rowCount()):


            checkbox = self.suppliers_table.cellWidget(row, 0)


            if checkbox:


                checkbox.setChecked(checked)





    def delete_selected_suppliers(self):


        selected_ids = []


        for row in range(self.suppliers_table.rowCount()):


            checkbox = self.suppliers_table.cellWidget(row, 0)


            if checkbox and checkbox.isChecked():


                supplier_id = int(self.suppliers_table.item(row, 1).text())


                selected_ids.append(supplier_id)





        if not selected_ids:


            QMessageBox.warning(self, "鎻愮ず", "璇峰厛閫夋嫨瑕佸垹闄ょ殑渚涘簲鍟?)


            return





        reply = QMessageBox.question(self, "纭鍒犻櫎", f"纭畾瑕佸垹闄ら€変腑鐨?{len(selected_ids)} 涓緵搴斿晢鍚楋紵",


                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)


        if reply == QMessageBox.StandardButton.No:


            return





        for supplier_id in selected_ids:


            try:


                self.api_client.delete_supplier(supplier_id)


            except Exception as e:


                QMessageBox.warning(self, "閿欒", f"鍒犻櫎渚涘簲鍟嗗け璐? {str(e)}")





        self.load_suppliers()


        self.select_all_checkbox.setChecked(False)





    def create_pi_tab(self):


        widget = QWidget()


        layout = QVBoxLayout()


        layout.setContentsMargins(20, 20, 20, 20)





        toolbar = QHBoxLayout()


        title = QLabel("PI璁㈠崟鍒楄〃")


        title.setFont(get_font(14, QFont.Weight.Bold))


        toolbar.addWidget(title)


        toolbar.addStretch()





        add_btn = QPushButton("+ 鏂板缓PI")


        add_btn.clicked.connect(self.add_pi)


        add_btn.setStyleSheet("""


            QPushButton {


                background-color: #2563eb;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #1d4ed8; }


        """)


        toolbar.addWidget(add_btn)


        


        self.pi_batch_delete_btn = QPushButton("鎵归噺鍒犻櫎")


        self.pi_batch_delete_btn.clicked.connect(self.batch_delete_pi)


        self.pi_batch_delete_btn.setStyleSheet("""


            QPushButton {


                background-color: #dc2626;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #b91c1c; }


        """)


        toolbar.addWidget(self.pi_batch_delete_btn)


        


        self.pi_batch_export_btn = QPushButton("鎵归噺瀵煎嚭")


        self.pi_batch_export_btn.clicked.connect(self.batch_export_pi)


        self.pi_batch_export_btn.setStyleSheet("""


            QPushButton {


                background-color: #059669;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #047857; }


        """)


        toolbar.addWidget(self.pi_batch_export_btn)


        


        refresh_btn = QPushButton("鍒锋柊")


        refresh_btn.clicked.connect(self.load_pi_orders_async)


        refresh_btn.setStyleSheet("""


            QPushButton {


                background-color: #e5e7eb;


                color: #374151;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #d1d5db; }


        """)


        toolbar.addWidget(refresh_btn)





        layout.addLayout(toolbar)





        self.pi_table = QTableWidget()


        self.pi_table.setColumnCount(10)


        self.pi_table.setHorizontalHeaderLabels(["", "ID", "PI鍙?, "閲戦", "甯佺", "鐘舵€?, "鍒涘缓鏃堕棿", "鎿嶄綔", "瀹屾垚", "瀵煎嚭"])


        self.pi_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)


        self.pi_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)


        self.pi_table.setColumnWidth(0, 40)


        self.pi_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)


        self.pi_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)


        self.pi_table.doubleClicked.connect(self.on_pi_double_click)


        self.setup_table_context_menu(self.pi_table, ["", "ID", "PI鍙?, "閲戦", "甯佺", "鐘舵€?, "鍒涘缓鏃堕棿", "鎿嶄綔", "瀹屾垚", "瀵煎嚭"])


        layout.addWidget(self.pi_table)





        widget.setLayout(layout)


        return widget





    def create_purchase_tab(self):


        widget = QWidget()


        layout = QVBoxLayout()


        layout.setContentsMargins(20, 20, 20, 20)





        toolbar = QHBoxLayout()


        title = QLabel("閲囪喘璁㈠崟鍒楄〃")


        title.setFont(get_font(14, QFont.Weight.Bold))


        toolbar.addWidget(title)


        toolbar.addStretch()





        add_btn = QPushButton("+ 鏂板缓閲囪喘鍗?)


        add_btn.clicked.connect(self.add_purchase)


        add_btn.setStyleSheet("""


            QPushButton {


                background-color: #2563eb;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #1d4ed8; }


        """)


        toolbar.addWidget(add_btn)





        refresh_btn = QPushButton("鍒锋柊")


        refresh_btn.clicked.connect(self.load_purchase_orders)


        refresh_btn.setStyleSheet("""


            QPushButton {


                background-color: #e5e7eb;


                color: #374151;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #d1d5db; }


        """)


        toolbar.addWidget(refresh_btn)





        layout.addLayout(toolbar)





        self.purchase_table = QTableWidget()


        self.purchase_table.setColumnCount(9)


        self.purchase_table.setHorizontalHeaderLabels(["ID", "閲囪喘鍗曞彿", "PI鍙?, "渚涘簲鍟?, "閲戦", "鐘舵€?, "鎿嶄綔", "纭", "鍏ュ簱"])


        self.purchase_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)


        self.purchase_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)


        self.purchase_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)


        self.setup_table_context_menu(self.purchase_table, ["ID", "閲囪喘鍗曞彿", "PI鍙?, "渚涘簲鍟?, "閲戦", "鐘舵€?, "鎿嶄綔"])


        layout.addWidget(self.purchase_table)





        widget.setLayout(layout)


        return widget





    def create_quotes_tab(self):


        """鎶ヤ环绠＄悊鏍囩椤?""


        widget = QWidget()


        layout = QVBoxLayout()


        layout.setContentsMargins(20, 20, 20, 20)





        toolbar = QHBoxLayout()


        title = QLabel("鎶ヤ环绠＄悊")


        title.setFont(get_font(14, QFont.Weight.Bold))


        toolbar.addWidget(title)


        toolbar.addStretch()





        add_btn = QPushButton("+ 鏂板缓鎶ヤ环鍗?)


        add_btn.clicked.connect(self.add_quote)


        add_btn.setStyleSheet("""


            QPushButton {


                background-color: #2563eb;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #1d4ed8; }


        """)


        toolbar.addWidget(add_btn)


        


        batch_delete_btn = QPushButton("鎵归噺鍒犻櫎")


        batch_delete_btn.clicked.connect(self.batch_delete_quotes)


        batch_delete_btn.setStyleSheet("""


            QPushButton {


                background-color: #dc2626;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #b91c1c; }


        """)


        toolbar.addWidget(batch_delete_btn)


        


        batch_export_btn = QPushButton("鎵归噺瀵煎嚭")


        batch_export_btn.clicked.connect(self.batch_export_quotes)


        batch_export_btn.setStyleSheet("""


            QPushButton {


                background-color: #059669;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #047857; }


        """)


        toolbar.addWidget(batch_export_btn)


        


        refresh_btn = QPushButton("鍒锋柊")


        refresh_btn.clicked.connect(self.load_quotes_async)


        refresh_btn.setStyleSheet("""


            QPushButton {


                background-color: #6b7280;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #4b5563; }


        """)


        toolbar.addWidget(refresh_btn)





        layout.addLayout(toolbar)





        # 鎶ヤ环鍗曞垪琛?        self.quote_table = QTableWidget()


        self.quote_table.setColumnCount(10)


        self.quote_table.setHorizontalHeaderLabels(["", "ID", "鎶ヤ环鍗曞彿", "瀹㈡埛", "閲戦", "甯佺", "鐘舵€?, "鏈夋晥鏈?, "澶囨敞", "鎿嶄綔"])


        self.quote_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)


        self.quote_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)


        self.quote_table.setColumnHidden(1, True)  # 闅愯棌ID鍒?        self.quote_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)


        self.quote_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)


        self.quote_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)


        self.quote_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)


        self.quote_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)


        self.quote_table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)


        self.quote_table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)


        self.quote_table.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeMode.ResizeToContents)


        self.quote_table.horizontalHeader().setSectionResizeMode(8, QHeaderView.ResizeMode.Stretch)


        self.quote_table.horizontalHeader().setSectionResizeMode(9, QHeaderView.ResizeMode.ResizeToContents)


        layout.addWidget(self.quote_table)





        widget.setLayout(layout)


        return widget





    def load_quotes_async(self):


        """寮傛鍔犺浇鎶ヤ环鍗?""


        self._load_async(


            self.api_client.get_quotes,


            self._update_quote_table,


            "鍔犺浇鎶ヤ环鍗曞け璐?


        )





    def _update_quote_table(self, quotes):


        """鏇存柊鎶ヤ环鍗曡〃鏍?""


        try:


            self.quote_table.setRowCount(len(quotes))


            status_map = {1: "鑽夌", 2: "宸插彂閫?, 3: "宸叉帴鍙?, 4: "宸叉嫆缁?}


            


            for row, q in enumerate(quotes):


                # 澶嶉€夋


                checkbox = QTableWidgetItem()


                checkbox.setFlags(checkbox.flags() | Qt.ItemFlag.ItemIsUserCheckable)


                checkbox.setCheckState(Qt.CheckState.Unchecked)


                checkbox.setData(Qt.ItemDataRole.UserRole, q.get('id'))


                self.quote_table.setItem(row, 0, checkbox)


                


                self.quote_table.setItem(row, 1, QTableWidgetItem(str(q.get('id', ''))))


                self.quote_table.setItem(row, 2, QTableWidgetItem(q.get('quote_no', '')))


                self.quote_table.setItem(row, 3, QTableWidgetItem(q.get('customer_name', '')))


                self.quote_table.setItem(row, 4, QTableWidgetItem(f"${q.get('total_amount', 0):,.2f}"))


                self.quote_table.setItem(row, 5, QTableWidgetItem(q.get('currency', 'USD')))


                self.quote_table.setItem(row, 6, QTableWidgetItem(status_map.get(q.get('status', 1), '鑽夌')))


                valid_until = q.get('valid_until')


                if valid_until:


                    valid_until = str(valid_until)[:10]


                self.quote_table.setItem(row, 7, QTableWidgetItem(valid_until or '-'))


                self.quote_table.setItem(row, 8, QTableWidgetItem(q.get('remark', '') or '-'))


                


                # 鎿嶄綔鎸夐挳


                btn_widget = QWidget()


                btn_layout = QHBoxLayout()


                btn_layout.setContentsMargins(0, 0, 0, 0)


                


                edit_btn = QPushButton("缂栬緫")


                edit_btn.setFixedWidth(50)


                edit_btn.clicked.connect(lambda _, qid=q.get('id'): self.edit_quote(qid))


                btn_layout.addWidget(edit_btn)


                


                pi_btn = QPushButton("杞琍I")


                pi_btn.setFixedWidth(50)


                pi_btn.setStyleSheet("color: #10b981;")


                pi_btn.clicked.connect(lambda _, qid=q.get('id'): self.convert_quote_to_pi(qid))


                btn_layout.addWidget(pi_btn)


                


                btn_widget.setLayout(btn_layout)


                self.quote_table.setCellWidget(row, 9, btn_widget)


        except Exception as e:


            print(f"鏇存柊鎶ヤ环鍗曡〃鏍煎け璐? {e}")


        finally:


            self.quote_table.viewport().update()


    


    def get_selected_quote_ids(self):


        """鑾峰彇閫変腑鐨勬姤浠峰崟ID鍒楄〃"""


        ids = []


        for row in range(self.quote_table.rowCount()):


            item = self.quote_table.item(row, 0)


            if item and item.checkState() == Qt.CheckState.Checked:


                quote_id = item.data(Qt.ItemDataRole.UserRole)


                if quote_id:


                    ids.append(quote_id)


        return ids


    


    def toggle_all_quotes(self, state):


        """鍏ㄩ€?鍙栨秷鍏ㄩ€?""


        check_state = Qt.CheckState.Checked if state else Qt.CheckState.Unchecked


        for row in range(self.quote_table.rowCount()):


            item = self.quote_table.item(row, 0)


            if item:


                item.setCheckState(check_state)


    


    def batch_delete_quotes(self):


        """鎵归噺鍒犻櫎鎶ヤ环鍗?""


        selected_ids = self.get_selected_quote_ids()


        if not selected_ids:


            QMessageBox.warning(self, "鎻愮ず", "璇峰厛閫夋嫨瑕佸垹闄ょ殑鎶ヤ环鍗?)


            return


        


        reply = QMessageBox.question(


            self, "纭鍒犻櫎", 


            f"纭畾瑕佸垹闄ら€変腑鐨?{len(selected_ids)} 涓姤浠峰崟鍚楋紵",


            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No


        )


        if reply != QMessageBox.StandardButton.Yes:


            return


        


        try:


            result = self.api_client.batch_delete_quotes(selected_ids)


            deleted = result.get('deleted', 0)


            errors = result.get('errors', [])


            if errors:


                QMessageBox.warning(self, "閮ㄥ垎鍒犻櫎澶辫触", f"鎴愬姛鍒犻櫎 {deleted} 涓猏n澶辫触: {len(errors)} 涓猏n{errors}")


            else:


                QMessageBox.information(self, "鎴愬姛", f"宸插垹闄?{deleted} 涓姤浠峰崟")


            self.load_quotes_async()


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"鍒犻櫎澶辫触: {str(e)}")


            self.load_quotes_async()


    


    def batch_export_quotes(self):


        """鎵归噺瀵煎嚭鎶ヤ环鍗?""


        selected_ids = self.get_selected_quote_ids()


        if not selected_ids:


            QMessageBox.warning(self, "鎻愮ず", "璇峰厛閫夋嫨瑕佸鍑虹殑鎶ヤ环鍗?)


            return


        


        try:


            quotes_data = []


            for quote_id in selected_ids:


                quote = self.api_client.get_quote(quote_id)


                quotes_data.append(quote)


            


            if not quotes_data:


                QMessageBox.information(self, "鎻愮ず", "娌℃湁鍙鍑虹殑鏁版嵁")


                return


            


            # 鏋勫缓瀵煎嚭鏁版嵁


            export_rows = []


            for q in quotes_data:


                for item in q.get('items', []):


                    export_rows.append({


                        '鎶ヤ环鍗曞彿': q.get('quote_no', ''),


                        '瀹㈡埛': q.get('customer_name', ''),


                        '甯佺': q.get('currency', 'USD'),


                        '鎬婚噾棰?: q.get('total_amount', 0),


                        '鏈夋晥鏈?: q.get('valid_until', ''),


                        '鐘舵€?: ['鑽夌', '宸插彂閫?, '宸叉帴鍙?, '宸叉嫆缁?][q.get('status', 1) - 1] if q.get('status', 1) <= 4 else '',


                        '澶囨敞': q.get('remark', ''),


                        '浜у搧缂栧彿': item.get('product_id', ''),


                        'OE鍙?: item.get('oe_number', ''),


                        '瀹㈡埛缂栧彿': item.get('customer_code', ''),


                        '浜у搧鎻忚堪': item.get('detail_desc', ''),


                        '鏁伴噺': item.get('quantity', 0),


                        '鍗曚环': item.get('unit_price', 0),


                        '鎬讳环': item.get('total_price', 0),


                        '鏄庣粏澶囨敞': item.get('remark', ''),


                    })


            


            # 鍒涘缓DataFrame骞跺鍑?            df = pd.DataFrame(export_rows)


            


            # 閫夋嫨淇濆瓨璺緞


            from PySide6.QtWidgets import QFileDialog


            file_path, _ = QFileDialog.getSaveFileName(


                self, "淇濆瓨鎶ヤ环鍗?, 


                f"鎶ヤ环鍗曞鍑篲{len(selected_ids)}涓?xlsx",


                "Excel Files (*.xlsx)"


            )


            if file_path:


                df.to_excel(file_path, index=False, engine='openpyxl')


                QMessageBox.information(self, "鎴愬姛", f"宸插鍑?{len(export_rows)} 鏉℃槑缁嗗埌:\n{file_path}")


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"瀵煎嚭澶辫触: {str(e)}")





    def add_quote(self):


        """鏂板缓鎶ヤ环鍗?""


        dialog = QuoteDialog(self, self.api_client, self.dept_id)


        if dialog.exec() == QDialog.DialogCode.Accepted:


            QMessageBox.information(self, "鎴愬姛", "鎶ヤ环鍗曞凡淇濆瓨")


            self.load_quotes_async()





    def edit_quote(self, quote_id):


        """缂栬緫鎶ヤ环鍗?""


        try:


            quote = self.api_client.get_quote(quote_id)


            dialog = QuoteDialog(self, self.api_client, self.dept_id, quote)


            if dialog.exec() == QDialog.DialogCode.Accepted:


                self.load_quotes_async()


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"鑾峰彇鎶ヤ环鍗曞け璐? {str(e)}")





    def convert_quote_to_pi(self, quote_id):


        """灏嗘姤浠峰崟杞负PI"""


        reply = QMessageBox.question(self, "纭", "纭畾瑕佸皢姝ゆ姤浠峰崟杞负PI鍚楋紵")


        if reply == QMessageBox.StandardButton.Yes:


            try:


                result = self.api_client.convert_quote_to_pi(quote_id)


                QMessageBox.information(self, "鎴愬姛", f"鎶ヤ环鍗曞凡杞负PI\nPI鍗曞彿: {result.get('pi_no', '')}")


                self.load_quotes_async()


            except Exception as e:


                QMessageBox.warning(self, "閿欒", f"杞崲澶辫触: {str(e)}")





    def create_shipment_tab(self):


        """鍑鸿揣绠＄悊鏍囩椤?- 涓讳粠琛ㄨ璁★紙鍙傝€冨簱瀛樼鐞嗭級"""


        widget = QWidget()


        layout = QVBoxLayout()


        layout.setContentsMargins(20, 20, 20, 20)





        toolbar = QHBoxLayout()


        title = QLabel("鍑鸿揣绠＄悊")


        title.setFont(get_font(14, QFont.Weight.Bold))


        toolbar.addWidget(title)


        toolbar.addStretch()





        add_btn = QPushButton("+ 鏂板缓鍑鸿揣")


        add_btn.clicked.connect(self.add_shipment)


        add_btn.setStyleSheet("""


            QPushButton {


                background-color: #2563eb;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #1d4ed8; }


        """)


        toolbar.addWidget(add_btn)





        refresh_btn = QPushButton("鍒锋柊")


        refresh_btn.clicked.connect(self.load_shipments_async)


        refresh_btn.setStyleSheet("""


            QPushButton {


                background-color: #e5e7eb;


                color: #374151;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #d1d5db; }


        """)


        toolbar.addWidget(refresh_btn)





        layout.addLayout(toolbar)





        # 涓昏〃锛氬嚭璐ф眹鎬?        self.shipment_table = QTableWidget()


        self.shipment_table.setColumnCount(9)


        self.shipment_table.setHorizontalHeaderLabels(["ID", "PI鍙?, "鎬婚噾棰?, "鎬荤鏁?, "浠樻鐘舵€?, "鍑鸿揣鐘舵€?, "闃舵鏁?, "鎿嶄綔", "纭"])


        self.shipment_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)


        self.shipment_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)


        self.shipment_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)


        self.setup_table_context_menu(self.shipment_table, ["ID", "PI鍙?, "鎬婚噾棰?, "鎬荤鏁?, "浠樻鐘舵€?, "鍑鸿揣鐘舵€?, "闃舵鏁?, "鎿嶄綔"])


        layout.addWidget(self.shipment_table)





        # 璇︽儏鏍囩


        detail_label = QLabel("馃搵 鍑鸿揣闃舵璇︽儏锛堣鐐瑰嚮涓婃柟璁板綍鏌ョ湅锛?)


        detail_label.setFont(get_font(12, QFont.Weight.Bold))


        layout.addWidget(detail_label)





        # 浠庤〃锛氬嚭璐ч樁娈垫槑缁?        self.shipment_stage_table = QTableWidget()


        self.shipment_stage_table.setColumnCount(9)


        self.shipment_stage_table.setHorizontalHeaderLabels(["闃舵", "鍑鸿揣鏃ユ湡", "鏌滃彿", "鎻愬崟鍙?, "鏁伴噺", "搴撳瓨", "瀛樻斁浣嶇疆", "浠樻鐘舵€?, "鎿嶄綔"])


        self.shipment_stage_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)


        self.shipment_stage_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)


        self.shipment_stage_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)


        self.shipment_stage_table.setMaximumHeight(250)


        layout.addWidget(self.shipment_stage_table)





        # 缁戝畾鐐瑰嚮浜嬩欢


        try:


            self.shipment_table.cellClicked.disconnect()


        except RuntimeError:


            pass


        self.shipment_table.cellClicked.connect(self.show_shipment_stages)





        widget.setLayout(layout)


        return widget





    def create_customer_payment_tab(self):


        widget = QWidget()


        layout = QVBoxLayout()


        layout.setContentsMargins(20, 20, 20, 20)





        toolbar = QHBoxLayout()


        title = QLabel("瀹㈡埛浠樻鍒楄〃")


        title.setFont(get_font(14, QFont.Weight.Bold))


        toolbar.addWidget(title)


        toolbar.addStretch()





        add_btn = QPushButton("+ 鏂板缓浠樻璁板綍")


        add_btn.clicked.connect(self.add_customer_payment)


        add_btn.setStyleSheet("""


            QPushButton {


                background-color: #2563eb;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #1d4ed8; }


        """)


        toolbar.addWidget(add_btn)





        refresh_btn = QPushButton("鍒锋柊")


        refresh_btn.clicked.connect(self.load_customer_payments)


        refresh_btn.setStyleSheet("""


            QPushButton {


                background-color: #e5e7eb;


                color: #374151;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #d1d5db; }


        """)


        toolbar.addWidget(refresh_btn)





        layout.addLayout(toolbar)





        self.customer_payment_table = QTableWidget()


        self.customer_payment_table.setColumnCount(6)


        self.customer_payment_table.setHorizontalHeaderLabels(["ID", "PI鍙?, "浠樻鏃ユ湡", "閲戦", "浠樻鏂瑰紡", "鎿嶄綔"])


        self.customer_payment_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)


        self.customer_payment_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)


        self.customer_payment_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)


        self.setup_table_context_menu(self.customer_payment_table, ["ID", "PI鍙?, "浠樻鏃ユ湡", "閲戦", "浠樻鏂瑰紡", "鎿嶄綔"])


        layout.addWidget(self.customer_payment_table)





        widget.setLayout(layout)


        return widget





    def create_supplier_payment_tab(self):


        """渚涘簲鍟嗕粯娆炬爣绛鹃〉 - 涓讳粠琛ㄨ璁★紙鍙傝€冨簱瀛樼鐞嗭級"""


        widget = QWidget()


        layout = QVBoxLayout()


        layout.setContentsMargins(20, 20, 20, 20)





        toolbar = QHBoxLayout()


        title = QLabel("渚涘簲鍟嗕粯娆剧鐞?)


        title.setFont(get_font(14, QFont.Weight.Bold))


        toolbar.addWidget(title)


        toolbar.addStretch()





        add_btn = QPushButton("+ 鏂板缓浠樻璁板綍")


        add_btn.clicked.connect(self.add_supplier_payment)


        add_btn.setStyleSheet("""


            QPushButton {


                background-color: #2563eb;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #1d4ed8; }


        """)


        toolbar.addWidget(add_btn)





        refresh_btn = QPushButton("鍒锋柊")


        refresh_btn.clicked.connect(self.load_supplier_payments_async)


        refresh_btn.setStyleSheet("""


            QPushButton {


                background-color: #e5e7eb;


                color: #374151;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #d1d5db; }


        """)


        toolbar.addWidget(refresh_btn)





        layout.addLayout(toolbar)





        # 涓昏〃锛氫粯娆炬眹鎬?        self.supplier_payment_table = QTableWidget()


        self.supplier_payment_table.setColumnCount(8)


        self.supplier_payment_table.setHorizontalHeaderLabels(["ID", "渚涘簲鍟?, "閲囪喘鍗?, "鎬婚噾棰?, "宸蹭粯閲戦", "鏈粯閲戦", "鐘舵€?, "鎿嶄綔"])


        self.supplier_payment_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)


        self.supplier_payment_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)


        self.supplier_payment_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)


        self.setup_table_context_menu(self.supplier_payment_table, ["ID", "渚涘簲鍟?, "閲囪喘鍗?, "鎬婚噾棰?, "宸蹭粯閲戦", "鏈粯閲戦", "鐘舵€?, "鎿嶄綔"])


        layout.addWidget(self.supplier_payment_table)





        # 璇︽儏鏍囩


        detail_label = QLabel("馃搵 浠樻闃舵璇︽儏锛堣鐐瑰嚮涓婃柟璁板綍鏌ョ湅锛?)


        detail_label.setFont(get_font(12, QFont.Weight.Bold))


        layout.addWidget(detail_label)





        # 浠庤〃锛氫粯娆鹃樁娈垫槑缁?        self.supplier_payment_stage_table = QTableWidget()


        self.supplier_payment_stage_table.setColumnCount(7)


        self.supplier_payment_stage_table.setHorizontalHeaderLabels(["闃舵", "搴斾粯閲戦", "宸蹭粯閲戦", "鐘舵€?, "浠樻鏃ユ湡", "鍑瘉", "鎿嶄綔"])


        self.supplier_payment_stage_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)


        self.supplier_payment_stage_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)


        self.supplier_payment_stage_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)


        self.supplier_payment_stage_table.setMaximumHeight(200)


        layout.addWidget(self.supplier_payment_stage_table)





        # 缁戝畾鐐瑰嚮浜嬩欢


        try:


            self.supplier_payment_table.cellClicked.disconnect()


        except RuntimeError:


            pass


        self.supplier_payment_table.cellClicked.connect(self.show_supplier_payment_stages)





        widget.setLayout(layout)


        return widget





    def create_inventory_tab(self):


        widget = QWidget()


        layout = QVBoxLayout()


        layout.setContentsMargins(20, 20, 20, 20)





        toolbar = QHBoxLayout()


        title = QLabel("搴撳瓨绠＄悊锛堟寜OE鍙峰垎缁勶級")


        title.setFont(get_font(14, QFont.Weight.Bold))


        toolbar.addWidget(title)


        toolbar.addStretch()





        add_btn = QPushButton("+ 鏂板缓搴撳瓨")


        add_btn.clicked.connect(self.add_inventory)


        add_btn.setStyleSheet("""


            QPushButton {


                background-color: #2563eb;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #1d4ed8; }


        """)


        toolbar.addWidget(add_btn)





        refresh_btn = QPushButton("鍒锋柊")


        refresh_btn.clicked.connect(self.load_inventories)


        refresh_btn.setStyleSheet("""


            QPushButton {


                background-color: #e5e7eb;


                color: #374151;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #d1d5db; }


        """)


        toolbar.addWidget(refresh_btn)





        layout.addLayout(toolbar)





        # 鎼滅储妗?        search_layout = QHBoxLayout()


        self.inventory_search_input = QLineEdit()


        self.inventory_search_input.setPlaceholderText("鎼滅储OE鍙?..")


        self.inventory_search_input.setFixedHeight(35)


        self.inventory_search_input.returnPressed.connect(self.search_inventory)


        search_layout.addWidget(self.inventory_search_input)


        


        search_btn = QPushButton("鎼滅储")


        search_btn.clicked.connect(self.search_inventory)


        search_btn.setStyleSheet("""


            QPushButton {


                background-color: #3b82f6;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


        """)


        search_layout.addWidget(search_btn)


        layout.addLayout(search_layout)





        # 涓昏〃鏍硷細鏄剧ずOE鍙凤紙浜у搧缁村害锛?        self.inventory_product_table = QTableWidget()


        self.inventory_product_table.setColumnCount(10)


        self.inventory_product_table.setHorizontalHeaderLabels(["OE鍙?, "浜у搧缂栧彿", "鎬诲簱瀛?, "渚涘簲鍟?, "瀹㈡埛", "鐘舵€佸垎甯?, "鏈€杩戝叆搴撲緵搴斿晢", "鏈€杩戝嚭搴撳鎴?, "鏈€杩戝彉鏇?, "鎿嶄綔"])


        self.inventory_product_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Interactive)


        self.inventory_product_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Interactive)


        self.inventory_product_table.setColumnWidth(0, 150)


        self.inventory_product_table.setColumnWidth(1, 110)


        self.inventory_product_table.setColumnWidth(2, 60)


        self.inventory_product_table.setColumnWidth(3, 60)


        self.inventory_product_table.setColumnWidth(4, 60)


        self.inventory_product_table.setColumnWidth(5, 100)


        self.inventory_product_table.setColumnWidth(6, 100)


        self.inventory_product_table.setColumnWidth(7, 100)


        self.inventory_product_table.setColumnWidth(8, 120)


        self.inventory_product_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)


        self.inventory_product_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)


        layout.addWidget(QLabel("馃摝 浜у搧搴撳瓨姹囨€伙紙鍙屽嚮灞曞紑鏌ョ湅璇︽儏锛?))


        layout.addWidget(self.inventory_product_table)


        


        # 鍒嗛殧绾?        line = QFrame()


        line.setFrameShape(QFrame.Shape.HLine)


        line.setStyleSheet("background-color: #e5e7eb;")


        line.setFixedHeight(2)


        layout.addWidget(line)


        


        # 瀛愯〃鏍硷細鏄剧ず閫変腑OE鍙风殑璇︾粏搴撳瓨璁板綍


        self.inventory_detail_label = QLabel("馃搵 搴撳瓨璇︽儏锛堣鍏堥€夋嫨涓婃柟浜у搧锛?)


        layout.addWidget(self.inventory_detail_label)


        


        self.inventory_detail_table = QTableWidget()


        self.inventory_detail_table.setColumnCount(9)


        self.inventory_detail_table.setHorizontalHeaderLabels(["ID", "渚涘簲鍟?, "瀹㈡埛", "鏁伴噺", "搴撲綅", "鐘舵€?, "澶囨敞", "鍒涘缓鏃堕棿", "鎿嶄綔"])


        self.inventory_detail_table.setColumnWidth(0, 40)


        self.inventory_detail_table.setColumnWidth(1, 100)


        self.inventory_detail_table.setColumnWidth(2, 100)


        self.inventory_detail_table.setColumnWidth(3, 60)


        self.inventory_detail_table.setColumnWidth(4, 80)


        self.inventory_detail_table.setColumnWidth(5, 50)


        self.inventory_detail_table.setColumnWidth(6, 120)


        self.inventory_detail_table.setColumnWidth(7, 130)


        self.inventory_detail_table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeMode.Stretch)


        self.inventory_detail_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)


        self.inventory_detail_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)


        layout.addWidget(self.inventory_detail_table)





        status_label = QLabel("鈼?榛勮壊: 閲囪喘鍦ㄩ€?| 鈼?钃濊壊: 寰呭叆搴?| 鈼?缁胯壊: 宸插叆搴?| 鈼?榛戣壊: 鍘嗗彶搴撳瓨")


        status_label.setStyleSheet("color: #6b7280; padding: 5px;")


        layout.addWidget(status_label)





        widget.setLayout(layout)


        return widget





    def load_data(self):


        """寮傛鍔犺浇鎵€鏈夋暟鎹紝閬垮厤闃诲UI"""


        self.update_user_info()


        


        # 鏄剧ず棰勫姞杞界姸鎬佹彁绀?        self._show_loading_tip("姝ｅ湪鍚屾鏈嶅姟鍣ㄦ暟鎹?..")


        


        # 鎵€鏈夋ā鍧楅兘浣跨敤寮傛鍔犺浇锛屽悓鏃惰Е鍙?        QTimer.singleShot(0, self.load_products_async)


        QTimer.singleShot(0, self.load_customers_async)


        QTimer.singleShot(0, self.load_suppliers_async)


        QTimer.singleShot(0, self.load_pi_orders_async)


        QTimer.singleShot(0, self.load_purchase_orders_async)


        QTimer.singleShot(0, self.load_shipments_async)


        QTimer.singleShot(0, self.load_customer_payments_async)


        QTimer.singleShot(0, self.load_supplier_payments_async)


        QTimer.singleShot(0, self.load_inventories_async)


    


    def _show_loading_tip(self, message: str):


        """鏄剧ず鍔犺浇鎻愮ず锛堢姸鎬佹爮鎴栦复鏃惰鐩栧眰锛?""


        # 鍦ㄤ富绐楀彛搴曢儴鐘舵€佹爮鏄剧ず鎻愮ず


        if not hasattr(self, '_status_label'):


            self._status_label = QLabel()


            self._status_label.setStyleSheet("""


                QLabel {


                    background-color: #2563eb;


                    color: white;


                    padding: 8px 20px;


                    font-size: 12px;


                }


            """)


            self.statusBar().addPermanentWidget(self._status_label)


        self._status_label.setText(f"  鈴?{message} ")


        # 3绉掑悗鑷姩闅愯棌


        QTimer.singleShot(3000, self._hide_loading_tip)


    


    def _hide_loading_tip(self):


        """闅愯棌鍔犺浇鎻愮ず"""


        if hasattr(self, '_status_label'):


            self._status_label.setText("")





    def update_user_info(self):


        """鏇存柊鐢ㄦ埛淇℃伅鏄剧ず"""


        if hasattr(self.api_client, 'current_user') and self.api_client.current_user:


            user = self.api_client.current_user


            self.user_info_label.setText(f"馃懁 {user.get('real_name', '鐢ㄦ埛')}")


            


            if user.get('is_admin'):


                self.admin_mode_label.setText("馃攽 绠＄悊鍛樻ā寮?)


            else:


                self.admin_mode_label.setText("馃懁 鏅€氱敤鎴?)


        else:


            self.user_info_label.setText("馃懁 鏈櫥褰?)


            self.admin_mode_label.setText("")


    


    def logout(self):


        """閫€鍑虹櫥褰?""


        reply = QMessageBox.question(


            self, 


            "纭閫€鍑?, 


            "纭畾瑕侀€€鍑虹櫥褰曞悧锛?,


            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No


        )


        


        if reply == QMessageBox.StandardButton.Yes:


            if hasattr(self.api_client, 'logout'):


                self.api_client.logout()


            self.close()





    def create_order_summary_tab(self):


        """鍒涘缓璁㈠崟鎬昏〃Tab - 姹囨€绘墍鏈夋ā鍧楁暟鎹?""


        widget = QWidget()


        layout = QVBoxLayout()


        layout.setContentsMargins(20, 20, 20, 20)


        


        # 宸ュ叿鏍?        toolbar = QHBoxLayout()


        title = QLabel("璁㈠崟绠＄悊鎬昏〃")


        title.setFont(get_font(14, QFont.Weight.Bold))


        toolbar.addWidget(title)


        toolbar.addStretch()


        


        # 鍒锋柊鎸夐挳


        refresh_btn = QPushButton("馃攧 鍒锋柊鏁版嵁")


        refresh_btn.clicked.connect(self.load_order_summary)


        refresh_btn.setStyleSheet("""


            QPushButton {


                background-color: #e5e7eb;


                color: #374151;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #d1d5db; }


        """)


        toolbar.addWidget(refresh_btn)


        


        # 瀵煎嚭鎸夐挳


        export_btn = QPushButton("馃摜 瀵煎嚭Excel")


        export_btn.clicked.connect(self.export_order_summary)


        export_btn.setStyleSheet("""


            QPushButton {


                background-color: #10b981;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #059669; }


        """)


        toolbar.addWidget(export_btn)


        


        layout.addLayout(toolbar)


        


        # 鎼滅储鏍?        search_layout = QHBoxLayout()


        self.order_summary_search = QLineEdit()


        self.order_summary_search.setPlaceholderText("鎼滅储璁㈠崟鍙?OE鍙?瀹㈡埛...")


        self.order_summary_search.setFixedHeight(35)


        self.order_summary_search.textChanged.connect(self.filter_order_summary)


        search_layout.addWidget(self.order_summary_search)


        


        # 鐘舵€佺瓫閫?        self.order_status_filter = QComboBox()


        self.order_status_filter.addItems(["鍏ㄩ儴", "杩涜涓?, "宸插畬鎴?, "宸插彇娑?])


        self.order_status_filter.currentTextChanged.connect(self.filter_order_summary)


        search_layout.addWidget(self.order_status_filter)


        


        # 鏃ユ湡绛涢€?        date_layout = QHBoxLayout()


        date_layout.addWidget(QLabel("浠?"))


        self.order_date_from = QDateEdit()


        self.order_date_from.setCalendarPopup(True)


        self.order_date_from.setDate(QDate.currentDate().addMonths(-1))


        date_layout.addWidget(self.order_date_from)


        date_layout.addWidget(QLabel("鑷?"))


        self.order_date_to = QDateEdit()


        self.order_date_to.setCalendarPopup(True)


        self.order_date_to.setDate(QDate.currentDate())


        date_layout.addWidget(self.order_date_to)


        search_layout.addLayout(date_layout)


        


        layout.addLayout(search_layout)


        


        # 涓昏〃鏍?- 41鍒楁寜妯℃澘璁捐


        self.order_summary_table = QTableWidget()


        self.order_summary_table.setColumnCount(41)


        self.order_summary_table.setHorizontalHeaderLabels([


            "璁㈠崟鏃ユ湡", "ORDER NO.", "瀹㈡埛浜у搧缂栧彿", "OE鍙?, "瀹㈡埛闇€姹傚娉?,


            "浜у搧鍚嶇О", "鍥剧墖", "瀹㈡埛鍨嬪彿", "OE鍙?1", "鏁伴噺",


            "鎶ヤ环(USD/RMB)", "鍚堣閲戦", "瀹㈡埛鏈€鏂板洖澶?, "瀹㈡埛棰勪粯娆?, "寰呮敹灏炬",


            "棰勪及缇庨噾鎶ヤ环", "棰勪及姣涘埄鐜?, "閲囪喘浠锋牸", "杩愯垂", "鏉傝垂",


            "鎬婚噾棰?, "宸ュ巶绠€绉?, "搴楅摵閾炬帴", "浜よ揣鏃ユ湡", "鏄惁宸叉敹璐?,


            "宸ュ巶璁㈤噾", "宸ュ巶灏炬", "鍏ュ簱鎿嶄綔", "鍏ュ簱鏁伴噺", "鍖呰鏂瑰紡",


            "閲囪喘閫夐」/鍚嶇О", "浜у搧缁嗚妭", "宸ュ巶缂栧彿", "绾哥灏哄", "鎵撳寘瑙勬牸",


            "绠辨暟", "棰勪及浣撶Н", "鏁寸姣涢噸", "鎬婚噸閲?, "鍝佺墝", "寮€绁ㄦ儏鍐?


        ])


        self.order_summary_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)


        self.order_summary_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)


        self.order_summary_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)


        self.order_summary_table.setAlternatingRowColors(True)


        


        # 鍔犺浇涓彁绀?        loading_label = QLabel("馃搳 姝ｅ湪鍔犺浇璁㈠崟鏁版嵁...")


        loading_label.setAlignment(Qt.AlignmentFlag.AlignCenter)


        loading_label.setStyleSheet("color: #6b7280; font-size: 14px; padding: 20px;")


        self.order_summary_table.setIndexWidget(QModelIndex(), loading_label)


        


        layout.addWidget(self.order_summary_table)


        


        # 鐘舵€佹爮鏄剧ず缁熻淇℃伅


        status_layout = QHBoxLayout()


        self.order_summary_status = QLabel("鍑嗗灏辩华")


        self.order_summary_status.setStyleSheet("color: #6b7280;")


        status_layout.addWidget(self.order_summary_status)


        status_layout.addStretch()


        layout.addLayout(status_layout)


        


        widget.setLayout(layout)


        


        # 寤惰繜鍔犺浇鏁版嵁


        QTimer.singleShot(100, self.load_order_summary)


        


        return widget


    


    def load_order_summary(self):


        """鍔犺浇璁㈠崟鎬昏〃鏁版嵁"""


        self._show_loading_tip("姝ｅ湪鍔犺浇璁㈠崟鎬昏〃...")


        


        def fetch():


            try:


                # 鑾峰彇鎵€鏈夌浉鍏虫暟鎹?                pi_list = self.api_client.get_pi_orders() or []


                purchase_list = self.api_client.get_purchase_orders() or []


                shipment_list = self.api_client.get_shipments() or []


                customer_payment_list = self.api_client.get_customer_payments() or []


                supplier_payment_list = self.api_client.get_supplier_payments() or []


                inventory_list = self.api_client.get_inventories() or []


                products = self.api_client.get_products() or []


                customers = self.api_client.get_customers() or []


                suppliers = self.api_client.get_suppliers() or []


                


                return {


                    'pi_list': pi_list,


                    'purchase_list': purchase_list,


                    'shipment_list': shipment_list,


                    'customer_payment_list': customer_payment_list,


                    'supplier_payment_list': supplier_payment_list,


                    'inventory_list': inventory_list,


                    'products': products,


                    'customers': customers,


                    'suppliers': suppliers


                }


            except Exception as e:


                print(f"鍔犺浇璁㈠崟鎬昏〃鏁版嵁澶辫触: {e}")


                return None


        


        from PySide6.QtCore import QThread


        


        class LoaderThread(QThread):


            def __init__(self, parent=None):


                super().__init__(parent)


                self.data = None


            


            def run(self):


                self.data = fetch()


        


        thread = LoaderThread(self)


        thread.finished.connect(self._on_order_summary_loaded)


        thread.start()


    


    def _on_order_summary_loaded(self, thread):


        """璁㈠崟鎬昏〃鏁版嵁鍔犺浇瀹屾垚"""


        # 鑾峰彇绾跨▼涓殑鏁版嵁


        class LoaderThread(QThread):


            pass


        


        # 閲嶆柊鎵ц鑾峰彇鏁版嵁


        try:


            pi_list = self.api_client.get_pi_orders() or []


            purchase_list = self.api_client.get_purchase_orders() or []


            shipment_list = self.api_client.get_shipments() or []


            customer_payment_list = self.api_client.get_customer_payments() or []


            supplier_payment_list = self.api_client.get_supplier_payments() or []


            


            # 鏋勫缓璁㈠崟鎬昏〃鏁版嵁


            orders = []


            for pi in pi_list:


                order = self._build_order_summary_row(pi, purchase_list, shipment_list, customer_payment_list, supplier_payment_list)


                orders.append(order)


            


            self._populate_order_summary_table(orders)


            self._hide_loading_tip()


            self.order_summary_status.setText(f"鍏?{len(orders)} 鏉¤鍗曡褰?)


        except Exception as e:


            print(f"澶勭悊璁㈠崟鎬昏〃鏁版嵁澶辫触: {e}")


            self._hide_loading_tip()


    


    def _build_order_summary_row(self, pi, purchase_list, shipment_list, customer_payment_list, supplier_payment_list):


        """鏋勫缓璁㈠崟鎬昏〃鍗曡鏁版嵁"""


        # 鍩虹淇℃伅


        order_date = pi.get('created_at', '')[:10] if pi.get('created_at') else ''


        order_no = pi.get('pi_no', '')


        


        # 鑾峰彇鍏宠仈鐨勪骇鍝佷俊鎭?        product_id = pi.get('product_id')


        product = None


        if product_id:


            products = self.api_client.get_products() or []


            product = next((p for p in products if p.get('id') == product_id), None)


        


        customer_id = pi.get('customer_id')


        customer = None


        if customer_id:


            customers = self.api_client.get_customers() or []


            customer = next((c for c in customers if c.get('id') == customer_id), None)


        


        # 鏌ユ壘閲囪喘淇℃伅


        purchase = next((p for p in purchase_list if p.get('pi_id') == pi.get('id')), None)


        supplier_id = purchase.get('supplier_id') if purchase else None


        supplier = None


        if supplier_id:


            suppliers = self.api_client.get_suppliers() or []


            supplier = next((s for s in suppliers if s.get('id') == supplier_id), None)


        


        # 鏌ユ壘鍑鸿揣淇℃伅


        shipment = next((s for s in shipment_list if s.get('pi_id') == pi.get('id')), None)


        


        # 鏌ユ壘瀹㈡埛浠樻


        customer_payment = next((cp for cp in customer_payment_list if cp.get('pi_id') == pi.get('id')), None)


        


        # 鏌ユ壘渚涘簲鍟嗕粯娆?        supplier_payment = next((sp for sp in supplier_payment_list if purchase and sp.get('purchase_id') == purchase.get('id')), None)


        


        # 鐘舵€佸垽鏂?        is_completed = pi.get('status') == 4


        


        return {


            'order_date': order_date,


            'order_no': order_no,


            'customer_product_code': pi.get('customer_code', ''),


            'oe_number': product.get('oe_number', '') if product else '',


            'customer_requirement': pi.get('remark', ''),


            'product_name': product.get('name', '') if product else '',


            'image': product.get('image_url', '') if product else '',


            'customer_model': pi.get('customer_code', ''),


            'quantity': pi.get('quantity', 0),


            'unit_price': pi.get('unit_price', 0),


            'total_amount': pi.get('total_amount', 0),


            'currency': pi.get('currency', 'USD'),


            'customer_prepayment': customer_payment.get('amount', 0) if customer_payment else 0,


            'remaining_payment': (pi.get('total_amount', 0) or 0) - (customer_payment.get('amount', 0) if customer_payment else 0),


            'purchase_price': purchase.get('total_amount', 0) if purchase else 0,


            'shipping_fee': purchase.get('shipping_fee', 0) if purchase else 0,


            'misc_fee': purchase.get('misc_fee', 0) if purchase else 0,


            'supplier_name': supplier.get('name', '') if supplier else '',


            'supplier_link': '',


            'delivery_date': purchase.get('expected_date', '')[:10] if purchase and purchase.get('expected_date') else '',


            'is_received': '鏄? if shipment else '鍚?,


            'supplier_deposit': supplier_payment.get('deposit_amount', 0) if supplier_payment else 0,


            'supplier_balance': supplier_payment.get('balance_amount', 0) if supplier_payment else 0,


            'storage_status': '宸插叆搴? if shipment else '鏈叆搴?,


            'storage_quantity': shipment.get('quantity', 0) if shipment else 0,


            'packaging': '',


            'factory_code': supplier.get('code', '') if supplier else '',


            'carton_size': '',


            'packing_spec': '',


            'carton_count': 0,


            'estimated_volume': '',


            'carton_gross_weight': '',


            'total_weight': '',


            'brand': product.get('brand', '') if product else '',


            'invoice_status': '',


            'status': '宸插畬鎴? if is_completed else '杩涜涓?


        }


    


    def _populate_order_summary_table(self, orders):


        """濉厖璁㈠崟鎬昏〃鏁版嵁"""


        self.order_summary_table.setRowCount(0)


        self.order_summary_table.setSortingEnabled(False)


        


        for order in orders:


            row = self.order_summary_table.rowCount()


            self.order_summary_table.insertRow(row)


            


            # 0: 璁㈠崟鏃ユ湡


            self.order_summary_table.setItem(row, 0, QTableWidgetItem(order['order_date']))


            


            # 1: ORDER NO.


            self.order_summary_table.setItem(row, 1, QTableWidgetItem(order['order_no']))


            


            # 2: 瀹㈡埛浜у搧缂栧彿


            self.order_summary_table.setItem(row, 2, QTableWidgetItem(order['customer_product_code']))


            


            # 3: OE鍙?            self.order_summary_table.setItem(row, 3, QTableWidgetItem(order['oe_number']))


            


            # 4: 瀹㈡埛闇€姹傚娉?            self.order_summary_table.setItem(row, 4, QTableWidgetItem(order['customer_requirement']))


            


            # 5: 浜у搧鍚嶇О


            self.order_summary_table.setItem(row, 5, QTableWidgetItem(order['product_name']))


            


            # 6: 鍥剧墖


            if order['image']:


                from PySide6.QtWidgets import QLabel


                img_label = QLabel("馃摲")


                img_label.setAlignment(Qt.AlignmentFlag.AlignCenter)


                self.order_summary_table.setCellWidget(row, 6, img_label)


            else:


                self.order_summary_table.setItem(row, 6, QTableWidgetItem(""))


            


            # 7: 瀹㈡埛鍨嬪彿


            self.order_summary_table.setItem(row, 7, QTableWidgetItem(order['customer_model']))


            


            # 8: OE鍙?1 (鍚孫E鍙?


            self.order_summary_table.setItem(row, 8, QTableWidgetItem(order['oe_number']))


            


            # 9: 鏁伴噺


            qty_item = QTableWidgetItem(str(order['quantity']))


            qty_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)


            self.order_summary_table.setItem(row, 9, qty_item)


            


            # 10: 鎶ヤ环(USD/RMB)


            price_text = f"{order['unit_price']} {order['currency']}"


            price_item = QTableWidgetItem(price_text)


            price_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)


            self.order_summary_table.setItem(row, 10, price_item)


            


            # 11: 鍚堣閲戦


            total_item = QTableWidgetItem(f"{order['total_amount']:.2f}")


            total_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)


            self.order_summary_table.setItem(row, 11, total_item)


            


            # 12: 瀹㈡埛鏈€鏂板洖澶?            self.order_summary_table.setItem(row, 12, QTableWidgetItem(""))


            


            # 13: 瀹㈡埛棰勪粯娆?            prepay_item = QTableWidgetItem(f"{order['customer_prepayment']:.2f}")


            prepay_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)


            self.order_summary_table.setItem(row, 13, prepay_item)


            


            # 14: 寰呮敹灏炬


            remaining_item = QTableWidgetItem(f"{order['remaining_payment']:.2f}")


            remaining_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)


            self.order_summary_table.setItem(row, 14, remaining_item)


            


            # 15-19: 閲囪喘鐩稿叧


            self.order_summary_table.setItem(row, 15, QTableWidgetItem(""))  # 棰勪及缇庨噾鎶ヤ环


            self.order_summary_table.setItem(row, 16, QTableWidgetItem(""))  # 棰勪及姣涘埄鐜?            purchase_price_item = QTableWidgetItem(f"{order['purchase_price']:.2f}")


            purchase_price_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)


            self.order_summary_table.setItem(row, 17, purchase_price_item)


            shipping_item = QTableWidgetItem(f"{order['shipping_fee']:.2f}")


            shipping_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)


            self.order_summary_table.setItem(row, 18, shipping_item)


            misc_item = QTableWidgetItem(f"{order['misc_fee']:.2f}")


            misc_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)


            self.order_summary_table.setItem(row, 19, misc_item)


            


            # 20-24: 宸ュ巶鐩稿叧


            total_amount = order['purchase_price'] + order['shipping_fee'] + order['misc_fee']


            total_item = QTableWidgetItem(f"{total_amount:.2f}")


            total_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)


            self.order_summary_table.setItem(row, 20, total_item)


            self.order_summary_table.setItem(row, 21, QTableWidgetItem(order['supplier_name']))


            self.order_summary_table.setItem(row, 22, QTableWidgetItem(order['supplier_link']))


            self.order_summary_table.setItem(row, 23, QTableWidgetItem(order['delivery_date']))


            self.order_summary_table.setItem(row, 24, QTableWidgetItem(order['is_received']))


            


            # 25-29: 浠樻鍜屽叆搴?            deposit_item = QTableWidgetItem(f"{order['supplier_deposit']:.2f}")


            deposit_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)


            self.order_summary_table.setItem(row, 25, deposit_item)


            balance_item = QTableWidgetItem(f"{order['supplier_balance']:.2f}")


            balance_item.setTextAlignment(Qt.AlignmentFlag.AlignRight)


            self.order_summary_table.setItem(row, 26, balance_item)


            self.order_summary_table.setItem(row, 27, QTableWidgetItem(order['storage_status']))


            storage_qty_item = QTableWidgetItem(str(order['storage_quantity']))


            storage_qty_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)


            self.order_summary_table.setItem(row, 28, storage_qty_item)


            self.order_summary_table.setItem(row, 29, QTableWidgetItem(order['packaging']))


            


            # 30-40: 鍏朵粬淇℃伅


            self.order_summary_table.setItem(row, 30, QTableWidgetItem(order['factory_code']))


            self.order_summary_table.setItem(row, 31, QTableWidgetItem(""))


            self.order_summary_table.setItem(row, 32, QTableWidgetItem(order['factory_code']))


            self.order_summary_table.setItem(row, 33, QTableWidgetItem(order['carton_size']))


            self.order_summary_table.setItem(row, 34, QTableWidgetItem(order['packing_spec']))


            carton_item = QTableWidgetItem(str(order['carton_count']))


            carton_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)


            self.order_summary_table.setItem(row, 35, carton_item)


            self.order_summary_table.setItem(row, 36, QTableWidgetItem(order['estimated_volume']))


            self.order_summary_table.setItem(row, 37, QTableWidgetItem(order['carton_gross_weight']))


            self.order_summary_table.setItem(row, 38, QTableWidgetItem(order['total_weight']))


            self.order_summary_table.setItem(row, 39, QTableWidgetItem(order['brand']))


            self.order_summary_table.setItem(row, 40, QTableWidgetItem(order['invoice_status']))


        


        self.order_summary_table.setSortingEnabled(True)


    


    def filter_order_summary(self):


        """绛涢€夎鍗曟€昏〃"""


        search_text = self.order_summary_search.text().lower()


        status_filter = self.order_status_filter.currentText()


        


        for row in range(self.order_summary_table.rowCount()):


            # 闅愯棌鎵€鏈夎锛屽啀鏍规嵁鏉′欢鏄剧ず


            match = True


            


            if search_text:


                match = False


                for col in range(min(10, self.order_summary_table.columnCount())):


                    item = self.order_summary_table.item(row, col)


                    if item and search_text in item.text().lower():


                        match = True


                        break


            


            # 鐘舵€佺瓫閫夛紙閫氳繃绗?1鍒楃殑棰滆壊鎴栫姸鎬佸垽鏂級


            # 鏍规嵁璁㈠崟鐘舵€佺瓫閫?            


            self.order_summary_table.setRowHidden(row, not match)


    


    def export_order_summary(self):


        """瀵煎嚭璁㈠崟鎬昏〃涓篍xcel"""


        try:


            from openpyxl import Workbook


            from openpyxl.styles import Font, Alignment, PatternFill


            from PySide6.QtWidgets import QFileDialog


            


            file_path, _ = QFileDialog.getSaveFileName(


                self,


                "淇濆瓨璁㈠崟鎬昏〃",


                f"璁㈠崟鎬昏〃_{QDate.currentDate().toString('yyyyMMdd')}.xlsx",


                "Excel Files (*.xlsx)"


            )


            


            if not file_path:


                return


            


            wb = Workbook()


            ws = wb.active


            ws.title = "璁㈠崟鎬昏〃"


            


            # 鍐欏叆琛ㄥご


            headers = [


                "璁㈠崟鏃ユ湡", "ORDER NO.", "瀹㈡埛浜у搧缂栧彿", "OE鍙?, "瀹㈡埛闇€姹傚娉?,


                "浜у搧鍚嶇О", "鍥剧墖", "瀹㈡埛鍨嬪彿", "OE鍙?1", "鏁伴噺",


                "鎶ヤ环(USD/RMB)", "鍚堣閲戦", "瀹㈡埛鏈€鏂板洖澶?, "瀹㈡埛棰勪粯娆?, "寰呮敹灏炬",


                "棰勪及缇庨噾鎶ヤ环", "棰勪及姣涘埄鐜?, "閲囪喘浠锋牸", "杩愯垂", "鏉傝垂",


                "鎬婚噾棰?, "宸ュ巶绠€绉?, "搴楅摵閾炬帴", "浜よ揣鏃ユ湡", "鏄惁宸叉敹璐?,


                "宸ュ巶璁㈤噾", "宸ュ巶灏炬", "鍏ュ簱鎿嶄綔", "鍏ュ簱鏁伴噺", "鍖呰鏂瑰紡",


                "閲囪喘閫夐」/鍚嶇О", "浜у搧缁嗚妭", "宸ュ巶缂栧彿", "绾哥灏哄", "鎵撳寘瑙勬牸",


                "绠辨暟", "棰勪及浣撶Н", "鏁寸姣涢噸", "鎬婚噸閲?, "鍝佺墝", "寮€绁ㄦ儏鍐?


            ]


            


            # 琛ㄥご鏍峰紡


            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")


            header_font = Font(bold=True, color="FFFFFF", size=10)


            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


            


            for col, header in enumerate(headers, 1):


                cell = ws.cell(row=1, column=col)


                cell.value = header


                cell.fill = header_fill


                cell.font = header_font


                cell.alignment = header_alignment


            


            # 鍐欏叆鏁版嵁


            for row in range(self.order_summary_table.rowCount()):


                if self.order_summary_table.isRowHidden(row):


                    continue


                for col in range(self.order_summary_table.columnCount()):


                    item = self.order_summary_table.item(row, col)


                    value = item.text() if item else ""


                    ws.cell(row=row + 2, column=col + 1, value=value)


            


            # 璁剧疆鍒楀


            for col in range(1, len(headers) + 1):


                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 12


            


            wb.save(file_path)


            QMessageBox.information(self, "瀵煎嚭鎴愬姛", f"璁㈠崟鎬昏〃宸插鍑哄埌:\n{file_path}")


            


        except Exception as e:


            QMessageBox.warning(self, "瀵煎嚭澶辫触", f"瀵煎嚭Excel鏃跺嚭閿?\n{str(e)}")





    def load_image_async(self, label, image_url):


        """寮傛鍔犺浇鍥剧墖 - 浣跨敤鍏ㄥ眬绾跨▼姹犲拰鍥剧墖缂撳瓨"""


        global _image_cache, _image_cache_lock


        


        # 鍏堟鏌ュ唴瀛樼紦瀛?        with _image_cache_lock:


            if image_url in _image_cache:


                pixmap = _image_cache[image_url]


                label.setPixmap(pixmap)


                return


        


        # 鏄剧ず鍔犺浇涓崰浣嶇


        label.setText("...")


        


        def fetch_image():


            try:


                # 浣跨敤缂撳瓨閬垮厤閲嶅涓嬭浇


                cache_key = f"img_{hash(image_url)}"


                cached_data = cache_manager.get(cache_key, max_age=3600)  # 鍥剧墖缂撳瓨1灏忔椂


                if cached_data:


                    return cached_data


                


                image_data = urllib.request.urlopen(image_url, timeout=3).read()


                cache_manager.set(cache_key, image_data)


                return image_data


            except Exception as e:


                print(f"鍥剧墖鍔犺浇澶辫触: {e}")


                return None


        


        def on_done(future):


            image_data = future.result()


            if image_data:


                try:


                    image = QImage.fromData(image_data)


                    pixmap = QPixmap.fromImage(image).scaled(56, 56, Qt.KeepAspectRatio, Qt.SmoothTransformation)


                    


                    # 瀛樺叆鍐呭瓨缂撳瓨


                    with _image_cache_lock:


                        if len(_image_cache) >= _MAX_IMAGE_CACHE_SIZE:


                            # 绠€鍗昄RU锛氬垹闄ゆ渶鏃╃殑


                            oldest_key = next(iter(_image_cache))


                            del _image_cache[oldest_key]


                        _image_cache[image_url] = pixmap


                    


                    label.setPixmap(pixmap)


                    return


                except Exception as e:


                    print(f"鍥剧墖澶勭悊澶辫触: {e}")


            label.setText("鏆傛棤鍥剧墖")


        


        # 浣跨敤鍏ㄥ眬绾跨▼姹犳墽琛?        future = _global_thread_pool.submit(fetch_image)


        future.add_done_callback(on_done)





    def load_products(self, use_cache: bool = True):


        try:


            # 灏濊瘯浠庣紦瀛樺姞杞戒骇鍝?            products = None


            if use_cache:


                products = cache_manager.get(CACHE_KEYS['PRODUCTS'], max_age=300)


            


            if products is None:


                print("DEBUG - 浠嶢PI鍔犺浇浜у搧...")


                products = self.api_client.get_products()


                # 淇濆瓨鍒扮紦瀛?                cache_manager.set(CACHE_KEYS['PRODUCTS'], products)


            else:


                print("DEBUG - 浠庣紦瀛樺姞杞戒骇鍝?)


            


            # 灏濊瘯浠庣紦瀛樺姞杞藉簱瀛樻眹鎬?            inventory_summary = cache_manager.get(CACHE_KEYS['INVENTORY_SUMMARY'], max_age=60)


            if inventory_summary is None:


                try:


                    print("DEBUG - 浠嶢PI鍔犺浇搴撳瓨...")


                    inventory_summary = self.api_client.get_all_inventory_summary()


                    cache_manager.set(CACHE_KEYS['INVENTORY_SUMMARY'], inventory_summary)


                except Exception as e:


                    print(f"鑾峰彇搴撳瓨澶辫触: {e}")


                    inventory_summary = {}


            else:


                print("DEBUG - 浠庣紦瀛樺姞杞藉簱瀛?)


            


            if products is None:


                products = []


            self.products_table.setRowCount(len(products))


            for row, p in enumerate(products):


                checkbox = QCheckBox()


                checkbox.setCheckState(Qt.CheckState.Unchecked)


                checkbox.setStyleSheet("margin-left: 50%;")


                self.products_table.setCellWidget(row, 0, checkbox)





                # 浜у搧缂栧彿


                self.products_table.setItem(row, 1, QTableWidgetItem(p.get('product_code', '')))





                # 鏄剧ず鍥剧墖缂╃暐鍥?                image_label = QLabel()


                image_label.setFixedSize(60, 60)


                image_label.setStyleSheet("border: 1px solid #e5e7eb;")


                image_label.setAlignment(Qt.AlignCenter)


                


                image_url = p.get('default_image_url')


                if image_url:


                    # 寮傛鍔犺浇鍥剧墖锛岄伩鍏嶉樆濉濽I


                    self.load_image_async(image_label, image_url)


                else:


                    image_label.setText("鏆傛棤鍥剧墖")


                


                image_label.mousePressEvent = lambda event, p=p: self.view_product_images(p.get('id'))


                image_label.setCursor(Qt.PointingHandCursor)


                self.products_table.setCellWidget(row, 2, image_label)





                # OE鍙?                self.products_table.setItem(row, 3, QTableWidgetItem(p.get('oe_number', '')))


                


                # 閲囪喘娓犻亾


                self.products_table.setItem(row, 4, QTableWidgetItem(p.get('purchase_channel', '') or '-'))


                


                # 搴撳瓨閲?                product_id = p.get('id')


                stock_qty = inventory_summary.get(product_id, 0)


                stock_item = QTableWidgetItem(str(int(stock_qty)) if stock_qty else '0')


                if stock_qty > 0:


                    stock_item.setForeground(QBrush(QColor("#10b981")))


                else:


                    stock_item.setForeground(QBrush(QColor("#ef4444")))


                self.products_table.setItem(row, 5, stock_item)


                


                # 鐘舵€?                status = p.get('status', 1)


                status_text = "鍚敤" if status == 1 else "绂佺敤"


                status_color = "#10b981" if status == 1 else "#ef4444"


                status_item = QTableWidgetItem(status_text)


                status_item.setForeground(QBrush(QColor(status_color)))


                self.products_table.setItem(row, 6, status_item)





                # 鎿嶄綔鏍?                def create_edit_callback(product):


                    return lambda: self.edit_product(product)


                


                action_widget = QWidget()


                action_layout = QHBoxLayout()


                action_layout.setContentsMargins(0, 0, 0, 0)


                


                edit_btn = QPushButton("缂栬緫")


                edit_btn.setFixedWidth(50)


                edit_btn.clicked.connect(create_edit_callback(p))


                action_layout.addWidget(edit_btn)


                


                action_widget.setLayout(action_layout)


                self.products_table.setCellWidget(row, 7, action_widget)


                


            self.select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"鍔犺浇浜у搧鏁版嵁澶辫触: {str(e)}")





    def view_product_images(self, product):


        """鏌ョ湅浜у搧鍥剧墖"""


        product_id = product.get('id')


        product_code = product.get('product_code', '')


        image_url = product.get('default_image_url')


        


        if not image_url:


            QMessageBox.information(self, "鍥剧墖鏌ョ湅", f"浜у搧 {product_code} 鏆傛棤鍥剧墖")


            return


        


        try:


            # 鍒涘缓鍥剧墖鏌ョ湅瀵硅瘽妗?            dialog = QDialog(self)


            dialog.setWindowTitle(f"浜у搧鍥剧墖 - {product_code}")


            dialog.setMinimumSize(600, 600)


            


            layout = QVBoxLayout()


            


            # 鍥剧墖鏍囩


            image_label = QLabel()


            image_label.setAlignment(Qt.AlignCenter)


            


            # 鍔犺浇鍥剧墖


            image_data = urllib.request.urlopen(image_url).read()


            image = QImage.fromData(image_data)


            pixmap = QPixmap.fromImage(image).scaled(580, 580, Qt.KeepAspectRatio, Qt.SmoothTransformation)


            image_label.setPixmap(pixmap)


            


            layout.addWidget(image_label)


            dialog.setLayout(layout)


            dialog.exec()


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"鍔犺浇鍥剧墖澶辫触: {str(e)}")


    


    def confirm_product_import(self, product):


        """纭浜у搧瀵煎叆"""


        print(f"DEBUG - confirm_product_import called for product: {product.get('product_code')}")





        product_id = product.get('id')


        product_code = product.get('product_code', '')





        # 妫€鏌ヤ骇鍝佹槸鍚﹀凡缁忓鍏?        if product.get('is_imported') == 1:


            QMessageBox.warning(self, "鎻愮ず", f"浜у搧 {product_code} 宸茬粡纭瀵煎叆锛屾棤闇€閲嶅鎿嶄綔")


            return





        reply = QMessageBox.question(


            self, "纭瀵煎叆",


            f"纭畾瑕佺‘璁や骇鍝?{product_code} 宸插鍏ュ悧锛熺‘璁ゅ悗灏嗘棤娉曚慨鏀癸紙鏅€氱敤鎴凤級",


            QMessageBox.Ok | QMessageBox.Cancel


        )





        if reply == QMessageBox.Ok:


            try:


                print(f"DEBUG - Calling API to confirm import for product_id: {product_id}")


                result = self.api_client.confirm_product_import(product_id)


                print(f"DEBUG - API response: {result}")


                QMessageBox.information(self, "鎴愬姛", "浜у搧瀵煎叆宸茬‘璁?)


                self.load_products()


            except Exception as e:


                print(f"DEBUG - Confirm import failed: {str(e)}")


                QMessageBox.warning(self, "閿欒", f"纭瀵煎叆澶辫触: {str(e)}")


    


    def cancel_product_import(self, product):


        """鍙栨秷浜у搧瀵煎叆纭锛堜粎绠＄悊鍛橈級"""


        print(f"DEBUG - cancel_product_import called for product: {product.get('product_code')}")


        


        product_id = product.get('id')


        product_code = product.get('product_code', '')


        


        reply = QMessageBox.question(


            self, "鍙栨秷瀵煎叆", 


            f"纭畾瑕佸彇娑堜骇鍝?{product_code} 鐨勫鍏ョ‘璁ゅ悧锛?,


            QMessageBox.Ok | QMessageBox.Cancel


        )


        


        if reply == QMessageBox.Ok:


            try:


                print(f"DEBUG - Calling API to cancel import for product_id: {product_id}")


                result = self.api_client.cancel_product_import(product_id)


                print(f"DEBUG - API response: {result}")


                QMessageBox.information(self, "鎴愬姛", "浜у搧瀵煎叆纭宸插彇娑?)


                self.load_products()


            except Exception as e:


                print(f"DEBUG - Cancel import failed: {str(e)}")


                QMessageBox.warning(self, "閿欒", f"鍙栨秷瀵煎叆澶辫触: {str(e)}")





    def batch_confirm_import_products(self):


        """鎵归噺纭浜у搧瀵煎叆"""


        print("DEBUG - batch_confirm_import_products called")


        


        selected_products = []


        already_imported = []


        


        # 鑾峰彇閫変腑鐨勪骇鍝?        for row in range(self.products_table.rowCount()):


            checkbox = self.products_table.cellWidget(row, 0)


            if checkbox and checkbox.isChecked():


                # 鑾峰彇浜у搧鏁版嵁


                product_id = int(self.products_table.item(row, 2).text())


                product_code = self.products_table.item(row, 9)  # 浜у搧缂栧彿鍦ㄧ10鍒楋紙绱㈠紩9锛?                if not product_code:


                    product_code = self.products_table.item(row, 10).text()  # 鎴栬€呬粠闅愯棌鍒楄幏鍙?                else:


                    product_code = product_code.text()


                


                # 妫€鏌ユ槸鍚﹀凡瀵煎叆锛堜粠鏁版嵁婧愯幏鍙栵級


                # 杩欓噷闇€瑕侀噸鏂拌幏鍙栦骇鍝佹暟鎹潵妫€鏌s_imported鐘舵€?                try:


                    product_detail = self.api_client.get_product_detail(product_id)


                    if product_detail.get('is_imported') == 1:


                        already_imported.append(product_code)


                    else:


                        selected_products.append({'id': product_id, 'code': product_code})


                except Exception as e:


                    print(f"DEBUG - Failed to get product detail for {product_id}: {str(e)}")


                    selected_products.append({'id': product_id, 'code': product_code})


        


        if not selected_products and not already_imported:


            QMessageBox.warning(self, "鎻愮ず", "璇峰厛閫夋嫨瑕佺‘璁ゅ鍏ョ殑浜у搧")


            return


        


        # 濡傛灉鏈夊凡瀵煎叆鐨勪骇鍝侊紝鏄剧ず鎻愮ず


        message = ""


        if already_imported:


            message = f"浠ヤ笅浜у搧宸茬‘璁ゅ鍏ワ紝灏嗚烦杩囷細\n{', '.join(already_imported)}\n\n"


        


        if not selected_products:


            QMessageBox.information(self, "鎻愮ず", message + "娌℃湁闇€瑕佺‘璁ゅ鍏ョ殑浜у搧")


            return


        


        reply = QMessageBox.question(


            self, "鎵归噺纭瀵煎叆",


            f"{message}纭畾瑕佺‘璁ら€変腑鐨?{len(selected_products)} 涓骇鍝佸凡瀵煎叆鍚楋紵\n纭鍚庡皢鏃犳硶淇敼锛堟櫘閫氱敤鎴凤級",


            QMessageBox.Ok | QMessageBox.Cancel


        )


        


        if reply == QMessageBox.Ok:


            success_count = 0


            failed_count = 0


            failed_list = []


            


            for product in selected_products:


                try:


                    print(f"DEBUG - Calling API to confirm import for product_id: {product['id']}")


                    result = self.api_client.confirm_product_import(product['id'])


                    print(f"DEBUG - API response: {result}")


                    success_count += 1


                except Exception as e:


                    print(f"DEBUG - Confirm import failed for {product['code']}: {str(e)}")


                    failed_count += 1


                    failed_list.append(product['code'])


            


            # 鏄剧ず缁撴灉


            result_msg = f"鎵归噺纭瀵煎叆瀹屾垚\n鎴愬姛锛歿success_count} 涓猏n澶辫触锛歿failed_count} 涓?


            if failed_list:


                result_msg += f"\n澶辫触浜у搧锛歿', '.join(failed_list)}"


            


            QMessageBox.information(self, "鎵归噺纭瀵煎叆缁撴灉", result_msg)


            self.load_products()





    def search_products(self):


        keyword = self.search_input.text().strip()


        category_id = self.category_filter.currentData()


        category_id = category_id if category_id != 0 else None


        


        try:


            # 鎼滅储鏃舵樉绀轰骇鍝佺紪鍙峰垪


            self.products_table.setColumnHidden(10, False)


            print(f"鎼滅储鍙傛暟 - 鍏抽敭璇? '{keyword}', 绫诲埆ID: {category_id}")


            products = self.api_client.search_products(keyword, category_id)


            print(f"鎼滅储缁撴灉鏁伴噺: {len(products)}")


            self.products_table.setRowCount(len(products))


            for row, p in enumerate(products):


                checkbox = QCheckBox()


                checkbox.setCheckState(Qt.CheckState.Unchecked)


                checkbox.setStyleSheet("margin-left: 50%;")


                self.products_table.setCellWidget(row, 0, checkbox)


                


                # 鏄剧ず鍥剧墖缂╃暐鍥?                image_label = QLabel()


                image_label.setFixedSize(60, 60)


                image_label.setStyleSheet("border: 1px solid #e5e7eb;")


                image_label.setAlignment(Qt.AlignCenter)


                


                image_url = p.get('default_image_url')


                if image_url:


                    try:


                        image_data = urllib.request.urlopen(image_url).read()


                        image = QImage.fromData(image_data)


                        pixmap = QPixmap.fromImage(image).scaled(56, 56, Qt.KeepAspectRatio, Qt.SmoothTransformation)


                        image_label.setPixmap(pixmap)


                    except Exception as e:


                        image_label.setText("鏆傛棤鍥剧墖")


                else:


                    image_label.setText("鏆傛棤鍥剧墖")


                


                image_label.mousePressEvent = lambda event, p=p: self.view_product_images(p.get('id'))


                image_label.setCursor(Qt.PointingHandCursor)


                self.products_table.setCellWidget(row, 1, image_label)





                self.products_table.setItem(row, 2, QTableWidgetItem(str(p.get('id', ''))))


                self.products_table.setItem(row, 3, QTableWidgetItem(p.get('factory_code', '')))


                self.products_table.setItem(row, 4, QTableWidgetItem(p.get('oe_number', '')))


                self.products_table.setItem(row, 5, QTableWidgetItem(p.get('brand', '')))


                self.products_table.setItem(row, 6, QTableWidgetItem(str(p.get('units_per_carton', ''))))


                


                unit_price = p.get('fob_price_excl') or p.get('exw_price_excl') or ''


                self.products_table.setItem(row, 7, QTableWidgetItem(str(unit_price)))


                


                status = p.get('status', 1)


                status_text = "鍚敤" if status == 1 else "绂佺敤"


                status_color = "#10b981" if status == 1 else "#ef4444"


                status_item = QTableWidgetItem(status_text)


                status_item.setForeground(QBrush(QColor(status_color)))


                self.products_table.setItem(row, 8, status_item)





                edit_btn = QPushButton("缂栬緫")


                edit_btn.setFixedWidth(50)


                edit_btn.clicked.connect(lambda _, p=p: self.edit_product(p))


                self.products_table.setCellWidget(row, 9, edit_btn)


                


                # 浜у搧缂栧彿鏀惧湪鏈€鍚庝竴鍒楋紙榛樿闅愯棌锛?                self.products_table.setItem(row, 10, QTableWidgetItem(p.get('product_code', '')))


                


            self.select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)


        except Exception as e:


            print(f"鎼滅储閿欒: {str(e)}")


            QMessageBox.warning(self, "閿欒", f"鎼滅储浜у搧澶辫触: {str(e)}")





    def reset_search(self):


        self.search_input.clear()


        self.category_filter.setCurrentIndex(0)


        self.load_products()





    def load_product_categories(self):


        """浠庢暟鎹簱鍔犺浇浜у搧绫诲埆"""


        try:


            categories = self.api_client.get("/product-categories")


            # 娓呯┖鐜版湁閫夐」锛堜繚鐣?鍏ㄩ儴鍒嗙被"锛?            self.category_filter.clear()


            self.category_filter.addItem("鍏ㄩ儴鍒嗙被", 0)


            for cat in categories:


                self.category_filter.addItem(cat.get('name', ''), cat.get('id', 0))


        except Exception as e:


            print(f"鍔犺浇浜у搧绫诲埆澶辫触: {str(e)}")





    def toggle_select_all_products(self, state):


        for row in range(self.products_table.rowCount()):


            checkbox = self.products_table.cellWidget(row, 0)


            if checkbox:


                checkbox.setCheckState(Qt.CheckState(state))





    def get_selected_product_ids(self):


        selected_ids = []


        for row in range(self.products_table.rowCount()):


            checkbox = self.products_table.cellWidget(row, 0)


            if checkbox and checkbox.isChecked():


                id_item = self.products_table.item(row, 2)


                if id_item:


                    selected_ids.append(int(id_item.text()))


        return selected_ids





    def batch_toggle_product_status(self):


        selected_ids = self.get_selected_product_ids()


        if not selected_ids:


            QMessageBox.warning(self, "璀﹀憡", "璇峰厛閫夋嫨瑕佹搷浣滅殑浜у搧")


            return





        reply = QMessageBox.question(


            self, "纭鎿嶄綔", 


            f"纭畾瑕佸垏鎹㈤€変腑鐨?{len(selected_ids)} 涓骇鍝佺殑鐘舵€佸悧锛?,


            QMessageBox.Ok | QMessageBox.Cancel


        )





        if reply == QMessageBox.Ok:


            try:


                for product_id in selected_ids:


                    self.api_client.update_product_status(product_id, None)


                QMessageBox.information(self, "鎴愬姛", f"宸叉垚鍔熷垏鎹?{len(selected_ids)} 涓骇鍝佺殑鐘舵€?)


                self.load_products()


            except Exception as e:


                QMessageBox.warning(self, "閿欒", f"鎵归噺鎿嶄綔澶辫触: {str(e)}")





    def batch_delete_products(self):


        selected_ids = self.get_selected_product_ids()


        if not selected_ids:


            QMessageBox.warning(self, "璀﹀憡", "璇峰厛閫夋嫨瑕佸垹闄ょ殑浜у搧")


            return





        reply = QMessageBox.question(


            self, "纭鍒犻櫎", 


            f"纭畾瑕佸垹闄ら€変腑鐨?{len(selected_ids)} 涓骇鍝佸悧锛熸鎿嶄綔涓嶅彲鎭㈠锛?,


            QMessageBox.Ok | QMessageBox.Cancel


        )





        if reply == QMessageBox.Ok:


            try:


                for product_id in selected_ids:


                    self.api_client.delete_product(product_id)


                QMessageBox.information(self, "鎴愬姛", f"宸叉垚鍔熷垹闄?{len(selected_ids)} 涓骇鍝?)


                self.load_products()


            except Exception as e:


                QMessageBox.warning(self, "閿欒", f"鎵归噺鍒犻櫎澶辫触: {str(e)}")





    def view_product_images(self, product_id):


        try:


            images = self.api_client.get_product_images(product_id)


            


            if not images:


                QMessageBox.information(self, "鎻愮ず", "璇ヤ骇鍝佹殏鏃犲浘鐗?)


                return





            dialog = QDialog(self)


            dialog.setWindowTitle("浜у搧鍥剧墖")


            dialog.setMinimumSize(800, 600)


            


            layout = QVBoxLayout()


            


            scroll_area = QScrollArea()


            scroll_area.setWidgetResizable(True)


            scroll_content = QWidget()


            scroll_layout = QVBoxLayout(scroll_content)


            


            for i, image_data in enumerate(images, 1):


                image_url = image_data.get('image_url', '')


                if image_url:


                    try:


                        response = requests.get(image_url)


                        response.raise_for_status()


                        


                        pixmap = QPixmap()


                        pixmap.loadFromData(response.content)


                        


                        if not pixmap.isNull():


                            label = QLabel(f"鍥剧墖 {i}")


                            label.setFont(get_font(12, QFont.Weight.Bold))


                            scroll_layout.addWidget(label)


                            


                            image_label = QLabel()


                            image_label.setPixmap(pixmap.scaled(700, 500, Qt.AspectRatioMode.KeepAspectRatio, Qt.TransformationMode.SmoothTransformation))


                            image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)


                            scroll_layout.addWidget(image_label)


                        else:


                            scroll_layout.addWidget(QLabel(f"鍥剧墖 {i} 鍔犺浇澶辫触"))


                    except Exception as e:


                        scroll_layout.addWidget(QLabel(f"鍥剧墖 {i} 鍔犺浇澶辫触: {str(e)}"))


            


            scroll_area.setWidget(scroll_content)


            layout.addWidget(scroll_area)


            


            close_btn = QPushButton("鍏抽棴")


            close_btn.clicked.connect(dialog.close)


            layout.addWidget(close_btn)


            


            dialog.setLayout(layout)


            dialog.exec()


            


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"鑾峰彇浜у搧鍥剧墖澶辫触: {str(e)}")





    def load_customers(self):


        try:


            customers = self.api_client.get_customers()


            if customers is None:


                customers = []


            self.customers_table.setRowCount(len(customers))


            


            countries = sorted(set([c.get('country', '') for c in customers if c.get('country')]))


            current_country = self.customer_country_filter.currentText()


            self.customer_country_filter.clear()


            self.customer_country_filter.addItem("鍏ㄩ儴鍥藉", 0)


            for country in countries:


                self.customer_country_filter.addItem(country, country)


            index = self.customer_country_filter.findText(current_country)


            if index >= 0:


                self.customer_country_filter.setCurrentIndex(index)


            


            # 鍏堟樉绀哄熀鏈俊鎭紙涓嶇瓑寰呰仈绯讳汉鍜屽湴鍧€锛?            for row, c in enumerate(customers):


                checkbox = QCheckBox()


                checkbox.setStyleSheet("margin-left: 50%;")


                self.customers_table.setCellWidget(row, 0, checkbox)


                


                self.customers_table.setItem(row, 1, QTableWidgetItem(str(c.get('id', ''))))


                self.customers_table.setItem(row, 2, QTableWidgetItem(c.get('customer_code', '')))


                self.customers_table.setItem(row, 3, QTableWidgetItem(c.get('customer_name', '')))


                self.customers_table.setItem(row, 4, QTableWidgetItem(c.get('country', '')))


                self.customers_table.setItem(row, 5, QTableWidgetItem(""))  # 榛樿鑱旂郴浜猴紙绋嶅悗濉厖锛?                self.customers_table.setItem(row, 6, QTableWidgetItem(""))  # 榛樿鍦板潃锛堢◢鍚庡～鍏咃級


                self.customers_table.setItem(row, 7, QTableWidgetItem(c.get('payment_terms', '')))


                


                status = c.get('status', 1)


                status_text = "鍚敤" if status == 1 else "绂佺敤"


                status_color = "#10b981" if status == 1 else "#ef4444"


                status_item = QTableWidgetItem(status_text)


                status_item.setForeground(QBrush(QColor(status_color)))


                self.customers_table.setItem(row, 8, status_item)





                action_bar = ActionBarFactory.create_customer_action_bar(


                    edit_callback=lambda _, c=c: self.edit_customer(c),


                    toggle_callback=lambda _, c=c: self.toggle_customer_status(c),


                    status=status


                )


                self.customers_table.setCellWidget(row, 9, action_bar)


            


            self.customer_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)


            


            # 寮傛鍔犺浇鑱旂郴浜哄拰鍦板潃锛堜笉闃诲UI锛?            QTimer.singleShot(0, lambda: self._load_customer_extra_info(customers))


            


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"鍔犺浇瀹㈡埛鏁版嵁澶辫触: {str(e)}")


    


    def _load_customer_extra_info(self, customers=None):


        """寮傛鍔犺浇瀹㈡埛鐨勮仈绯讳汉鍜屽湴鍧€淇℃伅"""


        try:


            # 濡傛灉娌℃湁浼犲叆瀹㈡埛鍒楄〃锛屽垯浠嶢PI鑾峰彇


            if customers is None:


                customers = self.api_client.get_customers()


            if not customers:


                return


            


            customer_contacts = {}


            customer_addresses = {}


            


            for c in customers:


                try:


                    contacts = self.api_client.get_customer_contacts(c['id'])


                    if contacts:


                        primary_contact = None


                        for contact in contacts:


                            if contact.get('is_primary', 0) == 1:


                                primary_contact = contact


                                break


                        if not primary_contact and contacts:


                            primary_contact = contacts[0]


                        if primary_contact:


                            customer_contacts[c['id']] = primary_contact.get('name', '')


                except Exception:


                    pass


                


                try:


                    addresses = self.api_client.get_customer_addresses(c['id'])


                    if addresses:


                        default_addr = None


                        for addr in addresses:


                            if addr.get('is_default', 0) == 1:


                                default_addr = addr


                                break


                        if not default_addr and addresses:


                            default_addr = addresses[0]


                        if default_addr:


                            addr_text = f"{default_addr.get('country', '')}-{default_addr.get('port', '')}"


                            customer_addresses[c['id']] = addr_text


                except Exception:


                    pass


            


            # 鏇存柊琛ㄦ牸涓殑鑱旂郴浜哄拰鍦板潃鍒?            for row in range(self.customers_table.rowCount()):


                id_item = self.customers_table.item(row, 1)


                if id_item:


                    cid = int(id_item.text())


                    contact = customer_contacts.get(cid, '')


                    address = customer_addresses.get(cid, '')


                    if contact:


                        self.customers_table.setItem(row, 5, QTableWidgetItem(contact))


                    if address:


                        self.customers_table.setItem(row, 6, QTableWidgetItem(address))


        except Exception as e:


            print(f"鍔犺浇瀹㈡埛棰濆淇℃伅澶辫触: {e}")





    def search_customers(self):


        keyword = self.customer_search_input.text().strip()


        country = self.customer_country_filter.currentData()


        country = country if country != 0 else None


        


        try:


            customers = self.api_client.search_customers(keyword, country)


            self.customers_table.setRowCount(len(customers))


            


            # 鍏堟樉绀哄熀鏈俊鎭?            for row, c in enumerate(customers):


                checkbox = QCheckBox()


                checkbox.setStyleSheet("margin-left: 50%;")


                self.customers_table.setCellWidget(row, 0, checkbox)


                


                self.customers_table.setItem(row, 1, QTableWidgetItem(str(c.get('id', ''))))


                self.customers_table.setItem(row, 2, QTableWidgetItem(c.get('customer_code', '')))


                self.customers_table.setItem(row, 3, QTableWidgetItem(c.get('customer_name', '')))


                self.customers_table.setItem(row, 4, QTableWidgetItem(c.get('country', '')))


                self.customers_table.setItem(row, 5, QTableWidgetItem(""))  # 榛樿鑱旂郴浜猴紙绋嶅悗濉厖锛?                self.customers_table.setItem(row, 6, QTableWidgetItem(""))  # 榛樿鍦板潃锛堢◢鍚庡～鍏咃級


                self.customers_table.setItem(row, 7, QTableWidgetItem(c.get('payment_terms', '')))


                


                status = c.get('status', 1)


                status_text = "鍚敤" if status == 1 else "绂佺敤"


                status_color = "#10b981" if status == 1 else "#ef4444"


                status_item = QTableWidgetItem(status_text)


                status_item.setForeground(QBrush(QColor(status_color)))


                self.customers_table.setItem(row, 8, status_item)





                action_bar = ActionBarFactory.create_customer_action_bar(


                    edit_callback=lambda _, c=c: self.edit_customer(c),


                    toggle_callback=lambda _, c=c: self.toggle_customer_status(c),


                    status=status


                )


                self.customers_table.setCellWidget(row, 9, action_bar)


            


            self.customer_select_all_checkbox.setCheckState(Qt.CheckState.Unchecked)


            


            # 寮傛鍔犺浇鑱旂郴浜哄拰鍦板潃


            QTimer.singleShot(0, lambda: self._load_customer_extra_info(customers))


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"鎼滅储瀹㈡埛澶辫触: {str(e)}")





    def toggle_customer_status(self, customer):


        try:


            result = self.api_client.toggle_customer_status(customer['id'])


            if result:


                self.load_customers()


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"鎿嶄綔澶辫触: {str(e)}")





    def on_customer_double_click(self, index):


        row = index.row()


        customer_id = self.customers_table.item(row, 1).text()


        try:


            customer = self.api_client.get_customer_detail(int(customer_id))


            self.view_customer_detail(customer)


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"鍔犺浇瀹㈡埛璇︽儏澶辫触: {str(e)}")





    def on_pi_double_click(self, index):


        """鍙屽嚮PI琛屾煡鐪嬭鎯?""


        row = index.row()


        pi_id = self.pi_table.item(row, 1).text()


        try:


            pi_detail = self.api_client.get_pi_detail(int(pi_id))


            self.view_pi_detail(pi_detail)


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"鍔犺浇PI璇︽儏澶辫触: {str(e)}")





    def view_pi_detail(self, pi):


        """鏌ョ湅PI璇︽儏锛堝彧璇伙級"""


        try:


            pi_id = pi.get('id')


            pi_detail = self.api_client.get_pi_detail(pi_id)


            dialog = PIDialog(self.api_client, self.dept_id, pi_detail, readonly=True)


            dialog.exec()


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"鏌ョ湅PI璇︽儏澶辫触: {str(e)}")





    def view_customer_detail(self, customer):


        dialog = CustomerDetailDialog(self.api_client, customer)


        dialog.exec()





    def setup_table_context_menu(self, table, headers):


        def show_context_menu(point):


            menu = QMenu()


            row = table.rowAt(point.y())


            if row < 0:


                return


            copy_cn = menu.addAction("澶嶅埗涓枃淇℃伅")


            copy_en = menu.addAction("Copy English Info")


            menu.addSeparator()


            copy_row = menu.addAction("澶嶅埗鏁磋鏁版嵁")


            action = menu.exec(table.mapToGlobal(point))


            if action == copy_cn or action == copy_en:


                texts = []


                for col in range(table.columnCount()):


                    item = table.item(row, col)


                    if item:


                        texts.append(str(item.text()))


                    else:


                        widget = table.cellWidget(row, col)


                        if widget:


                            if isinstance(widget, QPushButton):


                                texts.append(widget.text())


                            else:


                                texts.append("")


                        else:


                            texts.append("")


                text = " | ".join(texts) if action == copy_cn else " | ".join(texts)


                from PySide6.QtGui import QGuiApplication


                clipboard = QGuiApplication.clipboard()


                clipboard.setText(text)


            elif action == copy_row:


                row_data = []


                for col in range(table.columnCount()):


                    item = table.item(row, col)


                    if item:


                        row_data.append(str(item.text()))


                    else:


                        widget = table.cellWidget(row, col)


                        if widget and isinstance(widget, QPushButton):


                            row_data.append(widget.text())


                        else:


                            row_data.append("")


                from PySide6.QtGui import QGuiApplication


                clipboard = QGuiApplication.clipboard()


                clipboard.setText("\t".join(row_data))


        table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)


        table.customContextMenuRequested.connect(show_context_menu)





    def toggle_select_all_customers(self, state):


        for row in range(self.customers_table.rowCount()):


            checkbox = self.customers_table.cellWidget(row, 0)


            if checkbox:


                checkbox.setCheckState(Qt.CheckState(state))





    def load_suppliers(self):


        try:


            suppliers = self.api_client.get_suppliers()


            if suppliers is None:


                suppliers = []


            self.suppliers_table.setRowCount(len(suppliers))


            for row, s in enumerate(suppliers):


                checkbox = QCheckBox()


                checkbox.setStyleSheet("margin-left: 15px;")


                self.suppliers_table.setCellWidget(row, 0, checkbox)





                self.suppliers_table.setItem(row, 1, QTableWidgetItem(str(s.get('id', ''))))


                self.suppliers_table.setItem(row, 2, QTableWidgetItem(s.get('supplier_code', '')))


                self.suppliers_table.setItem(row, 3, QTableWidgetItem(s.get('supplier_name', '')))


                self.suppliers_table.setItem(row, 4, QTableWidgetItem(s.get('region', '')))


                self.suppliers_table.setItem(row, 5, QTableWidgetItem(s.get('contact_person', '')))


                self.suppliers_table.setItem(row, 6, QTableWidgetItem(s.get('phone', '')))





                edit_btn = QPushButton("缂栬緫")


                edit_btn.setFixedWidth(60)


                edit_btn.clicked.connect(lambda _, s=s: self.edit_supplier(s))


                self.suppliers_table.setCellWidget(row, 7, edit_btn)


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"鍔犺浇渚涘簲鍟嗘暟鎹け璐? {str(e)}")





    def load_pi_orders(self):


        try:


            pi_orders = self.api_client.get_pi_orders()


            if pi_orders is None:


                pi_orders = []


            self.pi_table.setRowCount(len(pi_orders))


            status_map = {1: "鑽夌", 2: "宸茬‘璁?, 3: "宸插彂璐?, 4: "宸插畬鎴?}


            for row, pi in enumerate(pi_orders):


                status = pi.get('status', 1)


                is_completed = status == 4  # 宸插畬鎴愮姸鎬佷笉鍙搷浣?                


                # 閫夋嫨妗嗭紙宸插畬鎴怭I涓嶅彲閫夛級


                if is_completed:


                    checkbox = QTableWidgetItem("鉁?)


                    checkbox.setFlags(checkbox.flags() & ~Qt.ItemFlag.ItemIsSelectable)


                    checkbox.setForeground(QColor("#9ca3af"))  # 鐏拌壊


                    self.pi_table.setItem(row, 0, checkbox)


                else:


                    checkbox = QTableWidgetItem()


                    checkbox.setFlags(checkbox.flags() | Qt.ItemFlag.ItemIsUserCheckable)


                    checkbox.setCheckState(Qt.CheckState.Unchecked)


                    checkbox.setData(Qt.ItemDataRole.UserRole, pi.get('id'))


                    self.pi_table.setItem(row, 0, checkbox)


                


                self.pi_table.setItem(row, 1, QTableWidgetItem(str(pi.get('id', ''))))


                self.pi_table.setItem(row, 2, QTableWidgetItem(pi.get('pi_no', '')))


                self.pi_table.setItem(row, 3, QTableWidgetItem(f"{pi.get('total_amount', 0):,.2f}"))


                self.pi_table.setItem(row, 4, QTableWidgetItem(pi.get('currency', 'USD')))


                status_text = status_map.get(status, "鏈煡")


                status_item = QTableWidgetItem(status_text)


                if is_completed:


                    status_item.setForeground(QColor("#6b7280"))  # 鐏拌壊


                self.pi_table.setItem(row, 5, status_item)


                self.pi_table.setItem(row, 6, QTableWidgetItem(str(pi.get('created_at', ''))[:19] if pi.get('created_at') else ''))





                # 鎿嶄綔鍒?7)銆佸畬鎴愬垪(8)銆佸鍑哄垪(9)


                if is_completed:


                    # 宸插畬鎴怭I锛氭搷浣滃垪鏄剧ず"-"锛屽畬鎴愬垪鏄剧ず"鉁?锛屽鍑哄垪鏄剧ず瀵煎嚭鎸夐挳


                    self.pi_table.setItem(row, 7, QTableWidgetItem("-"))


                    self.pi_table.item(row, 7).setForeground(QColor("#9ca3af"))


                    self.pi_table.setItem(row, 8, QTableWidgetItem("鉁?))


                    self.pi_table.item(row, 8).setForeground(QColor("#10b981"))


                    export_btn = QPushButton("瀵煎嚭")


                    export_btn.setFixedWidth(50)


                    export_btn.setStyleSheet("background-color: #10b981; color: white; border: none; border-radius: 4px; padding: 2px 4px;")


                    export_btn.clicked.connect(lambda _, p=pi: self.export_pi(p))


                    self.pi_table.setCellWidget(row, 9, export_btn)


                else:


                    # 鏈畬鎴怭I锛氭搷浣滃垪鏄剧ず缂栬緫鎸夐挳锛屽畬鎴愬垪鏄剧ず瀹屾垚鎸夐挳锛屽鍑哄垪鏄剧ず瀵煎嚭鎸夐挳


                    edit_btn = QPushButton("缂栬緫")


                    edit_btn.setFixedWidth(50)


                    edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 4px;")


                    edit_btn.clicked.connect(lambda _, p=pi: self.edit_pi(p))


                    self.pi_table.setCellWidget(row, 7, edit_btn)


                    


                    complete_btn = QPushButton("瀹屾垚")


                    complete_btn.setFixedWidth(40)


                    complete_btn.setStyleSheet("background-color: #8b5cf6; color: white; border: none; border-radius: 4px; padding: 2px 4px;")


                    complete_btn.clicked.connect(lambda _, p=pi: self.complete_pi(p))


                    self.pi_table.setCellWidget(row, 8, complete_btn)


                    


                    export_btn = QPushButton("瀵煎嚭")


                    export_btn.setFixedWidth(40)


                    export_btn.setStyleSheet("background-color: #10b981; color: white; border: none; border-radius: 4px; padding: 2px 4px;")


                    export_btn.clicked.connect(lambda _, p=pi: self.export_pi(p))


                    self.pi_table.setCellWidget(row, 9, export_btn)


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"鍔犺浇PI璁㈠崟鏁版嵁澶辫触: {str(e)}")





    def load_purchase_orders(self):


        try:


            purchase_orders = self.api_client.get_purchase_orders()


            if purchase_orders is None:


                purchase_orders = []


            self.purchase_table.setRowCount(len(purchase_orders))


            status_map = {1: "鑽夌", 2: "宸茬‘璁?, 3: "宸插叆搴?}


            for row, po in enumerate(purchase_orders):


                self.purchase_table.setItem(row, 0, QTableWidgetItem(str(po.get('id', ''))))


                self.purchase_table.setItem(row, 1, QTableWidgetItem(po.get('po_no', '')))


                self.purchase_table.setItem(row, 2, QTableWidgetItem(po.get('pi_no', '')))


                self.purchase_table.setItem(row, 3, QTableWidgetItem(po.get('supplier_name', '')))


                self.purchase_table.setItem(row, 4, QTableWidgetItem(str(po.get('total_amount', ''))))


                self.purchase_table.setItem(row, 5, QTableWidgetItem(status_map.get(po.get('status', 1), "鏈煡")))





                edit_btn = QPushButton("缂栬緫")


                edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 8px;")


                edit_btn.clicked.connect(lambda _, p=po: self.edit_purchase(p))


                self.purchase_table.setCellWidget(row, 6, edit_btn)


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"鍔犺浇閲囪喘璁㈠崟鏁版嵁澶辫触: {str(e)}")





    def load_shipments(self):


        try:


            shipments = self.api_client.get_shipments()


            if shipments is None:


                shipments = []


            self.shipment_table.setRowCount(len(shipments))


            status_map = {1: "寰呭嚭璐?, 2: "鍑鸿揣涓?, 3: "宸插嚭璐?, 4: "宸插埌杈?}


            for row, s in enumerate(shipments):


                self.shipment_table.setItem(row, 0, QTableWidgetItem(str(s.get('id', ''))))


                self.shipment_table.setItem(row, 1, QTableWidgetItem(s.get('pi_no', '')))


                self.shipment_table.setItem(row, 2, QTableWidgetItem(str(s.get('shipment_date', ''))[:10] if s.get('shipment_date') else ''))


                self.shipment_table.setItem(row, 3, QTableWidgetItem(s.get('container_no', '')))


                self.shipment_table.setItem(row, 4, QTableWidgetItem(s.get('bl_no', '')))


                self.shipment_table.setItem(row, 5, QTableWidgetItem(status_map.get(s.get('status', 1), "鏈煡")))





                edit_btn = QPushButton("缂栬緫")


                edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 8px;")


                edit_btn.clicked.connect(lambda _, s=s: self.edit_shipment(s))


                self.shipment_table.setCellWidget(row, 6, edit_btn)


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"鍔犺浇鍑鸿揣鏁版嵁澶辫触: {str(e)}")





    def load_customer_payments(self):


        try:


            payments = self.api_client.get_customer_payments()


            if payments is None:


                payments = []


            self.customer_payment_table.setRowCount(len(payments))


            for row, p in enumerate(payments):


                self.customer_payment_table.setItem(row, 0, QTableWidgetItem(str(p.get('id', ''))))


                self.customer_payment_table.setItem(row, 1, QTableWidgetItem(p.get('pi_no', '')))


                self.customer_payment_table.setItem(row, 2, QTableWidgetItem(str(p.get('payment_date', ''))[:10] if p.get('payment_date') else ''))


                self.customer_payment_table.setItem(row, 3, QTableWidgetItem(str(p.get('actual_amount', ''))))


                self.customer_payment_table.setItem(row, 4, QTableWidgetItem(p.get('payment_method', '')))





                edit_btn = QPushButton("缂栬緫")


                edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 8px;")


                edit_btn.clicked.connect(lambda _, p=p: self.edit_customer_payment(p))


                self.customer_payment_table.setCellWidget(row, 5, edit_btn)


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"鍔犺浇瀹㈡埛浠樻鏁版嵁澶辫触: {str(e)}")





    def load_supplier_payments(self):


        try:


            payments = self.api_client.get_supplier_payments()


            if payments is None:


                payments = []


            self.supplier_payment_table.setRowCount(len(payments))


            for row, p in enumerate(payments):


                self.supplier_payment_table.setItem(row, 0, QTableWidgetItem(str(p.get('id', ''))))


                self.supplier_payment_table.setItem(row, 1, QTableWidgetItem(p.get('supplier_name', '')))


                self.supplier_payment_table.setItem(row, 2, QTableWidgetItem(f"PO-{p.get('po_id', '')}"))


                self.supplier_payment_table.setItem(row, 3, QTableWidgetItem(p.get('payment_stage', '')))


                self.supplier_payment_table.setItem(row, 4, QTableWidgetItem(str(p.get('payment_date', ''))[:10] if p.get('payment_date') else ''))


                self.supplier_payment_table.setItem(row, 5, QTableWidgetItem(str(p.get('actual_amount', ''))))





                edit_btn = QPushButton("缂栬緫")


                edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 8px;")


                edit_btn.clicked.connect(lambda _, p=p: self.edit_supplier_payment(p))


                self.supplier_payment_table.setCellWidget(row, 6, edit_btn)


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"鍔犺浇渚涘簲鍟嗕粯娆炬暟鎹け璐? {str(e)}")





    # ========== 寮傛鍔犺浇鏂规硶 ==========





    def _show_loading_indicator(self, table, message="鍔犺浇涓?.."):


        """鏄剧ず鍔犺浇鐘舵€佹寚绀哄櫒"""


        table.setRowCount(1)


        table.setItem(0, 0, QTableWidgetItem(""))


        table.setItem(0, 1, QTableWidgetItem(message))


        for col in range(2, table.columnCount()):


            table.setItem(0, col, QTableWidgetItem(""))


        table.setEnabled(False)





    def _hide_loading_indicator(self, table):


        """闅愯棌鍔犺浇鐘舵€佹寚绀哄櫒"""


        table.setEnabled(True)





    def load_pi_orders_async(self, force_refresh=False):


        """寮傛鍔犺浇PI璁㈠崟锛堝甫缂撳瓨鍜屽姞杞芥寚绀哄櫒锛?""


        print("DEBUG - load_pi_orders_async started")


        # 鏄剧ず鍔犺浇鐘舵€侊紙鍦ㄤ富绾跨▼锛?        self._show_loading_indicator(self.pi_table, "姝ｅ湪鍔犺浇PI璁㈠崟...")


        


        # 浣跨敤QThread鏉ョ‘淇濅俊鍙峰湪涓荤嚎绋嬪鐞?        from PySide6.QtCore import QThread


        


        class PiLoaderThread(QThread):


            def __init__(self, api_client, force_refresh, parent=None):


                super().__init__(parent)


                self.api_client = api_client


                self.force_refresh = force_refresh


                self.result_data = []


                self.error_msg = None


            


            def run(self):


                try:


                    if self.force_refresh:


                        cache_manager.delete(CACHE_KEYS['PI_LIST'])


                    data = self.api_client.get_pi_orders()


                    if data:


                        cache_manager.set(CACHE_KEYS['PI_LIST'], data)


                    self.result_data = data if data else []


                except Exception as e:


                    print(f"鍔犺浇PI璁㈠崟澶辫触: {e}")


                    self.error_msg = str(e)


                    self.result_data = []


        


        self._pi_loader_thread = PiLoaderThread(self.api_client, force_refresh, self)


        self._pi_loader_thread.finished.connect(


            lambda: self._on_pi_load_finished(self._pi_loader_thread.result_data)


        )


        self._pi_loader_thread.start()


        print("DEBUG - thread started")


    


    def _on_pi_load_finished(self, data):


        """PI鍔犺浇瀹屾垚鍥炶皟"""


        print(f"DEBUG - _on_pi_load_finished with {len(data)} items")


        self._update_pi_table(data)





    def load_pi_orders(self):


        """鍚屾鍔犺浇PI璁㈠崟锛堥娆¤繘鍏ワ級"""


        try:


            # 鍏堝皾璇曚粠缂撳瓨鍔犺浇


            cached = cache_manager.get(CACHE_KEYS['PI_LIST'], max_age=120)


            if cached is not None:


                self._update_pi_table(cached)


                return


            # 缂撳瓨涓嶅瓨鍦紝浠嶢PI鍔犺浇


            pi_orders = self.api_client.get_pi_orders()


            if pi_orders:


                cache_manager.set(CACHE_KEYS['PI_LIST'], pi_orders)


            self._update_pi_table(pi_orders if pi_orders else [])


        except Exception as e:


            print(f"鍔犺浇PI璁㈠崟澶辫触: {e}")


            self._update_pi_table([])





    def _update_pi_table(self, pi_orders):


        try:


            print(f"DEBUG - _update_pi_table called with {len(pi_orders) if pi_orders else 0} orders")


            print(f"DEBUG - pi_orders type: {type(pi_orders)}")


            if pi_orders and len(pi_orders) > 0:


                print(f"DEBUG - first order: {pi_orders[0]}")


            


            # 闅愯棌鍔犺浇鎸囩ず鍣?            self._hide_loading_indicator(self.pi_table)


            


            # 纭繚鏁版嵁鏄垪琛?            if not pi_orders:


                pi_orders = []


            


            self.pi_table.setRowCount(len(pi_orders))


            status_map = {1: "鑽夌", 2: "宸茬‘璁?, 3: "宸插彂璐?, 4: "宸插畬鎴?}


            for row, pi in enumerate(pi_orders):


                status = pi.get('status', 1)


                is_completed = status == 4


                


                # 閫夋嫨妗嗭紙宸插畬鎴怭I涓嶅彲閫夛級


                if is_completed:


                    checkbox = QTableWidgetItem("鉁?)


                    checkbox.setFlags(checkbox.flags() & ~Qt.ItemFlag.ItemIsSelectable)


                    checkbox.setForeground(QColor("#9ca3af"))


                    self.pi_table.setItem(row, 0, checkbox)


                else:


                    checkbox = QTableWidgetItem()


                    checkbox.setFlags(checkbox.flags() | Qt.ItemFlag.ItemIsUserCheckable)


                    checkbox.setCheckState(Qt.CheckState.Unchecked)


                    checkbox.setData(Qt.ItemDataRole.UserRole, pi.get('id'))


                    self.pi_table.setItem(row, 0, checkbox)


                


                self.pi_table.setItem(row, 1, QTableWidgetItem(str(pi.get('id', ''))))


                self.pi_table.setItem(row, 2, QTableWidgetItem(pi.get('pi_no', '')))


                self.pi_table.setItem(row, 3, QTableWidgetItem(f"{pi.get('total_amount', 0):,.2f}"))


                self.pi_table.setItem(row, 4, QTableWidgetItem(pi.get('currency', 'USD')))


                status_text = status_map.get(status, "鏈煡")


                status_item = QTableWidgetItem(status_text)


                if is_completed:


                    status_item.setForeground(QColor("#6b7280"))


                self.pi_table.setItem(row, 5, status_item)


                self.pi_table.setItem(row, 6, QTableWidgetItem(str(pi.get('created_at', ''))[:19] if pi.get('created_at') else ''))


                


                # 鎿嶄綔鍒?7)銆佸畬鎴愬垪(8)銆佸鍑哄垪(9)


                if is_completed:


                    # 宸插畬鎴怭I锛氭搷浣滃垪鏄剧ず"-"锛屽畬鎴愬垪鏄剧ず"鉁?锛屽鍑哄垪鏄剧ず瀵煎嚭鎸夐挳


                    self.pi_table.setItem(row, 7, QTableWidgetItem("-"))


                    self.pi_table.item(row, 7).setForeground(QColor("#9ca3af"))


                    self.pi_table.setItem(row, 8, QTableWidgetItem("鉁?))


                    self.pi_table.item(row, 8).setForeground(QColor("#10b981"))


                    export_btn = QPushButton("瀵煎嚭")


                    export_btn.setFixedWidth(50)


                    export_btn.setStyleSheet("background-color: #10b981; color: white; border: none; border-radius: 4px; padding: 2px 4px;")


                    export_btn.clicked.connect(lambda _, p=pi: self.export_pi(p))


                    self.pi_table.setCellWidget(row, 9, export_btn)


                else:


                    # 鏈畬鎴怭I锛氭搷浣滃垪鏄剧ず缂栬緫鎸夐挳锛屽畬鎴愬垪鏄剧ず瀹屾垚鎸夐挳锛屽鍑哄垪鏄剧ず瀵煎嚭鎸夐挳


                    edit_btn = QPushButton("缂栬緫")


                    edit_btn.setFixedWidth(50)


                    edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 4px;")


                    edit_btn.clicked.connect(lambda _, p=pi: self.edit_pi(p))


                    self.pi_table.setCellWidget(row, 7, edit_btn)


                    


                    complete_btn = QPushButton("瀹屾垚")


                    complete_btn.setFixedWidth(40)


                    complete_btn.setStyleSheet("background-color: #8b5cf6; color: white; border: none; border-radius: 4px; padding: 2px 4px;")


                    complete_btn.clicked.connect(lambda _, p=pi: self.complete_pi(p))


                    self.pi_table.setCellWidget(row, 8, complete_btn)


                    


                    export_btn = QPushButton("瀵煎嚭")


                    export_btn.setFixedWidth(40)


                    export_btn.setStyleSheet("background-color: #10b981; color: white; border: none; border-radius: 4px; padding: 2px 4px;")


                    export_btn.clicked.connect(lambda _, p=pi: self.export_pi(p))


                    self.pi_table.setCellWidget(row, 9, export_btn)


        except Exception as e:


            print(f"鏇存柊PI璁㈠崟琛ㄦ牸澶辫触: {e}")


            import traceback


            traceback.print_exc()





    def load_purchase_orders_async(self):


        """寮傛鍔犺浇閲囪喘璁㈠崟"""


        self._load_async(


            self.api_client.get_purchase_orders,


            self._update_purchase_table,


            "鍔犺浇閲囪喘璁㈠崟澶辫触"


        )





    def _update_purchase_table(self, purchase_orders):


        try:


            self.purchase_table.setRowCount(len(purchase_orders))


            status_map = {1: "鑽夌", 2: "宸茬‘璁?, 3: "宸插叆搴?}


            for row, po in enumerate(purchase_orders):


                self.purchase_table.setItem(row, 0, QTableWidgetItem(str(po.get('id', ''))))


                self.purchase_table.setItem(row, 1, QTableWidgetItem(po.get('po_no', '')))


                self.purchase_table.setItem(row, 2, QTableWidgetItem(po.get('pi_no', '')))


                self.purchase_table.setItem(row, 3, QTableWidgetItem(po.get('supplier_name', '')))


                self.purchase_table.setItem(row, 4, QTableWidgetItem(str(po.get('total_amount', ''))))


                status = po.get('status', 1)


                status_item = QTableWidgetItem(status_map.get(status, "鏈煡"))


                if status == 3:


                    status_item.setForeground(QBrush(QColor("#10b981")))


                elif status == 2:


                    status_item.setForeground(QBrush(QColor("#f59e0b")))


                self.purchase_table.setItem(row, 5, status_item)


                edit_btn = QPushButton("缂栬緫")


                edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 8px;")


                edit_btn.clicked.connect(lambda _, p=po: self.edit_purchase(p))


                self.purchase_table.setCellWidget(row, 6, edit_btn)


                # 纭鎸夐挳


                confirm_btn = QPushButton("纭")


                confirm_btn.setStyleSheet("background-color: #f59e0b; color: white; border: none; border-radius: 4px; padding: 2px 8px;")


                confirm_btn.clicked.connect(lambda _, p=po: self.confirm_purchase_order(p))


                self.purchase_table.setCellWidget(row, 7, confirm_btn)


                # 鍏ュ簱鎸夐挳


                inbound_btn = QPushButton("鍏ュ簱")


                inbound_btn.setStyleSheet("background-color: #10b981; color: white; border: none; border-radius: 4px; padding: 2px 8px;")


                inbound_btn.clicked.connect(lambda _, p=po: self.inbound_purchase_order(p))


                self.purchase_table.setCellWidget(row, 8, inbound_btn)


        except Exception as e:


            print(f"鏇存柊閲囪喘璁㈠崟琛ㄦ牸澶辫触: {e}")





    def load_shipments_async(self):


        """寮傛鍔犺浇鍑鸿揣鏁版嵁"""


        self._load_async(


            self.api_client.get_shipments,


            self._update_shipment_table,


            "鍔犺浇鍑鸿揣鏁版嵁澶辫触"


        )





    def _update_shipment_table(self, shipments):


        """鏇存柊鍑鸿揣涓昏〃锛堟眹鎬讳俊鎭級"""


        try:


            self.shipment_table.setRowCount(len(shipments))


            status_map = {1: "寰呭嚭璐?, 2: "鍑鸿揣涓?, 3: "宸插嚭璐?, 4: "宸插埌杈?}


            payment_status_map = {1: "鏈敹娆?, 2: "閮ㄥ垎鏀舵", 3: "宸叉敹榻?}


            


            for row, s in enumerate(shipments):


                # ID


                self.shipment_table.setItem(row, 0, QTableWidgetItem(str(s.get('id', ''))))


                # PI鍙?                self.shipment_table.setItem(row, 1, QTableWidgetItem(s.get('pi_no', '')))


                # 鎬婚噾棰?                total_amount = s.get('total_amount', 0) or 0


                self.shipment_table.setItem(row, 2, QTableWidgetItem(f"{float(total_amount):,.2f}"))


                # 鎬荤鏁?                total_cartons = s.get('total_cartons', 0) or 0


                self.shipment_table.setItem(row, 3, QTableWidgetItem(str(total_cartons)))


                # 浠樻鐘舵€?                payment_status = s.get('payment_status', 1)


                payment_item = QTableWidgetItem(payment_status_map.get(payment_status, "鏈煡"))


                if payment_status == 3:


                    payment_item.setForeground(QBrush(QColor("#10b981")))


                elif payment_status == 2:


                    payment_item.setForeground(QBrush(QColor("#f59e0b")))


                self.shipment_table.setItem(row, 4, payment_item)


                # 鍑鸿揣鐘舵€?                status = s.get('status', 1)


                status_item = QTableWidgetItem(status_map.get(status, "鏈煡"))


                if status == 3:


                    status_item.setForeground(QBrush(QColor("#10b981")))


                elif status == 2:


                    status_item.setForeground(QBrush(QColor("#3b82f6")))


                self.shipment_table.setItem(row, 5, status_item)


                # 闃舵鏁?                stages_count = s.get('stages_count', 0) or 0


                self.shipment_table.setItem(row, 6, QTableWidgetItem(str(stages_count)))


                # 鎿嶄綔鎸夐挳


                edit_btn = QPushButton("缂栬緫")


                edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 8px;")


                edit_btn.clicked.connect(lambda _, s=s: self.edit_shipment(s))


                self.shipment_table.setCellWidget(row, 7, edit_btn)


                # 纭鍑鸿揣鎸夐挳


                confirm_btn = QPushButton("纭鍑鸿揣")


                confirm_btn.setStyleSheet("background-color: #10b981; color: white; border: none; border-radius: 4px; padding: 2px 8px;")


                confirm_btn.clicked.connect(lambda _, s=s: self.confirm_shipment_order(s))


                self.shipment_table.setCellWidget(row, 8, confirm_btn)


            


            # 娓呯┖闃舵琛?            self.shipment_stage_table.setRowCount(0)


        except Exception as e:


            print(f"鏇存柊鍑鸿揣琛ㄦ牸澶辫触: {e}")





    def show_shipment_stages(self, row, column):


        """鏄剧ず閫変腑鍑鸿揣璁板綍鐨勯樁娈垫槑缁?""


        try:


            shipment_id = int(self.shipment_table.item(row, 0).text())


            # 鑾峰彇鍑鸿揣璇︽儏锛堝寘鍚玸tages锛?            shipment = self.api_client.get_shipment(shipment_id)


            if not shipment:


                return


            stages = shipment.get('stages', [])


            self._update_shipment_stage_table(stages)


        except Exception as e:


            print(f"鏄剧ず鍑鸿揣闃舵澶辫触: {e}")





    def _update_shipment_stage_table(self, stages):


        """鏇存柊鍑鸿揣闃舵浠庤〃"""


        try:


            self.shipment_stage_table.setRowCount(len(stages))


            payment_status_map = {1: "鏈敹娆?, 2: "閮ㄥ垎鏀舵", 3: "宸叉敹榻?}


            


            for row, s in enumerate(stages):


                # 闃舵鍚嶇О


                self.shipment_stage_table.setItem(row, 0, QTableWidgetItem(s.get('stage_name', '')))


                # 鍑鸿揣鏃ユ湡


                shipment_date = s.get('shipment_date', '')


                if shipment_date:


                    shipment_date = str(shipment_date)[:10]


                self.shipment_stage_table.setItem(row, 1, QTableWidgetItem(shipment_date))


                # 鏌滃彿


                self.shipment_stage_table.setItem(row, 2, QTableWidgetItem(s.get('container_no', '')))


                # 鎻愬崟鍙?                self.shipment_stage_table.setItem(row, 3, QTableWidgetItem(s.get('bl_no', '')))


                # 鏁伴噺


                quantity = s.get('quantity', 0) or 0


                self.shipment_stage_table.setItem(row, 4, QTableWidgetItem(f"{float(quantity):,.0f}"))


                # 搴撳瓨


                inv_qty = s.get('inventory_quantity', 0) or 0


                inv_item = QTableWidgetItem(f"{float(inv_qty):,.0f}")


                if inv_qty > 0:


                    inv_item.setForeground(QBrush(QColor("#10b981")))


                self.shipment_stage_table.setItem(row, 5, inv_item)


                # 瀛樻斁浣嶇疆


                self.shipment_stage_table.setItem(row, 6, QTableWidgetItem(s.get('storage_location', '-')))


                # 浠樻鐘舵€?                payment_status = s.get('payment_status', 1)


                pay_item = QTableWidgetItem(payment_status_map.get(payment_status, "鏈煡"))


                if payment_status == 3:


                    pay_item.setForeground(QBrush(QColor("#10b981")))


                elif payment_status == 2:


                    pay_item.setForeground(QBrush(QColor("#f59e0b")))


                self.shipment_stage_table.setItem(row, 7, pay_item)


                # 鎿嶄綔鎸夐挳


                btn_widget = QWidget()


                btn_layout = QHBoxLayout()


                btn_layout.setContentsMargins(0, 0, 0, 0)


                


                ci_btn = QPushButton("CI")


                ci_btn.setFixedWidth(40)


                ci_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 3px;")


                ci_btn.clicked.connect(lambda _, s=s: self.view_ci_document(s))


                btn_layout.addWidget(ci_btn)


                


                pl_btn = QPushButton("PL")


                pl_btn.setFixedWidth(40)


                pl_btn.setStyleSheet("background-color: #10b981; color: white; border: none; border-radius: 3px;")


                pl_btn.clicked.connect(lambda _, s=s: self.view_pl_document(s))


                btn_layout.addWidget(pl_btn)


                


                btn_widget.setLayout(btn_layout)


                self.shipment_stage_table.setCellWidget(row, 8, btn_widget)


        except Exception as e:


            print(f"鏇存柊鍑鸿揣闃舵琛ㄦ牸澶辫触: {e}")





    def view_ci_document(self, stage):


        """鏌ョ湅CI鏂囨。"""


        ci_doc = stage.get('ci_document')


        if ci_doc:


            # TODO: 鎵撳紑CI鏂囨。鏌ョ湅鍣?            QMessageBox.information(self, "CI鏂囨。", f"CI鏂囨。璺緞: {ci_doc}")


        else:


            QMessageBox.information(self, "CI鏂囨。", "璇ラ樁娈垫殏鏃燙I鏂囨。")





    def view_pl_document(self, stage):


        """鏌ョ湅PL鏂囨。"""


        pl_doc = stage.get('pl_document')


        if pl_doc:


            # TODO: 鎵撳紑PL鏂囨。鏌ョ湅鍣?            QMessageBox.information(self, "PL鏂囨。", f"PL鏂囨。璺緞: {pl_doc}")


        else:


            QMessageBox.information(self, "PL鏂囨。", "璇ラ樁娈垫殏鏃燩L鏂囨。")





    def load_customer_payments_async(self):


        """寮傛鍔犺浇瀹㈡埛浠樻"""


        self._load_async(


            self.api_client.get_customer_payments,


            self._update_customer_payment_table,


            "鍔犺浇瀹㈡埛浠樻澶辫触"


        )





    def _update_customer_payment_table(self, payments):


        try:


            self.customer_payment_table.setRowCount(len(payments))


            for row, p in enumerate(payments):


                self.customer_payment_table.setItem(row, 0, QTableWidgetItem(str(p.get('id', ''))))


                self.customer_payment_table.setItem(row, 1, QTableWidgetItem(p.get('pi_no', '')))


                self.customer_payment_table.setItem(row, 2, QTableWidgetItem(str(p.get('payment_date', ''))[:10] if p.get('payment_date') else ''))


                self.customer_payment_table.setItem(row, 3, QTableWidgetItem(str(p.get('actual_amount', ''))))


                self.customer_payment_table.setItem(row, 4, QTableWidgetItem(p.get('payment_method', '')))


                edit_btn = QPushButton("缂栬緫")


                edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 8px;")


                edit_btn.clicked.connect(lambda _, p=p: self.edit_customer_payment(p))


                self.customer_payment_table.setCellWidget(row, 5, edit_btn)


        except Exception as e:


            print(f"鏇存柊瀹㈡埛浠樻琛ㄦ牸澶辫触: {e}")





    def load_supplier_payments_async(self):


        """寮傛鍔犺浇渚涘簲鍟嗕粯娆?""


        self._load_async(


            self.api_client.get_supplier_payments,


            self._update_supplier_payment_table,


            "鍔犺浇渚涘簲鍟嗕粯娆惧け璐?


        )





    def _update_supplier_payment_table(self, payments):


        """鏇存柊渚涘簲鍟嗕粯娆句富琛紙姹囨€讳俊鎭級"""


        try:


            self.supplier_payment_table.setRowCount(len(payments))


            for row, p in enumerate(payments):


                # ID


                self.supplier_payment_table.setItem(row, 0, QTableWidgetItem(str(p.get('id', ''))))


                # 渚涘簲鍟?                self.supplier_payment_table.setItem(row, 1, QTableWidgetItem(p.get('supplier_name', '')))


                # 閲囪喘鍗?                self.supplier_payment_table.setItem(row, 2, QTableWidgetItem(f"PO-{p.get('po_id', '')}"))


                # 鎬婚噾棰?                total = p.get('total_amount', 0) or 0


                self.supplier_payment_table.setItem(row, 3, QTableWidgetItem(f"{float(total):,.2f}"))


                # 宸蹭粯閲戦


                paid = p.get('paid_amount', 0) or 0


                paid_item = QTableWidgetItem(f"{float(paid):,.2f}")


                if paid > 0:


                    paid_item.setForeground(QBrush(QColor("#10b981")))


                self.supplier_payment_table.setItem(row, 4, paid_item)


                # 鏈粯閲戦


                unpaid = p.get('unpaid_amount', 0) or 0


                unpaid_item = QTableWidgetItem(f"{float(unpaid):,.2f}")


                if unpaid > 0:


                    unpaid_item.setForeground(QBrush(QColor("#ef4444")))


                self.supplier_payment_table.setItem(row, 5, unpaid_item)


                # 鐘舵€?                status_map = {1: "寰呬粯娆?, 2: "閮ㄥ垎浠樻", 3: "宸蹭粯娓?}


                status = p.get('status', 1)


                status_text = status_map.get(status, "鏈煡")


                status_item = QTableWidgetItem(status_text)


                if status == 3:


                    status_item.setForeground(QBrush(QColor("#10b981")))


                elif status == 2:


                    status_item.setForeground(QBrush(QColor("#f59e0b")))


                self.supplier_payment_table.setItem(row, 6, status_item)


                # 鎿嶄綔鎸夐挳


                edit_btn = QPushButton("缂栬緫")


                edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 8px;")


                edit_btn.clicked.connect(lambda _, p=p: self.edit_supplier_payment(p))


                self.supplier_payment_table.setCellWidget(row, 7, edit_btn)


            # 娓呯┖闃舵琛?            self.supplier_payment_stage_table.setRowCount(0)


        except Exception as e:


            print(f"鏇存柊渚涘簲鍟嗕粯娆捐〃鏍煎け璐? {e}")





    def show_supplier_payment_stages(self, row, column):


        """鏄剧ず閫変腑浠樻璁板綍鐨勯樁娈垫槑缁?""


        try:


            payment_id = int(self.supplier_payment_table.item(row, 0).text())


            # 鑾峰彇浠樻璇︽儏锛堝寘鍚玸tages锛?            payment = self.api_client.get_supplier_payment(payment_id)


            if not payment:


                return


            stages = payment.get('stages', [])


            self._update_supplier_payment_stage_table(stages)


        except Exception as e:


            print(f"鏄剧ず浠樻闃舵澶辫触: {e}")





    def _update_supplier_payment_stage_table(self, stages):


        """鏇存柊浠樻闃舵浠庤〃"""


        try:


            self.supplier_payment_stage_table.setRowCount(len(stages))


            status_map = {1: "寰呬粯", 2: "閮ㄥ垎浠?, 3: "宸蹭粯娓?}


            for row, s in enumerate(stages):


                # 闃舵鍚嶇О


                self.supplier_payment_stage_table.setItem(row, 0, QTableWidgetItem(s.get('stage_name', '')))


                # 搴斾粯閲戦


                amount = s.get('amount', 0) or 0


                self.supplier_payment_stage_table.setItem(row, 1, QTableWidgetItem(f"{float(amount):,.2f}"))


                # 宸蹭粯閲戦


                paid = s.get('paid_amount', 0) or 0


                paid_item = QTableWidgetItem(f"{float(paid):,.2f}")


                if paid > 0:


                    paid_item.setForeground(QBrush(QColor("#10b981")))


                self.supplier_payment_stage_table.setItem(row, 2, paid_item)


                # 鐘舵€?                status = s.get('status', 1)


                status_item = QTableWidgetItem(status_map.get(status, "鏈煡"))


                if status == 3:


                    status_item.setForeground(QBrush(QColor("#10b981")))


                elif status == 2:


                    status_item.setForeground(QBrush(QColor("#f59e0b")))


                self.supplier_payment_stage_table.setItem(row, 3, status_item)


                # 浠樻鏃ユ湡


                payment_date = s.get('payment_date', '')


                if payment_date:


                    payment_date = str(payment_date)[:10]


                self.supplier_payment_stage_table.setItem(row, 4, QTableWidgetItem(payment_date))


                # 鍑瘉


                has_proof = "鏈? if s.get('payment_proof') else "鏃?


                self.supplier_payment_stage_table.setItem(row, 5, QTableWidgetItem(has_proof))


                # 鎿嶄綔鎸夐挳


                pay_btn = QPushButton("浠樻")


                pay_btn.setStyleSheet("background-color: #10b981; color: white; border: none; border-radius: 4px; padding: 2px 8px;")


                pay_btn.clicked.connect(lambda _, s=s: self.pay_supplier_stage(s))


                self.supplier_payment_stage_table.setCellWidget(row, 6, pay_btn)


        except Exception as e:


            print(f"鏇存柊浠樻闃舵琛ㄦ牸澶辫触: {e}")





    def load_inventories_async(self):


        """寮傛鍔犺浇搴撳瓨鏁版嵁"""


        self._load_async(


            self.api_client.get_inventories,


            self._populate_inventories,


            "鍔犺浇搴撳瓨澶辫触"


        )





    def load_inventories(self):


        """鍔犺浇搴撳瓨鏁版嵁 - 鎸塐E鍙峰垎缁勬樉绀猴紙鍙傝€冧緵搴斿晢鏂规妯″紡锛?""


        try:


            inventories = self.api_client.get_inventories()


            if inventories is None:


                inventories = []


            self._populate_inventories(inventories)


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"鍔犺浇搴撳瓨鏁版嵁澶辫触: {str(e)}")





    def _populate_inventories(self, inventories):


        """濉厖搴撳瓨鏁版嵁鍒拌〃鏍?""


        try:


            # 鎸塐E鍙峰垎缁勭粺璁?            self.inventory_data = {}


            for inv in inventories:


                oe = inv.get('oe_number') or inv.get('product_code', '鏈煡')


                if oe not in self.inventory_data:


                    self.inventory_data[oe] = []


                self.inventory_data[oe].append(inv)


            


            self._load_inventory_product_table()


            self.inventory_detail_table.setRowCount(0)


            self.inventory_detail_label.setText("馃搵 搴撳瓨璇︽儏锛堣鐐瑰嚮涓婃柟浜у搧鏌ョ湅锛?)


        except Exception as e:


            print(f"濉厖搴撳瓨鏁版嵁澶辫触: {e}")





    def load_products_async(self):


        """寮傛鍔犺浇浜у搧鏁版嵁"""


        self._load_async(


            self.api_client.get_products,


            self.load_products_callback,


            "鍔犺浇浜у搧澶辫触"


        )





    def load_products_callback(self, products):


        """浜у搧鏁版嵁鍔犺浇瀹屾垚鍚庢洿鏂癠I"""


        try:


            self.load_products_with_data(products)


        except Exception as e:


            print(f"鏇存柊浜у搧琛ㄦ牸澶辫触: {e}")





    def load_products_with_data(self, products):


        """鐢ㄥ凡鏈夋暟鎹姞杞戒骇鍝佸垪琛紙寮傛鍥炶皟浣跨敤锛?""


        try:


            inventory_summary = {}


            try:


                inventory_summary = self.api_client.get_all_inventory_summary()


            except Exception:


                pass





            if products is None:


                products = []


            self.products_table.setRowCount(len(products))


            for row, p in enumerate(products):


                # 0: 閫夋嫨


                checkbox = QCheckBox()


                checkbox.setStyleSheet("margin-left: 50%;")


                self.products_table.setCellWidget(row, 0, checkbox)


                # 1: 浜у搧缂栧彿


                self.products_table.setItem(row, 1, QTableWidgetItem(p.get('product_code', '')))


                # 2: 鍥剧墖


                image_label = QLabel()


                image_label.setFixedSize(60, 60)


                image_label.setStyleSheet("border: 1px solid #e5e7eb;")


                image_label.setAlignment(Qt.AlignmentFlag.AlignCenter)


                image_url = p.get('default_image_url')


                if image_url:


                    self.load_image_async(image_label, image_url)


                else:


                    image_label.setText("鏆傛棤鍥剧墖")


                image_label.setCursor(Qt.CursorShape.PointingHandCursor)


                self.products_table.setCellWidget(row, 2, image_label)


                # 3: OE鍙?                self.products_table.setItem(row, 3, QTableWidgetItem(p.get('oe_number', '')))


                # 4: 閲囪喘娓犻亾


                self.products_table.setItem(row, 4, QTableWidgetItem(p.get('purchase_channel', '') or '-'))


                # 5: 搴撳瓨閲?                qty = inventory_summary.get(p.get('id'), 0)


                qty_item = QTableWidgetItem(str(int(qty)) if qty else '0')


                if qty > 0:


                    qty_item.setForeground(QBrush(QColor("#10b981")))


                else:


                    qty_item.setForeground(QBrush(QColor("#ef4444")))


                self.products_table.setItem(row, 5, qty_item)


                # 6: 鐘舵€?                status = p.get('status', 1)


                status_text = "鍚敤" if status == 1 else "绂佺敤"


                status_color = "#10b981" if status == 1 else "#ef4444"


                status_item = QTableWidgetItem(status_text)


                status_item.setForeground(QBrush(QColor(status_color)))


                self.products_table.setItem(row, 6, status_item)


                # 7: 鎿嶄綔


                action_widget = QWidget()


                action_layout = QHBoxLayout()


                action_layout.setContentsMargins(0, 0, 0, 0)


                


                edit_btn = QPushButton("缂栬緫")


                edit_btn.setFixedWidth(50)


                edit_btn.clicked.connect(lambda _, p=p: self.edit_product(p))


                action_layout.addWidget(edit_btn)


                


                action_widget.setLayout(action_layout)


                self.products_table.setCellWidget(row, 7, action_widget)


        except Exception as e:


            print(f"鏇存柊浜у搧琛ㄦ牸澶辫触: {e}")





    def _load_inventory_product_table(self, data=None):


        """鍔犺浇浜у搧搴撳瓨姹囨€昏〃"""


        if data is None:


            data = self.inventory_data


        


        # 鑾峰彇鏈€杩戝彉鏇存棩蹇?            product_logs = {}


            try:


                product_logs = self.api_client.get_product_logs() or {}


                print(f"DEBUG - 鑾峰彇鍒?{len(product_logs)} 鏉′骇鍝佹棩蹇?)


            except Exception as e:


                print(f"DEBUG - 鑾峰彇浜у搧鏃ュ織澶辫触: {e}")


            


            self.inventory_product_table.setRowCount(len(data))


        


        for row, (oe_number, records) in enumerate(data.items()):


            # 鑾峰彇浜у搧ID


            product_id = records[0].get('product_id')


            print(f"DEBUG - 浜у搧ID={product_id}, OE={oe_number}")


            


            # OE鍙凤紙缂╃煭鏄剧ず锛?            oe_display = str(oe_number)


            if len(oe_display) > 18:


                oe_display = oe_display[:15] + '...'


            oe_item = QTableWidgetItem(oe_display)


            oe_item.setToolTip(str(oe_number))  # 鎮仠鏄剧ず瀹屾暣OE鍙?            self.inventory_product_table.setItem(row, 0, oe_item)


            


            # 浜у搧缂栧彿锛堝彇绗竴鏉¤褰曠殑product_code锛?            product_code = records[0].get('product_code', '') or ''


            if len(str(product_code)) > 14:


                product_code = str(product_code)[:11] + '...'


            self.inventory_product_table.setItem(row, 1, QTableWidgetItem(str(product_code)))


            


            # 鎬诲簱瀛橀噺


            total_qty = sum(float(r.get('total_quantity', 0) or 0) for r in records)


            qty_item = QTableWidgetItem(str(int(total_qty)))


            # 搴撳瓨涓?鏃舵爣绾?            if total_qty == 0:


                qty_item.setForeground(QBrush(QColor("#ef4444")))


            self.inventory_product_table.setItem(row, 2, qty_item)


            


            # 渚涘簲鍟嗘暟


            suppliers = set(r.get('supplier_name') for r in records if r.get('supplier_name'))


            self.inventory_product_table.setItem(row, 3, QTableWidgetItem(str(len(suppliers))))


            


            # 瀹㈡埛鏁?            customers = set(r.get('customer_name') for r in records if r.get('customer_name'))


            self.inventory_product_table.setItem(row, 4, QTableWidgetItem(str(len(customers))))


            


            # 鐘舵€佸垎甯冿紙鐢ㄩ鑹茬偣琛ㄧず锛?            status_counts = {}


            for r in records:


                st = r.get('stock_type', 1)


                if isinstance(st, str):


                    try: st = int(st)


                    except ValueError: st = 1


                status_counts[st] = status_counts.get(st, 0) + 1


            


            status_parts = []


            if 1 in status_counts:


                status_parts.append(f"馃煛{status_counts[1]}")


            if 2 in status_counts:


                status_parts.append(f"馃數{status_counts[2]}")


            if 3 in status_counts:


                status_parts.append(f"馃煝{status_counts[3]}")


            if 4 in status_counts:


                status_parts.append(f"鈿珄status_counts[4]}")


            self.inventory_product_table.setItem(row, 5, QTableWidgetItem(" ".join(status_parts) if status_parts else "-"))


            


            # 鏈€杩戝叆搴撲緵搴斿晢


            # product_id 鏄暣鏁帮紝浣?product_logs 鐨勯敭鏄瓧绗︿覆


            log_info = product_logs.get(str(product_id), {}) or product_logs.get(product_id, {})


            print(f"DEBUG - 浜у搧{product_id}鐨勬棩蹇? {log_info}")


            supplier_name = log_info.get('supplier_name', '') or ''


            if len(supplier_name) > 10:


                supplier_name = supplier_name[:8] + '..'


            self.inventory_product_table.setItem(row, 6, QTableWidgetItem(supplier_name or '-'))


            


            # 鏈€杩戝嚭搴撳鎴?            customer_name = log_info.get('customer_name', '') or ''


            if len(customer_name) > 10:


                customer_name = customer_name[:8] + '..'


            self.inventory_product_table.setItem(row, 7, QTableWidgetItem(customer_name or '-'))


            


            # 鏈€杩戝彉鏇存椂闂?            last_change = log_info.get('last_change_time', '') or ''


            if last_change and len(str(last_change)) > 16:


                last_change = str(last_change)[:16].replace('T', ' ')


            self.inventory_product_table.setItem(row, 8, QTableWidgetItem(last_change))


            


            # 鎿嶄綔鎸夐挳锛堝睍寮€ + 娣诲姞锛?            op_widget = QWidget()


            op_layout = QHBoxLayout(op_widget)


            op_layout.setContentsMargins(2, 2, 2, 2)


            op_layout.setSpacing(4)


            


            expand_btn = QPushButton("灞曞紑")


            expand_btn.setFixedWidth(50)


            expand_btn.setStyleSheet("""


                QPushButton {


                    background-color: #10b981; color: white; border: none;


                    border-radius: 4px; padding: 4px 8px; font-size: 12px;


                }


                QPushButton:hover { background-color: #059669; }


            """)


            expand_btn.clicked.connect(lambda _, oe=oe_number: self.show_inventory_detail(oe))


            op_layout.addWidget(expand_btn)


            


            add_btn = QPushButton("+娣诲姞")


            add_btn.setFixedWidth(50)


            add_btn.setStyleSheet("""


                QPushButton {


                    background-color: #3b82f6; color: white; border: none;


                    border-radius: 4px; padding: 4px 8px; font-size: 12px;


                }


                QPushButton:hover { background-color: #2563eb; }


            """)


            add_btn.clicked.connect(lambda _, oe=oe_number, recs=records: self.add_inventory_for_oe(oe, recs))


            op_layout.addWidget(add_btn)


            


            self.inventory_product_table.setCellWidget(row, 9, op_widget)


        


        # 鍙屽嚮灞曞紑璇︽儏


        try:


            self.inventory_product_table.cellDoubleClicked.disconnect()


        except TypeError:


            pass


        self.inventory_product_table.cellDoubleClicked.connect(


            lambda r, c: self.show_inventory_detail(list(data.keys())[r])


        )


    


    def show_inventory_detail(self, oe_number):


        """鏄剧ず鎸囧畾OE鍙风殑搴撳瓨璇︽儏"""


        records = self.inventory_data.get(oe_number, [])


        self.inventory_detail_label.setText(f"馃搵 '{oe_number}' 鐨勫簱瀛樿鎯咃紙鍏?{len(records)} 鏉¤褰曪級")


        


        self.inventory_detail_table.setRowCount(len(records))


        


        for row, inv in enumerate(records):


            # ID


            self.inventory_detail_table.setItem(row, 0, QTableWidgetItem(str(inv.get('id', ''))))


            # 渚涘簲鍟嗭紙缂╃煭鏄剧ず锛?            supplier_name = inv.get('supplier_name', '-') or '-'


            if len(str(supplier_name)) > 10:


                supplier_name = str(supplier_name)[:8] + '..'


            supplier_item = QTableWidgetItem(supplier_name)


            supplier_item.setToolTip(inv.get('supplier_name', '-') or '-')


            self.inventory_detail_table.setItem(row, 1, supplier_item)


            # 瀹㈡埛锛堢缉鐭樉绀猴級


            customer_name = inv.get('customer_name', '-') or '-'


            if len(str(customer_name)) > 10:


                customer_name = str(customer_name)[:8] + '..'


            customer_item = QTableWidgetItem(customer_name)


            customer_item.setToolTip(inv.get('customer_name', '-') or '-')


            self.inventory_detail_table.setItem(row, 2, customer_item)


            # 鏁伴噺


            self.inventory_detail_table.setItem(row, 3, QTableWidgetItem(str(int(inv.get('total_quantity', 0) or 0))))


            # 搴撲綅


            self.inventory_detail_table.setItem(row, 4, QTableWidgetItem(inv.get('current_location', '') or ''))


            # 鐘舵€侀鑹?            stock_type = inv.get('stock_type', 1)


            if isinstance(stock_type, str):


                try:


                    stock_type = int(stock_type)


                except (ValueError, TypeError):


                    stock_type = 1


            color_map = {1: "#ffff00", 2: "#0000ff", 3: "#00cc00", 4: "#000000"}


            color_hex = color_map.get(stock_type, "#cccccc")


            color_label = QLabel()


            color_label.setFixedSize(20, 20)


            color_label.setStyleSheet(f"background-color: {color_hex}; border: 1px solid #999; border-radius: 3px;")


            type_text_map = {1: "閲囪喘鍦ㄩ€?, 2: "寰呭叆搴?, 3: "宸插叆搴?, 4: "鍘嗗彶搴撳瓨"}


            color_label.setToolTip(type_text_map.get(stock_type, str(stock_type)))


            self.inventory_detail_table.setCellWidget(row, 5, color_label)


            # 澶囨敞


            remark_item = QTableWidgetItem(inv.get('remark', '') or '')


            self.inventory_detail_table.setItem(row, 6, remark_item)


            # 鍒涘缓鏃堕棿


            created_at = inv.get('created_at', '') or ''


            if created_at and len(str(created_at)) > 16:


                created_at = str(created_at)[:16].replace('T', ' ')


            self.inventory_detail_table.setItem(row, 7, QTableWidgetItem(str(created_at)))


            # 缂栬緫鎸夐挳


            edit_btn = QPushButton("缂栬緫")


            edit_btn.setStyleSheet("background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 2px 8px;")


            edit_btn.clicked.connect(lambda _, inv=inv: self.edit_inventory(inv))


            self.inventory_detail_table.setCellWidget(row, 8, edit_btn)


    


    def add_inventory_for_oe(self, oe_number, existing_records):


        """涓烘寚瀹歄E鍙锋坊鍔犲簱瀛樿褰?""


        # 鑾峰彇浜у搧淇℃伅


        product_id = existing_records[0].get('product_id') if existing_records else None


        if product_id:


            self.add_inventory_with_product(product_id, oe_number)


        else:


            self.add_inventory()


    


    def search_inventory(self):


        """鎼滅储搴撳瓨"""


        keyword = self.inventory_search_input.text().strip().lower()


        if not keyword:


            self._load_inventory_product_table()


            return


        


        # 杩囨护鏁版嵁


        filtered_data = {}


        for oe_number, records in self.inventory_data.items():


            if keyword in str(oe_number).lower():


                filtered_data[oe_number] = records


            else:


                for r in records:


                    if (keyword in str(r.get('supplier_name', '')).lower() or 


                        keyword in str(r.get('customer_name', '')).lower()):


                        filtered_data[oe_number] = records


                        break


        


        self._load_inventory_product_table(filtered_data)


    


    def add_inventory_with_product(self, product_id, oe_number):


        """涓烘寚瀹氫骇鍝佹坊鍔犲簱瀛?""


        dialog = InventoryDialog(self.api_client, self.dept_id, product_id=product_id, oe_number=oe_number)


        if dialog.exec() == QDialog.DialogCode.Accepted:


            self.load_inventories()





    def toggle_product_status(self, product):


        product_id = product.get('id')


        product_code = product.get('product_code', '')


        status = product.get('status', 1)


        status_text = "绂佺敤" if status == 1 else "鍚敤"


        


        reply = QMessageBox.question(


            self, "纭鎿嶄綔", 


            f"纭畾瑕亄status_text}浜у搧 {product_code} 鍚楋紵",


            QMessageBox.Ok | QMessageBox.Cancel


        )


        


        if reply == QMessageBox.Ok:


            try:


                self.api_client.toggle_product_status(product_id)


                QMessageBox.information(self, "鎴愬姛", f"浜у搧宸瞷status_text}")


                self.load_products()


            except Exception as e:


                QMessageBox.warning(self, "閿欒", f"{status_text}浜у搧澶辫触: {str(e)}")





    def delete_product(self, product):


        product_id = product.get('id')


        product_code = product.get('product_code', '')


        


        reply = QMessageBox.question(


            self, "纭鍒犻櫎", 


            f"纭畾瑕佸垹闄や骇鍝?{product_code} 鍚楋紵姝ゆ搷浣滀笉鍙仮澶嶏紒",


            QMessageBox.Ok | QMessageBox.Cancel


        )


        


        if reply == QMessageBox.Ok:


            try:


                self.api_client.delete_product(product_id)


                QMessageBox.information(self, "鎴愬姛", "浜у搧宸插垹闄?)


                self.load_products()


            except Exception as e:


                QMessageBox.warning(self, "閿欒", f"鍒犻櫎浜у搧澶辫触: {str(e)}")





    def add_product(self):


        print("add_product called, dept_id:", self.dept_id)


        try:


            dialog = ProductDialog(self.api_client, self.dept_id)


            print("ProductDialog created successfully")


            if dialog.exec() == QDialog.DialogCode.Accepted:


                self.load_products()


        except Exception as e:


            print("Error in add_product:", str(e))


            QMessageBox.warning(self, "閿欒", f"鎵撳紑鏂板浜у搧瀵硅瘽妗嗗け璐? {str(e)}")





    def on_product_double_click(self, index):


        row = index.row()


        product_id = self.products_table.item(row, 2).text()


        try:


            product = self.api_client.get_product_detail(int(product_id))


            self.edit_product(product)


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"鍔犺浇浜у搧淇℃伅澶辫触: {str(e)}")





    def edit_product(self, product):


        dialog = ProductDialog(self.api_client, self.dept_id, product)


        if dialog.exec() == QDialog.DialogCode.Accepted:


            self.load_products()





    def import_products(self):


        print("DEBUG - import_products called")


        


        file_path, _ = QFileDialog.getOpenFileName(


            self, "閫夋嫨瀵煎叆鏂囦欢", "", "Excel鏂囦欢 (*.xlsx *.xls)"


        )


        


        print(f"DEBUG - Selected file: {file_path}")


        


        if not file_path:


            return


        


        try:


            df = pd.read_excel(file_path)


            print(f"DEBUG - Excel file loaded, rows: {len(df)}")


            


            products_data = []


            for idx, row in df.iterrows():


                product = {


                    "oe_number": str(row.get("OE鍙?, "")),


                    "factory_code": str(row.get("宸ュ巶缂栧彿", "")),


                    "brand": str(row.get("鍝佺墝", "")),


                    "detail_desc": str(row.get("缁嗚妭鎻忚堪", "")),


                    "category_id": int(row.get("绫诲埆ID", 1)),


                    "supplier_id": int(row.get("渚涘簲鍟咺D", 0)) if pd.notna(row.get("渚涘簲鍟咺D")) else None,


                    "exw_price_incl": float(row.get("EXW鍚◣浠?, 0)) if pd.notna(row.get("EXW鍚◣浠?)) else None,


                    "exw_price_excl": float(row.get("EXW涓嶅惈绋庝环", 0)) if pd.notna(row.get("EXW涓嶅惈绋庝环")) else None,


                    "fob_price_incl": float(row.get("FOB鍚◣浠?, 0)) if pd.notna(row.get("FOB鍚◣浠?)) else None,


                    "fob_price_excl": float(row.get("FOB涓嶅惈绋庝环", 0)) if pd.notna(row.get("FOB涓嶅惈绋庝环")) else None,


                    "freight": float(row.get("杩愯垂", 0)) if pd.notna(row.get("杩愯垂")) else None,


                    "packing_fee": float(row.get("鍖呰璐?, 0)) if pd.notna(row.get("鍖呰璐?)) else None,


                    "purchase_channel": str(row.get("閲囪喘娓犻亾", "")),


                    "carton_length_cm": float(row.get("绾哥闀?cm)", 0)) if pd.notna(row.get("绾哥闀?cm)")) else None,


                    "carton_width_cm": float(row.get("绾哥瀹?cm)", 0)) if pd.notna(row.get("绾哥瀹?cm)")) else None,


                    "carton_height_cm": float(row.get("绾哥楂?cm)", 0)) if pd.notna(row.get("绾哥楂?cm)")) else None,


                    "carton_volume_cbm": float(row.get("绾哥浣撶Н(CBM)", 0)) if pd.notna(row.get("绾哥浣撶Н(CBM)")) else None,


                    "carton_weight_kg": float(row.get("绾哥閲嶉噺(KG)", 0)) if pd.notna(row.get("绾哥閲嶉噺(KG)")) else None,


                    "pieces_per_carton": int(row.get("姣忕鏁伴噺", 0)) if pd.notna(row.get("姣忕鏁伴噺")) else None,


                    "unit": str(row.get("鍗曚綅", "浠?)),


                    "moq": int(row.get("鏈€灏忚捣璁㈤噺", 0)) if pd.notna(row.get("鏈€灏忚捣璁㈤噺")) else None,


                }


                products_data.append(product)


            


            print(f"DEBUG - Prepared {len(products_data)} products for import")


            print(f"DEBUG - First product data: {products_data[0] if products_data else 'None'}")


            


            result = self.api_client.import_products(products_data)


            print(f"DEBUG - Import API result: {result}")


            


            if result.get("success"):


                QMessageBox.information(self, "鎴愬姛", f"鎴愬姛瀵煎叆 {result.get('count', 0)} 涓骇鍝?)


            else:


                QMessageBox.warning(self, "瀵煎叆缁撴灉", f"瀵煎叆瀹屾垚锛岄儴鍒嗗け璐? {result.get('message', '')}")


            


            self.load_products()


            


        except Exception as e:


            print(f"DEBUG - Import failed: {str(e)}")


            QMessageBox.warning(self, "閿欒", f"瀵煎叆澶辫触: {str(e)}")





    def add_customer(self):


        dialog = CustomerDialog(self.api_client)


        if dialog.exec() == QDialog.DialogCode.Accepted:


            cache_manager.delete(CACHE_KEYS['CUSTOMERS'])


            self.load_customers()





    def edit_customer(self, customer):


        dialog = CustomerDialog(self.api_client, customer)


        if dialog.exec() == QDialog.DialogCode.Accepted:


            cache_manager.delete(CACHE_KEYS['CUSTOMERS'])


            self.load_customers()





    def import_suppliers(self):


        file_path, _ = QFileDialog.getOpenFileName(


            self, "閫夋嫨瀵煎叆鏂囦欢", "", 


            "Excel鏂囦欢 (*.xlsx *.xls);;CSV鏂囦欢 (*.csv)"


        )


        


        if not file_path:


            return


            


        try:


            if file_path.endswith('.csv'):


                df = pd.read_csv(file_path, encoding='utf-8')


            else:


                df = pd.read_excel(file_path)


            


            supplier_list = []


            province_codes = self.PROVINCE_CODES


            city_data = self.CITY_DATA


            


            for _, row in df.iterrows():


                supplier_data = {}


                


                for col in df.columns:


                    col_lower = str(col).strip().lower()


                    value = row[col]


                    


                    if pd.isna(value):


                        value = None


                    else:


                        value = str(value).strip()


                    


                    if '渚涘簲鍟? in col_lower or '鍚嶇О' in col_lower:


                        supplier_data['supplier_name'] = value


                    elif '鐪佷唤' in col_lower:


                        supplier_data['province'] = value


                    elif '鍩庡競' in col_lower or '甯? in col_lower:


                        supplier_data['city'] = value


                    elif '鑱旂郴浜? in col_lower:


                        supplier_data['contact_person'] = value


                    elif '鐢佃瘽' in col_lower or '鎵嬫満' in col_lower:


                        supplier_data['phone'] = value


                    elif '閭' in col_lower or 'email' in col_lower:


                        supplier_data['email'] = value


                    elif '鍦板潃' in col_lower:


                        supplier_data['address'] = value


                


                if 'supplier_name' in supplier_data and supplier_data['supplier_name']:


                    province = supplier_data.get('province')


                    city = supplier_data.get('city')


                    


                    if province and city:


                        province_code = province_codes.get(province)


                        if province_code:


                            cities = city_data.get(province_code, {})


                            city_code = cities.get(city)


                            if city_code:


                                supplier_data['city_code'] = province_code + city_code


                    


                    supplier_list.append(supplier_data)


            


            if not supplier_list:


                QMessageBox.warning(self, "璀﹀憡", "鏈壘鍒版湁鏁堢殑渚涘簲鍟嗘暟鎹?)


                return


                


            progress = QProgressDialog("姝ｅ湪瀵煎叆渚涘簲鍟?..", "鍙栨秷", 0, len(supplier_list), self)


            progress.setWindowModality(2)


            progress.show()


            


            def import_task():


                try:


                    result = self.api_client.post("/suppliers/batch", {"suppliers": supplier_list})


                    return result


                except Exception as e:


                    return {"error": str(e)}


            


            result = import_task()


            


            progress.close()


            


            if "error" in result:


                QMessageBox.critical(self, "瀵煎叆澶辫触", f"瀵煎叆杩囩▼涓彂鐢熼敊璇? {result['error']}")


            else:


                success = result.get("success", 0)


                failed = result.get("failed", 0)


                msg = f"瀵煎叆瀹屾垚锛乗n鎴愬姛: {success} 鏉n澶辫触: {failed} 鏉?


                if failed > 0:


                    failed_items = result.get("failed_items", [])


                    for item in failed_items[:5]:


                        msg += f"\n- {item['supplier_name']}: {item['error']}"


                    if len(failed_items) > 5:


                        msg += f"\n... 杩樻湁 {len(failed_items) - 5} 鏉″け璐ヨ褰?


                QMessageBox.information(self, "瀵煎叆瀹屾垚", msg)


                self.load_suppliers()


                


        except Exception as e:


            QMessageBox.critical(self, "瀵煎叆澶辫触", f"璇诲彇鏂囦欢鏃跺彂鐢熼敊璇? {str(e)}")





    def add_supplier(self):


        dialog = SupplierDialog(self.api_client)


        if dialog.exec() == QDialog.DialogCode.Accepted:


            cache_manager.delete(CACHE_KEYS['SUPPLIERS'])


            self.load_suppliers()





    def on_supplier_double_click(self, index):


        row = index.row()


        supplier_id = self.suppliers_table.item(row, 1).text()


        try:


            supplier = self.api_client.get_supplier_detail(int(supplier_id))


            self.edit_supplier(supplier)


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"鍔犺浇渚涘簲鍟嗕俊鎭け璐? {str(e)}")





    def edit_supplier(self, supplier):


        dialog = SupplierDialog(self.api_client, supplier)


        if dialog.exec() == QDialog.DialogCode.Accepted:


            cache_manager.delete(CACHE_KEYS['SUPPLIERS'])


            self.load_suppliers()





    def add_pi(self):


        dialog = PIDialog(self.api_client, self.dept_id)


        if dialog.exec() == QDialog.DialogCode.Accepted:


            self.load_pi_orders_async()





    def edit_pi(self, pi):


        """缂栬緫PI璁㈠崟"""


        try:


            # 鑾峰彇瀹屾暣PI璇︽儏


            pi_id = pi.get('id')


            if not pi_id:


                QMessageBox.warning(self, "閿欒", "鏃犳硶鑾峰彇PI ID")


                return


            pi_detail = self.api_client.get_pi_detail(pi_id)


            dialog = PIDialog(self.api_client, self.dept_id, pi_detail)


            if dialog.exec() == QDialog.DialogCode.Accepted:


                self.load_pi_orders_async()


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"鎵撳紑PI澶辫触: {str(e)}")





    def get_selected_pi_ids(self):


        """鑾峰彇閫変腑鐨凱I ID鍒楄〃"""


        ids = []


        for row in range(self.pi_table.rowCount()):


            item = self.pi_table.item(row, 0)


            if item and item.checkState() == Qt.CheckState.Checked:


                pi_id = item.data(Qt.ItemDataRole.UserRole)


                if pi_id:


                    ids.append(pi_id)


        return ids





    def batch_delete_pi(self):


        """鎵归噺鍒犻櫎PI璁㈠崟"""


        selected_ids = self.get_selected_pi_ids()


        if not selected_ids:


            QMessageBox.warning(self, "鎻愮ず", "璇峰厛閫夋嫨瑕佸垹闄ょ殑PI璁㈠崟")


            return


        


        reply = QMessageBox.question(


            self, "纭鍒犻櫎", 


            f"纭畾瑕佸垹闄ら€変腑鐨?{len(selected_ids)} 涓狿I璁㈠崟鍚楋紵",


            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No


        )


        if reply != QMessageBox.StandardButton.Yes:


            return


        


        try:


            result = self.api_client.batch_delete_pi(selected_ids)


            deleted = result.get('deleted', 0)


            errors = result.get('errors', [])


            if errors:


                QMessageBox.warning(self, "閮ㄥ垎鍒犻櫎澶辫触", f"鎴愬姛鍒犻櫎 {deleted} 涓猏n澶辫触: {len(errors)} 涓猏n{errors}")


            else:


                QMessageBox.information(self, "鎴愬姛", f"宸插垹闄?{deleted} 涓狿I璁㈠崟")


            self.load_pi_orders_async()


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"鍒犻櫎澶辫触: {str(e)}")


            self.load_pi_orders_async()





    def batch_export_pi(self):


        """鎵归噺瀵煎嚭PI璁㈠崟"""


        selected_ids = self.get_selected_pi_ids()


        if not selected_ids:


            QMessageBox.warning(self, "鎻愮ず", "璇峰厛閫夋嫨瑕佸鍑虹殑PI璁㈠崟")


            return


        


        try:


            pi_data_list = []


            for pi_id in selected_ids:


                pi_detail = self.api_client.get_pi_detail(pi_id)


                pi_data_list.append(pi_detail)


            


            if not pi_data_list:


                QMessageBox.information(self, "鎻愮ず", "娌℃湁鍙鍑虹殑鏁版嵁")


                return


            


            # 鏋勫缓瀵煎嚭鏁版嵁


            export_rows = []


            for pi in pi_data_list:


                for item in pi.get('items', []):


                    export_rows.append({


                        'PI鍙?: pi.get('pi_no', ''),


                        '瀹㈡埛ID': pi.get('customer_id', ''),


                        '鎬婚噾棰?: pi.get('total_amount', 0),


                        '甯佺': pi.get('currency', 'USD'),


                        '鐘舵€?: ['鑽夌', '宸茬‘璁?, '宸插彂璐?, '宸插畬鎴?][pi.get('status', 1) - 1] if pi.get('status', 1) <= 4 else '',


                        '鍒涘缓鏃堕棿': pi.get('created_at', ''),


                        '浜у搧缂栧彿': item.get('product_id', ''),


                        'OE鍙?: item.get('oe_number', ''),


                        '瀹㈡埛缂栧彿': item.get('customer_code', ''),


                        '浜у搧鎻忚堪': item.get('detail_desc', ''),


                        '鏁伴噺': item.get('quantity', 0),


                        '鍗曚环': item.get('unit_price', 0),


                        '鎬讳环': item.get('total_price', 0),


                        '澶囨敞': item.get('remark', ''),


                    })


            


            df = pd.DataFrame(export_rows)


            from PySide6.QtWidgets import QFileDialog


            file_path, _ = QFileDialog.getSaveFileName(


                self, "淇濆瓨PI璁㈠崟", 


                f"PI璁㈠崟瀵煎嚭_{len(selected_ids)}涓?xlsx",


                "Excel Files (*.xlsx)"


            )


            if file_path:


                df.to_excel(file_path, index=False, engine='openpyxl')


                QMessageBox.information(self, "鎴愬姛", f"宸插鍑?{len(export_rows)} 鏉℃槑缁嗗埌:\n{file_path}")


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"瀵煎嚭澶辫触: {str(e)}")





    def export_pi(self, pi):


        """瀵煎嚭PI涓篍xcel"""


        try:


            import tempfile


            content = self.api_client.export_pi_excel(pi.get('id'))


            # 淇濆瓨鍒颁复鏃舵枃浠?            filename = f"PI_{pi.get('pi_no', pi.get('id'))}.xlsx"


            filepath = tempfile.gettempdir() + "\\" + filename


            with open(filepath, 'wb') as f:


                f.write(content)


            # 鎵撳紑鏂囦欢


            import os


            os.startfile(filepath)


            QMessageBox.information(self, "鎴愬姛", f"PI宸插鍑? {filename}")


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"瀵煎嚭澶辫触: {str(e)}")





    def complete_pi(self, pi):


        """灏哖I鏍囪涓哄凡瀹屾垚"""


        reply = QMessageBox.question(


            self, "纭瀹屾垚",


            f"纭畾瑕佸皢PI鍗?{pi.get('pi_no', '')} 鏍囪涓哄凡瀹屾垚鍚楋紵\n瀹屾垚鍚庡皢涓嶅彲缂栬緫銆?,


            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,


            QMessageBox.StandardButton.No


        )


        if reply == QMessageBox.StandardButton.Yes:


            try:


                self.api_client.update_pi_status(pi.get('id'), 4)


                cache_manager.delete(CACHE_KEYS['PI_LIST'])


                self.load_pi_orders_async()


                QMessageBox.information(self, "鎴愬姛", f"PI鍗?{pi.get('pi_no', '')} 宸叉爣璁颁负瀹屾垚")


            except Exception as e:


                QMessageBox.warning(self, "閿欒", f"鎿嶄綔澶辫触: {str(e)}")





    def add_purchase(self):


        dialog = PurchaseDialog(self.api_client, self.dept_id)


        if dialog.exec() == QDialog.DialogCode.Accepted:


            self.load_purchase_orders()





    def edit_purchase(self, purchase):


        dialog = PurchaseDialog(self.api_client, self.dept_id, purchase)


        if dialog.exec() == QDialog.DialogCode.Accepted:


            self.load_purchase_orders_async()





    def confirm_purchase_order(self, purchase):


        """纭閲囪喘鍗?""


        try:


            self.api_client.confirm_purchase(purchase.get('id'))


            QMessageBox.information(self, "鎴愬姛", "閲囪喘鍗曞凡纭")


            self.load_purchase_orders_async()


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"纭澶辫触: {str(e)}")





    def inbound_purchase_order(self, purchase):


        """閲囪喘鍗曞叆搴?""


        try:


            self.api_client.inbound_purchase_order(purchase.get('id'))


            QMessageBox.information(self, "鎴愬姛", "閲囪喘鍗曞凡鍏ュ簱")


            self.load_purchase_orders_async()


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"鍏ュ簱澶辫触: {str(e)}")





    def add_shipment(self):


        dialog = ShipmentDialog(self.api_client, self.dept_id)


        if dialog.exec() == QDialog.DialogCode.Accepted:


            self.load_shipments()





    def edit_shipment(self, shipment):


        dialog = ShipmentDialog(self.api_client, self.dept_id, shipment)


        if dialog.exec() == QDialog.DialogCode.Accepted:


            self.load_shipments_async()





    def confirm_shipment_order(self, shipment):


        """纭鍑鸿揣"""


        try:


            self.api_client.confirm_shipment(shipment.get('id'))


            QMessageBox.information(self, "鎴愬姛", "鍑鸿揣宸茬‘璁?)


            self.load_shipments_async()


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"纭澶辫触: {str(e)}")





    def add_customer_payment(self):


        dialog = CustomerPaymentDialog(self.api_client, self.dept_id)


        if dialog.exec() == QDialog.DialogCode.Accepted:


            self.load_customer_payments()





    def edit_customer_payment(self, payment):


        dialog = CustomerPaymentDialog(self.api_client, self.dept_id, payment)


        if dialog.exec() == QDialog.DialogCode.Accepted:


            self.load_customer_payments()





    def add_supplier_payment(self):


        dialog = SupplierPaymentDialog(self.api_client, self.dept_id)


        if dialog.exec() == QDialog.DialogCode.Accepted:


            self.load_supplier_payments()





    def edit_supplier_payment(self, payment):


        dialog = SupplierPaymentDialog(self.api_client, self.dept_id, payment)


        if dialog.exec() == QDialog.DialogCode.Accepted:


            self.load_supplier_payments()





    def add_inventory(self):


        dialog = InventoryDialog(self.api_client, self.dept_id)


        if dialog.exec() == QDialog.DialogCode.Accepted:


            self.load_inventories()





    def edit_inventory(self, inventory):


        dialog = InventoryDialog(self.api_client, self.dept_id, inventory)


        if dialog.exec() == QDialog.DialogCode.Accepted:


            self.load_inventories()








class QuoteDialog(QDialog):


    """鎶ヤ环鍗曞璇濇"""


    def __init__(self, parent, api_client, dept_id, quote=None):


        super().__init__(parent)


        self.api_client = api_client


        self.dept_id = dept_id


        self.quote = quote


        self.is_edit = quote is not None


        self.customers = []


        self.products = []


        self.items = []


        self.init_ui()


        QTimer.singleShot(0, self.load_data)





    def load_data(self):


        try:


            # 鍏堝姞杞藉鎴峰垪琛?            self.customers = self.api_client.get_customers()


            self.customer_combo.clear()


            self.customer_combo.addItem("", "")


            for c in self.customers:


                self.customer_combo.addItem(f"{c.get('customer_code')} - {c.get('customer_name')}", c.get('id'))


            


            if self.quote:


                # 鍥炲～瀹㈡埛


                customer_id = self.quote.get('customer_id')


                idx = self.customer_combo.findData(customer_id)


                if idx >= 0:


                    self.customer_combo.setCurrentIndex(idx)


                


                # 鍥炲～甯佺


                currency = self.quote.get('currency', 'USD')


                idx = self.currency_combo.findText(currency)


                if idx >= 0:


                    self.currency_combo.setCurrentIndex(idx)


                


                # 鍥炲～鏈夋晥鏈?                valid_until = self.quote.get('valid_until')


                if valid_until:


                    parts = str(valid_until)[:10].split('-')


                    if len(parts) == 3:


                        from PySide6.QtCore import QDate


                        self.valid_until_input.setDate(QDate(int(parts[0]), int(parts[1]), int(parts[2])))


                


                # 鍥炲～澶囨敞


                remark = self.quote.get('remark', '')


                if remark:


                    self.remark_input.setText(str(remark))


                


                # 鍥炲～浜у搧鏄庣粏


                if 'items' in self.quote and self.quote['items']:


                    self.items = self.quote['items']


                    self.refresh_items_table()


        except Exception as e:


            print(f"鍔犺浇鏁版嵁澶辫触: {e}")





    def init_ui(self):


        self.setWindowTitle("缂栬緫鎶ヤ环鍗? if self.is_edit else "鏂板缓鎶ヤ环鍗?)


        self.setMinimumSize(900, 600)


        layout = QVBoxLayout()


        layout.setContentsMargins(20, 20, 20, 20)


        layout.setSpacing(12)





        basic_group = QGroupBox("鍩烘湰淇℃伅")


        basic_layout = QFormLayout()


        basic_layout.setSpacing(10)





        self.customer_combo = QComboBox()


        self.customer_combo.setFixedHeight(35)


        self.customer_combo.currentIndexChanged.connect(self.on_customer_changed)


        basic_layout.addRow("瀹㈡埛:", self.customer_combo)





        self.currency_combo = QComboBox()


        self.currency_combo.setFixedHeight(35)


        self.currency_combo.addItems(["USD", "EUR", "CNY", "GBP"])


        basic_layout.addRow("甯佺:", self.currency_combo)





        self.valid_until_input = QDateEdit()


        self.valid_until_input.setCalendarPopup(True)


        self.valid_until_input.setFixedHeight(35)


        self.valid_until_input.setDate(QDate.currentDate().addDays(30))


        basic_layout.addRow("鏈夋晥鏈熻嚦:", self.valid_until_input)





        self.remark_input = QLineEdit()


        self.remark_input.setFixedHeight(35)


        basic_layout.addRow("澶囨敞:", self.remark_input)





        basic_group.setLayout(basic_layout)


        layout.addWidget(basic_group)





        import_group = QGroupBox("浜у搧鏄庣粏")


        import_layout = QVBoxLayout()





        import_toolbar = QHBoxLayout()


        import_toolbar.addStretch()





        import_btn = QPushButton("瀵煎叆鍘嗗彶閲囪喘")


        import_btn.clicked.connect(self.import_customer_products)


        import_btn.setStyleSheet("""


            QPushButton { background-color: #10b981; color: white; border: none; border-radius: 4px; padding: 8px 16px; }


            QPushButton:hover { background-color: #059669; }


        """)


        import_toolbar.addWidget(import_btn)





        add_product_btn = QPushButton("+ 娣诲姞浜у搧")


        add_product_btn.clicked.connect(self.add_product)


        add_product_btn.setStyleSheet("""


            QPushButton { background-color: #3b82f6; color: white; border: none; border-radius: 4px; padding: 8px 16px; }


            QPushButton:hover { background-color: #2563eb; }


        """)


        import_toolbar.addWidget(add_product_btn)


        import_layout.addLayout(import_toolbar)





        self.items_table = QTableWidget()


        self.items_table.setColumnCount(8)


        self.items_table.setHorizontalHeaderLabels(["浜у搧缂栧彿", "OE鍙?, "瀹㈡埛缂栧彿", "浜у搧鎻忚堪", "鏁伴噺", "鍗曚环", "鎬讳环", "鎿嶄綔"])


        self.items_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)


        self.items_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)


        self.items_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)


        self.items_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)


        self.items_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)


        self.items_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)


        self.items_table.setMaximumHeight(250)


        import_layout.addWidget(self.items_table)





        summary_layout = QHBoxLayout()


        summary_layout.addStretch()


        self.total_label = QLabel("鎬婚噾棰? $0.00")


        self.total_label.setStyleSheet("font-weight: bold; font-size: 14px; color: #2563eb;")


        summary_layout.addWidget(self.total_label)


        import_layout.addLayout(summary_layout)





        import_group.setLayout(import_layout)


        layout.addWidget(import_group)





        buttons_layout = QHBoxLayout()


        buttons_layout.addStretch()


        save_btn = QPushButton("淇濆瓨")


        save_btn.clicked.connect(self.save_quote)


        save_btn.setStyleSheet("""


            QPushButton { background-color: #2563eb; color: white; border: none; border-radius: 6px; padding: 8px 24px; }


            QPushButton:hover { background-color: #1d4ed8; }


        """)


        buttons_layout.addWidget(save_btn)


        cancel_btn = QPushButton("鍙栨秷")


        cancel_btn.clicked.connect(self.reject)


        cancel_btn.setStyleSheet("""


            QPushButton { background-color: #e5e7eb; color: #374151; border: none; border-radius: 6px; padding: 8px 24px; }


            QPushButton:hover { background-color: #d1d5db; }


        """)


        buttons_layout.addWidget(cancel_btn)


        layout.addLayout(buttons_layout)


        self.setLayout(layout)





    def on_customer_changed(self):


        customer_id = self.customer_combo.currentData()


        if customer_id:


            for c in self.customers:


                if c.get('id') == customer_id:


                    currency = c.get('currency')


                    if currency:


                        idx = self.currency_combo.findText(currency)


                        if idx >= 0:


                            self.currency_combo.setCurrentIndex(idx)


                    break





    def import_customer_products(self):


        customer_id = self.customer_combo.currentData()


        if not customer_id:


            QMessageBox.warning(self, "璀﹀憡", "璇峰厛閫夋嫨瀹㈡埛")


            return


        try:


            products = self.api_client.get_customer_products(customer_id)


            if not products:


                QMessageBox.information(self, "鎻愮ず", "璇ュ鎴锋病鏈夐噰璐巻鍙茶褰?)


                return


            for p in products:


                item = {


                    'product_id': p.get('product_id'),


                    'product_code': p.get('product_code'),


                    'oe_number': p.get('oe_number'),


                    'customer_code': p.get('customer_code'),


                    'detail_desc': p.get('detail_desc'),


                    'quantity': p.get('last_quantity') or 1,


                    'unit_price': p.get('unit_price') or 0,


                    'total_price': (p.get('last_quantity') or 1) * (p.get('unit_price') or 0)


                }


                self.items.append(item)


            self.refresh_items_table()


            QMessageBox.information(self, "鎴愬姛", f"宸插鍏?{len(products)} 涓骇鍝?)


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"瀵煎叆澶辫触: {str(e)}")





    def add_product(self):


        customer = self.customer_combo.currentData()


        dialog = QuoteProductDialog(self, self.api_client, customer)


        if dialog.exec() == QDialog.DialogCode.Accepted:


            product = dialog.get_product()


            if product:


                self.items.append(product)


                self.refresh_items_table()





    def edit_item(self, index):


        if index < 0 or index >= len(self.items):


            return


        item = self.items[index]


        dialog = QuoteProductDialog(self, self.api_client, None, item)


        if dialog.exec() == QDialog.DialogCode.Accepted:


            self.items[index] = dialog.get_product()


            self.refresh_items_table()





    def delete_item(self, index):


        if index < 0 or index >= len(self.items):


            return


        reply = QMessageBox.question(self, "纭", "纭畾瑕佸垹闄ゆ浜у搧鍚楋紵")


        if reply == QMessageBox.StandardButton.Yes:


            self.items.pop(index)


            self.refresh_items_table()





    def refresh_items_table(self):


        self.items_table.setRowCount(len(self.items))


        total = 0


        for row, item in enumerate(self.items):


            self.items_table.setItem(row, 0, QTableWidgetItem(item.get('product_code', '')))


            self.items_table.setItem(row, 1, QTableWidgetItem(item.get('oe_number', '')))


            self.items_table.setItem(row, 2, QTableWidgetItem(item.get('customer_code', '')))


            self.items_table.setItem(row, 3, QTableWidgetItem(item.get('detail_desc', '')))


            self.items_table.setItem(row, 4, QTableWidgetItem(str(item.get('quantity', 0))))


            self.items_table.setItem(row, 5, QTableWidgetItem(f"${item.get('unit_price', 0):.2f}"))


            total_price = item.get('quantity', 0) * item.get('unit_price', 0)


            total += total_price


            self.items_table.setItem(row, 6, QTableWidgetItem(f"${total_price:.2f}"))


            btn_widget = QWidget()


            btn_layout = QHBoxLayout()


            btn_layout.setContentsMargins(0, 0, 0, 0)


            edit_btn = QPushButton("缂栬緫")


            edit_btn.setFixedWidth(40)


            edit_btn.clicked.connect(lambda _, r=row: self.edit_item(r))


            btn_layout.addWidget(edit_btn)


            del_btn = QPushButton("鍒犻櫎")


            del_btn.setFixedWidth(40)


            del_btn.setStyleSheet("color: #ef4444;")


            del_btn.clicked.connect(lambda _, r=row: self.delete_item(r))


            btn_layout.addWidget(del_btn)


            btn_widget.setLayout(btn_layout)


            self.items_table.setCellWidget(row, 7, btn_widget)


        self.total_label.setText(f"鎬婚噾棰? ${total:,.2f}")





    def save_quote(self):


        customer_id = self.customer_combo.currentData()


        if not customer_id:


            QMessageBox.warning(self, "璀﹀憡", "璇烽€夋嫨瀹㈡埛")


            return


        if not self.items:


            QMessageBox.warning(self, "璀﹀憡", "璇疯嚦灏戞坊鍔犱竴涓骇鍝?)


            return


        quote_data = {


            "dept_id": self.dept_id,


            "customer_id": customer_id,


            "currency": self.currency_combo.currentText(),


            "valid_until": self.valid_until_input.date().toString("yyyy-MM-dd"),


            "remark": self.remark_input.text().strip(),


            "items": [{"product_id": item.get('product_id'), "oe_number": item.get('oe_number'),


                "customer_code": item.get('customer_code'), "detail_desc": item.get('detail_desc'),


                "quantity": item.get('quantity', 0), "unit_price": item.get('unit_price', 0), "remark": ""}


                for item in self.items]


        }


        try:


            if self.is_edit:


                self.api_client.update_quote(self.quote.get('id'), quote_data)


            else:


                self.api_client.create_quote(quote_data)


            QMessageBox.information(self, "鎴愬姛", "鎶ヤ环鍗曞凡淇濆瓨")


            self.accept()


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"淇濆瓨澶辫触: {str(e)}")








class QuoteProductDialog(QDialog):


    """鎶ヤ环浜у搧瀵硅瘽妗?""


    def __init__(self, parent, api_client, customer, item=None):


        super().__init__(parent)


        self.api_client = api_client


        self.customer = customer


        self.item = item or {}


        self.products = []


        self.init_ui()


        self.load_products()





    def load_products(self):


        try:


            self.products = self.api_client.get_products()


            self.product_combo.clear()


            self.product_combo.addItem("", None)


            for p in self.products:


                self.product_combo.addItem(f"{p.get('product_code')} - {p.get('description', '')[:20]}", p)


            if self.item.get('product_id'):


                idx = self.product_combo.findData(self.item.get('product_id'))


                if idx >= 0:


                    self.product_combo.setCurrentIndex(idx)


        except Exception as e:


            print(f"鍔犺浇浜у搧澶辫触: {e}")





    def init_ui(self):


        self.setWindowTitle("缂栬緫浜у搧" if self.item else "娣诲姞浜у搧")


        self.setFixedSize(500, 400)


        layout = QVBoxLayout()


        layout.setContentsMargins(20, 20, 20, 20)


        layout.setSpacing(12)


        form_layout = QFormLayout()


        form_layout.setSpacing(10)


        self.product_combo = QComboBox()


        self.product_combo.setFixedHeight(35)


        self.product_combo.currentIndexChanged.connect(self.on_product_selected)


        form_layout.addRow("浜у搧:", self.product_combo)


        self.oe_number_input = QLineEdit()


        self.oe_number_input.setFixedHeight(35)


        form_layout.addRow("OE鍙?", self.oe_number_input)


        self.customer_code_input = QLineEdit()


        self.customer_code_input.setFixedHeight(35)


        if self.customer:


            self.customer_code_input.setText(self.customer.get('customer_code', ''))


        form_layout.addRow("瀹㈡埛缂栧彿:", self.customer_code_input)


        self.detail_desc_input = QLineEdit()


        self.detail_desc_input.setFixedHeight(35)


        form_layout.addRow("浜у搧鎻忚堪:", self.detail_desc_input)


        self.quantity_input = QLineEdit()


        self.quantity_input.setFixedHeight(35)


        self.quantity_input.setText(str(self.item.get('quantity', 1)))


        form_layout.addRow("鏁伴噺:", self.quantity_input)


        self.unit_price_input = QLineEdit()


        self.unit_price_input.setFixedHeight(35)


        self.unit_price_input.setText(str(self.item.get('unit_price', 0)))


        form_layout.addRow("鍗曚环:", self.unit_price_input)


        layout.addLayout(form_layout)


        buttons_layout = QHBoxLayout()


        buttons_layout.addStretch()


        ok_btn = QPushButton("纭畾")


        ok_btn.clicked.connect(self.validate_and_accept)


        ok_btn.setStyleSheet("""


            QPushButton { background-color: #2563eb; color: white; border: none; border-radius: 6px; padding: 8px 24px; }


            QPushButton:hover { background-color: #1d4ed8; }


        """)


        buttons_layout.addWidget(ok_btn)


        cancel_btn = QPushButton("鍙栨秷")


        cancel_btn.clicked.connect(self.reject)


        cancel_btn.setStyleSheet("""


            QPushButton { background-color: #e5e7eb; color: #374151; border: none; border-radius: 6px; padding: 8px 24px; }


            QPushButton:hover { background-color: #d1d5db; }


        """)


        buttons_layout.addWidget(cancel_btn)


        layout.addLayout(buttons_layout)


        self.setLayout(layout)





    def on_product_selected(self):


        product = self.product_combo.currentData()


        if product:


            if not self.oe_number_input.text():


                self.oe_number_input.setText(product.get('oe_number', ''))


            if not self.detail_desc_input.text():


                self.detail_desc_input.setText(product.get('description', ''))


            if self.customer and product.get('id'):


                try:


                    price_info = self.api_client.get_latest_price(self.customer.get('id'), product.get('id'))


                    if price_info and price_info.get('unit_price'):


                        self.unit_price_input.setText(str(price_info['unit_price']))


                except Exception:


                    pass





    def validate_and_accept(self):


        try:


            qty = float(self.quantity_input.text())


            price = float(self.unit_price_input.text())


            if qty <= 0:


                QMessageBox.warning(self, "璀﹀憡", "鏁伴噺蹇呴』澶т簬0")


                return


            if price < 0:


                QMessageBox.warning(self, "璀﹀憡", "鍗曚环涓嶈兘涓鸿礋")


                return


            self.accept()


        except ValueError:


            QMessageBox.warning(self, "璀﹀憡", "鏁伴噺鍜屽崟浠峰繀椤绘槸鏁板瓧")


            return





    def get_product(self):


        product = self.product_combo.currentData()


        return {


            'product_id': product.get('id') if product else None,


            'product_code': product.get('product_code', '') if product else '',


            'oe_number': self.oe_number_input.text().strip(),


            'customer_code': self.customer_code_input.text().strip(),


            'detail_desc': self.detail_desc_input.text().strip(),


            'quantity': float(self.quantity_input.text() or 0),


            'unit_price': float(self.unit_price_input.text() or 0),


            'total_price': float(self.quantity_input.text() or 0) * float(self.unit_price_input.text() or 0)


        }








class PIDialog(QDialog):


    def __init__(self, api_client, dept_id, pi=None, readonly=False):


        super().__init__()


        self.api_client = api_client


        self.dept_id = dept_id


        self.pi = pi


        self.is_edit = pi is not None


        self.readonly = readonly


        self.customers = []


        self.products = []


        self.items = []


        self.init_ui()


        QTimer.singleShot(0, self.load_data)


    


    def load_data(self):


        """鍔犺浇鏁版嵁锛堝鎴峰悓姝ュ姞杞戒繚璇佸洖濉紝浜у搧寮傛鍔犺浇锛?""


        # 瀹㈡埛鍒楄〃闇€瑕佸悓姝ュ姞杞斤紝鍥犱负鍚庨潰瑕佸洖濉€変腑椤?        try:


            cached_customers = cache_manager.get(CACHE_KEYS['CUSTOMERS'], max_age=300)


            if cached_customers is not None:


                self.customers = cached_customers


            else:


                self.customers = self.api_client.get_customers()


                cache_manager.set(CACHE_KEYS['CUSTOMERS'], self.customers)


            


            self.customer_combo.clear()


            self.customer_combo.addItem("", "")


            for c in self.customers:


                self.customer_combo.addItem(f"{c.get('customer_code')} - {c.get('customer_name')}", c.get('id'))


            


            if self.is_edit and not self.readonly:


                self.customer_combo.setEnabled(True)


                self.currency_combo.setEnabled(True)


        except Exception as e:


            print(f"鍔犺浇瀹㈡埛澶辫触: {e}")


        


        # 浜у搧鍒楄〃寮傛鍔犺浇锛堜笉闇€瑕佸洖濉級


        self.load_products()


        


        # 缂栬緫妯″紡锛氬洖濉幇鏈夋暟鎹紙鍦ㄥ鎴峰姞杞藉畬鎴愬悗鎵ц锛?        if self.is_edit and self.pi:


            self._fill_existing_data()


    


    def _fill_existing_data(self):


        """鍥炲～缂栬緫妯″紡涓嬬殑鐜版湁PI鏁版嵁"""


        # 鍥炲～瀹㈡埛


        customer_id = self.pi.get('customer_id')


        if customer_id:


            idx = self.customer_combo.findData(customer_id)


            if idx >= 0:


                self.customer_combo.setCurrentIndex(idx)


        


        # 鍥炲～甯佺


        currency = self.pi.get('currency', 'USD')


        idx = self.currency_combo.findText(currency)


        if idx >= 0:


            self.currency_combo.setCurrentIndex(idx)


        


        # 鍥炲～浜у搧鏄庣粏


        items = self.pi.get('items', [])


        if items:


            self.items = []


            for item in items:


                self.items.append({


                    "product_id": item.get('product_id'),


                    "product_code": item.get('product_code', ''),


                    "oe_number": item.get('oe_number', ''),


                    "quantity": item.get('quantity', 1),


                    "unit_price": item.get('unit_price', 0),


                    "customer_code": item.get('customer_code', ''),


                    "detail_desc": item.get('detail_desc', ''),


                    "remark": item.get('remark', '')


                })


            self.update_items_table()





    def load_customers(self):


        try:


            self.customers = self.api_client.get_customers()


            self.customer_combo.clear()


            self.customer_combo.addItem("", "")


            for c in self.customers:


                self.customer_combo.addItem(f"{c.get('customer_code')} - {c.get('customer_name')}", c.get('id'))


            


            # 缂栬緫妯″紡涓嬪惎鐢ㄥ鎴峰拰甯佺閫夋嫨


            if self.is_edit and not self.readonly:


                self.customer_combo.setEnabled(True)


                self.currency_combo.setEnabled(True)


        except Exception as e:


            print(f"鍔犺浇瀹㈡埛澶辫触: {e}")





    def load_products(self):


        """寮傛鍔犺浇浜у搧鍒楄〃"""


        # 鍏堟樉绀哄姞杞戒腑


        self.product_combo.blockSignals(True)


        self.product_combo.clear()


        self.product_combo.addItem("鍔犺浇涓?..", None)


        self.product_combo.blockSignals(False)


        


        def fetch():


            try:


                # 灏濊瘯浠庣紦瀛樺姞杞?                products = cache_manager.get(CACHE_KEYS['PRODUCTS'], max_age=300)


                if products is None:


                    products = self.api_client.get_products()


                    cache_manager.set(CACHE_KEYS['PRODUCTS'], products)


                


                self.products = products


                self.all_products = products.copy() if products else []


                


                # 鐢≦Timer鍒囨崲鍒颁富绾跨▼鏇存柊UI


                QTimer.singleShot(0, lambda: self.update_product_combo(self.all_products))


            except Exception as e:


                print(f"鍔犺浇浜у搧澶辫触: {e}")


                QTimer.singleShot(0, lambda: self.product_combo.setItemText(0, "鍔犺浇澶辫触"))


        


        # 浣跨敤鍏ㄥ眬绾跨▼姹犲紓姝ュ姞杞?        _global_thread_pool.submit(fetch)


    


    def update_product_combo(self, products):


        """鏇存柊浜у搧涓嬫媺妗嗭紙鍒嗘壒鍔犺浇閬垮厤鍗￠】锛?""


        self.product_combo.blockSignals(True)


        self.product_combo.clear()


        self.product_combo.addItem("", None)


        


        # 闄愬埗鏄剧ず鏁伴噺锛岄伩鍏嶈繃澶氫骇鍝佸鑷村崱椤?        max_display = 100


        display_products = products[:max_display] if len(products) > max_display else products


        


        for p in display_products:


            product_code = p.get('product_code', '')


            oe_number = p.get('oe_number', '')


            # 鎴柇闀挎枃鏈?            oe_display = oe_number[:15] + "..." if len(oe_number) > 15 else oe_number


            self.product_combo.addItem(f"{product_code} - {oe_display}", p)


        


        if len(products) > max_display:


            self.product_combo.addItem(f"...杩樻湁 {len(products) - max_display} 涓骇鍝侊紝璇蜂娇鐢ㄦ悳绱?, None)


        


        self.product_combo.blockSignals(False)


    


    def filter_products(self, text):


        """鏍规嵁鎼滅储鍏抽敭璇嶈繃婊や骇鍝?""


        if not text:


            self.update_product_combo(self.all_products)


            return


        text = text.lower()


        filtered = [


            p for p in self.all_products 


            if text in str(p.get('product_code', '')).lower() 


            or text in str(p.get('oe_number', '')).lower()


        ]


        self.update_product_combo(filtered)


    


    def on_product_selected(self, index):


        """浜у搧閫夋嫨鍙樺寲鏃舵洿鏂板浘鐗囬瑙堝拰鏄剧ず渚涘簲鍟嗘柟妗?""


        product = self.product_combo.currentData()


        if product:


            # 鏇存柊鍥剧墖棰勮


            image_url = product.get('default_image_url') or product.get('image_url')


            if image_url:


                try:


                    import urllib.request


                    image_data = urllib.request.urlopen(image_url).read()


                    image = QImage.fromData(image_data)


                    pixmap = QPixmap.fromImage(image).scaled(46, 46, Qt.KeepAspectRatio, Qt.SmoothTransformation)


                    self.product_image_preview.setPixmap(pixmap)


                except Exception:


                    self.product_image_preview.setText("鉂?)


            else:


                self.product_image_preview.setText("鍥?)


            


            # 鍔犺浇渚涘簲鍟嗘柟妗?            self.load_product_schemes(product.get('id'))


        else:


            self.product_image_preview.setText("鍥?)


            self.scheme_row.setEnabled(False)


    


    def load_product_schemes(self, product_id):


        """鍔犺浇浜у搧鐨勪緵搴斿晢鏂规锛堜粠PrdProductSupplier琛ㄨ鍙栵級"""


        try:


            schemes = self.api_client.get_product_schemes(product_id)


            self.scheme_combo.clear()


            self.scheme_combo.addItem("-- 閫夋嫨渚涘簲鍟嗘柟妗?--", None)


            if schemes:


                for s in schemes:


                    price = s.get('purchase_price', 0) or 0


                    customer = s.get('customer_name', '閫氱敤') or '閫氱敤'


                    label = (


                        f"銆恵s.get('supplier_name', '渚涘簲鍟?)}銆?


                        f" ({customer})"


                        f" 浠锋牸:{price:.2f}"


                    )


                    self.scheme_combo.addItem(label, s)


                self.scheme_row.setEnabled(True)


            else:


                self.scheme_row.setEnabled(False)


        except Exception as e:


            print(f"鍔犺浇渚涘簲鍟嗘柟妗堝け璐? {e}")


            self.scheme_row.setEnabled(False)


    


    def on_scheme_selected(self, index):


        """渚涘簲鍟嗘柟妗堥€夋嫨鍙樺寲"""


        pass  # 鍙互鍦ㄨ繖閲屾坊鍔犻瑙堝姛鑳?    


    def apply_scheme(self):


        """浣跨敤閫変腑鐨勪緵搴斿晢鏂规濉厖鍗曚环"""


        scheme = self.scheme_combo.currentData()


        if not scheme:


            QMessageBox.warning(self, "鎻愮ず", "璇峰厛閫夋嫨涓€涓緵搴斿晢鏂规")


            return


        


        # 浣跨敤purchase_price浣滀负鍗曚环


        price = scheme.get('purchase_price', 0) or 0


        if price:


            self.unit_price_input.setText(f"{price:.2f}")


            QMessageBox.information(self, "鎴愬姛", f"宸蹭娇鐢ㄣ€恵scheme.get('supplier_name', '渚涘簲鍟?)}銆戠殑鏂规\n閲囪喘浠? {price:.2f} 宸插～鍏?)


        else:


            QMessageBox.warning(self, "鎻愮ず", "璇ユ柟妗堝皻鏈缃环鏍硷紝璇峰湪浜у搧绠＄悊涓缃?)


    


    def create_new_scheme(self):


        """涓哄綋鍓嶄骇鍝佹柊寤轰緵搴斿晢鏂规锛堢洿鎺ュ脊鍑篠upplierSchemeDialog锛?""


        product = self.product_combo.currentData()


        if not product:


            QMessageBox.warning(self, "璀﹀憡", "璇峰厛閫夋嫨涓€涓骇鍝?)


            return


        


        try:


            # 鍔犺浇渚涘簲鍟嗗拰瀹㈡埛鍒楄〃


            suppliers = self.api_client.get_suppliers()


            customers = self.api_client.get_customers()


            


            # 鐩存帴寮瑰嚭渚涘簲鍟嗘柟妗堢紪杈戝脊绐?            dialog = SupplierSchemeDialog(self.api_client, suppliers, customers, parent=self)


            if dialog.exec() == QDialog.DialogCode.Accepted:


                scheme_data = dialog.get_scheme_data()


                if scheme_data:


                    # 閫氳繃API鍒涘缓渚涘簲鍟嗘柟妗?                    scheme_data['product_id'] = product.get('id')


                    self.api_client.create_product_scheme(product.get('id'), scheme_data)


                    QMessageBox.information(self, "鎴愬姛", "渚涘簲鍟嗘柟妗堝凡鍒涘缓")


                    # 鍒锋柊鏂规鍒楄〃


                    self.load_product_schemes(product.get('id'))


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"鍒涘缓渚涘簲鍟嗘柟妗堝け璐? {str(e)}")





    def init_ui(self):


        self.setWindowTitle("鏌ョ湅PI鍗? if self.readonly else ("缂栬緫PI鍗? if self.is_edit else "鏂板缓PI鍗?))


        self.setMinimumSize(850, 750)


        self.resize(850, 750)





        layout = QVBoxLayout()


        layout.setContentsMargins(15, 15, 15, 15)


        layout.setSpacing(10)





        # ===== 涓婇儴鍒嗭細鍩烘湰淇℃伅 =====


        basic_group = QGroupBox("鍩烘湰淇℃伅")


        basic_layout = QFormLayout()


        basic_layout.setSpacing(10)





        self.customer_combo = QComboBox()


        self.customer_combo.setFixedHeight(32)


        if self.readonly:


            self.customer_combo.setEnabled(False)


        basic_layout.addRow("瀹㈡埛:", self.customer_combo)





        self.currency_combo = QComboBox()


        self.currency_combo.setFixedHeight(32)


        self.currency_combo.addItems(["USD", "CNY", "EUR"])


        if self.readonly:


            self.currency_combo.setEnabled(False)


        basic_layout.addRow("璐у竵:", self.currency_combo)


        basic_group.setLayout(basic_layout)


        layout.addWidget(basic_group)





        # ===== 涓儴鍒嗭細娣诲姞浜у搧鍖哄煙 =====


        add_product_group = QGroupBox("娣诲姞浜у搧")


        add_product_layout = QVBoxLayout()


        add_product_layout.setSpacing(8)





        # 浜у搧鎼滅储鍜岄€夋嫨琛?        product_row = QHBoxLayout()


        product_row.addWidget(QLabel("鎼滅储浜у搧:"))


        


        self.product_search = QLineEdit()


        self.product_search.setPlaceholderText("杈撳叆浜у搧鍙?OE鍙?..")


        self.product_search.setFixedWidth(150)


        self.product_search.textChanged.connect(self.filter_products)


        product_row.addWidget(self.product_search)


        


        self.product_combo = QComboBox()


        self.product_combo.setMinimumWidth(180)


        self.product_combo.currentIndexChanged.connect(self.on_product_selected)


        product_row.addWidget(self.product_combo)


        


        self.product_image_preview = QLabel()


        self.product_image_preview.setFixedSize(50, 50)


        self.product_image_preview.setStyleSheet("border: 1px solid #d1d5db; border-radius: 4px; background-color: #f9fafb;")


        self.product_image_preview.setAlignment(Qt.AlignCenter)


        self.product_image_preview.setText("鍥?)


        product_row.addWidget(self.product_image_preview)


        


        product_row.addWidget(QLabel("鏁伴噺:"))


        self.quantity_input = QLineEdit()


        self.quantity_input.setPlaceholderText("0")


        self.quantity_input.setFixedWidth(80)


        product_row.addWidget(self.quantity_input)


        


        product_row.addWidget(QLabel("鍗曚环:"))


        self.unit_price_input = QLineEdit()


        self.unit_price_input.setPlaceholderText("0.00")


        self.unit_price_input.setFixedWidth(80)


        product_row.addWidget(self.unit_price_input)


        


        if not self.readonly:


            add_item_btn = QPushButton("+ 娣诲姞")


            add_item_btn.setStyleSheet("background-color: #2563eb; color: white; border: none; border-radius: 4px; padding: 6px 16px;")


            add_item_btn.clicked.connect(self.add_item)


            product_row.addWidget(add_item_btn)


        


        product_row.addStretch()


        add_product_layout.addLayout(product_row)





        # 渚涘簲鍟嗘柟妗堥€夋嫨琛岋紙閫夋嫨浜у搧鍚庢樉绀猴級


        self.scheme_row = QHBoxLayout()


        self.scheme_row.addWidget(QLabel("渚涘簲鍟嗘柟妗?"))


        self.scheme_combo = QComboBox()


        self.scheme_combo.setMinimumWidth(200)


        self.scheme_combo.currentIndexChanged.connect(self.on_scheme_selected)


        self.scheme_row.addWidget(self.scheme_combo)


        


        use_scheme_btn = QPushButton("浣跨敤鏂规")


        use_scheme_btn.setStyleSheet("background-color: #059669; color: white; border: none; border-radius: 4px; padding: 6px 12px;")


        use_scheme_btn.clicked.connect(self.apply_scheme)


        self.scheme_row.addWidget(use_scheme_btn)


        


        new_scheme_btn = QPushButton("+ 鏂板缓鏂规")


        new_scheme_btn.setStyleSheet("background-color: #f59e0b; color: white; border: none; border-radius: 4px; padding: 6px 12px;")


        new_scheme_btn.clicked.connect(self.create_new_scheme)


        self.scheme_row.addWidget(new_scheme_btn)


        


        self.scheme_row.addStretch()


        add_product_layout.addLayout(self.scheme_row)


        


        self.scheme_row_widget = None  # 鐢ㄤ簬鎺у埗鏄剧ず


        add_product_group.setLayout(add_product_layout)


        layout.addWidget(add_product_group)





        # ===== 涓嬮儴鍒嗭細浜у搧鏄庣粏鍒楄〃 =====


        items_group = QGroupBox("浜у搧鏄庣粏鍒楄〃")


        items_layout = QVBoxLayout()





        self.items_table = QTableWidget()


        self.items_table.setColumnCount(6)


        self.items_table.setHorizontalHeaderLabels(["浜у搧缂栧彿", "OE鍙?, "鏁伴噺", "鍗曚环", "鎬讳环", "鎿嶄綔"])


        self.items_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)


        self.items_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)


        self.items_table.setMinimumHeight(200)


        if self.readonly:


            self.items_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)


        items_layout.addWidget(self.items_table)


        items_group.setLayout(items_layout)


        layout.addWidget(items_group)





        # 搴曢儴鎸夐挳


        buttons_layout = QHBoxLayout()


        buttons_layout.addStretch()





        if not self.readonly:


            save_btn = QPushButton("淇濆瓨")


            save_btn.setFixedWidth(100)


            save_btn.clicked.connect(self.save_pi)


            save_btn.setStyleSheet("""


                QPushButton { background-color: #2563eb; color: white; border: none; border-radius: 6px; padding: 8px 24px; }


                QPushButton:hover { background-color: #1d4ed8; }


            """)


            buttons_layout.addWidget(save_btn)





        close_btn = QPushButton("鍏抽棴" if self.readonly else "鍙栨秷")


        close_btn.setFixedWidth(100)


        close_btn.clicked.connect(self.accept if self.readonly else self.reject)


        close_btn.setStyleSheet("""


            QPushButton { background-color: #e5e7eb; color: #374151; border: none; border-radius: 6px; padding: 8px 24px; }


            QPushButton:hover { background-color: #d1d5db; }


        """)


        buttons_layout.addWidget(close_btn)





        layout.addLayout(buttons_layout)


        self.setLayout(layout)





    def add_item(self):


        product = self.product_combo.currentData()


        quantity = self.quantity_input.text().strip()


        unit_price = self.unit_price_input.text().strip()





        if not product or not quantity or not unit_price:


            QMessageBox.warning(self, "璀﹀憡", "璇峰～鍐欏畬鏁翠俊鎭?)


            return





        try:


            quantity = int(quantity)


            unit_price = float(unit_price)


        except ValueError:


            QMessageBox.warning(self, "璀﹀憡", "鏁伴噺鍜屽崟浠峰繀椤绘槸鏁板瓧")


            return





        self.items.append({


            "product_id": product.get('id'),


            "product_code": product.get('product_code'),


            "oe_number": product.get('oe_number'),


            "quantity": quantity,


            "unit_price": unit_price,


            "customer_code": "",


            "detail_desc": "",


            "remark": ""


        })





        self.update_items_table()





    def update_items_table(self):


        self.items_table.setRowCount(len(self.items))


        for row, item in enumerate(self.items):


            quantity = item.get('quantity', 0)


            unit_price = item.get('unit_price', 0)


            total_price = quantity * unit_price


            


            self.items_table.setItem(row, 0, QTableWidgetItem(item.get('product_code', '')))


            self.items_table.setItem(row, 1, QTableWidgetItem(item.get('oe_number', '')))


            self.items_table.setItem(row, 2, QTableWidgetItem(str(quantity)))


            self.items_table.setItem(row, 3, QTableWidgetItem(f"{unit_price:.2f}"))


            self.items_table.setItem(row, 4, QTableWidgetItem(f"{total_price:.2f}"))





            delete_btn = QPushButton("鍒犻櫎")


            delete_btn.setFixedWidth(50)


            delete_btn.setStyleSheet("background-color: #dc2626; color: white; border: none; border-radius: 4px; padding: 2px;")


            delete_btn.clicked.connect(lambda _, r=row: self.remove_item(r))


            self.items_table.setCellWidget(row, 5, delete_btn)





    def remove_item(self, row):


        del self.items[row]


        self.update_items_table()





    def save_pi(self):


        """寮傛淇濆瓨PI鍗?""


        customer_id = self.customer_combo.currentData()


        currency = self.currency_combo.currentText()





        if not customer_id:


            QMessageBox.warning(self, "璀﹀憡", "璇烽€夋嫨瀹㈡埛")


            return





        if not self.items:


            QMessageBox.warning(self, "璀﹀憡", "璇锋坊鍔犱骇鍝佹槑缁?)


            return





        # 绂佺敤淇濆瓨鎸夐挳锛岄槻姝㈤噸澶嶆彁浜?        self.setEnabled(False)


        


        pi_data = {


            "dept_id": self.dept_id,


            "customer_id": customer_id,


            "currency": currency,


            "items": self.items,


            "payment_stages": []


        }





        def do_save():


            try:


                if self.is_edit:


                    result = self.api_client.update_pi(self.pi.get('id'), pi_data)


                else:


                    result = self.api_client.create_pi(pi_data)


                


                # 娓呴櫎PI鍒楄〃缂撳瓨


                cache_manager.delete(CACHE_KEYS['PI_LIST'])


                


                # 鍦ㄤ富绾跨▼鏇存柊UI


                from PySide6.QtCore import QMetaObject, Qt


                QMetaObject.invokeMethod(self, "_on_save_success",


                                        Qt.ConnectionType.QueuedConnection)


            except Exception as e:


                self._save_error_msg = str(e)


                from PySide6.QtCore import QMetaObject, Qt


                QMetaObject.invokeMethod(self, "_on_save_error",


                                        Qt.ConnectionType.QueuedConnection)


        


        # 浣跨敤鍏ㄥ眬绾跨▼姹犲紓姝ヤ繚瀛?        _global_thread_pool.submit(do_save)


    


    def _on_save_success(self):


        """淇濆瓨鎴愬姛鍥炶皟"""


        self.setEnabled(True)


        QMessageBox.information(self, "鎴愬姛", "PI鍗曞凡淇濆瓨")


        self.accept()


    


    def _on_save_error(self):


        """淇濆瓨澶辫触鍥炶皟"""


        self.setEnabled(True)


        error_msg = getattr(self, '_save_error_msg', '鏈煡閿欒')


        QMessageBox.warning(self, "閿欒", f"淇濆瓨澶辫触: {error_msg}")








class PurchaseDialog(QDialog):


    def __init__(self, api_client, dept_id, purchase=None):


        super().__init__()


        self.api_client = api_client


        self.dept_id = dept_id


        self.purchase = purchase


        self.is_edit = purchase is not None


        self.pi_orders = []


        self.suppliers = []


        self.products = []


        self.items = []


        self.init_ui()


        QTimer.singleShot(0, self.load_data)


    


    def load_data(self):


        self.load_pi_orders()


        self.load_suppliers()


        self.load_products()





    def load_pi_orders(self):


        try:


            self.pi_orders = self.api_client.get_pi_orders()


            self.pi_combo.clear()


            self.pi_combo.addItem("", "")


            for pi in self.pi_orders:


                self.pi_combo.addItem(f"{pi.get('pi_no')} - {pi.get('total_amount')} {pi.get('currency')}", pi)


        except Exception as e:


            print(f"鍔犺浇PI璁㈠崟澶辫触: {e}")





    def load_suppliers(self):


        try:


            self.suppliers = self.api_client.get_suppliers()


            self.supplier_combo.clear()


            self.supplier_combo.addItem("", "")


            for s in self.suppliers:


                self.supplier_combo.addItem(f"{s.get('supplier_code')} - {s.get('supplier_name')}", s.get('id'))


        except Exception as e:


            print(f"鍔犺浇渚涘簲鍟嗗け璐? {e}")





    def load_products(self):


        try:


            self.products = self.api_client.get_products()


            self.product_combo.clear()


            self.product_combo.addItem("", "")


            for p in self.products:


                self.product_combo.addItem(f"{p.get('product_code')} - {p.get('oe_number')}", p)


        except Exception as e:


            print(f"鍔犺浇浜у搧澶辫触: {e}")





    def init_ui(self):


        self.setWindowTitle("缂栬緫閲囪喘鍗? if self.is_edit else "鏂板缓閲囪喘鍗?)


        self.setFixedSize(700, 600)





        layout = QVBoxLayout()


        layout.setContentsMargins(20, 20, 20, 20)


        layout.setSpacing(12)





        form_layout = QFormLayout()


        form_layout.setSpacing(10)





        self.pi_combo = QComboBox()


        self.pi_combo.setFixedHeight(35)


        form_layout.addRow("鍏宠仈PI鍗?", self.pi_combo)





        self.supplier_combo = QComboBox()


        self.supplier_combo.setFixedHeight(35)


        form_layout.addRow("渚涘簲鍟?", self.supplier_combo)





        self.currency_combo = QComboBox()


        self.currency_combo.setFixedHeight(35)


        self.currency_combo.addItems(["CNY", "USD", "EUR"])


        form_layout.addRow("璐у竵:", self.currency_combo)





        layout.addLayout(form_layout)





        items_group = QGroupBox("閲囪喘鏄庣粏")


        items_layout = QVBoxLayout()





        toolbar = QHBoxLayout()


        self.product_combo = QComboBox()


        self.product_combo.setFixedWidth(150)


        toolbar.addWidget(self.product_combo)





        self.factory_code_input = QLineEdit()


        self.factory_code_input.setPlaceholderText("宸ュ巶缂栧彿")


        self.factory_code_input.setFixedWidth(100)


        toolbar.addWidget(self.factory_code_input)





        self.color_input = QLineEdit()


        self.color_input.setPlaceholderText("棰滆壊")


        self.color_input.setFixedWidth(80)


        toolbar.addWidget(self.color_input)





        self.quantity_input = QLineEdit()


        self.quantity_input.setPlaceholderText("鏁伴噺")


        self.quantity_input.setFixedWidth(80)


        toolbar.addWidget(self.quantity_input)





        self.unit_price_input = QLineEdit()


        self.unit_price_input.setPlaceholderText("鍗曚环")


        self.unit_price_input.setFixedWidth(80)


        toolbar.addWidget(self.unit_price_input)





        add_item_btn = QPushButton("+ 娣诲姞")


        add_item_btn.clicked.connect(self.add_item)


        toolbar.addWidget(add_item_btn)





        items_layout.addLayout(toolbar)





        self.items_table = QTableWidget()


        self.items_table.setColumnCount(10)


        self.items_table.setHorizontalHeaderLabels(["浜у搧缂栧彿", "OE鍙?, "宸ュ巶缂栧彿", "棰滆壊", "鏁伴噺", "鍗曚环", "鍑哄巶浠?, "鍑哄巶鍚◣浠?, "FOB浠?, "鎿嶄綔"])


        self.items_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)


        self.items_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Fixed)


        self.items_table.setColumnWidth(1, 120)


        self.items_table.setFixedHeight(200)


        items_layout.addWidget(self.items_table)





        items_group.setLayout(items_layout)


        layout.addWidget(items_group)





        buttons_layout = QHBoxLayout()


        buttons_layout.addStretch()





        save_btn = QPushButton("淇濆瓨")


        save_btn.clicked.connect(self.save_purchase)


        save_btn.setStyleSheet("""


            QPushButton {


                background-color: #2563eb;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px 24px;


            }


            QPushButton:hover { background-color: #1d4ed8; }


        """)


        buttons_layout.addWidget(save_btn)





        cancel_btn = QPushButton("鍙栨秷")


        cancel_btn.clicked.connect(self.reject)


        cancel_btn.setStyleSheet("""


            QPushButton {


                background-color: #e5e7eb;


                color: #374151;


                border: none;


                border-radius: 6px;


                padding: 8px 24px;


            }


            QPushButton:hover { background-color: #d1d5db; }


        """)


        buttons_layout.addWidget(cancel_btn)





        layout.addLayout(buttons_layout)





        self.setLayout(layout)





    def add_item(self):


        product = self.product_combo.currentData()


        factory_code = self.factory_code_input.text().strip()


        color = self.color_input.text().strip()


        quantity = self.quantity_input.text().strip()


        unit_price = self.unit_price_input.text().strip()





        if not product or not quantity or not unit_price:


            QMessageBox.warning(self, "璀﹀憡", "璇峰～鍐欏畬鏁翠俊鎭?)


            return





        try:


            quantity = int(quantity)


            unit_price = float(unit_price)


        except ValueError:


            QMessageBox.warning(self, "璀﹀憡", "鏁伴噺鍜屽崟浠峰繀椤绘槸鏁板瓧")


            return





        self.items.append({


            "product_id": product.get('id'),


            "product_code": product.get('product_code'),


            "oe_number": product.get('oe_number'),


            "factory_code": factory_code,


            "color": color,


            "detail_requirement": '',


            "quantity": quantity,


            "unit_price": unit_price,


            "price_ex_factory": product.get('exw_price_excl') or product.get('price_ex_factory'),


            "price_ex_factory_tax": product.get('exw_price_incl') or product.get('price_ex_factory_tax'),


            "price_fob": product.get('fob_price_excl') or product.get('price_fob'),


            "price_fob_tax": product.get('fob_price_incl') or product.get('price_fob_tax'),


            "remark": ""


        })





        self.factory_code_input.clear()


        self.color_input.clear()


        self.quantity_input.clear()


        self.unit_price_input.clear()





        self.update_items_table()





    def update_items_table(self):


        self.items_table.setRowCount(len(self.items))


        for row, item in enumerate(self.items):


            self.items_table.setItem(row, 0, QTableWidgetItem(item.get('product_code', '')))


            self.items_table.setItem(row, 1, QTableWidgetItem(item.get('oe_number', '')))


            self.items_table.setItem(row, 2, QTableWidgetItem(item.get('factory_code', '')))


            self.items_table.setItem(row, 3, QTableWidgetItem(item.get('color', '')))


            self.items_table.setItem(row, 4, QTableWidgetItem(str(item.get('quantity', ''))))


            self.items_table.setItem(row, 5, QTableWidgetItem(str(item.get('unit_price', ''))))


            self.items_table.setItem(row, 6, QTableWidgetItem(str(item.get('price_ex_factory', ''))))


            self.items_table.setItem(row, 7, QTableWidgetItem(str(item.get('price_ex_factory_tax', ''))))


            self.items_table.setItem(row, 8, QTableWidgetItem(str(item.get('price_fob', ''))))





            delete_btn = QPushButton("鍒犻櫎")


            delete_btn.setFixedWidth(50)


            delete_btn.setStyleSheet("background-color: #dc2626; color: white; border: none; border-radius: 4px; padding: 2px;")


            delete_btn.clicked.connect(lambda _, r=row: self.remove_item(r))


            self.items_table.setCellWidget(row, 9, delete_btn)





    def remove_item(self, row):


        del self.items[row]


        self.update_items_table()





    def save_purchase(self):


        pi = self.pi_combo.currentData()


        supplier_id = self.supplier_combo.currentData()


        currency = self.currency_combo.currentText()





        if not pi:


            QMessageBox.warning(self, "璀﹀憡", "璇烽€夋嫨鍏宠仈鐨凱I鍗?)


            return





        if not supplier_id:


            QMessageBox.warning(self, "璀﹀憡", "璇烽€夋嫨渚涘簲鍟?)


            return





        if not self.items:


            QMessageBox.warning(self, "璀﹀憡", "璇锋坊鍔犻噰璐槑缁?)


            return





        purchase_data = {


            "dept_id": self.dept_id,


            "pi_id": pi.get('id'),


            "supplier_id": supplier_id,


            "currency": currency,


            "items": self.items


        }





        try:


            if self.is_edit:


                self.api_client.update_purchase(self.purchase.get('id'), purchase_data)


            else:


                self.api_client.create_purchase(purchase_data)


            QMessageBox.information(self, "鎴愬姛", "閲囪喘鍗曞凡淇濆瓨")


            self.accept()


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"淇濆瓨澶辫触: {str(e)}")








class ShipmentDialog(QDialog):


    """鍑鸿揣瀵硅瘽妗?- 鏀寔澶氶樁娈电鐞?""


    def __init__(self, api_client, dept_id=None, shipment=None):


        super().__init__()


        self.api_client = api_client


        self.dept_id = dept_id or "S"


        self.shipment = shipment


        self.is_edit = shipment is not None


        self.pi_orders = []


        self.stages = []  # 鍑鸿揣闃舵鍒楄〃


        self.init_ui()


        QTimer.singleShot(0, self.load_data)





    def load_data(self):


        """鍔犺浇PI璁㈠崟鏁版嵁"""


        try:


            self.pi_orders = self.api_client.get_pi_orders()


            self.pi_combo.clear()


            self.pi_combo.addItem("", "")


            for pi in self.pi_orders:


                # 鏄剧ずPI鍙枫€佸鎴峰悕銆侀噾棰?                customer_name = pi.get('customer_name', '') or ''


                display_text = f"{pi.get('pi_no')} - {customer_name} - ${pi.get('total_amount', 0)}"


                self.pi_combo.addItem(display_text, pi)


            


            # 缂栬緫妯″紡锛氬洖濉暟鎹?            if self.shipment:


                idx = self.pi_combo.findData(self.shipment.get('pi_id'))


                if idx >= 0:


                    self.pi_combo.setCurrentIndex(idx)


                    self.pi_combo.setEnabled(False)  # 缂栬緫鏃朵笉鑳戒慨鏀筆I


                # 鍔犺浇宸叉湁鐨剆tages锛堝鏋滄湁锛?                if 'stages' in self.shipment and self.shipment['stages']:


                    self.stages = self.shipment['stages']


                    self.refresh_stages_table()


        except Exception as e:


            print(f"鍔犺浇PI璁㈠崟澶辫触: {e}")





    def init_ui(self):


        """鍒濆鍖朥I"""


        self.setWindowTitle("缂栬緫鍑鸿揣" if self.is_edit else "鏂板缓鍑鸿揣")


        self.setMinimumSize(800, 600)


        layout = QVBoxLayout()


        layout.setContentsMargins(20, 20, 20, 20)


        layout.setSpacing(12)





        # 鍩烘湰淇℃伅鍖哄煙


        basic_group = QGroupBox("鍩烘湰淇℃伅")


        basic_layout = QFormLayout()


        basic_layout.setSpacing(10)





        self.pi_combo = QComboBox()


        self.pi_combo.setFixedHeight(35)


        basic_layout.addRow("PI鍗?", self.pi_combo)





        basic_group.setLayout(basic_layout)


        layout.addWidget(basic_group)





        # 鍑鸿揣闃舵绠＄悊鍖哄煙


        stages_group = QGroupBox("鍑鸿揣闃舵绠＄悊")


        stages_layout = QVBoxLayout()





        # 闃舵鍒楄〃琛ㄦ牸


        self.stages_table = QTableWidget()


        self.stages_table.setColumnCount(6)


        self.stages_table.setHorizontalHeaderLabels(["闃舵鍚嶇О", "鍑鸿揣鏃ユ湡", "鏌滃彿", "鎻愬崟鍙?, "鏁伴噺", "鎿嶄綔"])


        self.stages_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)


        self.stages_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)


        self.stages_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)


        self.stages_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)


        self.stages_table.setMaximumHeight(250)


        stages_layout.addWidget(self.stages_table)





        # 娣诲姞闃舵鎸夐挳


        add_stage_layout = QHBoxLayout()


        add_stage_layout.addStretch()





        add_stage_btn = QPushButton("+ 娣诲姞鍑鸿揣闃舵")


        add_stage_btn.clicked.connect(self.add_stage)


        add_stage_btn.setStyleSheet("""


            QPushButton {


                background-color: #3b82f6;


                color: white;


                border: none;


                border-radius: 4px;


                padding: 8px 16px;


            }


            QPushButton:hover { background-color: #2563eb; }


        """)


        add_stage_layout.addWidget(add_stage_btn)


        stages_layout.addLayout(add_stage_layout)





        # 姹囨€讳俊鎭?        summary_layout = QHBoxLayout()


        summary_layout.addStretch()


        self.total_stages_label = QLabel("闃舵鏁? 0")


        self.total_stages_label.setStyleSheet("font-weight: bold; color: #374151;")


        summary_layout.addWidget(self.total_stages_label)


        summary_layout.addSpacing(20)


        self.total_qty_label = QLabel("鎬绘暟閲? 0")


        self.total_qty_label.setStyleSheet("font-weight: bold; color: #10b981;")


        summary_layout.addWidget(self.total_qty_label)


        stages_layout.addLayout(summary_layout)





        stages_group.setLayout(stages_layout)


        layout.addWidget(stages_group)





        # 鎸夐挳鍖哄煙


        buttons_layout = QHBoxLayout()


        buttons_layout.addStretch()





        save_btn = QPushButton("淇濆瓨")


        save_btn.clicked.connect(self.save_shipment)


        save_btn.setStyleSheet("""


            QPushButton {


                background-color: #2563eb;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px 24px;


            }


            QPushButton:hover { background-color: #1d4ed8; }


        """)


        buttons_layout.addWidget(save_btn)





        cancel_btn = QPushButton("鍙栨秷")


        cancel_btn.clicked.connect(self.reject)


        cancel_btn.setStyleSheet("""


            QPushButton {


                background-color: #e5e7eb;


                color: #374151;


                border: none;


                border-radius: 6px;


                padding: 8px 24px;


            }


            QPushButton:hover { background-color: #d1d5db; }


        """)


        buttons_layout.addWidget(cancel_btn)





        layout.addLayout(buttons_layout)


        self.setLayout(layout)





    def add_stage(self):


        """娣诲姞鍑鸿揣闃舵"""


        dialog = ShipmentStageDialog(self, len(self.stages))


        if dialog.exec() == QDialog.DialogCode.Accepted:


            stage_data = dialog.get_stage_data()


            self.stages.append(stage_data)


            self.refresh_stages_table()





    def edit_stage(self, index):


        """缂栬緫鍑鸿揣闃舵"""


        if index < 0 or index >= len(self.stages):


            return


        dialog = ShipmentStageDialog(self, index, self.stages[index])


        if dialog.exec() == QDialog.DialogCode.Accepted:


            self.stages[index] = dialog.get_stage_data()


            self.refresh_stages_table()





    def delete_stage(self, index):


        """鍒犻櫎鍑鸿揣闃舵"""


        if index < 0 or index >= len(self.stages):


            return


        reply = QMessageBox.question(self, "纭", f"纭畾瑕佸垹闄ら樁娈?'{self.stages[index].get('stage_name')}' 鍚楋紵")


        if reply == QMessageBox.StandardButton.Yes:


            self.stages.pop(index)


            self.refresh_stages_table()





    def refresh_stages_table(self):


        """鍒锋柊闃舵琛ㄦ牸"""


        self.stages_table.setRowCount(len(self.stages))


        


        total_qty = 0


        


        for row, stage in enumerate(self.stages):


            # 闃舵鍚嶇О


            self.stages_table.setItem(row, 0, QTableWidgetItem(stage.get('stage_name', '')))


            # 鍑鸿揣鏃ユ湡


            self.stages_table.setItem(row, 1, QTableWidgetItem(str(stage.get('shipment_date', ''))))


            # 鏌滃彿


            self.stages_table.setItem(row, 2, QTableWidgetItem(stage.get('container_no', '')))


            # 鎻愬崟鍙?            self.stages_table.setItem(row, 3, QTableWidgetItem(stage.get('bl_no', '')))


            # 鏁伴噺


            qty = stage.get('quantity', 0) or 0


            total_qty += float(qty)


            self.stages_table.setItem(row, 4, QTableWidgetItem(str(qty)))


            


            # 鎿嶄綔鎸夐挳


            btn_widget = QWidget()


            btn_layout = QHBoxLayout()


            btn_layout.setContentsMargins(0, 0, 0, 0)


            


            edit_btn = QPushButton("缂栬緫")


            edit_btn.setFixedWidth(50)


            edit_btn.clicked.connect(lambda _, r=row: self.edit_stage(r))


            btn_layout.addWidget(edit_btn)


            


            del_btn = QPushButton("鍒犻櫎")


            del_btn.setFixedWidth(50)


            del_btn.setStyleSheet("color: #ef4444;")


            del_btn.clicked.connect(lambda _, r=row: self.delete_stage(r))


            btn_layout.addWidget(del_btn)


            


            btn_widget.setLayout(btn_layout)


            self.stages_table.setCellWidget(row, 5, btn_widget)


        


        # 鏇存柊姹囨€?        self.total_stages_label.setText(f"闃舵鏁? {len(self.stages)}")


        self.total_qty_label.setText(f"鎬绘暟閲? {total_qty}")





    def save_shipment(self):


        """淇濆瓨鍑鸿揣璁板綍"""


        pi = self.pi_combo.currentData()


        if not pi:


            QMessageBox.warning(self, "璀﹀憡", "璇烽€夋嫨PI鍗?)


            return





        if not self.stages:


            QMessageBox.warning(self, "璀﹀憡", "璇疯嚦灏戞坊鍔犱竴涓嚭璐ч樁娈?)


            return





        # 鏋勫缓stages鏁版嵁


        stages_data = []


        for stage in self.stages:


            stages_data.append({


                'id': stage.get('id'),  # 缂栬緫鏃跺彲鑳芥湁id


                'stage_name': stage.get('stage_name'),


                'shipment_date': stage.get('shipment_date'),


                'container_no': stage.get('container_no'),


                'bl_no': stage.get('bl_no'),


                'quantity': stage.get('quantity'),


                'ci_document': stage.get('ci_document'),


                'pl_document': stage.get('pl_document'),


                'storage_location': stage.get('storage_location'),


                'remark': stage.get('remark')


            })





        shipment_data = {


            "dept_id": self.dept_id,


            "pi_id": pi.get('id'),


            "stages": stages_data,


            "items": []  # 鍑鸿揣鏄庣粏锛屾殏鏃朵负绌?        }





        try:


            if self.is_edit:


                self.api_client.update_shipment(self.shipment.get('id'), shipment_data)


            else:


                self.api_client.create_shipment(shipment_data)


            QMessageBox.information(self, "鎴愬姛", "鍑鸿揣璁板綍宸蹭繚瀛?)


            self.accept()


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"淇濆瓨澶辫触: {str(e)}")








class ShipmentStageDialog(QDialog):


    """鍑鸿揣闃舵瀵硅瘽妗?""


    def __init__(self, parent, stage_no, stage_data=None):


        super().__init__(parent)


        self.stage_no = stage_no


        self.stage_data = stage_data or {}


        self.is_edit = stage_data is not None


        self.init_ui()


        


        # 缂栬緫妯″紡鍥炲～鏁版嵁


        if self.is_edit:


            self.load_stage_data()





    def init_ui(self):


        self.setWindowTitle(f"缂栬緫鍑鸿揣闃舵" if self.is_edit else f"娣诲姞鍑鸿揣闃舵 #{self.stage_no + 1}")


        self.setFixedSize(500, 450)


        layout = QVBoxLayout()


        layout.setContentsMargins(20, 20, 20, 20)


        layout.setSpacing(12)





        form_layout = QFormLayout()


        form_layout.setSpacing(10)





        # 闃舵鍚嶇О


        self.stage_name_input = QLineEdit()


        self.stage_name_input.setFixedHeight(35)


        self.stage_name_input.setText(f"鍑鸿揣{self.stage_no + 1}")


        form_layout.addRow("闃舵鍚嶇О:", self.stage_name_input)





        # 鍑鸿揣鏃ユ湡


        self.shipment_date_input = QDateEdit()


        self.shipment_date_input.setCalendarPopup(True)


        self.shipment_date_input.setFixedHeight(35)


        self.shipment_date_input.setDate(QDate.currentDate())


        form_layout.addRow("鍑鸿揣鏃ユ湡:", self.shipment_date_input)





        # 鏌滃彿


        self.container_no_input = QLineEdit()


        self.container_no_input.setFixedHeight(35)


        self.container_no_input.setPlaceholderText("濡? MSKU1234567")


        form_layout.addRow("鏌滃彿:", self.container_no_input)





        # 鎻愬崟鍙?        self.bl_no_input = QLineEdit()


        self.bl_no_input.setFixedHeight(35)


        self.bl_no_input.setPlaceholderText("濡? BL123456789")


        form_layout.addRow("鎻愬崟鍙?", self.bl_no_input)





        # 鏁伴噺


        self.quantity_input = QLineEdit()


        self.quantity_input.setFixedHeight(35)


        self.quantity_input.setPlaceholderText("鍑鸿揣鏁伴噺")


        form_layout.addRow("鏁伴噺:", self.quantity_input)





        # 瀛樻斁浣嶇疆


        self.storage_location_input = QLineEdit()


        self.storage_location_input.setFixedHeight(35)


        self.storage_location_input.setPlaceholderText("濡? 涓婃捣娓?)


        form_layout.addRow("瀛樻斁浣嶇疆:", self.storage_location_input)





        # CI鏂囦欢璺緞


        self.ci_document_input = QLineEdit()


        self.ci_document_input.setFixedHeight(35)


        self.ci_document_input.setPlaceholderText("CI鏂囦欢璺緞鎴栫紪鍙?)


        form_layout.addRow("CI鏂囦欢:", self.ci_document_input)





        # PL鏂囦欢璺緞


        self.pl_document_input = QLineEdit()


        self.pl_document_input.setFixedHeight(35)


        self.pl_document_input.setPlaceholderText("PL鏂囦欢璺緞鎴栫紪鍙?)


        form_layout.addRow("PL鏂囦欢:", self.pl_document_input)





        # 澶囨敞


        self.remark_input = QLineEdit()


        self.remark_input.setFixedHeight(35)


        form_layout.addRow("澶囨敞:", self.remark_input)





        layout.addLayout(form_layout)





        # 鎸夐挳


        buttons_layout = QHBoxLayout()


        buttons_layout.addStretch()





        save_btn = QPushButton("纭畾")


        save_btn.clicked.connect(self.validate_and_accept)


        save_btn.setStyleSheet("""


            QPushButton {


                background-color: #2563eb;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px 24px;


            }


            QPushButton:hover { background-color: #1d4ed8; }


        """)


        buttons_layout.addWidget(save_btn)





        cancel_btn = QPushButton("鍙栨秷")


        cancel_btn.clicked.connect(self.reject)


        cancel_btn.setStyleSheet("""


            QPushButton {


                background-color: #e5e7eb;


                color: #374151;


                border: none;


                border-radius: 6px;


                padding: 8px 24px;


            }


            QPushButton:hover { background-color: #d1d5db; }


        """)


        buttons_layout.addWidget(cancel_btn)





        layout.addLayout(buttons_layout)


        self.setLayout(layout)





    def load_stage_data(self):


        """鍔犺浇闃舵鏁版嵁锛堢紪杈戞ā寮忥級"""


        self.stage_name_input.setText(self.stage_data.get('stage_name', ''))


        if self.stage_data.get('shipment_date'):


            date = QDate.fromString(str(self.stage_data['shipment_date'])[:10], "yyyy-MM-dd")


            if date.isValid():


                self.shipment_date_input.setDate(date)


        self.container_no_input.setText(self.stage_data.get('container_no', ''))


        self.bl_no_input.setText(self.stage_data.get('bl_no', ''))


        self.quantity_input.setText(str(self.stage_data.get('quantity', '')))


        self.storage_location_input.setText(self.stage_data.get('storage_location', ''))


        self.ci_document_input.setText(self.stage_data.get('ci_document', ''))


        self.pl_document_input.setText(self.stage_data.get('pl_document', ''))


        self.remark_input.setText(self.stage_data.get('remark', ''))





    def validate_and_accept(self):


        """楠岃瘉骞剁‘璁?""


        if not self.stage_name_input.text().strip():


            QMessageBox.warning(self, "璀﹀憡", "璇疯緭鍏ラ樁娈靛悕绉?)


            return


        


        try:


            qty = float(self.quantity_input.text() or 0)


            if qty <= 0:


                QMessageBox.warning(self, "璀﹀憡", "鏁伴噺蹇呴』澶т簬0")


                return


        except ValueError:


            QMessageBox.warning(self, "璀﹀憡", "鏁伴噺蹇呴』鏄暟瀛?)


            return





        self.accept()





    def get_stage_data(self):


        """鑾峰彇闃舵鏁版嵁"""


        return {


            'id': self.stage_data.get('id') if self.is_edit else None,


            'stage_name': self.stage_name_input.text().strip(),


            'shipment_date': self.shipment_date_input.date().toString("yyyy-MM-dd"),


            'container_no': self.container_no_input.text().strip(),


            'bl_no': self.bl_no_input.text().strip(),


            'quantity': float(self.quantity_input.text() or 0),


            'storage_location': self.storage_location_input.text().strip(),


            'ci_document': self.ci_document_input.text().strip(),


            'pl_document': self.pl_document_input.text().strip(),


            'remark': self.remark_input.text().strip()


        }








class CustomerPaymentDialog(QDialog):


    def __init__(self, api_client, dept_id=None, payment=None):


        super().__init__()


        self.api_client = api_client


        self.dept_id = dept_id or "S"


        self.payment = payment


        self.is_edit = payment is not None


        self.pi_orders = []


        self.init_ui()


        QTimer.singleShot(0, self.load_data)





    def load_data(self):


        try:


            self.pi_orders = self.api_client.get_pi_orders()


            self.pi_combo.clear()


            self.pi_combo.addItem("", "")


            for pi in self.pi_orders:


                self.pi_combo.addItem(f"{pi.get('pi_no')} - {pi.get('total_amount')}", pi)


            if self.payment:


                idx = self.pi_combo.findData(self.payment.get('pi_id'))


                if idx >= 0:


                    self.pi_combo.setCurrentIndex(idx)


        except Exception as e:


            print(f"鍔犺浇PI璁㈠崟澶辫触: {e}")





    def init_ui(self):


        self.setWindowTitle("缂栬緫瀹㈡埛浠樻" if self.is_edit else "鏂板缓瀹㈡埛浠樻")


        self.setFixedSize(500, 350)


        layout = QVBoxLayout()


        layout.setContentsMargins(20, 20, 20, 20)


        layout.setSpacing(12)





        form_layout = QFormLayout()


        form_layout.setSpacing(10)





        self.pi_combo = QComboBox()


        self.pi_combo.setFixedHeight(35)


        form_layout.addRow("PI鍗?", self.pi_combo)





        self.payment_date_input = QDateEdit()


        self.payment_date_input.setCalendarPopup(True)


        self.payment_date_input.setFixedHeight(35)


        self.payment_date_input.setDate(QDate.currentDate())


        form_layout.addRow("浠樻鏃ユ湡:", self.payment_date_input)





        self.amount_input = QLineEdit()


        self.amount_input.setFixedHeight(35)


        form_layout.addRow("浠樻閲戦:", self.amount_input)





        self.payment_method_combo = QComboBox()


        self.payment_method_combo.setFixedHeight(35)


        self.payment_method_combo.addItems(["閾惰杞处", "鐜伴噾", "鏀エ", "鍏朵粬"])


        form_layout.addRow("浠樻鏂瑰紡:", self.payment_method_combo)





        self.remark_input = QTextEdit()


        self.remark_input.setFixedHeight(80)


        form_layout.addRow("澶囨敞:", self.remark_input)





        layout.addLayout(form_layout)





        buttons_layout = QHBoxLayout()


        buttons_layout.addStretch()





        save_btn = QPushButton("淇濆瓨")


        save_btn.clicked.connect(self.save_payment)


        save_btn.setStyleSheet("""


            QPushButton {


                background-color: #2563eb;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px 24px;


            }


            QPushButton:hover { background-color: #1d4ed8; }


        """)


        buttons_layout.addWidget(save_btn)





        cancel_btn = QPushButton("鍙栨秷")


        cancel_btn.clicked.connect(self.reject)


        cancel_btn.setStyleSheet("""


            QPushButton {


                background-color: #e5e7eb;


                color: #374151;


                border: none;


                border-radius: 6px;


                padding: 8px 24px;


            }


            QPushButton:hover { background-color: #d1d5db; }


        """)


        buttons_layout.addWidget(cancel_btn)





        layout.addLayout(buttons_layout)


        self.setLayout(layout)





        if self.payment:


            self.payment_date_input.setDate(QDate.fromString(self.payment.get('payment_date', ''), "yyyy-MM-dd"))


            self.amount_input.setText(str(self.payment.get('actual_amount', '')))


            method_map = {"閾惰杞处": 0, "鐜伴噾": 1, "鏀エ": 2, "鍏朵粬": 3}


            self.payment_method_combo.setCurrentIndex(method_map.get(self.payment.get('payment_method', '閾惰杞处'), 0))


            self.remark_input.setPlainText(self.payment.get('remark', ''))





    def save_payment(self):


        pi = self.pi_combo.currentData()


        if not pi:


            QMessageBox.warning(self, "璀﹀憡", "璇烽€夋嫨PI鍗?)


            return





        amount = self.amount_input.text().strip()


        if not amount:


            QMessageBox.warning(self, "璀﹀憡", "璇疯緭鍏ヤ粯娆鹃噾棰?)


            return





        try:


            amount = float(amount)


        except ValueError:


            QMessageBox.warning(self, "璀﹀憡", "浠樻閲戦蹇呴』鏄暟瀛?)


            return





        payment_data = {


            "dept_id": self.dept_id,


            "pi_id": pi.get('id'),


            "customer_id": pi.get('customer_id'),


            "amount": amount,


            "actual_amount": amount,


            "payment_date": self.payment_date_input.date().toString("yyyy-MM-dd"),


            "payment_method": self.payment_method_combo.currentText(),


            "remark": self.remark_input.toPlainText()


        }





        try:


            if self.is_edit:


                self.api_client.update_customer_payment(self.payment.get('id'), payment_data)


            else:


                self.api_client.create_customer_payment(payment_data)


            QMessageBox.information(self, "鎴愬姛", "浠樻璁板綍宸蹭繚瀛?)


            self.accept()


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"淇濆瓨澶辫触: {str(e)}")








class SupplierPaymentDialog(QDialog):


    """渚涘簲鍟嗕粯娆惧璇濇 - 鏀寔澶氶樁娈电鐞?""


    def __init__(self, api_client, dept_id=None, payment=None):


        super().__init__()


        self.api_client = api_client


        self.dept_id = dept_id or "S"


        self.payment = payment


        self.is_edit = payment is not None


        self.suppliers = []


        self.purchases = []


        self.stages = []  # 浠樻闃舵鍒楄〃


        self.init_ui()


        QTimer.singleShot(0, self.load_data)





    def load_data(self):


        """鍔犺浇渚涘簲鍟嗗拰閲囪喘鍗曟暟鎹?""


        try:


            self.suppliers = self.api_client.get_suppliers()


            self.supplier_combo.clear()


            self.supplier_combo.addItem("", "")


            for s in self.suppliers:


                self.supplier_combo.addItem(f"{s.get('supplier_code')} - {s.get('supplier_name')}", s.get('id'))


            


            # 缂栬緫妯″紡锛氬洖濉暟鎹?            if self.payment:


                idx = self.supplier_combo.findData(self.payment.get('supplier_id'))


                if idx >= 0:


                    self.supplier_combo.setCurrentIndex(idx)


                # 鍔犺浇宸叉湁鐨剆tages


                if 'stages' in self.payment and self.payment['stages']:


                    self.stages = self.payment['stages']


                    self.refresh_stages_table()


        except Exception as e:


            print(f"鍔犺浇渚涘簲鍟嗗け璐? {e}")





    def load_purchases(self):


        """鍔犺浇渚涘簲鍟嗙殑閲囪喘鍗?""


        supplier_id = self.supplier_combo.currentData()


        if not supplier_id:


            return


        try:


            purchases = self.api_client.get_purchases_by_supplier(supplier_id)


            self.purchase_combo.clear()


            self.purchase_combo.addItem("", "")


            for p in purchases:


                self.purchase_combo.addItem(f"PO-{p.get('id')} - {p.get('total_amount', 0)}", p)


            # 缂栬緫妯″紡锛氬洖濉噰璐崟


            if self.payment and self.payment.get('po_id'):


                for i in range(self.purchase_combo.count()):


                    data = self.purchase_combo.itemData(i)


                    if data and data.get('id') == self.payment.get('po_id'):


                        self.purchase_combo.setCurrentIndex(i)


                        break


        except Exception as e:


            print(f"鍔犺浇閲囪喘鍗曞け璐? {e}")





    def init_ui(self):


        """鍒濆鍖朥I"""


        self.setWindowTitle("缂栬緫渚涘簲鍟嗕粯娆? if self.is_edit else "鏂板缓渚涘簲鍟嗕粯娆?)


        self.setMinimumSize(700, 600)


        layout = QVBoxLayout()


        layout.setContentsMargins(20, 20, 20, 20)


        layout.setSpacing(12)





        # 鍩烘湰淇℃伅鍖哄煙


        basic_group = QGroupBox("鍩烘湰淇℃伅")


        basic_layout = QFormLayout()


        basic_layout.setSpacing(10)





        self.supplier_combo = QComboBox()


        self.supplier_combo.setFixedHeight(35)


        self.supplier_combo.currentIndexChanged.connect(self.load_purchases)


        basic_layout.addRow("渚涘簲鍟?", self.supplier_combo)





        self.purchase_combo = QComboBox()


        self.purchase_combo.setFixedHeight(35)


        basic_layout.addRow("閲囪喘鍗?", self.purchase_combo)





        self.payment_method_combo = QComboBox()


        self.payment_method_combo.setFixedHeight(35)


        self.payment_method_combo.addItems(["閾惰杞处", "鐜伴噾", "鏀エ", "鍏朵粬"])


        basic_layout.addRow("浠樻鏂瑰紡:", self.payment_method_combo)





        basic_group.setLayout(basic_layout)


        layout.addWidget(basic_group)





        # 浠樻闃舵绠＄悊鍖哄煙


        stages_group = QGroupBox("浠樻闃舵绠＄悊")


        stages_layout = QVBoxLayout()





        # 闃舵鍒楄〃琛ㄦ牸


        self.stages_table = QTableWidget()


        self.stages_table.setColumnCount(5)


        self.stages_table.setHorizontalHeaderLabels(["闃舵鍚嶇О", "搴斾粯閲戦", "宸蹭粯閲戦", "鐘舵€?, "鎿嶄綔"])


        self.stages_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)


        self.stages_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)


        self.stages_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)


        self.stages_table.setMaximumHeight(200)


        stages_layout.addWidget(self.stages_table)





        # 娣诲姞闃舵鎸夐挳鍖哄煙


        add_stage_layout = QHBoxLayout()


        add_stage_layout.addStretch()





        add_deposit_btn = QPushButton("+ 娣诲姞瀹氶噾")


        add_deposit_btn.clicked.connect(lambda: self.add_stage('deposit'))


        add_deposit_btn.setStyleSheet("""


            QPushButton {


                background-color: #f59e0b;


                color: white;


                border: none;


                border-radius: 4px;


                padding: 6px 12px;


            }


            QPushButton:hover { background-color: #d97706; }


        """)


        add_stage_layout.addWidget(add_deposit_btn)





        add_balance_btn = QPushButton("+ 娣诲姞灏炬")


        add_balance_btn.clicked.connect(lambda: self.add_stage('balance'))


        add_balance_btn.setStyleSheet("""


            QPushButton {


                background-color: #3b82f6;


                color: white;


                border: none;


                border-radius: 4px;


                padding: 6px 12px;


            }


            QPushButton:hover { background-color: #2563eb; }


        """)


        add_stage_layout.addWidget(add_balance_btn)





        stages_layout.addLayout(add_stage_layout)





        # 姹囨€讳俊鎭?        summary_layout = QHBoxLayout()


        summary_layout.addStretch()


        self.total_label = QLabel("鎬婚噾棰? 0.00")


        self.total_label.setStyleSheet("font-weight: bold; color: #374151;")


        summary_layout.addWidget(self.total_label)


        summary_layout.addSpacing(20)


        self.paid_label = QLabel("宸蹭粯: 0.00")


        self.paid_label.setStyleSheet("font-weight: bold; color: #10b981;")


        summary_layout.addWidget(self.paid_label)


        summary_layout.addSpacing(20)


        self.unpaid_label = QLabel("鏈粯: 0.00")


        self.unpaid_label.setStyleSheet("font-weight: bold; color: #ef4444;")


        summary_layout.addWidget(self.unpaid_label)


        stages_layout.addLayout(summary_layout)





        stages_group.setLayout(stages_layout)


        layout.addWidget(stages_group)





        # 鎸夐挳鍖哄煙


        buttons_layout = QHBoxLayout()


        buttons_layout.addStretch()





        save_btn = QPushButton("淇濆瓨")


        save_btn.clicked.connect(self.save_payment)


        save_btn.setStyleSheet("""


            QPushButton {


                background-color: #2563eb;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px 24px;


            }


            QPushButton:hover { background-color: #1d4ed8; }


        """)


        buttons_layout.addWidget(save_btn)





        cancel_btn = QPushButton("鍙栨秷")


        cancel_btn.clicked.connect(self.reject)


        cancel_btn.setStyleSheet("""


            QPushButton {


                background-color: #e5e7eb;


                color: #374151;


                border: none;


                border-radius: 6px;


                padding: 8px 24px;


            }


            QPushButton:hover { background-color: #d1d5db; }


        """)


        buttons_layout.addWidget(cancel_btn)





        layout.addLayout(buttons_layout)


        self.setLayout(layout)





        # 缂栬緫妯″紡锛氬洖濉粯娆炬柟寮?        if self.payment:


            method_map = {"閾惰杞处": 0, "鐜伴噾": 1, "鏀エ": 2, "鍏朵粬": 3}


            self.payment_method_combo.setCurrentIndex(method_map.get(self.payment.get('payment_method', '閾惰杞处'), 0))





    def add_stage(self, stage_type):


        """娣诲姞浠樻闃舵"""


        dialog = SupplierPaymentStageDialog(self, stage_type, len(self.stages))


        if dialog.exec() == QDialog.DialogCode.Accepted:


            stage_data = dialog.get_stage_data()


            self.stages.append(stage_data)


            self.refresh_stages_table()





    def edit_stage(self, index):


        """缂栬緫浠樻闃舵"""


        if index < 0 or index >= len(self.stages):


            return


        dialog = SupplierPaymentStageDialog(self, self.stages[index].get('stage_type', 'balance'), 


                                           index, self.stages[index])


        if dialog.exec() == QDialog.DialogCode.Accepted:


            self.stages[index] = dialog.get_stage_data()


            self.refresh_stages_table()





    def delete_stage(self, index):


        """鍒犻櫎浠樻闃舵"""


        if index < 0 or index >= len(self.stages):


            return


        reply = QMessageBox.question(self, "纭", f"纭畾瑕佸垹闄ら樁娈?'{self.stages[index].get('stage_name')}' 鍚楋紵")


        if reply == QMessageBox.StandardButton.Yes:


            self.stages.pop(index)


            self.refresh_stages_table()





    def refresh_stages_table(self):


        """鍒锋柊闃舵琛ㄦ牸"""


        self.stages_table.setRowCount(len(self.stages))


        status_map = {1: "寰呬粯", 2: "閮ㄥ垎浠?, 3: "宸蹭粯娓?}


        


        total = 0


        paid = 0


        


        for row, stage in enumerate(self.stages):


            # 闃舵鍚嶇О


            self.stages_table.setItem(row, 0, QTableWidgetItem(stage.get('stage_name', '')))


            # 搴斾粯閲戦


            amount = stage.get('amount', 0) or 0


            total += float(amount)


            self.stages_table.setItem(row, 1, QTableWidgetItem(f"{float(amount):,.2f}"))


            # 宸蹭粯閲戦


            stage_paid = stage.get('paid_amount', 0) or 0


            paid += float(stage_paid)


            self.stages_table.setItem(row, 2, QTableWidgetItem(f"{float(stage_paid):,.2f}"))


            # 鐘舵€?            status = stage.get('status', 1)


            self.stages_table.setItem(row, 3, QTableWidgetItem(status_map.get(status, "鏈煡")))


            # 鎿嶄綔鎸夐挳


            btn_widget = QWidget()


            btn_layout = QHBoxLayout()


            btn_layout.setContentsMargins(0, 0, 0, 0)


            


            edit_btn = QPushButton("缂栬緫")


            edit_btn.setFixedWidth(50)


            edit_btn.clicked.connect(lambda _, r=row: self.edit_stage(r))


            btn_layout.addWidget(edit_btn)


            


            del_btn = QPushButton("鍒犻櫎")


            del_btn.setFixedWidth(50)


            del_btn.setStyleSheet("color: #ef4444;")


            del_btn.clicked.connect(lambda _, r=row: self.delete_stage(r))


            btn_layout.addWidget(del_btn)


            


            btn_widget.setLayout(btn_layout)


            self.stages_table.setCellWidget(row, 4, btn_widget)


        


        # 鏇存柊姹囨€?        unpaid = total - paid


        self.total_label.setText(f"鎬婚噾棰? {total:,.2f}")


        self.paid_label.setText(f"宸蹭粯: {paid:,.2f}")


        self.unpaid_label.setText(f"鏈粯: {unpaid:,.2f}")





    def save_payment(self):


        """淇濆瓨浠樻璁板綍"""


        supplier_id = self.supplier_combo.currentData()


        purchase = self.purchase_combo.currentData()





        if not supplier_id:


            QMessageBox.warning(self, "璀﹀憡", "璇烽€夋嫨渚涘簲鍟?)


            return





        if not purchase:


            QMessageBox.warning(self, "璀﹀憡", "璇烽€夋嫨閲囪喘鍗?)


            return





        if not self.stages:


            QMessageBox.warning(self, "璀﹀憡", "璇疯嚦灏戞坊鍔犱竴涓粯娆鹃樁娈?)


            return





        # 鏋勫缓stages鏁版嵁


        stages_data = []


        for stage in self.stages:


            stages_data.append({


                'id': stage.get('id'),  # 缂栬緫鏃跺彲鑳芥湁id


                'stage_type': stage.get('stage_type'),


                'stage_name': stage.get('stage_name'),


                'amount': stage.get('amount'),


                'paid_amount': stage.get('paid_amount', 0),


                'status': stage.get('status', 1),


                'payment_date': stage.get('payment_date'),


                'payment_proof': stage.get('payment_proof'),


                'remark': stage.get('remark')


            })





        payment_data = {


            "dept_id": self.dept_id,


            "supplier_id": supplier_id,


            "po_id": purchase.get('id'),


            "payment_method": self.payment_method_combo.currentText(),


            "stages": stages_data,


            "remark": ""


        }





        try:


            print(f"DEBUG - 淇濆瓨渚涘簲鍟嗕粯娆? {payment_data}")


            if self.is_edit:


                result = self.api_client.update_supplier_payment(self.payment.get('id'), payment_data)


            else:


                result = self.api_client.create_supplier_payment(payment_data)


            print(f"DEBUG - 淇濆瓨鎴愬姛: {result}")


            QMessageBox.information(self, "鎴愬姛", "浠樻璁板綍宸蹭繚瀛?)


            self.accept()


        except Exception as e:


            print(f"DEBUG - 淇濆瓨澶辫触: {e}")


            import traceback


            traceback.print_exc()


            QMessageBox.warning(self, "閿欒", f"淇濆瓨澶辫触: {str(e)}")








class SupplierPaymentStageDialog(QDialog):


    """浠樻闃舵缂栬緫瀵硅瘽妗?""


    def __init__(self, parent, stage_type, index, stage_data=None):


        super().__init__(parent)


        self.stage_type = stage_type


        self.index = index


        self.stage_data = stage_data or {}


        self.is_edit = stage_data is not None


        self.init_ui()





    def init_ui(self):


        stage_type_name = "瀹氶噾" if self.stage_type == 'deposit' else f"灏炬{self.index}"


        self.setWindowTitle(f"缂栬緫{stage_type_name}" if self.is_edit else f"娣诲姞{stage_type_name}")


        self.setFixedSize(400, 350)


        


        layout = QVBoxLayout()


        layout.setContentsMargins(20, 20, 20, 20)


        layout.setSpacing(12)





        form_layout = QFormLayout()


        form_layout.setSpacing(10)





        # 闃舵鍚嶇О


        self.name_input = QLineEdit()


        self.name_input.setFixedHeight(35)


        default_name = self.stage_data.get('stage_name', '')


        if not default_name:


            default_name = "瀹氶噾" if self.stage_type == 'deposit' else f"灏炬{self.index + 1}"


        self.name_input.setText(default_name)


        form_layout.addRow("闃舵鍚嶇О:", self.name_input)





        # 搴斾粯閲戦


        self.amount_input = QLineEdit()


        self.amount_input.setFixedHeight(35)


        self.amount_input.setText(str(self.stage_data.get('amount', '')))


        form_layout.addRow("搴斾粯閲戦:", self.amount_input)





        # 宸蹭粯閲戦


        self.paid_input = QLineEdit()


        self.paid_input.setFixedHeight(35)


        self.paid_input.setText(str(self.stage_data.get('paid_amount', '0')))


        form_layout.addRow("宸蹭粯閲戦:", self.paid_input)





        # 浠樻鏃ユ湡


        self.date_input = QDateEdit()


        self.date_input.setCalendarPopup(True)


        self.date_input.setFixedHeight(35)


        payment_date = self.stage_data.get('payment_date')


        if payment_date:


            self.date_input.setDate(QDate.fromString(str(payment_date)[:10], "yyyy-MM-dd"))


        else:


            self.date_input.setDate(QDate.currentDate())


        form_layout.addRow("浠樻鏃ユ湡:", self.date_input)





        # 澶囨敞


        self.remark_input = QTextEdit()


        self.remark_input.setFixedHeight(60)


        self.remark_input.setPlainText(self.stage_data.get('remark', ''))


        form_layout.addRow("澶囨敞:", self.remark_input)





        layout.addLayout(form_layout)





        # 鎸夐挳


        buttons_layout = QHBoxLayout()


        buttons_layout.addStretch()





        save_btn = QPushButton("纭畾")


        save_btn.clicked.connect(self.accept)


        save_btn.setStyleSheet("""


            QPushButton {


                background-color: #2563eb;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px 24px;


            }


        """)


        buttons_layout.addWidget(save_btn)





        cancel_btn = QPushButton("鍙栨秷")


        cancel_btn.clicked.connect(self.reject)


        cancel_btn.setStyleSheet("""


            QPushButton {


                background-color: #e5e7eb;


                color: #374151;


                border: none;


                border-radius: 6px;


                padding: 8px 24px;


            }


        """)


        buttons_layout.addWidget(cancel_btn)





        layout.addLayout(buttons_layout)


        self.setLayout(layout)





    def get_stage_data(self):


        """鑾峰彇闃舵鏁版嵁"""


        amount = float(self.amount_input.text() or 0)


        paid = float(self.paid_input.text() or 0)


        


        # 鑷姩璁＄畻鐘舵€?        status = 1  # 寰呬粯


        if paid >= amount and amount > 0:


            status = 3  # 宸蹭粯娓?        elif paid > 0:


            status = 2  # 閮ㄥ垎浠?


        return {


            'id': self.stage_data.get('id'),  # 缂栬緫鏃朵繚鐣欏師id


            'stage_type': self.stage_type,


            'stage_name': self.name_input.text(),


            'amount': amount,


            'paid_amount': paid,


            'status': status,


            'payment_date': self.date_input.date().toString("yyyy-MM-dd"),


            'remark': self.remark_input.toPlainText()


        }








class InventoryDialog(QDialog):


    def __init__(self, api_client, dept_id=None, inventory=None, product_id=None, oe_number=None):


        super().__init__()


        self.api_client = api_client


        self.dept_id = dept_id or "S"


        self.inventory = inventory


        self.preselected_product_id = product_id  # 棰勯€夌殑浜у搧ID


        self.preselected_oe_number = oe_number    # 棰勯€夌殑OE鍙?        self.is_edit = inventory is not None


        self.products = []


        self.init_ui()


        QTimer.singleShot(0, self.load_products)





    def load_products(self):


        try:


            # 灏濊瘯浠庣紦瀛樺姞杞?            self.products = cache_manager.get(CACHE_KEYS['PRODUCTS'], max_age=300)


            if self.products is None:


                self.products = self.api_client.get_products()


                cache_manager.set(CACHE_KEYS['PRODUCTS'], self.products)


            


            self.product_combo.clear()


            self.product_combo.addItem("璇烽€夋嫨浜у搧", None)


            


            # 濡傛灉鏈夐閫夌殑OE鍙凤紝鍙樉绀哄尮閰嶇殑浜у搧


            if self.preselected_oe_number:


                for p in self.products:


                    if p.get('oe_number') == self.preselected_oe_number:


                        self.product_combo.addItem(f"{p.get('product_code')} - {p.get('oe_number')}", p)


                        # 鑷姩閫変腑


                        self.product_combo.setCurrentIndex(1)


                        break


            else:


                for p in self.products:


                    self.product_combo.addItem(f"{p.get('product_code')} - {p.get('oe_number')}", p)


            


            # 缂栬緫妯″紡锛氶€変腑褰撳墠搴撳瓨鐨勪骇鍝?            if self.inventory:


                for i in range(self.product_combo.count()):


                    data = self.product_combo.itemData(i)


                    if data and data.get('id') == self.inventory.get('product_id'):


                        self.product_combo.setCurrentIndex(i)


                        break


            # 鏂板缓妯″紡锛氭湁棰勯€変骇鍝両D


            elif self.preselected_product_id:


                for i in range(self.product_combo.count()):


                    data = self.product_combo.itemData(i)


                    if data and data.get('id') == self.preselected_product_id:


                        self.product_combo.setCurrentIndex(i)


                        break


        except Exception as e:


            print(f"鍔犺浇浜у搧澶辫触: {e}")


    


    def load_suppliers(self):


        """鍔犺浇渚涘簲鍟嗗垪琛?""


        try:


            # 灏濊瘯浠庣紦瀛樺姞杞?            suppliers = cache_manager.get(CACHE_KEYS['SUPPLIERS'], max_age=300)


            if suppliers is None:


                suppliers = self.api_client.get_suppliers()


                cache_manager.set(CACHE_KEYS['SUPPLIERS'], suppliers)


            


            self.supplier_combo.clear()


            self.supplier_combo.addItem("璇烽€夋嫨渚涘簲鍟?, None)


            for s in suppliers:


                self.supplier_combo.addItem(f"{s.get('supplier_code')} - {s.get('supplier_name')}", s)


            


            # 濡傛灉鏄紪杈戯紝璁剧疆閫変腑鐨勪緵搴斿晢


            if self.inventory and self.inventory.get('supplier_id'):


                for i in range(self.supplier_combo.count()):


                    data = self.supplier_combo.itemData(i)


                    if data and data.get('id') == self.inventory.get('supplier_id'):


                        self.supplier_combo.setCurrentIndex(i)


                        break


        except Exception as e:


            print(f"鍔犺浇渚涘簲鍟嗗け璐? {e}")


    


    def load_customers(self):


        """鍔犺浇瀹㈡埛鍒楄〃"""


        try:


            # 灏濊瘯浠庣紦瀛樺姞杞?            customers = cache_manager.get(CACHE_KEYS['CUSTOMERS'], max_age=300)


            if customers is None:


                customers = self.api_client.get_customers()


                cache_manager.set(CACHE_KEYS['CUSTOMERS'], customers)


            


            self.customer_combo.clear()


            self.customer_combo.addItem("璇烽€夋嫨瀹㈡埛", None)


            for c in customers:


                self.customer_combo.addItem(f"{c.get('customer_code')} - {c.get('customer_name')}", c)


            


            # 濡傛灉鏄紪杈戯紝璁剧疆閫変腑鐨勫鎴?            if self.inventory and self.inventory.get('customer_id'):


                for i in range(self.customer_combo.count()):


                    data = self.customer_combo.itemData(i)


                    if data and data.get('id') == self.inventory.get('customer_id'):


                        self.customer_combo.setCurrentIndex(i)


                        break


        except Exception as e:


            print(f"鍔犺浇瀹㈡埛澶辫触: {e}")





    def init_ui(self):


        self.setWindowTitle("缂栬緫搴撳瓨" if self.is_edit else "鏂板缓搴撳瓨")


        self.setFixedSize(550, 550)


        layout = QVBoxLayout()


        layout.setContentsMargins(20, 20, 20, 20)


        layout.setSpacing(12)





        form_layout = QFormLayout()


        form_layout.setSpacing(10)





        self.product_combo = QComboBox()


        self.product_combo.setFixedHeight(35)


        form_layout.addRow("浜у搧 *:", self.product_combo)





        # 渚涘簲鍟嗛€夋嫨


        self.supplier_combo = QComboBox()


        self.supplier_combo.setFixedHeight(35)


        form_layout.addRow("渚涘簲鍟?*:", self.supplier_combo)


        # 鍔犺浇渚涘簲鍟嗗垪琛?        self.load_suppliers()





        # 瀹㈡埛閫夋嫨


        self.customer_combo = QComboBox()


        self.customer_combo.setFixedHeight(35)


        form_layout.addRow("瀹㈡埛 *:", self.customer_combo)


        # 鍔犺浇瀹㈡埛鍒楄〃


        self.load_customers()





        self.stock_type_combo = QComboBox()


        self.stock_type_combo.setFixedHeight(35)


        self.stock_type_combo.addItems(["閲囪喘鍦ㄩ€?, "寰呭叆搴?, "宸插叆搴?, "鍘嗗彶搴撳瓨"])


        form_layout.addRow("搴撳瓨绫诲瀷:", self.stock_type_combo)





        self.quantity_input = QLineEdit()


        self.quantity_input.setFixedHeight(35)


        form_layout.addRow("鏁伴噺:", self.quantity_input)





        self.location_input = QLineEdit()


        self.location_input.setFixedHeight(35)


        form_layout.addRow("搴撲綅:", self.location_input)





        self.remark_input = QTextEdit()


        self.remark_input.setFixedHeight(80)


        form_layout.addRow("澶囨敞:", self.remark_input)





        layout.addLayout(form_layout)





        status_group = QGroupBox("搴撳瓨鐘舵€侀鑹?)


        status_layout = QHBoxLayout()


        status_layout.addWidget(QLabel("鈼?榛勮壊: 閲囪喘鍦ㄩ€?))


        status_layout.addWidget(QLabel("鈼?钃濊壊: 寰呭叆搴?))


        status_layout.addWidget(QLabel("鈼?缁胯壊: 宸插叆搴?))


        status_layout.addWidget(QLabel("鈼?榛戣壊: 鍘嗗彶搴撳瓨"))


        status_group.setLayout(status_layout)


        layout.addWidget(status_group)





        buttons_layout = QHBoxLayout()


        buttons_layout.addStretch()





        save_btn = QPushButton("淇濆瓨")


        save_btn.clicked.connect(self.save_inventory)


        save_btn.setStyleSheet("""


            QPushButton {


                background-color: #2563eb;


                color: white;


                border: none;


                border-radius: 6px;


                padding: 8px 24px;


            }


            QPushButton:hover { background-color: #1d4ed8; }


        """)


        buttons_layout.addWidget(save_btn)





        cancel_btn = QPushButton("鍙栨秷")


        cancel_btn.clicked.connect(self.reject)


        cancel_btn.setStyleSheet("""


            QPushButton {


                background-color: #e5e7eb;


                color: #374151;


                border: none;


                border-radius: 6px;


                padding: 8px 24px;


            }


            QPushButton:hover { background-color: #d1d5db; }


        """)


        buttons_layout.addWidget(cancel_btn)





        layout.addLayout(buttons_layout)


        self.setLayout(layout)





        if self.inventory:


            # 鍥炲～鏁伴噺锛堜娇鐢?total_quantity 鎴?quantity锛?            qty = self.inventory.get('total_quantity') or self.inventory.get('quantity', '')


            self.quantity_input.setText(str(qty))


            # 鍥炲～搴撲綅


            self.location_input.setText(self.inventory.get('current_location', '') or '')


            # 鍥炲～澶囨敞


            self.remark_input.setPlainText(self.inventory.get('remark', '') or '')


            # 鍥炲～搴撳瓨绫诲瀷


            stock_type = self.inventory.get('stock_type', 2)


            # stock_type 杞笅鎷夋绱㈠紩


            index_map = {1: 0, 2: 1, 3: 2, 4: 3}  # 1->閲囪喘鍦ㄩ€?0), 2->寰呭叆搴?1), 3->宸插叆搴?2), 4->鍘嗗彶搴撳瓨(3)


            if isinstance(stock_type, str):


                try:


                    stock_type = int(stock_type)


                except ValueError:


                    stock_type = 2


            combo_index = index_map.get(stock_type, 0)


            self.stock_type_combo.setCurrentIndex(combo_index)





    def save_inventory(self):


        product = self.product_combo.currentData()


        if not product:


            QMessageBox.warning(self, "璀﹀憡", "璇烽€夋嫨浜у搧")


            return





        quantity = self.quantity_input.text().strip()


        if not quantity:


            QMessageBox.warning(self, "璀﹀憡", "璇疯緭鍏ユ暟閲?)


            return





        try:


            # 鍏堣浆涓烘诞鐐规暟锛屽啀杞负鏁存暟锛堝鐞嗗悗绔繑鍥炵殑 123.0 鏍煎紡锛?            quantity = int(float(quantity))


        except ValueError:


            QMessageBox.warning(self, "璀﹀憡", "鏁伴噺蹇呴』鏄暣鏁?)


            return





        # 搴撳瓨绫诲瀷鏄犲皠锛氫笅鎷夋鏂囨湰 -> stock_type


        type_map = {"閲囪喘鍦ㄩ€?: 1, "寰呭叆搴?: 2, "宸插叆搴?: 3, "鍘嗗彶搴撳瓨": 4}


        stock_type_text = self.stock_type_combo.currentText()


        stock_type = type_map.get(stock_type_text, 1)





        # 鑾峰彇渚涘簲鍟嗭紙浠庝緵搴斿晢閫夋嫨涓嬫媺妗嗭級


        supplier = self.supplier_combo.currentData()


        supplier_id = supplier.get('id') if supplier else None


        


        # 鑾峰彇瀹㈡埛锛堜粠瀹㈡埛閫夋嫨涓嬫媺妗嗭級


        customer = self.customer_combo.currentData()


        if not customer:


            QMessageBox.warning(self, "璀﹀憡", "璇烽€夋嫨瀹㈡埛")


            return


        customer_id = customer.get('id')


        


        inventory_data = {


            "dept_id": self.dept_id,


            "product_id": product.get('id'),


            "supplier_id": supplier_id,


            "customer_id": customer_id,


            "quantity": quantity,


            "current_location": self.location_input.text().strip() or 'WAREHOUSE',


            "stock_type": stock_type,


            "remark": self.remark_input.toPlainText().strip()


        }





        try:


            if self.is_edit:


                self.api_client.update_inventory(self.inventory.get('id'), inventory_data)


            else:


                self.api_client.create_inventory(inventory_data)


            # 娓呴櫎搴撳瓨缂撳瓨


            cache_manager.delete(CACHE_KEYS['INVENTORY_SUMMARY'])


            print("DEBUG - 宸叉竻闄ゅ簱瀛樼紦瀛?)


            QMessageBox.information(self, "鎴愬姛", "搴撳瓨璁板綍宸蹭繚瀛?)


            self.accept()


        except Exception as e:


            QMessageBox.warning(self, "閿欒", f"淇濆瓨澶辫触: {str(e)}")








def main():


    ctypes.windll.kernel32.SetConsoleOutputCP(65001)


    


    # 鍚敤楂楧PI缂╂斁锛堜娇鐢ㄦ帹鑽愭柟寮忥級


    os.environ["QT_ENABLE_HIGHDPI_SCALING"] = "1"


    os.environ["QT_AUTO_SCREEN_SCALE_FACTOR"] = "0"


    


    app = QApplication(sys.argv)


    


    # 璁剧疆鍏ㄥ眬瀛椾綋


    font = QFont()


    font.setFamily("Microsoft YaHei")


    font.setPointSize(10)


    app.setFont(font)





    # 浣跨敤甯︾紦瀛樺姛鑳界殑API瀹㈡埛绔?    api_client = CachedApiClient()





    try:


        login = LoginWindow(api_client)


        if login.exec() != QDialog.DialogCode.Accepted:


            return





        dept_id = login.get_selected_department() or "S"


        window = MainWindow(api_client, dept_id)


        window.show()





        sys.exit(app.exec())


    except Exception as e:


        print("Error during application execution:", str(e))


        traceback.print_exc()


        # 鏄剧ず閿欒瀵硅瘽妗?        error_dialog = QDialog()


        error_dialog.setWindowTitle("鍚姩閿欒")


        error_dialog.resize(400, 200)


        layout = QVBoxLayout()


        layout.addWidget(QLabel(f"鍚姩澶辫触: {str(e)}"))


        layout.addWidget(QLabel("璇锋鏌ユ棩蹇楄幏鍙栬缁嗕俊鎭?))


        error_dialog.setLayout(layout)


        error_dialog.exec()


        sys.exit(1)








if __name__ == "__main__":


    main()


