"""Native API 实现：文件/Excel/系统通知"""
import os
import openpyxl
from PySide6.QtWidgets import QFileDialog, QMessageBox
from PySide6.QtCore import QObject


class NativeAPI:
    """PyQt 本地能力封装"""

    def __init__(self, parent: QObject = None):
        self._bridge = None

    def set_bridge(self, bridge):
        self._bridge = bridge

    def select_file(self, filter_str: str = "All Files (*)") -> str:
        """打开文件选择对话框，返回文件路径"""
        dialog = QFileDialog()
        dialog.setNameFilter(filter_str)
        dialog.setFileMode(QFileDialog.ExistingFile)
        if dialog.exec():
            files = dialog.selectedFiles()
            return files[0] if files else ""
        return ""

    def save_file(self, default_name: str = "") -> str:
        """打开保存文件对话框，返回保存路径"""
        dialog = QFileDialog()
        dialog.setAcceptMode(QFileDialog.AcceptSave)
        dialog.selectFile(default_name)
        if dialog.exec():
            files = dialog.selectedFiles()
            return files[0] if files else ""
        return ""

    def read_excel(self, file_path: str) -> list:
        """读取 Excel 文件，返回 JSON 列表"""
        if not file_path or not os.path.exists(file_path):
            return []
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    rows.append(dict(zip(headers, row)))
            return rows
        except Exception as e:
            print(f"[NativeAPI] read_excel error: {e}")
            return []

    def write_excel(self, file_path: str, data: list) -> bool:
        """将数据写入 Excel 文件"""
        if not data:
            return False
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            headers = list(data[0].keys())
            ws.append(headers)
            for row in data:
                ws.append([row.get(h, "") for h in headers])
            wb.save(file_path)
            return True
        except Exception as e:
            print(f"[NativeAPI] write_excel error: {e}")
            return False

    def show_notification(self, message: str):
        """显示系统通知"""
        QMessageBox.information(None, "通知", message)

    def get_app_version(self) -> str:
        """获取桌面壳版本"""
        return "1.0.0"

    def get_app_name(self) -> str:
        """获取应用名称"""
        return "PI Manager"
